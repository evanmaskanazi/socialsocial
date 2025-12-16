#!/usr/bin/env python
"""
Complete app.py for Social Social Platform - Phase 6013 (Version 2004)
With Flask-Migrate and SQLAlchemy 2.0 style queries
Auto-migrates on startup for seamless deployment

PJ6013 Changes (v2004):
- FRONTEND FIX: Mobile notification toggles now load correctly from backend
- ROOT CAUSE: loadNotificationSettings() only set desktop toggles (emailOnAlertToggle, etc.)
  but did NOT set mobile toggles (mobileEmailOnAlertToggle, mobileEmailOnNotificationToggle, 
  mobileDailyDiaryReminderToggle)
- SYMPTOM: On mobile, toggles would show incorrect state (usually OFF even when ON in backend)
- FIX: Added code to also set mobile toggle states after setting desktop toggles
- This is a FRONTEND-ONLY fix in indexsend3.html - no backend changes

PJ6012 Changes (v2003):
- CRITICAL FIX: Consolidated batch emails now actually sent from TRIGGER SCHEDULER
- ROOT CAUSE: The run_background_trigger_check_for_watcher() function was creating alerts
  with create_alert_with_email() which correctly skipped individual emails (PJ6009),
  but it NEVER called send_consolidated_wellness_alert_email() to send the batch email
- The batch/consolidated email was only being sent from the job queue path (when user saves diary),
  NOT from the trigger scheduler path (5-minute background check)
- FIX: Added code to collect triggered_params_by_user and call send_consolidated_wellness_alert_email()
  after alerts are created in run_background_trigger_check_for_watcher()
- Now batch emails are sent from BOTH paths:
  1. Job queue path (process_parameter_triggers_async) - when user saves diary
  2. Trigger scheduler path (run_background_trigger_check_for_watcher) - every 5 minutes

PJ6011 Changes (v2002):
- CRITICAL FIX: Stopped sending individual "You have a new alert" emails for trigger alerts
- ROOT CAUSE: When email_on_alert is turned ON, the notification settings endpoint was 
  sending individual emails for ALL unread alerts, including trigger alerts
- This bypassed the consolidated batch email logic (send_consolidated_wellness_alert_email)
- FIX: Added filter `Alert.alert_category != 'trigger'` to exclude trigger alerts
- Trigger alerts should ONLY send batch/consolidated emails, never individual ones
- Individual email format: "You have a new alert" with single item (BAD for triggers)
- Batch email format: "We noticed some concerning wellness patterns" with list (CORRECT)

PJ6010 Changes (v2001):
- MOBILE FIX: About and Support links now visible on mobile devices
- ROOT CAUSE: .nav-link CSS had display:none in @media max-width:768px
- FIX: Changed to display:inline-block with responsive font-size (14px) and padding
- This is a frontend-only fix - no backend changes required
- Paired with indexmobile.html which contains the actual CSS fix

PJ6009 Changes:
- FIX 1: Stopped sending duplicate "TheraSocial - New Alert" individual emails for trigger alerts
  - Only batch emails "TheraSocial - Wellness Alert for {username}" should be sent
  - Modified create_alert_with_email() to skip email for alert_category='trigger'
  - Batch emails are sent via send_consolidated_wellness_alert_email()
- FIX 2: Batch wellness alert emails use watcher's preferred_language (already correct)
- FIX 3: Progress insights now properly translated based on frontend language
  - Backend now reads 'lang' query parameter sent by frontend
  - Falls back to user.preferred_language if lang param not provided

PJ5002 Changes (v2000):
- Mobile Navigation Consolidation (frontend-only changes in index5002.html)
- ROOT CAUSE: Too many tabs (10) on mobile made navigation confusing and touch targets too small
- FIX 1: Consolidated into 5 primary tabs: Home, Feed, Diary, Messages, More
- FIX 2: Fixed Feed icon from house to journal/post icon
- FIX 3: Added proper Home tab for dashboard/progress page
- FIX 4: Created "More" slide-up menu containing: Profile, Circles, Social section, Alerts
- FIX 5: Alert badge shows on both "More" button and inside More menu
- No backend changes required - all fixes are in the frontend HTML/JS

PJ5001 Changes (v1900):
- CRITICAL FIX: Mobile Alerts View (frontend-only changes in index5001.html)
- ROOT CAUSE: .right-sidebar was hidden on mobile (display:none in @media max-width:768px)
- This caused ALL alerts to be invisible on iPhone and similar small screens
- FIX 1: Added dedicated Alerts tab to mobile navigation bar
- FIX 2: Created alertsView section with mobile-optimized layout
- FIX 3: loadAlerts() now populates both desktop and mobile alert lists
- FIX 4: dismissAlert() syncs removal between desktop and mobile lists
- FIX 5: Added mobile alerts badge showing unread count
- FIX 6: Added syncMobileNotificationSettings() for toggle sync
- No backend changes required - all fixes are in the frontend HTML/JS

PJ816 Changes (v1800):
- CRITICAL FIX: Trigger alert emails now sent WITHOUT requiring watcher login
- ROOT CAUSE: check_parameter_triggers() required @login_required and only ran on polling
  - Emails were only sent when the WATCHER logged in
  - Expected behavior: Emails sent when WATCHED user's data triggers alerts
- FIX 1: Added background trigger scheduler that runs every 5 minutes
  - Checks triggers for ALL watchers, not just logged-in users
  - Creates alerts and sends emails automatically
  - Works like diary reminders and message notifications
- FIX 2: New function run_background_trigger_check() processes all triggers
  - Iterates through all watchers with triggers configured
  - Uses same logic as check_parameter_triggers() but no login required
  - Comprehensive duplicate detection prevents spam
- FIX 3: Integrated into diary scheduler with 5-minute offset
  - Runs at :00, :05, :10, etc. minutes (diary runs at :00, :01, :02)
  - Staggered to avoid resource conflicts
- LOGGING: [TRIGGER SCHEDULER] logs for background trigger processing

PJ815 Changes (v1705):
- CRITICAL FIX: Reverted broken trigger deduplication from v1704
  - v1704 broke triggers by merging them upfront - lost all parameter flags
  - ROOT CAUSE: Triggers use OLD schema (parameter_name) not NEW schema (mood_alert flags)
  - When merging with bool(t.mood_alert) where t.mood_alert=None, got False
- FIX 1: Process each trigger row individually (no upfront deduplication)
  - Each trigger row processed with its actual flags/parameter_name
  - Handles BOTH old schema and new schema triggers correctly
- FIX 2: Use patterns_seen SET to deduplicate RESULTS, not inputs
  - Key = (username, param_name, start_date_iso, end_date_iso)
  - Prevents duplicate alerts from multiple trigger rows for same pattern
- FIX 3: Fixed dates array generation in OLD schema code
  - Now builds proper dates array from streak, not just end_date
  - Patterns have unique date ranges for proper duplicate detection
- FIX 4: Fixed process_parameter_triggers for email sending
  - Handles old/new schema properly
  - Sends emails when watched user saves parameters
- EXTENSIVE DEBUG LOGGING:
  - [PJ815 DEBUG] - Shows each trigger's flags and schema type
  - [PJ815 PATTERN] - Shows each pattern found with date range

PJ813 Changes (v1703):
- CRITICAL FIX: Trigger alerts now create SEPARATE alerts for each date range
  - Previously: Only 5 alerts created (one per parameter type) even with 75 patterns
  - Now: Each distinct date range gets its own alert (e.g., mood Nov 10-12, mood Nov 15-17, etc.)
- FIX 1: check_parameter_triggers() now uses date-specific duplicate detection
  - Old: Checked for any "laura's mood" alert in 24 hours
  - New: Checks for alert with EXACT date range "(Nov 10 - Nov 12)" in content
- FIX 2: process_parameter_triggers() now finds ALL matching patterns
  - Old: Found only ONE pattern per parameter (the longest/most recent)
  - New: Finds ALL consecutive streaks that meet the threshold
- FIX 3: Emails sent when watched user saves (without requiring watcher login)
  - process_parameter_triggers runs on every parameter save
  - Creates alerts and sends emails for ALL watchers with triggers
  - Works like message notifications - delivered immediately on save

PJ812 Changes (v1702):
- Increased alerts limit from 50 to 100 for full month display
- Fixed process_parameter_triggers to include privacy checks (matching check_triggers)
- Fixed trigger emails to be sent when watched user saves parameters (no login required)
- Added detailed logging to process_parameter_triggers for debugging
- Fixed date formatting in trigger alert content

PJ501 Changes:
- Updated check-blocked endpoint to return blockedBy field
- Added allow_preview parameter to get_user_profile for recommended users
- Added block check to profile endpoint returning 403 with account_not_available

PJ601 Changes:
- No backend changes needed for block/unblock toggle, Follow button width, and navigation fix
- All fixes are in the frontend HTML/JS files

PJ602 Changes:
- No backend changes needed - all fixes in frontend HTML/JS files
- Fixed Follow button width, action buttons on own profile, double 403 message

PJ701 Changes:
- Fixed email notifications for all alert types (wellness, follow, invite)
- All alert creation now uses create_alert_with_email to send emails when user has email_on_alert enabled
- This applies to: wellness alerts, new follower alerts, and invite alerts

PJ703 Changes:
- No backend changes needed - all fixes in frontend HTML/JS files
- Fixed Block button translation key, Followers tab profile viewing, Invite section follow tracking

PJ704 Changes:
- CRITICAL FIX: Fixed get_alerts filtering that was hiding follow/invite alerts
- The filter was incorrectly hiding ALL alerts with source_user_id unless you followed that user
- Now only 'trigger' and 'feed' category alerts are filtered by following status
- Follow alerts (alert_category='follow') and invite alerts now always show to the recipient
- DIAGNOSTIC: Added comprehensive logging to all email sending functions
- Logs now show exactly why emails are or aren't being sent
- Check Render logs for [ALERT EMAIL], [MESSAGE EMAIL], [SMTP ALERT], [SMTP MESSAGE] tags

PJ705 Changes:
- MIGRATED FROM SENDGRID TO RESEND.COM for email delivery
- Replaced SendGrid API library with standard Python SMTP library
- All email functions now use SMTP_SSL with Resend.com (or any SMTP provider)
- Environment variables: SMTP_SERVER, SMTP_PORT, SMTP_USERNAME, SMTP_PASSWORD, FROM_EMAIL
- Default configuration uses Resend.com with onboarding@resend.dev sender
- No domain verification required on Resend free tier (3000 emails/month)

PJ806 Changes:
- CRITICAL BUG FIX: Fixed repeated email spam from trigger alerts
- ROOT CAUSE: check_parameter_triggers() endpoint was creating duplicate alerts every 60 seconds
  because duplicate detection used inconsistent parameter names ('sleep_quality' vs 'sleep quality')
- FIX 1: Made /api/parameters/check-triggers endpoint READ-ONLY - no longer creates alerts
  This endpoint now only returns computed trigger patterns for frontend display
- FIX 2: Only process_parameter_triggers() (called on parameter save) creates alerts now
- FIX 3: Improved duplicate detection to handle both underscore and space in parameter names

PJ807 Changes:
- Reduced duplicate detection window from 7 days to 1 day
- Added comprehensive logging to process_parameter_triggers()
- FRONTEND FIX: Fixed JavaScript error in displayParameterAlerts()

PJ808 Changes:
- Removed 24-hour cooldown that was blocking ALL legitimate alerts
- Added detailed trigger logging showing watcher, consecutive_days, and which alerts are enabled

PJ809 Changes (version 1500):
- ENHANCED LOGGING: Added [SAVE PARAMS] logs when parameters are saved
  - Shows user_id, date, and all parameter values when saved
  - Confirms process_parameter_triggers is being called
- REDUCED DUPLICATE WINDOW: Changed from 1 day to 4 hours
  - This allows more frequent alerts while still preventing rapid spam
  - Old spam alerts no longer block new legitimate alerts
- Check Render logs for:
  - [SAVE PARAMS] - When parameters are saved
  - [TRIGGER PROCESS] - Trigger processing lifecycle
  - [PJ809] - When duplicates are skipped (within 4-hour window)

PJ810 Changes (version 1600):
- FIXED: Double message sending when pressing Enter key
  - Added isMessageSending flag to prevent double submissions
  - Button is disabled during send to prevent clicks during API call
- FRONTEND CONFIG: Added TRIGGER_ALERT_DISPLAY_MODE setting
  - 'overlay' = Yellow floating alerts (original behavior)
  - 'standard' = Add to Alerts section like other notifications (default)
  - 'disabled' = Don't show polling alerts (real DB alerts still work)
- CLARIFICATION: Trigger emails are ONLY sent when parameters are saved
  - The yellow overlay alerts come from polling (READ-ONLY endpoint)
  - Real alerts + emails happen when process_parameter_triggers() runs on save
  - If no [SAVE PARAMS] logs in Render, no parameters were saved = no trigger emails

PJ811 Changes (version 1700):
- CRITICAL FIX: Trigger alerts now persist in database and don't vanish
- ROOT CAUSE: check-triggers returned computed patterns but didn't create DB alerts
  Alerts appeared briefly from DOM insertion, then vanished when loadAlerts() ran
- FIX 1: check_parameter_triggers() now creates ONE-TIME database alerts with emails
  - Uses 24-hour duplicate detection per (watcher, watched_user, parameter) combo
  - Creates alerts with proper alert_category='trigger' for filtering
  - Sends email notifications if user has email_on_alert enabled
- FIX 2: Added comprehensive logging throughout trigger lifecycle:
  - [TRIGGER CHECK] - Polling endpoint activity
  - [TRIGGER CREATE] - Alert creation with email
  - [TRIGGER DUPLICATE] - When duplicate detection prevents creation

PJ812 Changes (version 1701):
- FIX: Date formatting in trigger alerts now shows nice dates (e.g., "Dec 05 - Dec 07")
  instead of ISO format dates (e.g., "2025-12-05")
- FIX: Both new schema (with dates array) and old schema (with end_date) now format nicely
- FIX: Frontend now calls checkParameterAlerts() after login completes
- This ensures triggers are checked immediately when user logs in, not just on page load
  - [TRIGGER EMAIL] - Email sending attempts and results
- FIX 3: Emails now sent even when user is not logged in
  - Background trigger processing sends emails to watchers
  - Works like message notifications - alerts delivered regardless of login status
- Check Render logs for [TRIGGER CHECK], [TRIGGER CREATE], [TRIGGER DUPLICATE], [TRIGGER EMAIL]
"""

import os
import sys
import traceback
import json
import uuid
import redis
import logging
import threading
import pytz
from datetime import datetime, timedelta, date
from functools import wraps
import time
from collections import defaultdict

# Cache busting timestamp - updates on every app restart
CACHE_BUST_VERSION = str(int(time.time()))

from flask import (
    Flask, request, jsonify, session,
    render_template, send_from_directory, redirect, url_for
)
from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate
from flask_cors import CORS
from flask_session import Session
from werkzeug.security import generate_password_hash, check_password_hash
from sqlalchemy.exc import IntegrityError, SQLAlchemyError
from sqlalchemy import select, and_, or_, desc, func, inspect, text
# SMTP email (Resend.com compatible)
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

import mimetypes
import io
import tempfile

# Add MIME type for CSS files
mimetypes.add_type('text/css', '.css')
mimetypes.add_type('text/javascript', '.js')

# PDF and Excel report generation
from flask import send_file, Response
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import inch
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.colors import HexColor, black, white
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    REPORTLAB_AVAILABLE = True
except ImportError as e:
    REPORTLAB_AVAILABLE = False
    print(f"WARNING: reportlab not available - PDF generation disabled: {e}")

# WeasyPrint for better Unicode/RTL support (preferred method)
try:
    from weasyprint import HTML, CSS
    WEASYPRINT_AVAILABLE = True
    print("INFO: WeasyPrint available - using for PDF generation with full Unicode support")
except ImportError as e:
    WEASYPRINT_AVAILABLE = False
    print(f"WARNING: weasyprint not available - falling back to reportlab: {e}")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError as e:
    OPENPYXL_AVAILABLE = False
    print(f"WARNING: openpyxl not available - Excel generation disabled: {e}")

# Import security functions
from security import (
    sanitize_input, validate_email, validate_username,
    validate_password_strength, encrypt_field, decrypt_field,
    generate_token, rate_limit, content_moderator
)

# Initialize Flask app
app = Flask(__name__, static_folder='static', template_folder='templates')

# =====================
# CONFIGURATION
# =====================

# Environment detection - using modern Flask approach
is_production = os.environ.get('FLASK_DEBUG', 'False').lower() == 'false'
app.config['DEBUG'] = not is_production

# Secret key configuration
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production')
if is_production and app.config['SECRET_KEY'] == 'dev-secret-key-change-in-production':
    print("WARNING: Using default SECRET_KEY in production!")

# Database configuration
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL', 'sqlite:///social.db')

# Fix for Render PostgreSQL URL
if app.config['SQLALCHEMY_DATABASE_URI'].startswith('postgres://'):
    app.config['SQLALCHEMY_DATABASE_URI'] = app.config['SQLALCHEMY_DATABASE_URI'].replace(
        'postgres://', 'postgresql://', 1
    )

# SQLAlchemy configuration
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_size': 20,  # Increased from 10
    'max_overflow': 30,  # Allow temporary overflow connections
    'pool_recycle': 3600,
    'pool_pre_ping': True,
    'pool_timeout': 30,  # Wait up to 30 seconds for connection
}

# Session configuration
app.config['SESSION_TYPE'] = 'redis' if os.environ.get('REDIS_URL') else 'filesystem'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=7)
app.config['SESSION_COOKIE_SECURE'] = is_production
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_NAME'] = 'thera_session'
app.config['SESSION_KEY_PREFIX'] = 'thera_social:'
app.config['SESSION_USE_SIGNER'] = True

# Redis configuration for sessions
REDIS_URL = os.environ.get('REDIS_URL')
if REDIS_URL:
    app.config['SESSION_REDIS'] = redis.from_url(REDIS_URL)

# Initialize extensions
db = SQLAlchemy(app)
migrate = Migrate(app, db, render_as_batch=True)  # render_as_batch for SQLite

CORS(app, supports_credentials=True)
Session(app)

# =====================
# RATE LIMITING
# =====================

# Simple in-memory rate limiter (use Redis in production for distributed systems)
rate_limit_store = defaultdict(list)


def check_rate_limit(user_id, max_requests=100, window=60):
    """
    Check if user has exceeded rate limit
    max_requests: maximum requests allowed
    window: time window in seconds
    """
    now = time.time()
    user_requests = rate_limit_store[user_id]

    # Remove old requests outside window
    user_requests[:] = [req_time for req_time in user_requests if now - req_time < window]

    if len(user_requests) >= max_requests:
        return False, len(user_requests)

    user_requests.append(now)
    return True, len(user_requests)


def rate_limit_endpoint(max_requests=100, window=60):
    """
    Decorator for rate limiting endpoints
    Usage: @rate_limit_endpoint(max_requests=60, window=60)
    """

    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            user_id = session.get('user_id')
            if user_id:
                allowed, count = check_rate_limit(user_id, max_requests, window)
                if not allowed:
                    return jsonify({'error': 'Rate limit exceeded', 'retry_after': window}), 429
            return f(*args, **kwargs)

        return decorated_function

    return decorator


# Setup logging
logging.basicConfig(
    level=logging.INFO if is_production else logging.DEBUG,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger('thera_social')

# Email configuration - using SMTP with Resend.com
# NOTE: Using standard SMTP library, compatible with Resend

# SMTP configuration for Resend.com (or other SMTP providers)
app.config['MAIL_SERVER'] = os.environ.get('SMTP_SERVER', 'smtp.resend.com')
app.config['MAIL_PORT'] = int(os.environ.get('SMTP_PORT', 465))
app.config['MAIL_USE_TLS'] = False
app.config['MAIL_USE_SSL'] = True
app.config['MAIL_USERNAME'] = os.environ.get('SMTP_USERNAME', 'resend')
app.config['MAIL_PASSWORD'] = os.environ.get('SMTP_PASSWORD')
app.config['MAIL_DEFAULT_SENDER'] = os.environ.get('FROM_EMAIL', 'TheraSocial <onboarding@resend.dev>')

# Flask-Mail not used - using standard SMTP library directly


def get_email_translations(language='en'):
    """Get email translations based on language"""
    translations = {
        'en': {
            'subject': 'TheraSocial - Password Reset Request',
            'hello': 'Hello',
            'request_text': 'You requested to reset your password for TheraSocial.',
            'click_button': 'Click the button below to reset your password:',
            'button_text': 'Reset Password',
            'copy_link': 'Or copy and paste this link into your browser:',
            'expire_text': 'This link will expire in 1 hour.',
            'ignore_text': 'If you did not request this reset, please ignore this email.',
            'regards': 'Best regards',
            'team': 'TheraSocial Team'
        },
        'he': {
            'subject': 'TheraSocial - בקשת איפוס סיסמה',
            'hello': 'שלום',
            'request_text': 'ביקשת לאפס את הסיסמה שלך עבור TheraSocial.',
            'click_button': 'לחץ על הכפתור למטה כדי לאפס את הסיסמה:',
            'button_text': 'איפוס סיסמה',
            'copy_link': 'או העתק והדבק את הקישור הזה בדפדפן שלך:',
            'expire_text': 'קישור זה יפוג תוך שעה.',
            'ignore_text': 'אם לא ביקשת איפוס זה, אנא התעלם מאימייל זה.',
            'regards': 'בברכה',
            'team': 'צוות TheraSocial'
        },
        'ar': {
            'subject': 'TheraSocial - طلب إعادة تعيين كلمة المرور',
            'hello': 'مرحبا',
            'request_text': 'لقد طلبت إعادة تعيين كلمة المرور الخاصة بك لـ TheraSocial.',
            'click_button': 'انقر على الزر أدناه لإعادة تعيين كلمة المرور:',
            'button_text': 'إعادة تعيين كلمة المرور',
            'copy_link': 'أو انسخ والصق هذا الرابط في المتصفح:',
            'expire_text': 'ستنتهي صلاحية هذا الرابط خلال ساعة واحدة.',
            'ignore_text': 'إذا لم تطلب إعادة التعيين، يرجى تجاهل هذا البريد.',
            'regards': 'مع أطيب التحيات',
            'team': 'فريق TheraSocial'
        },
        'ru': {
            'subject': 'TheraSocial - Запрос на сброс пароля',
            'hello': 'Здравствуйте',
            'request_text': 'Вы запросили сброс пароля для TheraSocial.',
            'click_button': 'Нажмите кнопку ниже, чтобы сбросить пароль:',
            'button_text': 'Сбросить пароль',
            'copy_link': 'Или скопируйте эту ссылку в ваш браузер:',
            'expire_text': 'Эта ссылка истечет через 1 час.',
            'ignore_text': 'Если вы не запрашивали сброс, проигнорируйте это письмо.',
            'regards': 'С наилучшими пожеланиями',
            'team': 'Команда TheraSocial'
        }
    }
    return translations.get(language, translations['en'])


def send_password_reset_email(user_email, reset_token, user_language='en'):
    """Send password reset email in user's preferred language using Flask-Mail"""
    try:
        t = get_email_translations(user_language)
        reset_link = f"{os.environ.get('APP_URL', 'http://localhost:5000')}?reset_token={reset_token}"

        is_rtl = user_language in ['he', 'ar']
        text_dir = 'rtl' if is_rtl else 'ltr'
        text_align = 'right' if is_rtl else 'left'

        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
        </head>
        <body style="font-family: Arial, sans-serif; direction: {text_dir}; text-align: {text_align}; background-color: #f5f5f5; margin: 0; padding: 20px;">
            <div style="max-width: 600px; margin: 0 auto; background: white; border-radius: 8px; overflow: hidden; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
                <div style="background: #667eea; padding: 30px; text-align: center;">
                    <h1 style="color: white; margin: 0; font-size: 24px;">TheraSocial</h1>
                </div>
                <div style="padding: 40px 30px;">
                    <h2 style="color: #333; margin-top: 0;">{t['subject'].replace('TheraSocial - ', '')}</h2>
                    <p style="color: #666; line-height: 1.6;">{t['hello']},</p>
                    <p style="color: #666; line-height: 1.6;">{t['request_text']}</p>
                    <p style="color: #666; line-height: 1.6;">{t['click_button']}</p>
                    <div style="text-align: center; margin: 30px 0;">
                        <a href="{reset_link}" style="background: #667eea; color: white; padding: 14px 40px; text-decoration: none; border-radius: 5px; display: inline-block; font-weight: bold;">
                            {t['button_text']}
                        </a>
                    </div>
                    <p style="color: #666; font-size: 14px; line-height: 1.6;">{t['copy_link']}</p>
                    <p style="word-break: break-all; color: #667eea; background: #f8f9fa; padding: 12px; border-radius: 4px; font-size: 12px;">{reset_link}</p>
                    <p style="color: #999; font-size: 13px; margin-top: 30px;">{t['expire_text']}</p>
                    <p style="color: #999; font-size: 13px;">{t['ignore_text']}</p>
                </div>
                <div style="background: #f8f9fa; padding: 20px 30px; border-top: 1px solid #eee;">
                    <p style="color: #999; font-size: 12px; margin: 0;">
                        {t['regards']},<br>{t['team']}
                    </p>
                    <p style="color: #999; font-size: 11px; margin-top: 15px; line-height: 1.5;">
                        This email was sent because a password reset was requested for your TheraSocial account.
                        <br>If you did not request this, you can safely ignore this email.
                    </p>
                </div>
            </div>
        </body>
        </html>
        """

        text_content = f"""
        {t['hello']},
        {t['request_text']}
        {t['click_button']}
        {reset_link}
        {t['expire_text']}
        {t['ignore_text']}
        {t['regards']},
        {t['team']}
        """

        try:
            logger.info(f"[SMTP] Sending password reset email...")
            logger.info(f"[SMTP] MAIL_DEFAULT_SENDER: {app.config.get('MAIL_DEFAULT_SENDER', 'NOT SET')}")
            logger.info(f"[SMTP] SMTP_SERVER: {os.environ.get('SMTP_SERVER', 'smtp.resend.com')}")
            
            msg = MIMEMultipart('alternative')
            msg['Subject'] = t['subject']
            msg['From'] = app.config['MAIL_DEFAULT_SENDER']
            msg['To'] = user_email
            msg.attach(MIMEText(text_content, 'plain'))
            msg.attach(MIMEText(html_content, 'html'))

            smtp_server = os.environ.get('SMTP_SERVER', 'smtp.resend.com')
            smtp_port = int(os.environ.get('SMTP_PORT', '465'))
            smtp_user = os.environ.get('SMTP_USERNAME', 'resend')
            smtp_pass = os.environ.get('SMTP_PASSWORD', app.config.get('MAIL_PASSWORD', ''))
            
            with smtplib.SMTP_SSL(smtp_server, smtp_port) as server:
                server.login(smtp_user, smtp_pass)
                server.sendmail(msg['From'], user_email, msg.as_string())
            
            logging.info(f'[SMTP] Password reset email sent to {user_email}')
        except Exception as e:
            logging.error(f'[SMTP] Failed to send password reset email: {str(e)}')
            raise

        logger.info(f"Password reset email sent to {user_email} in {user_language}")
        return True

    except Exception as e:
        logger.error(f"Failed to send password reset email: {e}")
        return False


def send_magic_link_email(user_email, magic_token, user_language='en'):
    """Send magic link login email using Flask-Mail"""
    try:
        translations = {
            'en': {
                'subject': 'TheraSocial - Magic Link Sign In',
                'hello': 'Hello',
                'request_text': 'Click the link below to sign in to TheraSocial:',
                'button_text': 'Sign In',
                'expire_text': 'This link will expire in 5 years or until a new one is generated',
                'ignore_text': 'If you did not request this, please ignore this email.'
            },
            'he': {
                'subject': 'TheraSocial - התחברות בקישור קסם',
                'hello': 'שלום',
                'request_text': 'לחץ על הקישור למטה כדי להתחבר:',
                'button_text': 'התחבר',
                'expire_text': 'קישור זה יפוג תוקפו בעוד 5 שנים או עד שייווצר קישור חדש',
                'ignore_text': 'אם לא ביקשת זאת, התעלם מאימייל זה.'
            },
            'ar': {
                'subject': 'TheraSocial - تسجيل الدخول بالرابط السحري',
                'hello': 'مرحبا',
                'request_text': 'انقر على الرابط أدناه لتسجيل الدخول:',
                'button_text': 'تسجيل الدخول',
                'expire_text': 'سينتهي هذا الرابط خلال 5 سنوات أو حتى يتم إنشاء رابط جديد',
                'ignore_text': 'إذا لم تطلب هذا، تجاهل هذا البريد.'
            },
            'ru': {
                'subject': 'TheraSocial - Вход по волшебной ссылке',
                'hello': 'Здравствуйте',
                'request_text': 'Нажмите на ссылку ниже для входа:',
                'button_text': 'Войти',
                'expire_text': 'Эта ссылка истечет через 5 лет или пока не будет создана новая.',
                'ignore_text': 'Если вы не запрашивали это, проигнорируйте.'
            }
        }

        t = translations.get(user_language, translations['en'])
        magic_link = f"{os.environ.get('APP_URL', 'http://localhost:5000')}?magic_token={magic_token}"

        html_content = f"""<html><body>
            <div style="max-width: 600px; margin: 0 auto; padding: 20px;">
                <h2>{t['subject']}</h2>
                <p>{t['hello']},</p>
                <p>{t['request_text']}</p>
                <a href="{magic_link}" style="background: #667eea; color: white; padding: 12px 30px; text-decoration: none; border-radius: 5px; display: inline-block;">
                    {t['button_text']}
                </a>
                <p>{t['expire_text']}</p>
                <p>{t['ignore_text']}</p>
            </div>
        </body></html>"""

        try:
            logger.info(f"[SMTP] Sending magic link email...")
            
            msg = MIMEMultipart('alternative')
            msg['Subject'] = t['subject']
            msg['From'] = app.config['MAIL_DEFAULT_SENDER']
            msg['To'] = user_email
            msg.attach(MIMEText(html_content, 'html'))

            smtp_server = os.environ.get('SMTP_SERVER', 'smtp.resend.com')
            smtp_port = int(os.environ.get('SMTP_PORT', '465'))
            smtp_user = os.environ.get('SMTP_USERNAME', 'resend')
            smtp_pass = os.environ.get('SMTP_PASSWORD', app.config.get('MAIL_PASSWORD', ''))
            
            with smtplib.SMTP_SSL(smtp_server, smtp_port) as server:
                server.login(smtp_user, smtp_pass)
                server.sendmail(msg['From'], user_email, msg.as_string())
            
            logging.info(f'[SMTP] Magic link email sent to {user_email}')
        except Exception as e:
            logging.error(f'[SMTP] Failed to send magic link email: {str(e)}')
            raise

        logger.info(f"Magic link email sent to {user_email} in {user_language}")
        return True

    except Exception as e:
        logger.error(f"Failed to send magic link email: {e}")
        return False


def get_new_message_email_translations(language='en'):
    """Get email translations for new message notifications"""
    translations = {
        'en': {
            'subject': 'TheraSocial - New Message from {sender}',
            'hello': 'Hello',
            'new_message': 'You have received a new message from {sender}:',
            'view_message': 'View Message',
            'preview': 'Message preview:',
            'login_to_view': 'Log in to TheraSocial to view the full message and reply.',
            'regards': 'Best regards',
            'team': 'TheraSocial Team'
        },
        'he': {
            'subject': 'TheraSocial - הודעה חדשה מ-{sender}',
            'hello': 'שלום',
            'new_message': 'קיבלת הודעה חדשה מ-{sender}:',
            'view_message': 'צפה בהודעה',
            'preview': 'תצוגה מקדימה:',
            'login_to_view': 'התחבר ל-TheraSocial כדי לצפות בהודעה המלאה ולהשיב.',
            'regards': 'בברכה',
            'team': 'צוות TheraSocial'
        },
        'ar': {
            'subject': 'TheraSocial - رسالة جديدة من {sender}',
            'hello': 'مرحبا',
            'new_message': 'لقد تلقيت رسالة جديدة من {sender}:',
            'view_message': 'عرض الرسالة',
            'preview': 'معاينة الرسالة:',
            'login_to_view': 'سجل الدخول إلى TheraSocial لعرض الرسالة الكاملة والرد.',
            'regards': 'مع أطيب التحيات',
            'team': 'فريق TheraSocial'
        },
        'ru': {
            'subject': 'TheraSocial - Новое сообщение от {sender}',
            'hello': 'Здравствуйте',
            'new_message': 'Вы получили новое сообщение от {sender}:',
            'view_message': 'Просмотреть сообщение',
            'preview': 'Предварительный просмотр:',
            'login_to_view': 'Войдите в TheraSocial, чтобы прочитать полное сообщение и ответить.',
            'regards': 'С наилучшими пожеланиями',
            'team': 'Команда TheraSocial'
        }
    }
    return translations.get(language, translations['en'])


def send_new_message_notification_email(recipient_email, sender_name, message_preview, user_language='en'):
    """Send email notification when user receives a new message"""
    logger.info(f"[SMTP MESSAGE] send_new_message_notification_email called")
    logger.info(f"[SMTP MESSAGE] to: {recipient_email}, sender: {sender_name}, language: {user_language}")
    try:
        t = get_new_message_email_translations(user_language)
        app_url = os.environ.get('APP_URL', 'http://localhost:5000')
        messages_link = f"{app_url}/#messages"

        is_rtl = user_language in ['he', 'ar']
        text_dir = 'rtl' if is_rtl else 'ltr'
        text_align = 'right' if is_rtl else 'left'

        subject = t['subject'].format(sender=sender_name)

        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
        </head>
        <body style="font-family: Arial, sans-serif; direction: {text_dir}; text-align: {text_align}; background-color: #f5f5f5; margin: 0; padding: 20px;">
            <div style="max-width: 600px; margin: 0 auto; background: white; border-radius: 8px; overflow: hidden; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
                <div style="background: #667eea; padding: 30px; text-align: center;">
                    <h1 style="color: white; margin: 0; font-size: 24px;">TheraSocial</h1>
                </div>
                <div style="padding: 40px 30px;">
                    <h2 style="color: #333; margin-top: 0;">{t['new_message'].format(sender=sender_name)}</h2>
                    <p style="color: #666; line-height: 1.6;">{t['hello']},</p>
                    <p style="color: #666; line-height: 1.6;">{t['preview']}</p>
                    <div style="background: #f8f9fa; padding: 15px; border-radius: 8px; margin: 20px 0; border-left: 4px solid #667eea;">
                        <p style="color: #555; margin: 0; font-style: italic;">"{message_preview[:200]}{'...' if len(message_preview) > 200 else ''}"</p>
                    </div>
                    <p style="color: #666; line-height: 1.6;">{t['login_to_view']}</p>
                    <div style="text-align: center; margin: 30px 0;">
                        <a href="{messages_link}" style="background: #667eea; color: white; padding: 14px 40px; text-decoration: none; border-radius: 5px; display: inline-block; font-weight: bold;">
                            {t['view_message']}
                        </a>
                    </div>
                </div>
                <div style="background: #f8f9fa; padding: 20px 30px; border-top: 1px solid #eee;">
                    <p style="color: #999; font-size: 12px; margin: 0;">
                        {t['regards']},<br>{t['team']}
                    </p>
                </div>
            </div>
        </body>
        </html>
        """

        try:
            logger.info(f"[SMTP MESSAGE] Creating email...")
            logger.info(f"[SMTP MESSAGE] MAIL_DEFAULT_SENDER: {app.config.get('MAIL_DEFAULT_SENDER', 'NOT SET')}")
            logger.info(f"[SMTP MESSAGE] SMTP_SERVER: {os.environ.get('SMTP_SERVER', 'NOT SET')}")
            
            msg = MIMEMultipart('alternative')
            msg['Subject'] = subject
            msg['From'] = app.config['MAIL_DEFAULT_SENDER']
            msg['To'] = recipient_email
            msg.attach(MIMEText(html_content, 'html'))

            smtp_server = os.environ.get('SMTP_SERVER', 'smtp.resend.com')
            smtp_port = int(os.environ.get('SMTP_PORT', '465'))
            smtp_user = os.environ.get('SMTP_USERNAME', 'resend')
            smtp_pass = os.environ.get('SMTP_PASSWORD', app.config.get('MAIL_PASSWORD', ''))
            
            with smtplib.SMTP_SSL(smtp_server, smtp_port) as server:
                server.login(smtp_user, smtp_pass)
                server.sendmail(msg['From'], recipient_email, msg.as_string())
            
            logger.info(f'[SMTP MESSAGE] New message notification email sent to {recipient_email}')
            return True
        except Exception as e:
            logger.error(f'[SMTP MESSAGE] Failed to send new message notification email: {str(e)}')
            logger.error(f'[SMTP MESSAGE] Traceback: {traceback.format_exc()}')
            return False

    except Exception as e:
        logger.error(f"[SMTP MESSAGE] Failed to send new message notification email: {e}")
        logger.error(f"[SMTP MESSAGE] Traceback: {traceback.format_exc()}")
        return False


def get_alert_notification_email_translations(language='en'):
    """Get email translations for alert notifications"""
    translations = {
        'en': {
            'subject': 'TheraSocial - New Alert',
            'hello': 'Hello',
            'new_alert': 'You have a new alert:',
            'view_alerts': 'View Alerts',
            'login_to_view': 'Log in to TheraSocial to view all your alerts.',
            'regards': 'Best regards',
            'team': 'TheraSocial Team'
        },
        'he': {
            'subject': 'TheraSocial - התראה חדשה',
            'hello': 'שלום',
            'new_alert': 'יש לך התראה חדשה:',
            'view_alerts': 'צפה בהתראות',
            'login_to_view': 'התחבר ל-TheraSocial כדי לצפות בכל ההתראות שלך.',
            'regards': 'בברכה',
            'team': 'צוות TheraSocial'
        },
        'ar': {
            'subject': 'TheraSocial - تنبيه جديد',
            'hello': 'مرحبا',
            'new_alert': 'لديك تنبيه جديد:',
            'view_alerts': 'عرض التنبيهات',
            'login_to_view': 'سجل الدخول إلى TheraSocial لعرض جميع تنبيهاتك.',
            'regards': 'مع أطيب التحيات',
            'team': 'فريق TheraSocial'
        },
        'ru': {
            'subject': 'TheraSocial - Новое уведомление',
            'hello': 'Здравствуйте',
            'new_alert': 'У вас новое уведомление:',
            'view_alerts': 'Просмотреть уведомления',
            'login_to_view': 'Войдите в TheraSocial, чтобы просмотреть все уведомления.',
            'regards': 'С наилучшими пожеланиями',
            'team': 'Команда TheraSocial'
        }
    }
    return translations.get(language, translations['en'])


def get_invite_alert_translations(language='en'):
    """
    PJ706: Get translations for invite alert emails.
    These translate the invite.alert_title and invite.alert_content keys.
    """
    translations = {
        'en': {
            'invite.alert_title': 'New Invitation',
            'invite.alert_content': '{username} has invited you to follow them'
        },
        'he': {
            'invite.alert_title': 'הזמנה חדשה',
            'invite.alert_content': '{username} הזמין/ה אותך לעקוב אחריו/ה'
        },
        'ar': {
            'invite.alert_title': 'دعوة جديدة',
            'invite.alert_content': 'دعاك {username} لمتابعته'
        },
        'ru': {
            'invite.alert_title': 'Новое приглашение',
            'invite.alert_content': '{username} пригласил(а) вас подписаться'
        }
    }
    return translations.get(language, translations['en'])


def get_notification_translations(language='en'):
    """
    PJ6003: Get translations for all notification types (follow, message, invite).
    """
    translations = {
        'en': {
            # Follow notifications
            'follow.alert_title': '{username} started following you',
            'follow.alert_content': 'You have a new follower!',
            'follow.new_follower': 'You have a new follower!',
            # Message notifications
            'message.alert_title': 'New message from {username}',
            'message.alert_content': 'You have received a new message',
            # Invite notifications  
            'invite.alert_title': 'New Invitation',
            'invite.alert_content': '{username} has invited you to follow them'
        },
        'he': {
            # Follow notifications
            'follow.alert_title': '{username} התחיל/ה לעקוב אחריך',
            'follow.alert_content': 'יש לך עוקב/ת חדש/ה!',
            'follow.new_follower': 'יש לך עוקב/ת חדש/ה!',
            # Message notifications
            'message.alert_title': 'הודעה חדשה מ-{username}',
            'message.alert_content': 'קיבלת הודעה חדשה',
            # Invite notifications
            'invite.alert_title': 'הזמנה חדשה',
            'invite.alert_content': '{username} הזמין/ה אותך לעקוב אחריו/ה'
        },
        'ar': {
            # Follow notifications
            'follow.alert_title': 'بدأ {username} بمتابعتك',
            'follow.alert_content': 'لديك متابع جديد!',
            'follow.new_follower': 'لديك متابع جديد!',
            # Message notifications
            'message.alert_title': 'رسالة جديدة من {username}',
            'message.alert_content': 'لقد تلقيت رسالة جديدة',
            # Invite notifications
            'invite.alert_title': 'دعوة جديدة',
            'invite.alert_content': 'دعاك {username} لمتابعته'
        },
        'ru': {
            # Follow notifications
            'follow.alert_title': '{username} подписался на вас',
            'follow.alert_content': 'У вас новый подписчик!',
            'follow.new_follower': 'У вас новый подписчик!',
            # Message notifications
            'message.alert_title': 'Новое сообщение от {username}',
            'message.alert_content': 'Вы получили новое сообщение',
            # Invite notifications
            'invite.alert_title': 'Новое приглашение',
            'invite.alert_content': '{username} пригласил(а) вас подписаться'
        }
    }
    return translations.get(language, translations['en'])


def translate_alert_content(title, content, user_language='en'):
    """
    PJ706/PJ6003: Translate alert/notification title and content if they contain translation keys.
    Handles formats like:
    - title: "invite.alert_title"
    - content: "username|invite.alert_content"
    - title: "follow.alert_title" with {username} placeholder
    """
    invite_translations = get_invite_alert_translations(user_language)
    notification_translations = get_notification_translations(user_language)
    
    # Merge all translations
    all_translations = {**invite_translations, **notification_translations}
    
    translated_title = title
    translated_content = content
    
    # Translate title if it's a known key
    if title in all_translations:
        translated_title = all_translations[title]
    
    # Handle content format: "username|key" or just "key"
    if content and '|' in content:
        parts = content.split('|', 1)
        username = parts[0]
        content_key = parts[1] if len(parts) > 1 else ''
        
        if content_key in all_translations:
            translated_content = all_translations[content_key].replace('{username}', username)
        else:
            translated_content = content
    elif content in all_translations:
        translated_content = all_translations[content]
    
    # Handle {username} placeholder in title if content contains username
    if '{username}' in translated_title and content and '|' in content:
        username = content.split('|')[0]
        translated_title = translated_title.replace('{username}', username)
    
    logger.info(f"[PJ6003] Translated notification - title: '{title}' -> '{translated_title}', content: '{content}' -> '{translated_content}'")
    
    return translated_title, translated_content


def send_alert_notification_email(user_email, alert_title, alert_content, user_language='en'):
    """Send email notification when user gets a new alert"""
    logger.info(f"[SMTP ALERT] send_alert_notification_email called")
    logger.info(f"[SMTP ALERT] to: {user_email}, title: {alert_title}, language: {user_language}")
    try:
        t = get_alert_notification_email_translations(user_language)
        app_url = os.environ.get('APP_URL', 'http://localhost:5000')

        # PJ706: Translate alert title and content if they are translation keys
        translated_title, translated_content = translate_alert_content(alert_title, alert_content, user_language)
        alert_title = translated_title
        alert_content = translated_content

        is_rtl = user_language in ['he', 'ar']
        text_dir = 'rtl' if is_rtl else 'ltr'
        text_align = 'right' if is_rtl else 'left'

        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
        </head>
        <body style="font-family: Arial, sans-serif; direction: {text_dir}; text-align: {text_align}; background-color: #f5f5f5; margin: 0; padding: 20px;">
            <div style="max-width: 600px; margin: 0 auto; background: white; border-radius: 8px; overflow: hidden; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
                <div style="background: #667eea; padding: 30px; text-align: center;">
                    <h1 style="color: white; margin: 0; font-size: 24px;">TheraSocial</h1>
                </div>
                <div style="padding: 40px 30px;">
                    <p style="color: #666; line-height: 1.6;">{t['hello']},</p>
                    <p style="color: #666; line-height: 1.6;">{t['new_alert']}</p>
                    <div style="background: #fff3cd; padding: 15px; border-radius: 8px; margin: 20px 0; border-left: 4px solid #ffc107;">
                        <h3 style="color: #856404; margin: 0 0 10px 0;">{alert_title}</h3>
                        <p style="color: #856404; margin: 0;">{alert_content}</p>
                    </div>
                    <p style="color: #666; line-height: 1.6;">{t['login_to_view']}</p>
                    <div style="text-align: center; margin: 30px 0;">
                        <a href="{app_url}" style="background: #667eea; color: white; padding: 14px 40px; text-decoration: none; border-radius: 5px; display: inline-block; font-weight: bold;">
                            {t['view_alerts']}
                        </a>
                    </div>
                </div>
                <div style="background: #f8f9fa; padding: 20px 30px; border-top: 1px solid #eee;">
                    <p style="color: #999; font-size: 12px; margin: 0;">
                        {t['regards']},<br>{t['team']}
                    </p>
                </div>
            </div>
        </body>
        </html>
        """

        try:
            logger.info(f"[SMTP ALERT] Creating email...")
            logger.info(f"[SMTP ALERT] MAIL_DEFAULT_SENDER: {app.config.get('MAIL_DEFAULT_SENDER', 'NOT SET')}")
            logger.info(f"[SMTP ALERT] SMTP_SERVER: {os.environ.get('SMTP_SERVER', 'NOT SET')}")
            
            msg = MIMEMultipart('alternative')
            msg['Subject'] = t['subject']
            msg['From'] = app.config['MAIL_DEFAULT_SENDER']
            msg['To'] = user_email
            msg.attach(MIMEText(html_content, 'html'))

            smtp_server = os.environ.get('SMTP_SERVER', 'smtp.resend.com')
            smtp_port = int(os.environ.get('SMTP_PORT', '465'))
            smtp_user = os.environ.get('SMTP_USERNAME', 'resend')
            smtp_pass = os.environ.get('SMTP_PASSWORD', app.config.get('MAIL_PASSWORD', ''))
            
            with smtplib.SMTP_SSL(smtp_server, smtp_port) as server:
                server.login(smtp_user, smtp_pass)
                server.sendmail(msg['From'], user_email, msg.as_string())
            
            logger.info(f'[SMTP ALERT] Alert notification email sent to {user_email}')
            return True
        except Exception as e:
            logger.error(f'[SMTP ALERT] Failed to send alert notification email: {str(e)}')
            logger.error(f'[SMTP ALERT] Traceback: {traceback.format_exc()}')
            return False

    except Exception as e:
        logger.error(f"[SMTP ALERT] Failed to send alert notification email: {e}")
        logger.error(f"[SMTP ALERT] Traceback: {traceback.format_exc()}")
        return False


def get_daily_diary_reminder_translations(language='en'):
    """Get email translations for daily diary reminder"""
    translations = {
        'en': {
            'subject': 'TheraSocial - Daily Wellness Check-in Reminder',
            'hello': 'Hello',
            'reminder': "Don't forget to log your wellness parameters for today!",
            'description': 'Taking a few moments to track your mood, energy, sleep, and other wellness factors helps you understand your patterns and make positive changes.',
            'fill_diary': 'Fill Out Daily Diary',
            'regards': 'Best regards',
            'team': 'TheraSocial Team',
            'unsubscribe': 'To stop receiving these reminders, turn off daily email reminders in your alert settings.'
        },
        'he': {
            'subject': 'TheraSocial - תזכורת יומית למילוי יומן',
            'hello': 'שלום',
            'reminder': 'אל תשכח למלא את פרמטרי הבריאות שלך להיום!',
            'description': 'הקדשת כמה רגעים למעקב אחר מצב הרוח, האנרגיה, השינה וגורמי בריאות אחרים עוזרת לך להבין את הדפוסים שלך ולבצע שינויים חיוביים.',
            'fill_diary': 'מלא יומן יומי',
            'regards': 'בברכה',
            'team': 'צוות TheraSocial',
            'unsubscribe': 'כדי להפסיק לקבל תזכורות אלה, כבה את התזכורות היומיות בהגדרות ההתראות שלך.'
        },
        'ar': {
            'subject': 'TheraSocial - تذكير يومي بملء اليومية',
            'hello': 'مرحبا',
            'reminder': 'لا تنس تسجيل معلومات صحتك لهذا اليوم!',
            'description': 'قضاء بضع لحظات لتتبع مزاجك وطاقتك ونومك وعوامل الصحة الأخرى يساعدك على فهم أنماطك وإجراء تغييرات إيجابية.',
            'fill_diary': 'ملء اليومية اليومية',
            'regards': 'مع أطيب التحيات',
            'team': 'فريق TheraSocial',
            'unsubscribe': 'لإيقاف تلقي هذه التذكيرات، قم بإيقاف تشغيل التذكيرات اليومية في إعدادات التنبيهات.'
        },
        'ru': {
            'subject': 'TheraSocial - Ежедневное напоминание о дневнике здоровья',
            'hello': 'Здравствуйте',
            'reminder': 'Не забудьте записать ваши показатели здоровья за сегодня!',
            'description': 'Несколько минут на отслеживание настроения, энергии, сна и других факторов помогут вам понять ваши паттерны и внести позитивные изменения.',
            'fill_diary': 'Заполнить дневник',
            'regards': 'С наилучшими пожеланиями',
            'team': 'Команда TheraSocial',
            'unsubscribe': 'Чтобы перестать получать эти напоминания, отключите ежедневные напоминания в настройках уведомлений.'
        }
    }
    return translations.get(language, translations['en'])


def send_daily_diary_reminder_email(user_email, user_language='en'):
    """Send daily reminder email to fill out wellness diary"""
    logger.info(f"[DAILY REMINDER] ========================================")
    logger.info(f"[DAILY REMINDER] Starting send_daily_diary_reminder_email")
    logger.info(f"[DAILY REMINDER] user_email: {user_email}")
    logger.info(f"[DAILY REMINDER] user_language: {user_language}")
    
    try:
        t = get_daily_diary_reminder_translations(user_language)
        logger.info(f"[DAILY REMINDER] Got translations for language: {user_language}")
        
        app_url = os.environ.get('APP_URL', 'http://localhost:5000')
        diary_link = f"{app_url}/parameters"
        logger.info(f"[DAILY REMINDER] App URL: {app_url}")
        logger.info(f"[DAILY REMINDER] Diary link: {diary_link}")

        is_rtl = user_language in ['he', 'ar']
        text_dir = 'rtl' if is_rtl else 'ltr'
        text_align = 'right' if is_rtl else 'left'

        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
        </head>
        <body style="font-family: Arial, sans-serif; direction: {text_dir}; text-align: {text_align}; background-color: #f5f5f5; margin: 0; padding: 20px;">
            <div style="max-width: 600px; margin: 0 auto; background: white; border-radius: 8px; overflow: hidden; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
                <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); padding: 30px; text-align: center;">
                    <h1 style="color: white; margin: 0; font-size: 24px;">📝 TheraSocial</h1>
                </div>
                <div style="padding: 40px 30px;">
                    <p style="color: #666; line-height: 1.6;">{t['hello']},</p>
                    <h2 style="color: #667eea; margin-top: 0;">{t['reminder']}</h2>
                    <p style="color: #666; line-height: 1.6;">{t['description']}</p>
                    <div style="text-align: center; margin: 30px 0;">
                        <a href="{diary_link}" style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 14px 40px; text-decoration: none; border-radius: 25px; display: inline-block; font-weight: bold; box-shadow: 0 4px 15px rgba(102, 126, 234, 0.4);">
                            {t['fill_diary']}
                        </a>
                    </div>
                </div>
                <div style="background: #f8f9fa; padding: 20px 30px; border-top: 1px solid #eee;">
                    <p style="color: #999; font-size: 12px; margin: 0;">
                        {t['regards']},<br>{t['team']}
                    </p>
                    <p style="color: #999; font-size: 10px; margin-top: 15px; font-style: italic;">
                        {t['unsubscribe']}
                    </p>
                </div>
            </div>
        </body>
        </html>
        """

        logger.info(f"[DAILY REMINDER] HTML content prepared")
        logger.info(f"[DAILY REMINDER] Checking SMTP configuration...")
        
        smtp_pass = os.environ.get('SMTP_PASSWORD', app.config.get('MAIL_PASSWORD', ''))
        if not smtp_pass:
            logger.error(f"[DAILY REMINDER] ERROR: SMTP_PASSWORD is not configured!")
            return False
        
        logger.info(f"[DAILY REMINDER] SMTP credentials found")
        logger.info(f"[DAILY REMINDER] From email: {app.config.get('MAIL_DEFAULT_SENDER')}")
        
        try:
            msg = MIMEMultipart('alternative')
            msg['Subject'] = t['subject']
            msg['From'] = app.config['MAIL_DEFAULT_SENDER']
            msg['To'] = user_email
            msg.attach(MIMEText(html_content, 'html'))
            logger.info(f"[DAILY REMINDER] Email message created successfully")

            smtp_server = os.environ.get('SMTP_SERVER', 'smtp.resend.com')
            smtp_port = int(os.environ.get('SMTP_PORT', '465'))
            smtp_user = os.environ.get('SMTP_USERNAME', 'resend')
            logger.info(f"[DAILY REMINDER] Connecting to {smtp_server}:{smtp_port}")
            
            with smtplib.SMTP_SSL(smtp_server, smtp_port) as server:
                server.login(smtp_user, smtp_pass)
                server.sendmail(msg['From'], user_email, msg.as_string())
            
            logger.info(f'[DAILY REMINDER] SUCCESS: Daily diary reminder email sent to {user_email}')
            logger.info(f"[DAILY REMINDER] ========================================")
            return True
                
        except Exception as e:
            logger.error(f'[DAILY REMINDER] ERROR: Failed to send daily diary reminder email: {str(e)}')
            logger.error(f'[DAILY REMINDER] Exception type: {type(e).__name__}')
            logger.error(f'[DAILY REMINDER] Traceback: {traceback.format_exc()}')
            return False

    except Exception as e:
        logger.error(f"[DAILY REMINDER] CRITICAL ERROR: {e}")
        logger.error(f"[DAILY REMINDER] Traceback: {traceback.format_exc()}")
        logger.info(f"[DAILY REMINDER] ========================================")
        return False


def create_alert_with_email(user_id, title, content, alert_type='info', source_user_id=None, alert_category='general'):
    """
    Create an alert and optionally send email notification if user has email_on_alert enabled.
    
    Args:
        user_id: The user ID to create alert for
        title: Alert title (can be JSON string with translation key)
        content: Alert content/message
        alert_type: Type of alert ('info', 'warning', 'success', 'error')
        source_user_id: PJ401 - ID of user this alert is about (for filtering based on following)
        alert_category: PJ401 - Category: 'trigger', 'feed', 'message', 'follow', 'general'
    
    Returns:
        The created Alert object
    """
    logger.info(f"[ALERT EMAIL] ========================================")
    logger.info(f"[ALERT EMAIL] create_alert_with_email called")
    logger.info(f"[ALERT EMAIL] user_id: {user_id}, alert_category: {alert_category}")
    logger.info(f"[ALERT EMAIL] title: {title[:100] if title else 'None'}...")
    
    try:
        # Create the alert with source_user_id for following-based filtering
        alert = Alert(
            user_id=user_id,
            title=title,
            content=content,
            alert_type=alert_type,
            source_user_id=source_user_id,
            alert_category=alert_category
        )
        db.session.add(alert)
        db.session.flush()  # Get the alert ID without committing
        logger.info(f"[ALERT EMAIL] Alert created with ID: {alert.id}")
        
        # Check if user has email notifications enabled
        try:
            settings = NotificationSettings.query.filter_by(user_id=user_id).first()
            logger.info(f"[ALERT EMAIL] NotificationSettings found: {settings is not None}")
            
            if settings:
                logger.info(f"[ALERT EMAIL] email_on_alert setting: {settings.email_on_alert}")
            else:
                logger.info(f"[ALERT EMAIL] No NotificationSettings for user {user_id}, email will NOT be sent")
            
            if settings and settings.email_on_alert:
                # PJ6009: Skip individual emails for trigger alerts - they use batch emails via send_consolidated_wellness_alert_email
                if alert_category == 'trigger':
                    logger.info(f"[ALERT EMAIL] Skipping individual email for trigger alert - batch email will be sent instead")
                else:
                    user = db.session.get(User, user_id)
                    logger.info(f"[ALERT EMAIL] User found: {user is not None}, has email: {user.email if user else 'N/A'}")
                    
                    if user and user.email:
                        # Parse alert title for email
                        email_title = title
                        try:
                            title_data = json.loads(title)
                            if isinstance(title_data, dict) and 'key' in title_data:
                                username = title_data.get('params', {}).get('username', '')
                                key = title_data.get('key', '')
                                if 'new_message' in key:
                                    email_title = f"New message from {username}"
                                elif 'started_following' in key:
                                    email_title = f"{username} started following you"
                                elif 'invitation' in key.lower():
                                    email_title = "New invitation"
                                else:
                                    email_title = username or 'New Alert'
                        except:
                            pass  # Not JSON, use as-is
                        
                        logger.info(f"[ALERT EMAIL] Sending alert email to {user.email} with title: {email_title}")
                        user_language = user.preferred_language or 'en'
                        result = send_alert_notification_email(user.email, email_title, content or '', user_language)
                        logger.info(f"[ALERT EMAIL] Email send result: {result}")
                    else:
                        logger.info(f"[ALERT EMAIL] Skipping email - user not found or no email address")
            else:
                logger.info(f"[ALERT EMAIL] Skipping email - email_on_alert is disabled or no settings")
                
        except Exception as email_err:
            logger.error(f"[ALERT EMAIL] Error sending alert email notification: {str(email_err)}")
            logger.error(f"[ALERT EMAIL] Traceback: {traceback.format_exc()}")
            # Don't fail the alert creation if email fails
        
        logger.info(f"[ALERT EMAIL] ========================================")
        return alert
    except Exception as e:
        logger.error(f"[ALERT EMAIL] Error creating alert: {str(e)}")
        logger.error(f"[ALERT EMAIL] Traceback: {traceback.format_exc()}")
        raise


def create_alert_no_email(user_id, title, content, alert_type='info', source_user_id=None, alert_category='general'):
    """
    PJ6007: Create an alert WITHOUT sending email.
    Used for individual trigger alerts when we want to send a consolidated email later.
    
    Args:
        user_id: The user ID to create alert for
        title: Alert title
        content: Alert content/message
        alert_type: Type of alert
        source_user_id: ID of user this alert is about
        alert_category: Category of alert
    
    Returns:
        The created Alert object
    """
    try:
        alert = Alert(
            user_id=user_id,
            title=title,
            content=content,
            alert_type=alert_type,
            source_user_id=source_user_id,
            alert_category=alert_category
        )
        db.session.add(alert)
        db.session.flush()
        logger.info(f"[ALERT NO EMAIL] Created alert ID {alert.id} for user {user_id} (no email)")
        return alert
    except Exception as e:
        logger.error(f"[ALERT NO EMAIL] Error creating alert: {str(e)}")
        raise


def send_consolidated_wellness_alert_email(watcher_id, watched_username, triggered_params, user_language='en'):
    """
    PJ6007: Send ONE consolidated email for multiple wellness alerts.
    Instead of 7 separate emails, sends one email listing all triggered parameters.
    
    Args:
        watcher_id: User ID of the person receiving the alert
        watched_username: Username of the person being watched
        triggered_params: List of dicts with 'param_name', 'days', 'date_range'
        user_language: Language for email content
    
    Returns:
        True if email sent successfully, False otherwise
    """
    logger.info(f"[CONSOLIDATED EMAIL] ========================================")
    logger.info(f"[CONSOLIDATED EMAIL] Sending consolidated alert to watcher_id={watcher_id}")
    logger.info(f"[CONSOLIDATED EMAIL] watched_username={watched_username}, params={len(triggered_params)}")
    
    try:
        # Check if user has email notifications enabled
        settings = NotificationSettings.query.filter_by(user_id=watcher_id).first()
        if not settings or not settings.email_on_alert:
            logger.info(f"[CONSOLIDATED EMAIL] Skipping - email_on_alert disabled for user {watcher_id}")
            return False
        
        watcher = db.session.get(User, watcher_id)
        if not watcher or not watcher.email:
            logger.info(f"[CONSOLIDATED EMAIL] Skipping - no email for user {watcher_id}")
            return False
        
        # Build the consolidated content
        translations = {
            'en': {
                'subject': f'TheraSocial - Wellness Alert for {watched_username}',
                'hello': 'Hello',
                'intro': f"We noticed some concerning wellness patterns for {watched_username}:",
                'param_line': '{param} has been at concerning levels for {days} consecutive days ({date_range})',
                'recommendation': 'Consider reaching out to check in on them.',
                'view_details': 'View Details',
                'regards': 'Best regards',
                'team': 'TheraSocial Team',
                'mood': 'Mood',
                'energy': 'Energy',
                'sleep_quality': 'Sleep quality',
                'physical_activity': 'Physical activity',
                'anxiety': 'Anxiety'
            },
            'he': {
                'subject': f'TheraSocial - התראת בריאות עבור {watched_username}',
                'hello': 'שלום',
                'intro': f"שמנו לב לדפוסי בריאות מדאיגים עבור {watched_username}:",
                'param_line': '{param} היה ברמות מדאיגות במשך {days} ימים רצופים ({date_range})',
                'recommendation': 'שקול/י ליצור קשר כדי לבדוק את מצבם.',
                'view_details': 'צפה בפרטים',
                'regards': 'בברכה',
                'team': 'צוות TheraSocial',
                'mood': 'מצב רוח',
                'energy': 'אנרגיה',
                'sleep_quality': 'איכות שינה',
                'physical_activity': 'פעילות גופנית',
                'anxiety': 'חרדה'
            },
            'ar': {
                'subject': f'TheraSocial - تنبيه صحي لـ {watched_username}',
                'hello': 'مرحباً',
                'intro': f"لاحظنا بعض أنماط الصحة المقلقة لـ {watched_username}:",
                'param_line': '{param} كان عند مستويات مقلقة لمدة {days} أيام متتالية ({date_range})',
                'recommendation': 'فكر في التواصل للاطمئنان عليهم.',
                'view_details': 'عرض التفاصيل',
                'regards': 'مع أطيب التحيات',
                'team': 'فريق TheraSocial',
                'mood': 'المزاج',
                'energy': 'الطاقة',
                'sleep_quality': 'جودة النوم',
                'physical_activity': 'النشاط البدني',
                'anxiety': 'القلق'
            },
            'ru': {
                'subject': f'TheraSocial - Оповещение о здоровье {watched_username}',
                'hello': 'Здравствуйте',
                'intro': f"Мы заметили тревожные показатели здоровья у {watched_username}:",
                'param_line': '{param} был на тревожном уровне {days} дней подряд ({date_range})',
                'recommendation': 'Рассмотрите возможность связаться с ними.',
                'view_details': 'Подробнее',
                'regards': 'С уважением',
                'team': 'Команда TheraSocial',
                'mood': 'Настроение',
                'energy': 'Энергия',
                'sleep_quality': 'Качество сна',
                'physical_activity': 'Физическая активность',
                'anxiety': 'Тревожность'
            }
        }
        
        t = translations.get(user_language, translations['en'])
        app_url = os.environ.get('APP_URL', 'http://localhost:5000')
        is_rtl = user_language in ['he', 'ar']
        text_dir = 'rtl' if is_rtl else 'ltr'
        text_align = 'right' if is_rtl else 'left'
        
        # Build parameter list HTML
        param_items = []
        for p in triggered_params:
            param_display = t.get(p['param_name'], p['param_name'])
            line = t['param_line'].replace('{param}', param_display).replace('{days}', str(p['days'])).replace('{date_range}', p['date_range'])
            param_items.append(f'<li style="margin-bottom: 8px;">{line}</li>')
        
        params_html = '\n'.join(param_items)
        
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
        </head>
        <body style="font-family: Arial, sans-serif; direction: {text_dir}; text-align: {text_align}; background-color: #f5f5f5; margin: 0; padding: 20px;">
            <div style="max-width: 600px; margin: 0 auto; background: white; border-radius: 8px; overflow: hidden; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
                <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); padding: 30px; text-align: center;">
                    <h1 style="color: white; margin: 0; font-size: 24px;">TheraSocial</h1>
                </div>
                <div style="padding: 40px 30px;">
                    <p style="color: #666; line-height: 1.6;">{t['hello']},</p>
                    <p style="color: #666; line-height: 1.6;">{t['intro']}</p>
                    <div style="background: #fff3cd; padding: 20px; border-radius: 8px; margin: 20px 0; border-left: 4px solid #ffc107;">
                        <ul style="color: #856404; margin: 0; padding-left: 20px;">
                            {params_html}
                        </ul>
                    </div>
                    <p style="color: #666; line-height: 1.6;">{t['recommendation']}</p>
                    <div style="text-align: center; margin: 30px 0;">
                        <a href="{app_url}" style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 14px 40px; text-decoration: none; border-radius: 5px; display: inline-block; font-weight: bold;">
                            {t['view_details']}
                        </a>
                    </div>
                </div>
                <div style="background: #f8f9fa; padding: 20px 30px; border-top: 1px solid #eee;">
                    <p style="color: #999; font-size: 12px; margin: 0;">
                        {t['regards']},<br>{t['team']}
                    </p>
                </div>
            </div>
        </body>
        </html>
        """
        
        # Send the email
        try:
            msg = MIMEMultipart('alternative')
            msg['Subject'] = t['subject']
            msg['From'] = app.config['MAIL_DEFAULT_SENDER']
            msg['To'] = watcher.email
            msg.attach(MIMEText(html_content, 'html'))
            
            smtp_server = os.environ.get('SMTP_SERVER', 'smtp.resend.com')
            smtp_port = int(os.environ.get('SMTP_PORT', '465'))
            smtp_user = os.environ.get('SMTP_USERNAME', 'resend')
            smtp_pass = os.environ.get('SMTP_PASSWORD', app.config.get('MAIL_PASSWORD', ''))
            
            with smtplib.SMTP_SSL(smtp_server, smtp_port) as server:
                server.login(smtp_user, smtp_pass)
                server.sendmail(msg['From'], watcher.email, msg.as_string())
            
            logger.info(f'[CONSOLIDATED EMAIL] Successfully sent to {watcher.email} with {len(triggered_params)} params')
            logger.info(f"[CONSOLIDATED EMAIL] ========================================")
            return True
            
        except Exception as e:
            logger.error(f'[CONSOLIDATED EMAIL] Failed to send: {str(e)}')
            logger.error(f"[CONSOLIDATED EMAIL] ========================================")
            return False
            
    except Exception as e:
        logger.error(f"[CONSOLIDATED EMAIL] Error: {str(e)}")
        logger.error(f"[CONSOLIDATED EMAIL] Traceback: {traceback.format_exc()}")
        return False


def get_notification_email_translations(language='en'):
    """PJ6001: Get email translations for notification emails (messages, followers, invites)"""
    translations = {
        'en': {
            'subject': 'TheraSocial - New Notification',
            'hello': 'Hello',
            'new_notification': 'You have a new notification:',
            'view_notifications': 'View Notifications',
            'login_to_view': 'Log in to TheraSocial to view all your notifications.',
            'regards': 'Best regards',
            'team': 'TheraSocial Team'
        },
        'he': {
            'subject': 'TheraSocial - הודעה חדשה',
            'hello': 'שלום',
            'new_notification': 'יש לך הודעה חדשה:',
            'view_notifications': 'צפה בהודעות',
            'login_to_view': 'התחבר ל-TheraSocial כדי לצפות בכל ההודעות שלך.',
            'regards': 'בברכה',
            'team': 'צוות TheraSocial'
        },
        'ar': {
            'subject': 'TheraSocial - إشعار جديد',
            'hello': 'مرحبا',
            'new_notification': 'لديك إشعار جديد:',
            'view_notifications': 'عرض الإشعارات',
            'login_to_view': 'سجل الدخول إلى TheraSocial لعرض جميع إشعاراتك.',
            'regards': 'مع أطيب التحيات',
            'team': 'فريق TheraSocial'
        },
        'ru': {
            'subject': 'TheraSocial - Новое уведомление',
            'hello': 'Здравствуйте',
            'new_notification': 'У вас новое уведомление:',
            'view_notifications': 'Просмотреть уведомления',
            'login_to_view': 'Войдите в TheraSocial, чтобы просмотреть все уведомления.',
            'regards': 'С наилучшими пожеланиями',
            'team': 'Команда TheraSocial'
        }
    }
    return translations.get(language, translations['en'])


def send_notification_email(user_email, notification_title, notification_content, user_language='en'):
    """PJ6001: Send email for notifications (messages, followers, invites) - separate from alerts"""
    try:
        t = get_notification_email_translations(user_language)
        app_url = os.environ.get('APP_URL', 'http://localhost:5000')

        # PJ6003: Translate notification title and content if they contain translation keys
        translated_title, translated_content = translate_alert_content(notification_title, notification_content, user_language)
        notification_title = translated_title
        notification_content = translated_content

        is_rtl = user_language in ['he', 'ar']
        text_dir = 'rtl' if is_rtl else 'ltr'
        text_align = 'right' if is_rtl else 'left'

        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
        </head>
        <body style="font-family: Arial, sans-serif; direction: {text_dir}; text-align: {text_align}; background-color: #f5f5f5; margin: 0; padding: 20px;">
            <div style="max-width: 600px; margin: 0 auto; background: white; border-radius: 8px; overflow: hidden; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
                <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); padding: 30px; text-align: center;">
                    <h1 style="color: white; margin: 0; font-size: 24px;">TheraSocial</h1>
                </div>
                <div style="padding: 40px 30px;">
                    <p style="color: #666; line-height: 1.6;">{t['hello']},</p>
                    <h2 style="color: #667eea; margin-top: 0;">{t['new_notification']}</h2>
                    <div style="background: #f8f9fa; padding: 15px; border-radius: 8px; margin: 20px 0; border-left: 4px solid #667eea;">
                        <p style="color: #333; font-weight: 600; margin: 0 0 8px 0;">{notification_title}</p>
                        <p style="color: #666; margin: 0;">{notification_content}</p>
                    </div>
                    <p style="color: #666; line-height: 1.6;">{t['login_to_view']}</p>
                    <div style="text-align: center; margin: 30px 0;">
                        <a href="{app_url}" style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 14px 40px; text-decoration: none; border-radius: 25px; display: inline-block; font-weight: bold; box-shadow: 0 4px 15px rgba(102, 126, 234, 0.4);">
                            {t['view_notifications']}
                        </a>
                    </div>
                </div>
                <div style="background: #f8f9fa; padding: 20px 30px; border-top: 1px solid #eee;">
                    <p style="color: #999; font-size: 12px; margin: 0;">
                        {t['regards']},<br>{t['team']}
                    </p>
                </div>
            </div>
        </body>
        </html>
        """

        try:
            logger.info(f"[SMTP NOTIFICATION] Sending notification email to {user_email}...")
            
            msg = MIMEMultipart('alternative')
            msg['Subject'] = t['subject']
            msg['From'] = app.config['MAIL_DEFAULT_SENDER']
            msg['To'] = user_email
            msg.attach(MIMEText(html_content, 'html'))

            smtp_server = os.environ.get('SMTP_SERVER', 'smtp.resend.com')
            smtp_port = int(os.environ.get('SMTP_PORT', '465'))
            smtp_user = os.environ.get('SMTP_USERNAME', 'resend')
            smtp_pass = os.environ.get('SMTP_PASSWORD', app.config.get('MAIL_PASSWORD', ''))
            
            with smtplib.SMTP_SSL(smtp_server, smtp_port) as server:
                server.login(smtp_user, smtp_pass)
                server.sendmail(msg['From'], user_email, msg.as_string())
            
            logger.info(f'[SMTP NOTIFICATION] Notification email sent to {user_email}')
            return True
        except Exception as e:
            logger.error(f'[SMTP NOTIFICATION] Failed to send notification email: {str(e)}')
            return False

    except Exception as e:
        logger.error(f"[SMTP NOTIFICATION] Failed to send notification email: {e}")
        return False


def create_notification_with_email(user_id, title, content, alert_type='info', source_user_id=None, alert_category='notification'):
    """
    PJ6001: Create a notification and optionally send email if user has email_on_notification enabled.
    Used for messages, new followers, invites - NOT wellness alerts.
    
    Args:
        user_id: The user ID to create notification for
        title: Notification title
        content: Notification content/message
        alert_type: Type of notification
        source_user_id: ID of user this notification is about
        alert_category: Category: 'message', 'follow', 'invite', 'notification'
    
    Returns:
        The created Alert object (stored in alerts table but shown in notifications list)
    """
    logger.info(f"[NOTIFICATION EMAIL] ========================================")
    logger.info(f"[NOTIFICATION EMAIL] create_notification_with_email called")
    logger.info(f"[NOTIFICATION EMAIL] user_id: {user_id}, alert_category: {alert_category}")
    
    try:
        # Create the notification (stored in alerts table with notification category)
        alert = Alert(
            user_id=user_id,
            title=title,
            content=content,
            alert_type=alert_type,
            source_user_id=source_user_id,
            alert_category=alert_category
        )
        db.session.add(alert)
        db.session.flush()
        logger.info(f"[NOTIFICATION EMAIL] Notification created with ID: {alert.id}")
        
        # Check if user has email_on_notification enabled
        try:
            settings = NotificationSettings.query.filter_by(user_id=user_id).first()
            
            # Check email_on_notification setting (default to True if not set)
            email_enabled = settings.email_on_notification if settings and hasattr(settings, 'email_on_notification') else True
            
            if email_enabled:
                user = db.session.get(User, user_id)
                
                if user and user.email:
                    # Parse notification title for email
                    email_title = title
                    try:
                        title_data = json.loads(title)
                        if isinstance(title_data, dict) and 'key' in title_data:
                            username = title_data.get('params', {}).get('username', '')
                            key = title_data.get('key', '')
                            if 'new_message' in key:
                                email_title = f"New message from {username}"
                            elif 'started_following' in key:
                                email_title = f"{username} started following you"
                            elif 'invitation' in key.lower():
                                email_title = "New invitation"
                            else:
                                email_title = username or 'New Notification'
                    except:
                        pass
                    
                    logger.info(f"[NOTIFICATION EMAIL] Sending notification email to {user.email}")
                    user_language = user.preferred_language or 'en'
                    result = send_notification_email(user.email, email_title, content or '', user_language)
                    logger.info(f"[NOTIFICATION EMAIL] Email send result: {result}")
            else:
                logger.info(f"[NOTIFICATION EMAIL] Skipping email - email_on_notification is disabled")
                
        except Exception as email_err:
            logger.error(f"[NOTIFICATION EMAIL] Error sending notification email: {str(email_err)}")
        
        logger.info(f"[NOTIFICATION EMAIL] ========================================")
        return alert
    except Exception as e:
        logger.error(f"[NOTIFICATION EMAIL] Error creating notification: {str(e)}")
        raise


def ensure_notification_settings_schema():
    """Ensure notification_settings table has all required columns - runs on startup"""
    # Guard: Skip if already run in this process
    if hasattr(ensure_notification_settings_schema, '_completed'):
        return

    try:
        with app.app_context():
            # Check if table exists
            inspector = inspect(db.engine)
            if 'notification_settings' not in inspector.get_table_names():
                logger.info("notification_settings table doesn't exist yet, will be created by migrations")
                return

            # Get existing columns
            existing_columns = {col['name'] for col in inspector.get_columns('notification_settings')}
            logger.info(f"Existing columns in notification_settings: {existing_columns}")

            # Check if we're using PostgreSQL or SQLite
            is_postgres = 'postgresql' in str(db.engine.url)

            # Define required columns with their types
            required_columns = {
                'email_on_alert': 'BOOLEAN DEFAULT FALSE',
                'email_on_notification': 'BOOLEAN DEFAULT TRUE',  # PJ6001: Email for notifications
                'email_daily_diary_reminder': 'BOOLEAN DEFAULT FALSE',
                'email_on_new_message': 'BOOLEAN DEFAULT TRUE'
            }

            # Add missing columns
            missing_columns = set(required_columns.keys()) - existing_columns

            if missing_columns:
                logger.info(f"Adding missing columns to notification_settings: {missing_columns}")

                with db.engine.connect() as connection:
                    for column_name in missing_columns:
                        column_type = required_columns[column_name]

                        if is_postgres:
                            # PostgreSQL syntax with IF NOT EXISTS
                            alter_query = text(
                                f"ALTER TABLE notification_settings ADD COLUMN IF NOT EXISTS {column_name} {column_type}")
                        else:
                            # SQLite syntax (no IF NOT EXISTS)
                            alter_query = text(
                                f"ALTER TABLE notification_settings ADD COLUMN {column_name} {column_type}")

                        try:
                            connection.execute(alter_query)
                            connection.commit()
                            logger.info(f"Added column to notification_settings: {column_name}")
                        except Exception as e:
                            logger.error(f"Error adding column {column_name}: {e}")

                logger.info("Successfully added all missing columns to notification_settings")
            else:
                logger.info("All required columns exist in notification_settings")

        # Mark as completed for this process
        ensure_notification_settings_schema._completed = True

    except Exception as e:
        logger.error(f"Error ensuring notification_settings schema: {str(e)}")
        # Don't raise - allow app to start even if this fails


def ensure_saved_parameters_schema():
    """Ensure saved_parameters table has all required columns - runs on startup"""
    # Guard: Skip if already run in this process
    if hasattr(ensure_saved_parameters_schema, '_completed'):
        return

    try:
        with app.app_context():
            # Check if table exists
            inspector = inspect(db.engine)
            if 'saved_parameters' not in inspector.get_table_names():
                logger.info("saved_parameters table doesn't exist yet, will be created by migrations")
                return

            # Get existing columns
            existing_columns = {col['name'] for col in inspector.get_columns('saved_parameters')}
            logger.debug(f"Existing columns in saved_parameters: {existing_columns}")  # Changed to debug

            # Check if we're using PostgreSQL or SQLite
            is_postgres = 'postgresql' in str(db.engine.url)

            # Define required columns with their types
            required_columns = {
                'sleep_quality': 'INTEGER',
                'physical_activity': 'INTEGER',
                'anxiety': 'INTEGER',
                'updated_at': 'TIMESTAMP DEFAULT CURRENT_TIMESTAMP'
            }

            # Privacy columns (will be added if missing, with private default)
            privacy_columns = {
                'mood_privacy': 'VARCHAR(20) DEFAULT \'private\'',
                'energy_privacy': 'VARCHAR(20) DEFAULT \'private\'',
                'sleep_quality_privacy': 'VARCHAR(20) DEFAULT \'private\'',
                'physical_activity_privacy': 'VARCHAR(20) DEFAULT \'private\'',
                'anxiety_privacy': 'VARCHAR(20) DEFAULT \'private\''
            }

            # Combine all required columns
            all_required = {**required_columns, **privacy_columns}

            # Add missing columns
            missing_columns = set(all_required.keys()) - existing_columns

            if missing_columns:
                logger.info(f"Adding missing columns to saved_parameters: {missing_columns}")

                with db.engine.connect() as connection:
                    for column_name in missing_columns:
                        column_type = all_required[column_name]

                        if is_postgres:
                            # PostgreSQL syntax with IF NOT EXISTS
                            alter_query = text(
                                f"ALTER TABLE saved_parameters ADD COLUMN IF NOT EXISTS {column_name} {column_type}")
                        else:
                            # SQLite syntax (no IF NOT EXISTS)
                            alter_query = text(
                                f"ALTER TABLE saved_parameters ADD COLUMN {column_name} {column_type}")

                        try:
                            connection.execute(alter_query)
                            connection.commit()
                            logger.info(f"Added column: {column_name}")
                        except Exception as e:
                            # Column might already exist in SQLite (which doesn't support IF NOT EXISTS)
                            logger.debug(f"Column {column_name} might already exist: {e}")

                logger.info("Successfully added all missing columns to saved_parameters")
            else:
                logger.debug("All required columns exist in saved_parameters")  # Changed to debug

        # Mark as completed for this process
        ensure_saved_parameters_schema._completed = True

    except Exception as e:
        logger.error(f"Error ensuring saved_parameters schema: {str(e)}")
        # Don't raise - allow app to start even if this fails


# Initialize Redis client (optional, for caching)
try:
    redis_client = redis.from_url(REDIS_URL) if REDIS_URL else None
    if redis_client:
        redis_client.ping()
        logger.info("Redis connected successfully")
except Exception as e:
    redis_client = None
    logger.warning(f"Redis not available: {e}")


def parse_date_as_local(date_string):
    """Parse date string as local date without timezone conversion"""
    from datetime import datetime
    try:
        # This ensures the date is treated as-is without timezone shifts
        return datetime.strptime(date_string, '%Y-%m-%d').date()
    except ValueError:
        # Fallback for other formats
        return datetime.fromisoformat(date_string.split('T')[0]).date()


def get_db():
    """Get a direct database connection for raw SQL queries"""
    import psycopg2
    from psycopg2.extras import RealDictCursor

    # Get the database URI from SQLAlchemy config
    db_uri = app.config['SQLALCHEMY_DATABASE_URI']

    # For PostgreSQL (Render uses PostgreSQL)
    if db_uri.startswith('postgresql://') or db_uri.startswith('postgres://'):
        # Fix postgres:// to postgresql:// if needed
        if db_uri.startswith('postgres://'):
            db_uri = db_uri.replace('postgres://', 'postgresql://', 1)

        try:
            conn = psycopg2.connect(db_uri, cursor_factory=RealDictCursor)
            return conn
        except Exception as e:
            logger.error(f"Failed to connect to PostgreSQL: {e}")
            raise

    # For SQLite (local development)
    elif db_uri.startswith('sqlite:///'):
        import sqlite3
        db_path = db_uri.replace('sqlite:///', '')
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        return conn

    else:
        raise ValueError(f"Unsupported database type: {db_uri}")

def auto_migrate_database():
    """Automatically update database schema on startup"""
    logger.info("Starting auto_migrate_database...")
    with app.app_context():
        try:
            # Create all tables first
            db.create_all()

            # Now add missing columns
            from sqlalchemy import inspect, text
            inspector = inspect(db.engine)

            with db.engine.connect() as conn:
                # Check if we're using PostgreSQL or SQLite
                is_postgres = 'postgresql' in str(db.engine.url)

                # Update users table
                if 'users' in inspector.get_table_names():
                    columns = [col['name'] for col in inspector.get_columns('users')]

                    if 'has_completed_onboarding' not in columns:
                        if is_postgres:
                            conn.execute(
                                text("ALTER TABLE users ADD COLUMN has_completed_onboarding BOOLEAN DEFAULT FALSE"))
                        else:
                            conn.execute(
                                text("ALTER TABLE users ADD COLUMN has_completed_onboarding INTEGER DEFAULT 0"))
                        conn.commit()

                    if 'onboarding_dismissed' not in columns:
                        if is_postgres:
                            conn.execute(
                                text("ALTER TABLE users ADD COLUMN onboarding_dismissed BOOLEAN DEFAULT FALSE"))
                        else:
                            conn.execute(text("ALTER TABLE users ADD COLUMN onboarding_dismissed INTEGER DEFAULT 0"))
                        conn.commit()

                    if 'shareable_link_token' not in columns:
                        conn.execute(text("ALTER TABLE users ADD COLUMN shareable_link_token VARCHAR(100)"))
                        conn.commit()

                # Update parameters/saved_parameters table
                params_table = 'saved_parameters' if 'saved_parameters' in inspector.get_table_names() else 'parameters'
                if params_table in inspector.get_table_names():
                    columns = [col['name'] for col in inspector.get_columns(params_table)]

                    privacy_columns = ['mood_privacy', 'energy_privacy', 'sleep_quality_privacy',
                                       'physical_activity_privacy', 'anxiety_privacy']

                    for col in privacy_columns:
                        if col not in columns:
                            conn.execute(
                                text(f"ALTER TABLE {params_table} ADD COLUMN {col} VARCHAR(20) DEFAULT 'public'"))
                            conn.commit()

                # Create password_reset_tokens table if it doesn't exist
                if 'password_reset_tokens' not in inspector.get_table_names():
                    logger.info("Creating password_reset_tokens table...")
                    if is_postgres:
                        conn.execute(text("""
                            CREATE TABLE password_reset_tokens (
                                id SERIAL PRIMARY KEY,
                                user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                                token VARCHAR(100) UNIQUE NOT NULL,
                                expires_at TIMESTAMP NOT NULL,
                                used BOOLEAN DEFAULT FALSE,
                                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                            )
                        """))
                    else:
                        conn.execute(text("""
                            CREATE TABLE password_reset_tokens (
                                id INTEGER PRIMARY KEY AUTOINCREMENT,
                                user_id INTEGER NOT NULL,
                                token VARCHAR(100) UNIQUE NOT NULL,
                                expires_at TIMESTAMP NOT NULL,
                                used INTEGER DEFAULT 0,
                                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                                FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
                            )
                        """))
                    conn.commit()
                    logger.info("✓ Created password_reset_tokens table")

                # Create magic_login_tokens table
                if 'magic_login_tokens' not in inspector.get_table_names():
                    logger.info("Creating magic_login_tokens table...")
                    if is_postgres:
                        conn.execute(text("""
                            CREATE TABLE magic_login_tokens (
                                id SERIAL PRIMARY KEY,
                                user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                                token VARCHAR(100) UNIQUE NOT NULL,
                                expires_at TIMESTAMP NOT NULL,
                                used BOOLEAN DEFAULT FALSE,
                                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                            )
                        """))
                    else:
                        conn.execute(text("""
                            CREATE TABLE magic_login_tokens (
                                id INTEGER PRIMARY KEY AUTOINCREMENT,
                                user_id INTEGER NOT NULL,
                                token VARCHAR(100) UNIQUE NOT NULL,
                                expires_at TIMESTAMP NOT NULL,
                                used INTEGER DEFAULT 0,
                                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                                FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
                            )
                        """))
                    conn.commit()
                    logger.info("✓ Created magic_login_tokens table")

                # Create user_consents table
                if 'user_consents' not in inspector.get_table_names():
                    logger.info("Creating user_consents table...")
                    if is_postgres:
                        conn.execute(text("""
                            CREATE TABLE user_consents (
                                id SERIAL PRIMARY KEY,
                                user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                                email_updates BOOLEAN DEFAULT FALSE,
                                privacy_accepted BOOLEAN DEFAULT FALSE,
                                research_data BOOLEAN DEFAULT FALSE,
                                team_declaration BOOLEAN DEFAULT FALSE,
                                responsible_use BOOLEAN DEFAULT FALSE,
                                waiver_claims BOOLEAN DEFAULT FALSE,
                                consent_language VARCHAR(5) DEFAULT 'en',
                                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                            )
                        """))
                    else:
                        conn.execute(text("""
                            CREATE TABLE user_consents (
                                id INTEGER PRIMARY KEY AUTOINCREMENT,
                                user_id INTEGER NOT NULL,
                                email_updates INTEGER DEFAULT 0,
                                privacy_accepted INTEGER DEFAULT 0,
                                research_data INTEGER DEFAULT 0,
                                team_declaration INTEGER DEFAULT 0,
                                responsible_use INTEGER DEFAULT 0,
                                waiver_claims INTEGER DEFAULT 0,
                                consent_language VARCHAR(5) DEFAULT 'en',
                                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                                FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
                            )
                        """))
                    conn.commit()
                    logger.info("✓ Created user_consents table")

                # Create blocked_users table
                if 'blocked_users' not in inspector.get_table_names():
                    logger.info("Creating blocked_users table...")
                    if is_postgres:
                        conn.execute(text("""
                            CREATE TABLE blocked_users (
                                id SERIAL PRIMARY KEY,
                                blocker_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                                blocked_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                                UNIQUE(blocker_id, blocked_id)
                            )
                        """))
                    else:
                        conn.execute(text("""
                            CREATE TABLE blocked_users (
                                id INTEGER PRIMARY KEY AUTOINCREMENT,
                                blocker_id INTEGER NOT NULL,
                                blocked_id INTEGER NOT NULL,
                                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                                UNIQUE(blocker_id, blocked_id),
                                FOREIGN KEY(blocker_id) REFERENCES users(id) ON DELETE CASCADE,
                                FOREIGN KEY(blocked_id) REFERENCES users(id) ON DELETE CASCADE
                            )
                        """))
                    conn.commit()
                    logger.info("✓ Created blocked_users table")

                # Add diary reminder columns to notification_settings table
                if 'notification_settings' in inspector.get_table_names():
                    columns = [col['name'] for col in inspector.get_columns('notification_settings')]
                    if 'diary_reminder_time' not in columns:
                        logger.info("Adding diary_reminder_time column to notification_settings table...")
                        conn.execute(text("ALTER TABLE notification_settings ADD COLUMN diary_reminder_time VARCHAR(5) DEFAULT '09:00'"))
                        conn.commit()
                        logger.info("✓ Added diary_reminder_time column to notification_settings table")
                    if 'diary_reminder_timezone' not in columns:
                        logger.info("Adding diary_reminder_timezone column to notification_settings table...")
                        conn.execute(text("ALTER TABLE notification_settings ADD COLUMN diary_reminder_timezone VARCHAR(100) DEFAULT 'UTC'"))
                        conn.commit()
                        logger.info("✓ Added diary_reminder_timezone column to notification_settings table")

                # Add follow_note column to follows table
                if 'follows' in inspector.get_table_names():
                    columns = [col['name'] for col in inspector.get_columns('follows')]
                    if 'follow_note' not in columns:
                        logger.info("Adding follow_note column to follows table...")
                        conn.execute(text("ALTER TABLE follows ADD COLUMN follow_note VARCHAR(300)"))
                        conn.commit()
                        logger.info("✓ Added follow_note column to follows table")

                # Create or update parameter_triggers table
                if 'parameter_triggers' not in inspector.get_table_names():
                    logger.info("Creating parameter_triggers table...")
                    if is_postgres:
                        conn.execute(text("""
                            CREATE TABLE parameter_triggers (
                                id SERIAL PRIMARY KEY,
                                watcher_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                                watched_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                                mood_alert BOOLEAN DEFAULT FALSE,
                                energy_alert BOOLEAN DEFAULT FALSE,
                                sleep_alert BOOLEAN DEFAULT FALSE,
                                physical_alert BOOLEAN DEFAULT FALSE,
                                anxiety_alert BOOLEAN DEFAULT FALSE,
                                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                                UNIQUE(watcher_id, watched_id)
                            )
                        """))
                    else:
                        conn.execute(text("""
                            CREATE TABLE parameter_triggers (
                                id INTEGER PRIMARY KEY AUTOINCREMENT,
                                watcher_id INTEGER NOT NULL,
                                watched_id INTEGER NOT NULL,
                                mood_alert INTEGER DEFAULT 0,
                                energy_alert INTEGER DEFAULT 0,
                                sleep_alert INTEGER DEFAULT 0,
                                physical_alert INTEGER DEFAULT 0,
                                anxiety_alert INTEGER DEFAULT 0,
                                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                                FOREIGN KEY(watcher_id) REFERENCES users(id) ON DELETE CASCADE,
                                FOREIGN KEY(watched_id) REFERENCES users(id) ON DELETE CASCADE,
                                UNIQUE(watcher_id, watched_id)
                            )
                        """))
                    conn.commit()
                    logger.info("✓ Created parameter_triggers table")
                else:
                    # Table exists, check for missing columns
                    columns = [col['name'] for col in inspector.get_columns('parameter_triggers')]

                    alert_columns = ['mood_alert', 'energy_alert', 'sleep_alert',
                                     'physical_alert', 'anxiety_alert']

                    for col in alert_columns:
                        if col not in columns:
                            logger.info(f"Adding {col} column to parameter_triggers table...")
                            if is_postgres:
                                conn.execute(
                                    text(f"ALTER TABLE parameter_triggers ADD COLUMN {col} BOOLEAN DEFAULT FALSE"))
                            else:
                                conn.execute(text(f"ALTER TABLE parameter_triggers ADD COLUMN {col} INTEGER DEFAULT 0"))
                            conn.commit()
                            logger.info(f"✓ Added {col} column to parameter_triggers table")

                # Auto-migration: Remove posts circle_id foreign key constraint
                if 'posts' in inspector.get_table_names():
                    try:
                        # Check if the constraint exists (PostgreSQL only)
                        if is_postgres:
                            result = conn.execute(text("""
                                SELECT constraint_name 
                                FROM information_schema.table_constraints 
                                WHERE table_name = 'posts' 
                                AND constraint_name = 'posts_circle_id_fkey'
                            """))

                            if result.fetchone():
                                logger.info("Removing posts_circle_id_fkey constraint...")
                                conn.execute(text('ALTER TABLE posts DROP CONSTRAINT posts_circle_id_fkey'))
                                conn.commit()
                                logger.info("✓ Removed posts_circle_id_fkey constraint")
                    except Exception as e:
                        logger.warning(f"Posts constraint removal skipped: {e}")

                    # Clean up posts with invalid circle_id
                    try:
                        result = conn.execute(text("""
                            UPDATE posts 
                            SET circle_id = NULL 
                            WHERE circle_id IS NOT NULL 
                            AND circle_id NOT IN (SELECT id FROM circles)
                        """))
                        conn.commit()
                        rows_updated = result.rowcount
                        if rows_updated > 0:
                            logger.info(f"✓ Cleaned up {rows_updated} posts with invalid circle_id")
                    except Exception as e:
                        logger.warning(f"Posts cleanup skipped: {e}")

                    # Migrate old posts to use visibility field instead of circle_id
                    try:
                        result = conn.execute(text("""
                            UPDATE posts 
                            SET visibility = CASE 
                                WHEN circle_id = 1 THEN 'general'
                                WHEN circle_id = 2 THEN 'close_friends'
                                WHEN circle_id = 3 THEN 'family'
                                ELSE 'private'
                            END
                            WHERE visibility IS NULL OR visibility = ''
                        """))
                        conn.commit()
                        rows_updated = result.rowcount
                        if rows_updated > 0:
                            logger.info(f"✓ Migrated {rows_updated} posts to use visibility field")
                    except Exception as e:
                        logger.warning(f"Posts visibility migration skipped: {e}")

                logger.info("Database auto-migration completed successfully")

        except Exception as e:
            logger.warning(f"Auto-migration error (may be normal if columns exist): {e}")


# Call auto-migration on startup
auto_migrate_database()


# =====================
# DATABASE MODELS
# =====================
class User(db.Model):
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False, index=True)
    email = db.Column(db.String(120), unique=True, nullable=False, index=True)
    password_hash = db.Column(db.String(255), nullable=False)
    role = db.Column(db.String(50), default='user')
    is_active = db.Column(db.Boolean, default=True)
    preferred_language = db.Column(db.String(5), default='en')
    selected_city = db.Column(db.String(100), default='Jerusalem, Israel')
    # ADD THESE THREE NEW FIELDS:
    has_completed_onboarding = db.Column(db.Boolean, default=False)
    onboarding_dismissed = db.Column(db.Boolean, default=False)
    shareable_link_token = db.Column(db.String(100), unique=True)
    circles_privacy = db.Column(db.String(20), default='private')
    birth_year = db.Column(db.Integer, default=1985)  # PJ6001: Birth year field
    created_at = db.Column(db.DateTime, default=datetime.utcnow, index=True)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    last_login = db.Column(db.DateTime)

    # Keep ALL existing relationships
    profile = db.relationship('Profile', backref='user', uselist=False, cascade='all, delete-orphan')
    sent_messages = db.relationship('Message', foreign_keys='Message.sender_id', backref='sender',
                                    cascade='all, delete-orphan')
    received_messages = db.relationship('Message', foreign_keys='Message.recipient_id', backref='recipient',
                                        cascade='all, delete-orphan')
    circles = db.relationship('Circle', foreign_keys='Circle.user_id', backref='owner', cascade='all, delete-orphan')
    saved_parameters = db.relationship('SavedParameters', backref='user', cascade='all, delete-orphan')
    posts = db.relationship('Post', backref='author', cascade='all, delete-orphan')
    alerts = db.relationship('Alert', foreign_keys='Alert.user_id', backref='user', cascade='all, delete-orphan')
    activities = db.relationship('Activity', backref='user', cascade='all, delete-orphan')

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

    # KEEP THESE METHODS - THEY'RE ESSENTIAL!
    def follow(self, user, note=None):
        """Follow another user with optional note"""
        if not self.is_following(user):
            follow = Follow(follower_id=self.id, followed_id=user.id, follow_note=note)
            db.session.add(follow)

    def unfollow(self, user):
        """Unfollow a user"""
        follow = Follow.query.filter_by(
            follower_id=self.id,
            followed_id=user.id
        ).first()
        if follow:
            db.session.delete(follow)

    def is_following(self, user):
        """Check if following a user"""
        return Follow.query.filter_by(
            follower_id=self.id,
            followed_id=user.id
        ).first() is not None

    def block_user(self, user):
        """Block another user"""
        if not self.has_blocked(user):
            from sqlalchemy import select
            block = BlockedUser(blocker_id=self.id, blocked_id=user.id)
            db.session.add(block)
            return True
        return False

    def unblock_user(self, user):
        """Unblock a user"""
        block = BlockedUser.query.filter_by(
            blocker_id=self.id,
            blocked_id=user.id
        ).first()
        if block:
            db.session.delete(block)
            return True
        return False

    def has_blocked(self, user):
        """Check if this user has blocked another user"""
        return BlockedUser.query.filter_by(
            blocker_id=self.id,
            blocked_id=user.id
        ).first() is not None

    def is_blocked_by(self, user):
        """Check if this user is blocked by another user"""
        return BlockedUser.query.filter_by(
            blocker_id=user.id,
            blocked_id=self.id
        ).first() is not None

    def to_dict(self):
        return {
            'id': self.id,
            'username': self.username,
            'email': self.email,
            'role': self.role,
            'is_active': self.is_active,
            'preferred_language': self.preferred_language or 'en',
            'created_at': self.created_at.isoformat() if self.created_at else None,
            'last_login': self.last_login.isoformat() if self.last_login else None,
            # ADD THESE TO to_dict():
            'has_completed_onboarding': self.has_completed_onboarding,
            'shareable_link_token': self.shareable_link_token,
            'circles_privacy': self.circles_privacy or 'private',
            'birth_year': self.birth_year or 1985  # PJ6001: Birth year field
        }


class PasswordResetToken(db.Model):
    __tablename__ = 'password_reset_tokens'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    token = db.Column(db.String(100), unique=True, nullable=False)
    expires_at = db.Column(db.DateTime, nullable=False)
    used = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    user = db.relationship('User', backref='reset_tokens')


class MagicLoginToken(db.Model):
    __tablename__ = 'magic_login_tokens'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    token = db.Column(db.String(100), unique=True, nullable=False)
    expires_at = db.Column(db.DateTime, nullable=False)
    used = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    user_relation = db.relationship('User', backref='magic_tokens')


class UserConsent(db.Model):
    __tablename__ = 'user_consents'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    email_updates = db.Column(db.Boolean, default=False)
    privacy_accepted = db.Column(db.Boolean, default=False)
    research_data = db.Column(db.Boolean, default=False)
    team_declaration = db.Column(db.Boolean, default=False)
    responsible_use = db.Column(db.Boolean, default=False)
    waiver_claims = db.Column(db.Boolean, default=False)
    consent_language = db.Column(db.String(5), default='en')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    user_relation = db.relationship('User', backref='consent', uselist=False)


class ParameterTrigger(db.Model):
    __tablename__ = 'parameter_triggers'
    id = db.Column(db.Integer, primary_key=True)
    watcher_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    watched_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)

    # New style columns (what you actually use)
    mood_alert = db.Column(db.Boolean, default=False)
    energy_alert = db.Column(db.Boolean, default=False)
    sleep_alert = db.Column(db.Boolean, default=False)
    physical_alert = db.Column(db.Boolean, default=False)
    anxiety_alert = db.Column(db.Boolean, default=False)

    # Old style columns (keep for compatibility with existing DB)
    parameter_name = db.Column(db.String(50), nullable=True)
    trigger_condition = db.Column(db.String(50), nullable=True)
    trigger_value = db.Column(db.Float, nullable=True)
    consecutive_days = db.Column(db.Integer, nullable=True)
    is_active = db.Column(db.Boolean, default=True)
    last_triggered = db.Column(db.DateTime, nullable=True)

    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    watcher = db.relationship('User', foreign_keys=[watcher_id], backref='watching_triggers')
    watched = db.relationship('User', foreign_keys=[watched_id], backref='watched_by_triggers')

    __table_args__ = (db.UniqueConstraint('watcher_id', 'watched_id', name='unique_trigger'),)


class Profile(db.Model):
    __tablename__ = 'profiles'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), unique=True)
    bio = db.Column(db.Text)
    interests = db.Column(db.Text)
    occupation = db.Column(db.String(200))
    goals = db.Column(db.Text)
    favorite_hobbies = db.Column(db.Text)
    avatar_url = db.Column(db.String(500))
    mood_status = db.Column(db.String(50))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class Post(db.Model):
    __tablename__ = 'posts'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False, index=True)
    content = db.Column(db.Text, nullable=False)
    image_url = db.Column(db.String(500))
    likes = db.Column(db.Integer, default=0)
    circle_id = db.Column(db.Integer, nullable=True)
    visibility = db.Column(db.String(50), default='general')
    is_published = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, index=True)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    # Relationships
    comments = db.relationship('Comment', backref='post', cascade='all, delete-orphan')
    reactions = db.relationship('Reaction', backref='post', cascade='all, delete-orphan')


class Comment(db.Model):
    __tablename__ = 'comments'
    id = db.Column(db.Integer, primary_key=True)
    post_id = db.Column(db.Integer, db.ForeignKey('posts.id'), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    content = db.Column(db.Text, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    author = db.relationship('User', backref='comments')


class Reaction(db.Model):
    __tablename__ = 'reactions'
    id = db.Column(db.Integer, primary_key=True)
    post_id = db.Column(db.Integer, db.ForeignKey('posts.id'), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    type = db.Column(db.String(20), nullable=False)  # like, love, support, etc.
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    __table_args__ = (
        db.UniqueConstraint('post_id', 'user_id', name='unique_post_reaction'),
    )


class Circle(db.Model):
    __tablename__ = 'circles'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), index=True)
    circle_user_id = db.Column(db.Integer, db.ForeignKey('users.id'), index=True)
    circle_type = db.Column(db.String(50))  # 'public', 'class_b', 'class_a'
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    circle_user = db.relationship('User', foreign_keys=[circle_user_id])

    __table_args__ = (
        db.UniqueConstraint('user_id', 'circle_user_id', 'circle_type', name='_user_circle_uc'),
    )


class Message(db.Model):
    __tablename__ = 'messages'
    id = db.Column(db.Integer, primary_key=True)
    sender_id = db.Column(db.Integer, db.ForeignKey('users.id'), index=True)
    recipient_id = db.Column(db.Integer, db.ForeignKey('users.id'), index=True)
    content = db.Column(db.Text)
    is_read = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, index=True)


class SavedParameters(db.Model):
    __tablename__ = 'saved_parameters'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), index=True)
    date = db.Column(db.String(10), index=True)  # String date like '2025-11-03'
    mood = db.Column(db.Integer)
    energy = db.Column(db.Integer)
    sleep_quality = db.Column(db.Integer)
    physical_activity = db.Column(db.Integer)
    anxiety = db.Column(db.Integer)
    mood_privacy = db.Column(db.String(20), default='private')
    energy_privacy = db.Column(db.String(20), default='private')
    sleep_quality_privacy = db.Column(db.String(20), default='private')
    physical_activity_privacy = db.Column(db.String(20), default='private')
    anxiety_privacy = db.Column(db.String(20), default='private')
    notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow)
    # REMOVED: privacy = db.Column(db.JSON)  # This line should be removed/commented

    __table_args__ = (db.UniqueConstraint('user_id', 'date', name='_user_date_uc'),)

    def to_dict(self, viewer_id=None, privacy_level=None):
        base_dict = {
            'id': self.id,
            'date': self.date,
            'created_at': self.created_at.isoformat() if self.created_at else None,
            'updated_at': self.updated_at.isoformat() if self.updated_at else None
        }

        if viewer_id == self.user_id:
            base_dict.update({
                'mood': self.mood,
                'energy': self.energy,
                'sleep_quality': self.sleep_quality,
                'physical_activity': self.physical_activity,
                'anxiety': self.anxiety,
                'mood_privacy': self.mood_privacy,
                'energy_privacy': self.energy_privacy,
                'sleep_quality_privacy': self.sleep_quality_privacy,
                'physical_activity_privacy': self.physical_activity_privacy,
                'anxiety_privacy': self.anxiety_privacy,
                'notes': self.notes
            })
        else:
            for param in ['mood', 'energy', 'sleep_quality', 'physical_activity', 'anxiety']:
                param_privacy = getattr(self, f"{param}_privacy", 'public')
                # Only show if public, or if viewer has proper access level (NOT private)
                if param_privacy == 'public' or \
                        (param_privacy == 'class_b' and privacy_level in ['class_b', 'class_a']) or \
                        (param_privacy == 'class_a' and privacy_level == 'class_a'):
                    # Note: 'private' params are excluded - only owner can see
                    base_dict[param] = getattr(self, param)

        return base_dict


class Alert(db.Model):
    __tablename__ = 'alerts'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), index=True)
    title = db.Column(db.String(200))
    content = db.Column(db.Text)  # FIXED: Changed from 'message' to 'content'
    alert_type = db.Column(db.String(50))  # 'info', 'warning', 'success', 'error'
    is_read = db.Column(db.Boolean, default=False, index=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, index=True)
    # PJ401: Source user ID for filtering alerts based on following status
    # Alerts with source_user_id only show if current user follows source_user
    source_user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True, index=True)
    # Alert category for filtering: 'trigger', 'feed', 'message', 'follow', 'general'
    alert_category = db.Column(db.String(50), default='general')
    
    # PJ401: Relationship to source user (the user this alert is about)
    source_user = db.relationship('User', foreign_keys=[source_user_id], backref='triggered_alerts')

    def to_dict(self):
        return {
            'id': self.id,
            'title': self.title,
            'message': self.content,  # Return as 'message' for backward compatibility
            'type': self.alert_type,  # Map alert_type to type for API compatibility
            'is_read': self.is_read,
            'created_at': self.created_at.isoformat() if self.created_at else None,
            'source_user_id': self.source_user_id,
            'alert_category': self.alert_category
        }


class Activity(db.Model):
    """Store activity feed data by date for calendar functionality"""
    __tablename__ = 'activities'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    activity_date = db.Column(db.Date, nullable=False)
    post_count = db.Column(db.Integer, default=0)
    comment_count = db.Column(db.Integer, default=0)
    message_count = db.Column(db.Integer, default=0)
    mood_entries = db.Column(db.JSON, default=list)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    __table_args__ = (
        db.UniqueConstraint('user_id', 'activity_date', name='unique_user_activity_date'),
    )


class Follow(db.Model):
    __tablename__ = 'follows'
    id = db.Column(db.Integer, primary_key=True)
    follower_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    followed_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    follow_note = db.Column(db.String(300))  # New field for follow notes
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    follower = db.relationship('User', foreign_keys=[follower_id], backref='following')
    followed = db.relationship('User', foreign_keys=[followed_id], backref='followers')

    __table_args__ = (db.UniqueConstraint('follower_id', 'followed_id', name='unique_follow'),)


class FollowRequest(db.Model):
    __tablename__ = 'follow_requests'
    id = db.Column(db.Integer, primary_key=True)
    requester_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    target_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    privacy_level = db.Column(db.String(20), default='public')
    status = db.Column(db.String(20), default='pending')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    responded_at = db.Column(db.DateTime)

    requester = db.relationship('User', foreign_keys=[requester_id], backref='sent_follow_requests')
    target = db.relationship('User', foreign_keys=[target_id], backref='received_follow_requests')

    __table_args__ = (db.UniqueConstraint('requester_id', 'target_id', name='unique_follow_request'),)


class NotificationSettings(db.Model):
    __tablename__ = 'notification_settings'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), unique=True, nullable=False)
    follow_requests = db.Column(db.Boolean, default=True)
    parameter_triggers = db.Column(db.Boolean, default=True)
    daily_reminder = db.Column(db.Boolean, default=False)
    weekly_summary = db.Column(db.Boolean, default=False)
    # New email notification settings
    email_on_alert = db.Column(db.Boolean, default=False)  # Email for wellness alerts only
    email_on_notification = db.Column(db.Boolean, default=True)  # PJ6001: Email for notifications (messages, followers, invites)
    email_daily_diary_reminder = db.Column(db.Boolean, default=False)  # Daily reminder to fill diary
    email_on_new_message = db.Column(db.Boolean, default=True)  # Email on new message (default True)
    # Daily diary reminder time settings (24-hour format, e.g., "09:00" or "21:30")
    diary_reminder_time = db.Column(db.String(5), default='09:00')  # Default 9 AM
    diary_reminder_timezone = db.Column(db.String(100), default='UTC')  # Timezone based on selected city
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    user = db.relationship('User', backref=db.backref('notification_settings', uselist=False))


class BlockedUser(db.Model):
    """Model for tracking blocked users"""
    __tablename__ = 'blocked_users'
    id = db.Column(db.Integer, primary_key=True)
    blocker_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False, index=True)
    blocked_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False, index=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    blocker = db.relationship('User', foreign_keys=[blocker_id], backref='blocked_by_me')
    blocked = db.relationship('User', foreign_keys=[blocked_id], backref='blocked_me')

    __table_args__ = (db.UniqueConstraint('blocker_id', 'blocked_id', name='unique_block'),)


class BackgroundJob(db.Model):
    """
    Database-backed job queue for reliable background processing.
    Replaces simple threading with persistent, retry-capable jobs.
    
    Benefits over threading:
    - Jobs survive server restarts
    - Automatic retries on failure
    - Prevents thread exhaustion under load
    - Auditable job history
    """
    __tablename__ = 'background_jobs'
    id = db.Column(db.Integer, primary_key=True)
    job_type = db.Column(db.String(50), nullable=False, index=True)  # 'trigger_processing', 'send_email', etc.
    payload = db.Column(db.JSON, nullable=False)  # Job-specific data
    status = db.Column(db.String(20), default='pending', index=True)  # pending, processing, completed, failed
    priority = db.Column(db.Integer, default=0)  # Higher = processed first
    attempts = db.Column(db.Integer, default=0)
    max_attempts = db.Column(db.Integer, default=3)
    error_message = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, index=True)
    started_at = db.Column(db.DateTime)
    completed_at = db.Column(db.DateTime)
    
    # For job locking to prevent double-processing
    locked_by = db.Column(db.String(100))  # Worker ID
    locked_at = db.Column(db.DateTime)
    
    def __repr__(self):
        return f'<BackgroundJob {self.id} {self.job_type} {self.status}>'


def ensure_database_schema():
    """Automatically ensure all required columns exist"""
    # Guard: Skip if already run in this process
    if hasattr(ensure_database_schema, '_completed'):
        return

    try:
        with db.engine.connect() as conn:
            # Check if we're using PostgreSQL or SQLite
            is_postgres = 'postgresql' in str(db.engine.url)

            if is_postgres:
                # PostgreSQL - Check and add visibility column to posts table
                result = conn.execute(text("""
                    SELECT column_name
                    FROM information_schema.columns 
                    WHERE table_name = 'posts' 
                    AND table_schema = 'public'
                """))
                columns = [row[0] for row in result]

                if 'visibility' not in columns:
                    logger.info("Adding visibility column to posts table...")
                    conn.execute(text("ALTER TABLE posts ADD COLUMN visibility VARCHAR(50) DEFAULT 'general'"))
                    conn.commit()
                    logger.info("Visibility column added successfully")

                # PostgreSQL - Check and add circles_privacy column to users table
                result = conn.execute(text("""
                    SELECT column_name
                    FROM information_schema.columns 
                    WHERE table_name = 'users' 
                    AND table_schema = 'public'
                """))
                user_columns = [row[0] for row in result]

                if 'circles_privacy' not in user_columns:
                    logger.info("Adding circles_privacy column to users table...")
                    conn.execute(text("ALTER TABLE users ADD COLUMN circles_privacy VARCHAR(20) DEFAULT 'private'"))
                    conn.commit()
                    logger.info("circles_privacy column added successfully")
                else:
                    logger.debug("✓ circles_privacy column already exists")  # Changed to debug

            else:
                # SQLite - Check and add visibility column to posts table
                result = conn.execute(text("PRAGMA table_info(posts)"))
                columns = [row[1] for row in result]

                if 'visibility' not in columns:
                    logger.info("Adding visibility column to posts table...")
                    conn.execute(text("ALTER TABLE posts ADD COLUMN visibility VARCHAR(50) DEFAULT 'general'"))
                    conn.commit()
                    logger.info("Visibility column added successfully")

                # SQLite - Check and add circles_privacy column to users table
                result = conn.execute(text("PRAGMA table_info(users)"))
                user_columns = [row[1] for row in result]

                if 'circles_privacy' not in user_columns:
                    logger.info("Adding circles_privacy column to users table...")
                    conn.execute(text("ALTER TABLE users ADD COLUMN circles_privacy VARCHAR(20) DEFAULT 'private'"))
                    conn.commit()
                    logger.info("circles_privacy column added successfully")
                else:
                    logger.debug("✓ circles_privacy column already exists")  # Changed to debug

        # Mark as completed for this process
        ensure_database_schema._completed = True

    except Exception as e:
        logger.error(f"Database schema check error: {str(e)}")


def ensure_background_jobs_schema():
    """Ensure the background_jobs table exists for the job queue system"""
    try:
        with db.engine.connect() as conn:
            is_postgres = 'postgresql' in str(db.engine.url)
            
            if is_postgres:
                # Check if table exists
                result = conn.execute(text("""
                    SELECT EXISTS (
                        SELECT FROM information_schema.tables 
                        WHERE table_schema = 'public' 
                        AND table_name = 'background_jobs'
                    )
                """))
                exists = result.scalar()
                
                if not exists:
                    logger.info("[JOB QUEUE] Creating background_jobs table...")
                    conn.execute(text("""
                        CREATE TABLE background_jobs (
                            id SERIAL PRIMARY KEY,
                            job_type VARCHAR(50) NOT NULL,
                            payload JSON NOT NULL,
                            status VARCHAR(20) DEFAULT 'pending',
                            priority INTEGER DEFAULT 0,
                            attempts INTEGER DEFAULT 0,
                            max_attempts INTEGER DEFAULT 3,
                            error_message TEXT,
                            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                            started_at TIMESTAMP,
                            completed_at TIMESTAMP,
                            locked_by VARCHAR(100),
                            locked_at TIMESTAMP
                        )
                    """))
                    conn.execute(text("CREATE INDEX idx_bg_jobs_status ON background_jobs(status)"))
                    conn.execute(text("CREATE INDEX idx_bg_jobs_type ON background_jobs(job_type)"))
                    conn.execute(text("CREATE INDEX idx_bg_jobs_created ON background_jobs(created_at)"))
                    conn.commit()
                    logger.info("[JOB QUEUE] background_jobs table created successfully")
                else:
                    logger.debug("[JOB QUEUE] background_jobs table already exists")
            else:
                # SQLite
                result = conn.execute(text("""
                    SELECT name FROM sqlite_master 
                    WHERE type='table' AND name='background_jobs'
                """))
                exists = result.fetchone() is not None
                
                if not exists:
                    logger.info("[JOB QUEUE] Creating background_jobs table...")
                    conn.execute(text("""
                        CREATE TABLE background_jobs (
                            id INTEGER PRIMARY KEY AUTOINCREMENT,
                            job_type VARCHAR(50) NOT NULL,
                            payload JSON NOT NULL,
                            status VARCHAR(20) DEFAULT 'pending',
                            priority INTEGER DEFAULT 0,
                            attempts INTEGER DEFAULT 0,
                            max_attempts INTEGER DEFAULT 3,
                            error_message TEXT,
                            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                            started_at TIMESTAMP,
                            completed_at TIMESTAMP,
                            locked_by VARCHAR(100),
                            locked_at TIMESTAMP
                        )
                    """))
                    conn.commit()
                    logger.info("[JOB QUEUE] background_jobs table created successfully")
                    
    except Exception as e:
        logger.error(f"[JOB QUEUE] Schema error: {str(e)}")


# =====================
# DATABASE INITIALIZATION
# =====================

def init_database():
    """Initialize database with migrations and fixes"""
    with app.app_context():
        try:
            # Check if database exists and has tables
            inspector = inspect(db.engine)
            tables = inspector.get_table_names()

            if not tables:
                logger.info("No tables found, creating database schema...")
                db.create_all()
                ensure_database_schema()
                ensure_saved_parameters_schema()  # ← ADDED
                ensure_notification_settings_schema()  # ← ADDED for email notification columns
                ensure_background_jobs_schema()  # ← ADDED for job queue
                logger.info("Database schema created successfully")
                create_admin_user()
                create_test_users()
                create_test_follows()
                create_parameters_table()
            else:
                logger.info(f"Found {len(tables)} existing tables")

                # Fix all schema issues
                fix_all_schema_issues()
                ensure_database_schema()
                ensure_saved_parameters_schema()  # ← ADDED
                ensure_notification_settings_schema()  # ← ADDED for email notification columns
                ensure_background_jobs_schema()  # ← ADDED for job queue
                create_test_users()
                create_test_follows()
                create_parameters_table()

                # Only try migrations if migrations folder exists
                if os.path.exists('migrations'):
                    try:
                        from flask_migrate import upgrade
                        logger.info("Checking for pending migrations...")
                        upgrade()
                        logger.info("Database migrations applied successfully")
                    except Exception as e:
                        logger.warning(f"Migration error (non-critical): {e}")
                        logger.info("Using existing database schema")
                else:
                    logger.info("No migrations folder found, using existing schema")

            # Verify database connection
            db.session.execute(select(func.count()).select_from(User))
            db.session.commit()
            logger.info("Database connection verified")

            # Run one-time cleanup of stale trigger alerts
            try:
                removed = cleanup_all_stale_trigger_alerts()
                logger.info(f"Startup cleanup: Removed {removed} stale trigger alerts")
            except Exception as cleanup_err:
                logger.warning(f"Cleanup warning (non-critical): {cleanup_err}")

            # Start background diary reminder scheduler
            try:
                start_diary_reminder_scheduler()
            except Exception as scheduler_err:
                logger.warning(f"Diary scheduler warning (non-critical): {scheduler_err}")

            # PJ816: Start background trigger scheduler (emails without login)
            try:
                start_trigger_scheduler()
            except Exception as scheduler_err:
                logger.warning(f"Trigger scheduler warning (non-critical): {scheduler_err}")

            # Start background job queue scheduler
            try:
                start_job_queue_scheduler()
            except Exception as scheduler_err:
                logger.warning(f"Job queue scheduler warning (non-critical): {scheduler_err}")

        except Exception as e:
            logger.error(f"Database initialization error: {e}")
            # Try to create tables as fallback
            try:
                db.create_all()
                logger.info("Created database tables as fallback")
                ensure_saved_parameters_schema()  # ← ADDED
                ensure_notification_settings_schema()  # ← ADDED for email notification columns
                ensure_background_jobs_schema()  # ← ADDED for job queue
                create_admin_user()
                create_test_users()
                create_test_follows()
                create_parameters_table()
                # Start background diary reminder scheduler
                try:
                    start_diary_reminder_scheduler()
                except Exception as scheduler_err:
                    logger.warning(f"Diary scheduler warning (non-critical): {scheduler_err}")
                # PJ816: Start background trigger scheduler (emails without login)
                try:
                    start_trigger_scheduler()
                except Exception as scheduler_err:
                    logger.warning(f"Trigger scheduler warning (non-critical): {scheduler_err}")
                # Start background job queue scheduler
                try:
                    start_job_queue_scheduler()
                except Exception as scheduler_err:
                    logger.warning(f"Job queue scheduler warning (non-critical): {scheduler_err}")
            except Exception as e2:
                logger.error(f"Failed to create tables: {e2}")
                if not is_production:
                    raise


def fix_all_schema_issues():
    """Fix all known database schema issues"""
    try:
        with db.engine.connect() as conn:
            # Check if we're using PostgreSQL or SQLite
            is_postgres = 'postgresql' in str(db.engine.url)
            
            # PJ401: Add source_user_id and alert_category columns to alerts table
            inspector = inspect(db.engine)
            if 'alerts' in inspector.get_table_names():
                alert_columns = [col['name'] for col in inspector.get_columns('alerts')]
                
                if 'source_user_id' not in alert_columns:
                    logger.info("Adding source_user_id column to alerts table...")
                    if is_postgres:
                        conn.execute(text("ALTER TABLE alerts ADD COLUMN IF NOT EXISTS source_user_id INTEGER REFERENCES users(id)"))
                    else:
                        conn.execute(text("ALTER TABLE alerts ADD COLUMN source_user_id INTEGER"))
                    conn.commit()
                    logger.info("✓ Added source_user_id column to alerts")
                
                if 'alert_category' not in alert_columns:
                    logger.info("Adding alert_category column to alerts table...")
                    if is_postgres:
                        conn.execute(text("ALTER TABLE alerts ADD COLUMN IF NOT EXISTS alert_category VARCHAR(50) DEFAULT 'general'"))
                    else:
                        conn.execute(text("ALTER TABLE alerts ADD COLUMN alert_category VARCHAR(50) DEFAULT 'general'"))
                    conn.commit()
                    logger.info("✓ Added alert_category column to alerts")

            # 1. Fix alerts table (message -> content)
            try:
                if is_postgres:
                    result = conn.execute(text(
                        """SELECT column_name 
                        FROM information_schema.columns 
                        WHERE table_name='alerts' 
                        AND column_name IN ('message', 'content')"""
                    ))
                    columns = [row[0] for row in result]

                    if 'message' in columns and 'content' not in columns:
                        logger.info("Renaming alerts.message to alerts.content...")
                        conn.execute(text("ALTER TABLE alerts RENAME COLUMN message TO content"))
                        conn.commit()
                        logger.info("✓ Fixed alerts.message column")
                    elif 'content' not in columns and 'message' not in columns:
                        logger.info("Adding missing content column...")
                        conn.execute(text("ALTER TABLE alerts ADD COLUMN content TEXT"))
                        conn.commit()
                        logger.info("✓ Added alerts.content column")
                    else:
                        logger.info("✓ Alerts table schema is correct")
                else:
                    # SQLite handling
                    result = conn.execute(text("PRAGMA table_info(alerts)"))
                    columns = [row[1] for row in result]

                    if 'message' in columns and 'content' not in columns:
                        logger.info("Migrating alerts table for SQLite...")
                        conn.execute(text("""
                            CREATE TABLE alerts_new (
                                id INTEGER PRIMARY KEY,
                                user_id INTEGER,
                                title VARCHAR(200),
                                content TEXT,
                                type VARCHAR(50),
                                is_read BOOLEAN DEFAULT 0,
                                created_at DATETIME,
                                FOREIGN KEY(user_id) REFERENCES users(id)
                            )
                        """))
                        conn.execute(text("""
                            INSERT INTO alerts_new (id, user_id, title, content, type, is_read, created_at)
                            SELECT id, user_id, title, message, type, is_read, created_at FROM alerts
                        """))
                        conn.execute(text("DROP TABLE alerts"))
                        conn.execute(text("ALTER TABLE alerts_new RENAME TO alerts"))
                        conn.commit()
                        logger.info("✓ Migrated alerts table schema")
            except Exception as e:
                logger.warning(f"Could not fix alerts table: {e}")

            # 2. Fix circles table - ensure circle_user_id exists
            try:
                if is_postgres:
                    result = conn.execute(text(
                        """SELECT column_name 
                        FROM information_schema.columns 
                        WHERE table_name='circles'"""
                    ))
                    existing_columns = [row[0] for row in result]

                    if existing_columns:  # Table exists
                        if 'circle_user_id' not in existing_columns:
                            logger.info("Adding missing circle_user_id column to circles table...")
                            conn.execute(text("""
                                ALTER TABLE circles 
                                ADD COLUMN circle_user_id INTEGER 
                                REFERENCES users(id) ON DELETE CASCADE
                            """))
                            conn.commit()
                            logger.info("✓ Added circle_user_id column to circles table")
                        else:
                            logger.info("✓ Circles table has circle_user_id column")
                else:
                    # SQLite
                    result = conn.execute(text("PRAGMA table_info(circles)"))
                    existing_columns = [row[1] for row in result]

                    if existing_columns and 'circle_user_id' not in existing_columns:
                        logger.info("Recreating circles table for SQLite with circle_user_id...")
                        conn.execute(text("""
                            CREATE TABLE circles_new (
                                id INTEGER PRIMARY KEY,
                                user_id INTEGER NOT NULL,
                                circle_user_id INTEGER NOT NULL,
                                circle_type VARCHAR(50),
                                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                                UNIQUE(user_id, circle_user_id, circle_type),
                                FOREIGN KEY(user_id) REFERENCES users(id),
                                FOREIGN KEY(circle_user_id) REFERENCES users(id)
                            )
                        """))
                        conn.execute(text("""
                            INSERT INTO circles_new (id, user_id, circle_type, created_at)
                            SELECT id, user_id, circle_type, created_at FROM circles
                        """))
                        conn.execute(text("DROP TABLE circles"))
                        conn.execute(text("ALTER TABLE circles_new RENAME TO circles"))
                        conn.commit()
                        logger.info("✓ Recreated circles table with circle_user_id")

            except Exception as e:
                logger.warning(f"Could not fix circles table: {e}")

            # 3. Fix profiles table - add missing columns
            try:
                if is_postgres:
                    result = conn.execute(text(
                        """SELECT column_name 
                        FROM information_schema.columns 
                        WHERE table_name='profiles'"""
                    ))
                    existing_columns = [row[0] for row in result]
                else:
                    # SQLite
                    result = conn.execute(text("PRAGMA table_info(profiles)"))
                    existing_columns = [row[1] for row in result]

                # Define all columns that should exist in profiles table
                required_columns = [
                    ('mood_status', 'VARCHAR(50)'),
                    ('avatar_url', 'VARCHAR(500)'),
                    ('interests', 'TEXT'),
                    ('occupation', 'VARCHAR(200)'),
                    ('goals', 'TEXT'),
                    ('favorite_hobbies', 'TEXT')
                ]

                for col_name, col_type in required_columns:
                    if col_name not in existing_columns:
                        logger.info(f"Adding profiles.{col_name} column...")
                        conn.execute(text(f"ALTER TABLE profiles ADD COLUMN {col_name} {col_type}"))
                        conn.commit()
                        logger.info(f"✓ Added profiles.{col_name} column")

            except Exception as e:
                logger.warning(f"Could not fix profiles table: {e}")

            # 4. Ensure activities table exists
            try:
                if is_postgres:
                    result = conn.execute(text(
                        """SELECT table_name 
                        FROM information_schema.tables 
                        WHERE table_name='activities'"""
                    ))
                    table_exists = result.fetchone() is not None
                else:
                    # SQLite
                    result = conn.execute(text(
                        "SELECT name FROM sqlite_master WHERE type='table' AND name='activities'"
                    ))
                    table_exists = result.fetchone() is not None

                if not table_exists:
                    logger.info("Creating activities table...")
                    if is_postgres:
                        conn.execute(text("""
                            CREATE TABLE activities (
                                id SERIAL PRIMARY KEY,
                                user_id INTEGER NOT NULL REFERENCES users(id),
                                activity_date DATE NOT NULL,
                                post_count INTEGER DEFAULT 0,
                                comment_count INTEGER DEFAULT 0,
                                message_count INTEGER DEFAULT 0,
                                mood_entries JSON,
                                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                                UNIQUE(user_id, activity_date)
                            )
                        """))
                    else:
                        # SQLite version
                        conn.execute(text("""
                            CREATE TABLE activities (
                                id INTEGER PRIMARY KEY AUTOINCREMENT,
                                user_id INTEGER NOT NULL,
                                activity_date DATE NOT NULL,
                                post_count INTEGER DEFAULT 0,
                                comment_count INTEGER DEFAULT 0,
                                message_count INTEGER DEFAULT 0,
                                mood_entries TEXT,
                                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                                UNIQUE(user_id, activity_date),
                                FOREIGN KEY(user_id) REFERENCES users(id)
                            )
                        """))
                    conn.commit()
                    logger.info("✓ Created activities table")
                else:
                    logger.info("✓ Activities table already exists")

            except Exception as e:
                logger.warning(f"Could not create activities table: {e}")

            # 5. Ensure comments table exists
            try:
                if is_postgres:
                    result = conn.execute(text(
                        """SELECT table_name 
                        FROM information_schema.tables 
                        WHERE table_name='comments'"""
                    ))
                    table_exists = result.fetchone() is not None
                else:
                    result = conn.execute(text(
                        "SELECT name FROM sqlite_master WHERE type='table' AND name='comments'"
                    ))
                    table_exists = result.fetchone() is not None

                if not table_exists:
                    logger.info("Creating comments table...")
                    if is_postgres:
                        conn.execute(text("""
                            CREATE TABLE comments (
                                id SERIAL PRIMARY KEY,
                                post_id INTEGER NOT NULL REFERENCES posts(id),
                                user_id INTEGER NOT NULL REFERENCES users(id),
                                content TEXT NOT NULL,
                                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                            )
                        """))
                    else:
                        conn.execute(text("""
                            CREATE TABLE comments (
                                id INTEGER PRIMARY KEY AUTOINCREMENT,
                                post_id INTEGER NOT NULL,
                                user_id INTEGER NOT NULL,
                                content TEXT NOT NULL,
                                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                                FOREIGN KEY(post_id) REFERENCES posts(id),
                                FOREIGN KEY(user_id) REFERENCES users(id)
                            )
                        """))
                    conn.commit()
                    logger.info("✓ Created comments table")

            except Exception as e:
                logger.warning(f"Could not create comments table: {e}")

            # 6. Ensure reactions table exists
            try:
                if is_postgres:
                    result = conn.execute(text(
                        """SELECT table_name 
                        FROM information_schema.tables 
                        WHERE table_name='reactions'"""
                    ))
                    table_exists = result.fetchone() is not None
                else:
                    result = conn.execute(text(
                        "SELECT name FROM sqlite_master WHERE type='table' AND name='reactions'"
                    ))
                    table_exists = result.fetchone() is not None

                if not table_exists:
                    logger.info("Creating reactions table...")
                    if is_postgres:
                        conn.execute(text("""
                            CREATE TABLE reactions (
                                id SERIAL PRIMARY KEY,
                                post_id INTEGER NOT NULL REFERENCES posts(id),
                                user_id INTEGER NOT NULL REFERENCES users(id),
                                type VARCHAR(20) NOT NULL,
                                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                                UNIQUE(post_id, user_id)
                            )
                        """))
                    else:
                        conn.execute(text("""
                            CREATE TABLE reactions (
                                id INTEGER PRIMARY KEY AUTOINCREMENT,
                                post_id INTEGER NOT NULL,
                                user_id INTEGER NOT NULL,
                                type VARCHAR(20) NOT NULL,
                                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                                UNIQUE(post_id, user_id),
                                FOREIGN KEY(post_id) REFERENCES posts(id),
                                FOREIGN KEY(user_id) REFERENCES users(id)
                            )
                        """))
                    conn.commit()
                    logger.info("✓ Created reactions table")

            except Exception as e:
                logger.warning(f"Could not create reactions table: {e}")

            # 7. CRITICAL FIX: Handle posts table with encrypted columns
            try:
                if is_postgres:
                    # Check what columns exist
                    result = conn.execute(text(
                        """SELECT column_name, is_nullable
                        FROM information_schema.columns 
                        WHERE table_name='posts'"""
                    ))
                    column_info = {row[0]: row[1] for row in result}

                    # If encrypted columns exist and are NOT NULL, make them nullable
                    encrypted_cols = ['content_encrypted', 'image_url_encrypted']
                    for col in encrypted_cols:
                        if col in column_info and column_info[col] == 'NO':
                            logger.info(f"Making {col} nullable...")
                            conn.execute(text(f"ALTER TABLE posts ALTER COLUMN {col} DROP NOT NULL"))
                            conn.commit()
                            logger.info(f"✓ Made {col} nullable")

                    # Add missing plain columns
                    required_columns = [
                        ('content', 'TEXT'),
                        ('image_url', 'VARCHAR(500)'),
                        ('likes', 'INTEGER DEFAULT 0'),
                        ('circle_id', 'INTEGER'),
                        ('is_published', 'BOOLEAN DEFAULT TRUE'),
                        ('updated_at', 'TIMESTAMP DEFAULT CURRENT_TIMESTAMP')
                    ]

                    for col_name, col_type in required_columns:
                        if col_name not in column_info:
                            logger.info(f"Adding {col_name} column to posts...")
                            conn.execute(text(f"ALTER TABLE posts ADD COLUMN {col_name} {col_type}"))
                            conn.commit()
                            logger.info(f"✓ Added {col_name} column")
                else:
                    # SQLite - check and add columns
                    result = conn.execute(text("PRAGMA table_info(posts)"))
                    existing_columns = [row[1] for row in result]

                    required_columns = [
                        ('content', 'TEXT'),
                        ('image_url', 'VARCHAR(500)'),
                        ('likes', 'INTEGER DEFAULT 0'),
                        ('circle_id', 'INTEGER'),
                        ('is_published', 'BOOLEAN DEFAULT 1'),
                        ('updated_at', 'TIMESTAMP DEFAULT CURRENT_TIMESTAMP')
                    ]

                    for col_name, col_type in required_columns:
                        if col_name not in existing_columns:
                            logger.info(f"Adding {col_name} column to posts...")
                            conn.execute(text(f"ALTER TABLE posts ADD COLUMN {col_name} {col_type}"))
                            conn.commit()
                            logger.info(f"✓ Added {col_name} column")

            except Exception as e:
                logger.warning(f"Could not fix posts table: {e}")

            # 8. Fix alerts table - ensure type column exists
            try:
                if is_postgres:
                    result = conn.execute(text(
                        """SELECT column_name 
                        FROM information_schema.columns 
                        WHERE table_name='alerts'"""
                    ))
                    existing_columns = [row[0] for row in result]

                    if existing_columns and 'type' not in existing_columns:
                        logger.info("Adding missing type column to alerts table...")
                        conn.execute(text("ALTER TABLE alerts ADD COLUMN type VARCHAR(50) DEFAULT 'info'"))
                        conn.commit()
                        logger.info("✓ Added type column to alerts table")
                else:
                    # SQLite
                    result = conn.execute(text("PRAGMA table_info(alerts)"))
                    existing_columns = [row[1] for row in result]

                    if existing_columns and 'type' not in existing_columns:
                        logger.info("Adding type column to alerts table...")
                        conn.execute(text("ALTER TABLE alerts ADD COLUMN type VARCHAR(50) DEFAULT 'info'"))
                        conn.commit()
                        logger.info("✓ Added type column to alerts table")

            except Exception as e:
                logger.warning(f"Could not add type column to alerts table: {e}")

            # 9. Add preferred_language column to users table
            try:
                if is_postgres:
                    result = conn.execute(text(
                        """SELECT column_name 
                        FROM information_schema.columns 
                        WHERE table_name='users'"""
                    ))
                    existing_columns = [row[0] for row in result]

                    if existing_columns and 'preferred_language' not in existing_columns:
                        logger.info("Adding missing preferred_language column to users table...")
                        conn.execute(text("ALTER TABLE users ADD COLUMN preferred_language VARCHAR(5) DEFAULT 'en'"))
                        conn.commit()
                        logger.info("✓ Added preferred_language column to users table")
                    elif 'preferred_language' in existing_columns:
                        logger.info("✓ Users table already has preferred_language column")
                else:
                    # SQLite
                    result = conn.execute(text("PRAGMA table_info(users)"))
                    existing_columns = [row[1] for row in result]

                    if existing_columns and 'preferred_language' not in existing_columns:
                        logger.info("Adding preferred_language column to users table...")
                        conn.execute(text("ALTER TABLE users ADD COLUMN preferred_language VARCHAR(5) DEFAULT 'en'"))
                        conn.commit()
                        logger.info("✓ Added preferred_language column to users table")
                    elif 'preferred_language' in existing_columns:
                        logger.info("✓ Users table already has preferred_language column")

            except Exception as e:
                logger.warning(f"Could not add preferred_language column to users table: {e}")

            try:
                if is_postgres:
                    result = conn.execute(text(
                        """SELECT column_name 
                        FROM information_schema.columns 
                        WHERE table_name='users'"""
                    ))
                    existing_columns = [row[0] for row in result]

                    if existing_columns and 'selected_city' not in existing_columns:
                        logger.info("Adding missing selected_city column to users table...")
                        conn.execute(
                            text("ALTER TABLE users ADD COLUMN selected_city VARCHAR(100) DEFAULT 'Jerusalem, Israel'"))
                        conn.commit()
                        logger.info("✓ Added selected_city column to users table")
                    elif 'selected_city' in existing_columns:
                        logger.info("✓ Users table already has selected_city column")

                    # PJ6001: Add birth_year column to users table (PostgreSQL)
                    if existing_columns and 'birth_year' not in existing_columns:
                        logger.info("Adding birth_year column to users table (PostgreSQL)...")
                        conn.execute(
                            text("ALTER TABLE users ADD COLUMN birth_year INTEGER DEFAULT 1985"))
                        conn.commit()
                        logger.info("✓ Added birth_year column to users table")
                    elif 'birth_year' in existing_columns:
                        logger.info("✓ Users table already has birth_year column")
                else:
                    # SQLite
                    result = conn.execute(text("PRAGMA table_info(users)"))
                    existing_columns = [row[1] for row in result]

                    if existing_columns and 'selected_city' not in existing_columns:
                        logger.info("Adding selected_city column to users table...")
                        conn.execute(
                            text("ALTER TABLE users ADD COLUMN selected_city VARCHAR(100) DEFAULT 'Jerusalem, Israel'"))
                        conn.commit()
                        logger.info("✓ Added selected_city column to users table")
                    elif 'selected_city' in existing_columns:
                        logger.info("✓ Users table already has selected_city column")

                    # PJ6001: Add birth_year column to users table (SQLite)
                    if existing_columns and 'birth_year' not in existing_columns:
                        logger.info("Adding birth_year column to users table (SQLite)...")
                        conn.execute(
                            text("ALTER TABLE users ADD COLUMN birth_year INTEGER DEFAULT 1985"))
                        conn.commit()
                        logger.info("✓ Added birth_year column to users table")
                    elif 'birth_year' in existing_columns:
                        logger.info("✓ Users table already has birth_year column")

            except Exception as e:
                logger.warning(f"Could not add selected_city/birth_year column to users table: {e}")

            logger.info("✓ All schema fixes complete")

    except Exception as e:
        logger.error(f"Error in fix_all_schema_issues: {e}")
        # Don't fail the entire initialization for schema fixes
        pass


def create_test_users():
    """Create 12 test users for the social app"""
    try:
        logger.info("Checking for test users...")

        test_users = [
            {'username': 'alice', 'email': 'alice@example.com', 'password': 'password123',
             'bio': 'Love hiking and photography', 'interests': 'Photography, Hiking, Travel',
             'occupation': 'Photographer', 'goals': 'Travel to 50 countries', 'hobbies': 'Reading, Yoga'},

            {'username': 'bob', 'email': 'bob@example.com', 'password': 'password123',
             'bio': 'Software developer and gamer', 'interests': 'Gaming, Programming, Tech',
             'occupation': 'Software Engineer', 'goals': 'Build a successful startup', 'hobbies': 'Gaming, Coding'},

            {'username': 'charlie', 'email': 'charlie@example.com', 'password': 'password123',
             'bio': 'Chef and food enthusiast', 'interests': 'Cooking, Food, Wine',
             'occupation': 'Chef', 'goals': 'Open my own restaurant', 'hobbies': 'Cooking, Wine tasting'},

            {'username': 'diana', 'email': 'diana@example.com', 'password': 'password123',
             'bio': 'Artist and creative soul', 'interests': 'Art, Music, Dance',
             'occupation': 'Graphic Designer', 'goals': 'Have an art exhibition', 'hobbies': 'Painting, Dancing'},

            {'username': 'edward', 'email': 'edward@example.com', 'password': 'password123',
             'bio': 'Fitness coach and athlete', 'interests': 'Fitness, Sports, Nutrition',
             'occupation': 'Personal Trainer', 'goals': 'Complete an Ironman', 'hobbies': 'Running, Swimming'},

            {'username': 'fiona', 'email': 'fiona@example.com', 'password': 'password123',
             'bio': 'Teacher and bookworm', 'interests': 'Education, Literature, History',
             'occupation': 'High School Teacher', 'goals': 'Write a novel', 'hobbies': 'Reading, Writing'},

            {'username': 'george', 'email': 'george@example.com', 'password': 'password123',
             'bio': 'Musician and composer', 'interests': 'Music, Guitar, Jazz',
             'occupation': 'Music Teacher', 'goals': 'Record an album', 'hobbies': 'Guitar, Piano'},

            {'username': 'helen', 'email': 'helen@example.com', 'password': 'password123',
             'bio': 'Entrepreneur and innovator', 'interests': 'Business, Marketing, Innovation',
             'occupation': 'Marketing Manager', 'goals': 'Launch a successful product',
             'hobbies': 'Networking, Reading'},

            {'username': 'ivan', 'email': 'ivan@example.com', 'password': 'password123',
             'bio': 'Doctor and health advocate', 'interests': 'Medicine, Health, Research',
             'occupation': 'Physician', 'goals': 'Contribute to medical research', 'hobbies': 'Tennis, Chess'},

            {'username': 'julia', 'email': 'julia@example.com', 'password': 'password123',
             'bio': 'Environmental scientist', 'interests': 'Environment, Science, Sustainability',
             'occupation': 'Environmental Consultant', 'goals': 'Make a positive environmental impact',
             'hobbies': 'Gardening, Hiking'},

            {'username': 'kevin', 'email': 'kevin@example.com', 'password': 'password123',
             'bio': 'Film director and storyteller', 'interests': 'Film, Cinema, Storytelling',
             'occupation': 'Video Producer', 'goals': 'Direct a feature film', 'hobbies': 'Photography, Film'},

            {'username': 'laura', 'email': 'laura@example.com', 'password': 'password123',
             'bio': 'Psychologist and mindfulness coach', 'interests': 'Psychology, Mindfulness, Wellness',
             'occupation': 'Clinical Psychologist', 'goals': 'Help 1000 people improve their mental health',
             'hobbies': 'Meditation, Yoga'}
        ]

        created_count = 0
        for user_data in test_users:
            # Check if user exists
            existing_user = User.query.filter_by(email=user_data['email']).first()
            if not existing_user:
                # Create user
                user = User(
                    username=user_data['username'],
                    email=user_data['email']
                )
                user.set_password(user_data['password'])
                db.session.add(user)
                db.session.flush()  # Get the user ID

                # Create profile
                profile = Profile(
                    user_id=user.id,
                    bio=user_data['bio'],
                    interests=user_data['interests'],
                    occupation=user_data['occupation'],
                    goals=user_data['goals'],
                    favorite_hobbies=user_data['hobbies']
                )
                db.session.add(profile)
                created_count += 1
                logger.info(f"Created test user: {user_data['username']}")

        if created_count > 0:
            db.session.commit()
            logger.info(f"✓ Created {created_count} test users")
        else:
            logger.info("✓ Test users already exist")

    except Exception as e:
        logger.error(f"Error creating test users: {e}")
        db.session.rollback()


def create_test_follows():
    """Create follow relationships between test users and main user"""
    try:
        logger.info("Setting up test follow relationships...")

        # Get the main user (emaskanazi_1)
        main_user = User.query.filter_by(username='emaskanazi_1').first()
        if not main_user:
            logger.info("Main user emaskanazi_1 not found, skipping test follows")
            return

        # Get test users
        test_usernames = ['alice', 'bob', 'charlie', 'diana', 'edward', 'fiona',
                          'george', 'helen', 'ivan', 'julia', 'kevin', 'laura']

        test_users = User.query.filter(User.username.in_(test_usernames)).all()

        if not test_users:
            logger.info("No test users found, skipping test follows")
            return

        # Count existing follows
        existing_follows = Follow.query.filter_by(followed_id=main_user.id).count()

        if existing_follows >= len(test_users):
            logger.info(f"✓ Test follows already exist ({existing_follows} followers)")
            return

        # Create follows: each test user follows the main user
        created_count = 0
        for test_user in test_users:
            # Check if follow already exists
            existing = Follow.query.filter_by(
                follower_id=test_user.id,
                followed_id=main_user.id
            ).first()

            if not existing:
                follow = Follow(
                    follower_id=test_user.id,
                    followed_id=main_user.id
                )
                db.session.add(follow)
                created_count += 1

        # Also make main user follow some test users back
        for test_user in test_users[:6]:  # Follow back half of them
            existing = Follow.query.filter_by(
                follower_id=main_user.id,
                followed_id=test_user.id
            ).first()

            if not existing:
                follow = Follow(
                    follower_id=main_user.id,
                    followed_id=test_user.id
                )
                db.session.add(follow)

        db.session.commit()
        logger.info(f"✓ Created {created_count} new follow relationships")

        # Verify
        follower_count = Follow.query.filter_by(followed_id=main_user.id).count()
        following_count = Follow.query.filter_by(follower_id=main_user.id).count()
        logger.info(f"Main user now has {follower_count} followers and is following {following_count} users")

    except Exception as e:
        logger.error(f"Error creating test follows: {e}")
        db.session.rollback()


def create_parameters_table():
    """Create parameters table if it doesn't exist with correct schema"""
    try:
        logger.info("Checking parameters table...")

        conn = get_db()
        cursor = conn.cursor()

        # Check if parameters table exists and has correct columns
        try:
            cursor.execute("""
                SELECT column_name 
                FROM information_schema.columns 
                WHERE table_name = 'parameters'
                ORDER BY ordinal_position
            """)
            columns = [row['column_name'] for row in cursor.fetchall()]

            # Check if we have the mood/energy columns (our actual schema)
            required_columns = ['date', 'mood', 'energy', 'sleep_quality', 'user_id']

            if columns and all(col in columns for col in required_columns):
                logger.info("✓ Parameters table exists with correct schema")
                cursor.close()
                conn.close()
                return
            elif columns:
                # Table exists but has wrong schema
                logger.info(f"Parameters table has wrong columns: {columns}")
                logger.info("Attempting to drop and recreate parameters table...")

                # Try to drop with timeout
                cursor.execute("SET statement_timeout = '5s'")  # 5 second timeout
                try:
                    cursor.execute("DROP TABLE IF EXISTS parameters CASCADE")
                    conn.commit()
                    logger.info("Old parameters table dropped successfully")
                except Exception as drop_error:
                    logger.error(f"Failed to drop parameters table: {drop_error}")
                    # Try to rename instead of drop
                    try:
                        cursor.execute("ALTER TABLE parameters RENAME TO parameters_old")
                        conn.commit()
                        logger.info("Renamed old parameters table to parameters_old")
                    except:
                        logger.error("Could not drop or rename old table, continuing anyway...")
                        cursor.close()
                        conn.close()
                        return  # Give up, don't block startup

        except Exception as e:
            logger.info(f"Parameters table check: {e}")
            # Table doesn't exist, continue to create it

        # Create the new parameters table with correct schema
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS parameters (
                id SERIAL PRIMARY KEY,
                user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                date DATE NOT NULL,
                mood INTEGER CHECK (mood >= 1 AND mood <= 4),
                energy INTEGER CHECK (energy >= 1 AND energy <= 4),
                sleep_quality INTEGER CHECK (sleep_quality >= 1 AND sleep_quality <= 4),
                physical_activity INTEGER CHECK (physical_activity >= 1 AND physical_activity <= 4),
                anxiety INTEGER CHECK (anxiety >= 1 AND anxiety <= 4),
                notes TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(user_id, date)
            )
        ''')

        conn.commit()
        logger.info("✓ Parameters table created with correct schema")

        cursor.close()
        conn.close()

    except Exception as e:
        logger.error(f"Error in create_parameters_table: {e}")
        # Don't crash the app startup over this
        try:
            if cursor:
                cursor.close()
            if conn:
                conn.close()
        except:
            pass


def create_admin_user():
    """Create default admin user if it doesn't exist"""
    try:
        admin_email = os.environ.get('ADMIN_EMAIL', 'admin@example.com')
        admin_password = os.environ.get('ADMIN_PASSWORD', 'admin123')

        # Check if admin exists - SQLAlchemy 2.0 style
        stmt = select(User).filter_by(email=admin_email)
        admin = db.session.execute(stmt).scalar_one_or_none()

        if not admin:
            admin = User(
                username='admin',
                email=admin_email,
                role='admin',
                is_active=True
            )
            admin.set_password(admin_password)
            db.session.add(admin)
            db.session.flush()

            # Create admin profile
            profile = Profile(user_id=admin.id)
            db.session.add(profile)

            # Create welcome alert
            alert = Alert(
                user_id=admin.id,
                title='Welcome Admin!',
                content='Your admin account has been created.',
                alert_type='success'
            )
            db.session.add(alert)

            db.session.commit()
            logger.info(f"Admin user created: {admin_email}")
        else:
            logger.info("Admin user already exists")

    except Exception as e:
        logger.error(f"Error creating admin user: {e}")
        db.session.rollback()


# =====================
# DECORATORS
# =====================

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'error': 'Authentication required'}), 401

        # Verify user still exists and is active
        user = db.session.get(User, session['user_id'])
        if not user or not user.is_active:
            session.clear()
            return jsonify({'error': 'Invalid session'}), 401

        return f(*args, **kwargs)

    return decorated_function


def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'error': 'Authentication required'}), 401

        user = db.session.get(User, session['user_id'])
        if not user or user.role != 'admin':
            return jsonify({'error': 'Admin privileges required'}), 403

        return f(*args, **kwargs)

    return decorated_function


# =====================
# REQUEST HANDLERS
# =====================

@app.before_request
def before_request():
    """Log request details"""
    # Skip everything for health check - maximum speed
    if request.path == '/healthz':
        return
    request.request_id = str(uuid.uuid4())
    if not request.path.startswith('/static'):
        logger.info(f"Request started: {request.method} {request.path}")


@app.after_request
def after_request(response):
    """Log response details and set security headers"""
    # Skip logging for health check and static files
    if request.path != '/healthz' and not request.path.startswith('/static'):
        logger.info(f"Request completed: {response.status_code}")

    # Security headers
    if is_production:
        response.headers['X-Content-Type-Options'] = 'nosniff'
        response.headers['X-Frame-Options'] = 'DENY'
        response.headers['X-XSS-Protection'] = '1; mode=block'
        response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'

    return response


# =====================
# BASIC ROUTES
# =====================

@app.route('/')
def index():
    """Main landing page"""
    return render_template('index.html')


@app.route('/favicon.ico')
def favicon():
    """Serve favicon - handle missing file gracefully"""
    try:
        return send_from_directory(os.path.join(app.root_path, 'static'),
                                   'favicon.ico', mimetype='image/vnd.microsoft.icon')
    except:
        return '', 204


@app.route('/about')
def about_page():
    """About page"""
    return render_template('about.html')


@app.route('/support')
def support_page():
    """Support page"""
    return render_template('support.html')


@app.route('/profile')
@login_required
def profile_page():
    """Profile page"""
    return render_template('profile.html')


@app.route('/circles')
@login_required
def circles_page():
    """Circles page"""
    return render_template('circles.html', cache_bust=CACHE_BUST_VERSION)


@app.route('/messages')
@login_required
def messages_page():
    """Messages page"""
    return render_template('messages.html')


@app.route('/parameters')
@login_required
def parameters_page():
    return render_template('parameters.html')


@app.route('/healthz')
def healthz():
    """Ultra-lightweight health check for Render deployment.
    Returns immediately with no DB calls, no logging, no middleware overhead.
    Configure Render to use this endpoint: Settings -> Health Check Path -> /healthz
    """
    return 'OK', 200


@app.route('/api/health')
def health_check():
    """Health check endpoint"""
    status = {'status': 'healthy', 'timestamp': datetime.utcnow().isoformat()}

    try:
        # Check database
        db.session.execute(text('SELECT 1'))
        status['database'] = 'OK'
    except Exception as e:
        status['database'] = f'Error: {str(e)}'
        status['status'] = 'unhealthy'

    # Check Redis if configured
    try:
        if redis_client:
            redis_client.ping()
            status['redis'] = 'OK'
        else:
            status['redis'] = 'Not configured'
    except Exception as e:
        status['redis'] = f'Error: {str(e)}'

    return jsonify(status), 200 if status['status'] == 'healthy' else 503


# =====================
# AUTHENTICATION ROUTES
# =====================

@app.route('/api/auth/register', methods=['POST'])
@rate_limit(max_attempts=5, window_minutes=15)
def register():
    """Register new user"""
    try:
        data = request.json

        # Validate input
        username = sanitize_input(data.get('username', '').strip())
        email = sanitize_input(data.get('email', '').strip().lower())
        password = data.get('password', '')

        # Default username to email local part if not provided
        if not username and email:
            username = email.split('@')[0]
            # Sanitize the generated username too
            username = sanitize_input(username)

        # Validation - now password and email are required, username will have a default
        if not email or not password:
            return jsonify({'error': 'Email and password are required'}), 400

        # Check username after defaulting
        if not username:
            return jsonify({'error': 'Username could not be generated from email'}), 400

        if not validate_email(email):
            return jsonify({'error': 'Invalid email format'}), 400

        if not validate_username(username):
            return jsonify({'error': 'Invalid username format'}), 400

        valid, msg = validate_password_strength(password)
        if not valid:
            return jsonify({'error': msg}), 400

        # Check if user exists - SQLAlchemy 2.0 style
        existing_email = db.session.execute(
            select(User).filter_by(email=email)
        ).scalar_one_or_none()

        if existing_email:
            return jsonify({'error': 'Email already registered'}), 400

        existing_username = db.session.execute(
            select(User).filter_by(username=username)
        ).scalar_one_or_none()

        if existing_username:
            return jsonify({'error': 'Username already taken'}), 400

        # Create user
        user = User(
            username=username,
            email=email,
            preferred_language=data.get('preferred_language', 'en')
        )
        user.set_password(password)
        db.session.add(user)
        db.session.flush()

        # Create profile
        profile = Profile(user_id=user.id)
        db.session.add(profile)

        # Create welcome alert
        alert = Alert(
            user_id=user.id,
            title='alerts.welcome_title',
            content='alerts.welcome_message',
            alert_type='success'
        )
        db.session.add(alert)

        db.session.commit()

        # Log user in
        session['user_id'] = user.id
        session['username'] = user.username
        session.permanent = True
        
        # FIX: Set flag for one-time diary redirect check after registration
        session['diary_redirect_pending'] = True

        logger.info(f"User registered: {username}")

        return jsonify({
            'success': True,
            'user': user.to_dict()
        }), 200

    except IntegrityError as e:
        logger.error(f"Registration integrity error: {str(e)}")
        db.session.rollback()
        return jsonify({'error': 'Username or email already exists'}), 400
    except Exception as e:
        logger.error(f"Registration error: {str(e)}")
        db.session.rollback()
        return jsonify({'error': 'Registration failed'}), 500


@app.route('/api/auth/login', methods=['POST'])
@rate_limit(max_attempts=10, window_minutes=15)
def login():
    """User login"""
    try:
        data = request.json
        email = data.get('email', '').strip().lower()
        password = data.get('password', '')

        if not email or not password:
            return jsonify({'error': 'Email and password required'}), 400

        # Find user - SQLAlchemy 2.0 style
        user = db.session.execute(
            select(User).filter_by(email=email)
        ).scalar_one_or_none()

        if not user or not user.check_password(password):
            return jsonify({'error': 'Invalid credentials'}), 401

        if not user.is_active:
            return jsonify({'error': 'Account deactivated'}), 403

        # Update last login
        user.last_login = datetime.utcnow()
        db.session.commit()

        # Create session
        session['user_id'] = user.id
        session['username'] = user.username
        session['role'] = user.role
        session.permanent = True
        
        # FIX: Set flag for one-time diary redirect check after login
        session['diary_redirect_pending'] = True

        logger.info(f"Login successful: {user.username}")

        # Prepare user data with language preference
        user_data = user.to_dict()
        user_data['preferred_language'] = user.preferred_language or 'en'

        return jsonify({
            'success': True,
            'user': user_data
        }), 200

    except Exception as e:
        logger.error(f"Login error: {str(e)}")
        return jsonify({'error': 'Login failed'}), 500


@app.route('/api/auth/logout', methods=['POST'])
def logout():
    """User logout"""
    user_id = session.get('user_id')
    session.clear()
    logger.info(f"User logged out: {user_id}")
    return jsonify({'success': True}), 200


@app.route('/api/onboarding/status', methods=['GET'])
@login_required
def get_onboarding_status():
    user = User.query.get(session['user_id'])
    return jsonify({
        'needs_onboarding': not user.has_completed_onboarding and not user.onboarding_dismissed,
        'has_completed': user.has_completed_onboarding,
        'was_dismissed': user.onboarding_dismissed
    })


@app.route('/api/onboarding/complete', methods=['POST'])
@login_required
def complete_onboarding():
    user = User.query.get(session['user_id'])
    user.has_completed_onboarding = True
    db.session.commit()
    return jsonify({'message': 'Onboarding completed'}), 200


@app.route('/api/onboarding/dismiss', methods=['POST'])
@login_required
def dismiss_onboarding():
    user = User.query.get(session['user_id'])
    user.onboarding_dismissed = True
    db.session.commit()
    return jsonify({'message': 'Onboarding dismissed'}), 200


@app.route('/api/auth/session', methods=['GET'])
def check_session():
    """Check if user is logged in"""
    if 'user_id' in session:
        user = db.session.get(User, session['user_id'])
        if user and user.is_active:
            return jsonify({
                'authenticated': True,
                'user': user.to_dict()
            })

    return jsonify({'authenticated': False}), 401


# =====================
# USER & PROFILE ROUTES
# =====================

@app.route('/api/user/profile', methods=['GET'])
@login_required
def get_current_user():
    """Get current user info"""
    user = db.session.get(User, session['user_id'])
    if not user:
        return jsonify({'error': 'User not found'}), 404

    return jsonify(user.to_dict())


@app.route('/api/user/language', methods=['POST'])
def update_user_language():
    """
    Update user's language preference
    Works both for authenticated and unauthenticated users
    """
    logger.info("[LANG API DEBUG] ========================================")
    logger.info("[LANG API DEBUG] POST /api/user/language called")
    
    try:
        data = request.get_json()
        logger.info(f"[LANG API DEBUG] Request data: {data}")
        logger.info(f"[LANG API DEBUG] User ID in session: {session.get('user_id')}")
        
        if not data:
            logger.warning("[LANG API DEBUG] No data provided in request")
            return jsonify({
                'success': False,
                'message': 'No data provided'
            }), 400

        # FIXED: Accept both 'language' and 'preferred_language' for compatibility
        language = data.get('preferred_language') or data.get('language', 'en')
        logger.info(f"[LANG API DEBUG] Language from request: {language}")
        logger.info(f"[LANG API DEBUG]   preferred_language field: {data.get('preferred_language')}")
        logger.info(f"[LANG API DEBUG]   language field: {data.get('language')}")

        # Validate language code
        valid_languages = ['en', 'he', 'ar', 'ru']
        if language not in valid_languages:
            logger.warning(f"[LANG API DEBUG] Invalid language code: {language}")
            return jsonify({
                'success': False,
                'message': 'Invalid language code'
            }), 400

        # If user is authenticated, save to database
        if 'user_id' in session:
            user_id = session['user_id']
            logger.info(f"[LANG API DEBUG] User {user_id} is authenticated")

            try:
                user = db.session.get(User, user_id)
                if user:
                    old_language = user.preferred_language
                    logger.info(f"[LANG API DEBUG] Current language in DB: {old_language}")
                    logger.info(f"[LANG API DEBUG] New language to set: {language}")
                    
                    # FIXED: Use correct column name 'preferred_language'
                    user.preferred_language = language
                    db.session.commit()

                    logger.info(f"[LANG API DEBUG] SUCCESS: Updated language for user {user_id}: {old_language} -> {language}")
                    logger.info("[LANG API DEBUG] ========================================")
                    return jsonify({
                        'success': True,
                        'message': 'Language preference saved',
                        'language': language  # Return the saved language
                    }), 200
                else:
                    logger.error(f"[LANG API DEBUG] User {user_id} not found in database")
                    return jsonify({
                        'success': False,
                        'message': 'User not found'
                    }), 404

            except Exception as e:
                db.session.rollback()
                logger.error(f"[LANG API DEBUG] Error updating language preference: {e}")
                return jsonify({
                    'success': False,
                    'message': 'Database error'
                }), 500
        else:
            # For unauthenticated users, store in session
            session['preferred_language'] = language  # FIXED: Consistent naming

            logger.info(f"Stored language {language} in session (unauthenticated)")
            return jsonify({
                'success': True,
                'message': 'Language preference saved in session',
                'language': language
            }), 200

    except Exception as e:
        logger.error(f"Error in update_user_language: {e}")
        return jsonify({
            'success': False,
            'message': 'Server error'
        }), 500


@app.route('/api/user/update-city', methods=['POST'])
@login_required
def update_user_city():
    """Update user's selected city"""
    try:
        data = request.json
        selected_city = data.get('selected_city')

        if not selected_city:
            return jsonify({'error': 'City required'}), 400

        user = db.session.get(User, session['user_id'])
        if not user:
            return jsonify({'error': 'User not found'}), 404

        user.selected_city = selected_city
        db.session.commit()

        return jsonify({
            'success': True,
            'message': 'City updated successfully'
        }), 200

    except Exception as e:
        logger.error(f"Error updating city: {e}")
        db.session.rollback()
        return jsonify({'error': 'Failed to update city'}), 500


@app.route('/api/user/change-password', methods=['POST'])
@login_required
def change_password():
    """Change user password"""
    try:
        data = request.json
        current_password = data.get('current_password')
        new_password = data.get('new_password')

        if not current_password or not new_password:
            return jsonify({'error': 'Current and new passwords required'}), 400

        user = db.session.get(User, session['user_id'])
        if not user:
            return jsonify({'error': 'User not found'}), 404

        # Verify current password
        if not user.check_password(current_password):
            return jsonify({'error': 'Current password is incorrect'}), 401

        # Validate new password
        if len(new_password) < 12:
            return jsonify({'error': 'Password must be at least 12 characters long'}), 400

        # Check complexity
        import re
        has_upper = bool(re.search(r'[A-Z]', new_password))
        has_lower = bool(re.search(r'[a-z]', new_password))
        has_digit = bool(re.search(r'[0-9]', new_password))
        has_special = bool(re.search(r'[!@#$%^&*(),.?":{}|<>]', new_password))

        if not (has_upper and has_lower and has_digit and has_special):
            return jsonify({'error': 'Password must contain uppercase, lowercase, number, and special character'}), 400

        # Update password
        user.set_password(new_password)
        db.session.commit()

        logger.info(f"Password changed for user {user.id}")

        return jsonify({
            'success': True,
            'message': 'Password updated successfully'
        }), 200

    except Exception as e:
        logger.error(f"Error changing password: {e}")
        db.session.rollback()
        return jsonify({'error': 'Failed to change password'}), 500


@app.route('/api/auth/request-magic-link', methods=['POST'])
@rate_limit(max_attempts=5, window_minutes=15)
def request_magic_link():
    """Request magic link for email-only login"""
    try:
        data = request.json
        email = data.get('email', '').strip().lower()
        language = data.get('language', 'en')

        if not email:
            return jsonify({'error': 'Email required'}), 400

        user = db.session.execute(
            select(User).filter_by(email=email)
        ).scalar_one_or_none()

        if not user:
            import secrets
            # Extract username from email (part before @)
            email_username = email.split('@')[0]

            # Check if this username is taken
            existing = db.session.execute(
                select(User).filter_by(username=email_username)
            ).scalar_one_or_none()

            # If taken, add random suffix
            if existing:
                temp_username = f"{email_username}_{secrets.token_hex(4)}"
            else:
                temp_username = email_username

            user = User(
                username=temp_username,
                email=email,
                preferred_language=language
            )
            user.set_password(secrets.token_urlsafe(32))
            db.session.add(user)
            db.session.flush()

            profile = Profile(user_id=user.id)
            db.session.add(profile)

        import secrets
        magic_token = secrets.token_urlsafe(32)
        expires_at = datetime.utcnow() + timedelta(days=1825)

        token_record = MagicLoginToken(
            user_id=user.id,
            token=magic_token,
            expires_at=expires_at
        )
        db.session.add(token_record)
        db.session.commit()

        user_language = user.preferred_language or language
        send_magic_link_email(user.email, magic_token, user_language)

        return jsonify({'success': True, 'message': 'Magic link sent'}), 200

    except Exception as e:
        logger.error(f"Magic link error: {e}")
        db.session.rollback()
        return jsonify({'error': 'Failed to send magic link'}), 500


@app.route('/api/auth/verify-magic-link', methods=['POST'])
def verify_magic_link():
    """Verify magic link and log user in"""
    try:
        data = request.json
        magic_token = data.get('token', '')

        if not magic_token:
            return jsonify({'error': 'Token required'}), 400

        token_record = db.session.execute(
            select(MagicLoginToken).filter_by(
                token=magic_token
                # Removed: used=False check - allow unlimited use
            )
        ).scalar_one_or_none()

        if not token_record or token_record.expires_at < datetime.utcnow():
            return jsonify({'error': 'Invalid or expired token'}), 400

        # DO NOT mark as used - allow unlimited clicks until new token generated
        # token_record.used = True  # COMMENTED OUT
        user = db.session.get(User, token_record.user_id)

        # Check if username needs confirmation (not if it's from email)
        email_prefix = user.email.split('@')[0]
        needs_username = False  # Don't force username change if it matches email

        consent = db.session.execute(
            select(UserConsent).filter_by(user_id=user.id)
        ).scalar_one_or_none()
        needs_consent = consent is None

        user.last_login = datetime.utcnow()
        db.session.commit()

        session['user_id'] = user.id
        session['username'] = user.username
        session['role'] = user.role
        session.permanent = True
        
        # FIX: Set flag for one-time diary redirect check after magic link login
        session['diary_redirect_pending'] = True

        return jsonify({
            'success': True,
            'needs_username': needs_username,
            'needs_consent': needs_consent,
            'user': user.to_dict(),
            'suggested_username': email_prefix  # Send suggestion to frontend
        }), 200

    except Exception as e:
        logger.error(f"Magic link verification error: {e}")
        return jsonify({'error': 'Verification failed'}), 500


@app.route('/api/auth/set-username', methods=['POST'])
@login_required
def set_username():
    """Set username for magic link users (optional)"""
    try:
        data = request.json
        new_username = data.get('username', '').strip()
        skip_username = data.get('skip', False)

        user = db.session.get(User, session['user_id'])

        # If skipping, keep the email-based username
        if skip_username:
            # Username already set from email, just return success
            return jsonify({'success': True, 'username': user.username}), 200

        if not new_username:
            # If blank, use email prefix
            new_username = user.email.split('@')[0]

        # Validate username
        if not validate_username(new_username):
            return jsonify({'error': 'Invalid username format'}), 400

        # Check if username taken (excluding current user)
        existing = db.session.execute(
            select(User).filter(
                User.username == new_username,
                User.id != user.id
            )
        ).scalar_one_or_none()

        if existing:
            return jsonify({'error': 'Username already taken'}), 400

        user.username = new_username
        session['username'] = new_username
        db.session.commit()

        return jsonify({'success': True, 'username': new_username}), 200

    except Exception as e:
        logger.error(f"Set username error: {e}")
        return jsonify({'error': 'Failed to set username'}), 500


@app.route('/api/auth/save-consent', methods=['POST'])
@login_required
def save_consent():
    """Save user consent preferences with optional username and birth_year"""
    try:
        data = request.json
        user = db.session.get(User, session['user_id'])

        # Handle username if provided with consent
        username = data.get('username', '').strip()
        if username:
            # If blank, use email prefix
            if not username:
                username = user.email.split('@')[0]

            # Check if username is available
            existing = db.session.execute(
                select(User).filter(
                    User.username == username,
                    User.id != user.id
                )
            ).scalar_one_or_none()

            if not existing:
                user.username = username
                session['username'] = username

        # PJ6001: Handle birth_year
        birth_year = data.get('birth_year')
        if birth_year:
            try:
                birth_year = int(birth_year)
                if 1900 <= birth_year <= 2025:
                    user.birth_year = birth_year
            except (ValueError, TypeError):
                pass  # Ignore invalid birth year

        # Check if consent already exists
        consent = db.session.execute(
            select(UserConsent).filter_by(user_id=session['user_id'])
        ).scalar_one_or_none()

        if not consent:
            consent = UserConsent(user_id=session['user_id'])
            db.session.add(consent)

        # Update consent fields
        consent.email_updates = data.get('email_updates', False)
        consent.privacy_accepted = data.get('privacy_accepted', False)
        consent.research_data = data.get('research_data', False)
        consent.team_declaration = data.get('team_declaration', False)
        consent.responsible_use = data.get('responsible_use', False)
        consent.waiver_claims = data.get('waiver_claims', False)
        consent.consent_language = data.get('language', 'en')

        # All required consents must be true
        required = ['privacy_accepted', 'team_declaration', 'responsible_use', 'waiver_claims']
        all_accepted = all(data.get(field, False) for field in required)

        if not all_accepted:
            return jsonify({'error': 'All required consents must be accepted'}), 400

        db.session.commit()
        return jsonify({'success': True}), 200

    except Exception as e:
        logger.error(f"Save consent error: {e}")
        db.session.rollback()
        return jsonify({'error': 'Failed to save consent'}), 500


@app.route('/api/user/delete-account', methods=['POST'])
@login_required
def delete_account():
    """Delete user account and all associated data"""
    try:
        user_id = session['user_id']
        user = db.session.get(User, user_id)

        if not user:
            return jsonify({'error': 'User not found'}), 404

        # Log the deletion
        logger.info(f"Deleting account for user {user.id} ({user.username})")

        # Delete user (cascade will handle related data)
        db.session.delete(user)
        db.session.commit()

        # Clear session
        session.clear()

        return jsonify({
            'success': True,
            'message': 'Account deleted successfully'
        }), 200

    except Exception as e:
        logger.error(f"Error deleting account: {e}")
        db.session.rollback()
        return jsonify({'error': 'Failed to delete account'}), 500


@app.route('/api/auth/forgot-password', methods=['POST'])
@rate_limit(max_attempts=5, window_minutes=60)
def forgot_password():
    """Request password reset with language support"""
    try:
        data = request.json
        email = data.get('email', '').strip().lower()
        language = data.get('language', 'en')

        if not email:
            return jsonify({'error': 'Email required'}), 400

        user = db.session.execute(
            select(User).filter_by(email=email)
        ).scalar_one_or_none()

        if not user:
            logger.info(f"Password reset requested for non-existent email: {email}")
            return jsonify({
                'success': True,
                'message': 'If an account exists with that email, a reset link has been sent'
            }), 200

        import secrets
        reset_token = secrets.token_urlsafe(32)
        expires_at = datetime.utcnow() + timedelta(hours=1)

        token_record = PasswordResetToken(
            user_id=user.id,
            token=reset_token,
            expires_at=expires_at
        )
        db.session.add(token_record)
        db.session.commit()

        user_language = user.preferred_language or language
        email_sent = send_password_reset_email(user.email, reset_token, user_language)

        return jsonify({
            'success': True,
            'message': 'If an account exists with that email, a reset link has been sent'
        }), 200

    except Exception as e:
        logger.error(f"Forgot password error: {e}")
        db.session.rollback()
        return jsonify({'error': 'Failed to process request'}), 500


@app.route('/api/auth/reset-password', methods=['POST'])
def reset_password():
    """Reset password using token"""
    try:
        data = request.json
        token = data.get('token')
        new_password = data.get('new_password')

        if not token or not new_password:
            return jsonify({'error': 'Token and new password required'}), 400

        # Find token
        token_record = db.session.execute(
            select(PasswordResetToken).filter_by(token=token, used=False)
        ).scalar_one_or_none()

        if not token_record:
            return jsonify({'error': 'Invalid or expired reset token'}), 400

        # Check if token expired
        if datetime.utcnow() > token_record.expires_at:
            return jsonify({'error': 'Reset token has expired'}), 400

        # Validate new password
        if len(new_password) < 12:
            return jsonify({'error': 'Password must be at least 12 characters long'}), 400

        # Check complexity
        import re
        has_upper = bool(re.search(r'[A-Z]', new_password))
        has_lower = bool(re.search(r'[a-z]', new_password))
        has_digit = bool(re.search(r'[0-9]', new_password))
        has_special = bool(re.search(r'[!@#$%^&*(),.?":{}|<>]', new_password))

        if not (has_upper and has_lower and has_digit and has_special):
            return jsonify({'error': 'Password must contain uppercase, lowercase, number, and special character'}), 400

        # Get user and update password
        user = db.session.get(User, token_record.user_id)
        if not user:
            return jsonify({'error': 'User not found'}), 404

        user.set_password(new_password)
        token_record.used = True
        db.session.commit()

        logger.info(f"Password reset successful for user {user.id}")

        return jsonify({
            'success': True,
            'message': 'Password reset successfully'
        }), 200

    except Exception as e:
        logger.error(f"Reset password error: {e}")
        db.session.rollback()
        return jsonify({'error': 'Failed to reset password'}), 500


@app.route('/api/user/language', methods=['GET'])
def get_user_language():
    """Get user's language preference"""
    try:
        # Check if user is authenticated
        if 'user_id' in session:
            user_id = session['user_id']
            user = db.session.get(User, user_id)

            # FIXED: Use correct column name 'preferred_language'
            if user and hasattr(user, 'preferred_language') and user.preferred_language:
                return jsonify({
                    'success': True,
                    'language': user.preferred_language
                }), 200

        # Check session for unauthenticated users
        # FIXED: Check both old and new session keys for backwards compatibility
        session_lang = session.get('preferred_language') or session.get('language')
        if session_lang:
            return jsonify({
                'success': True,
                'language': session_lang
            }), 200

        # Default to English
        return jsonify({
            'success': True,
            'language': 'en'
        }), 200

    except Exception as e:
        logger.error(f"Error getting language preference: {e}")
        return jsonify({
            'success': True,  # Don't fail hard on language errors
            'language': 'en'
        }), 200


@app.route('/api/admin/migrate-language', methods=['POST'])
def migrate_language_column():
    """
    Migration endpoint to copy 'language' to 'preferred_language' if needed
    Only accessible to admins or in development
    """
    # Add authentication check here if needed

    try:
        # Check if old 'language' column exists
        from sqlalchemy import inspect
        inspector = inspect(db.engine)
        user_columns = [col['name'] for col in inspector.get_columns('user')]

        if 'language' in user_columns:
            # Migrate data from 'language' to 'preferred_language'
            users = User.query.filter(User.language.isnot(None)).all()
            count = 0
            for user in users:
                if hasattr(user, 'language') and user.language:
                    user.preferred_language = user.language
                    count += 1

            db.session.commit()
            return jsonify({
                'success': True,
                'message': f'Migrated {count} users'
            }), 200
        else:
            return jsonify({
                'success': True,
                'message': 'No migration needed - language column does not exist'
            }), 200

    except Exception as e:
        db.session.rollback()
        logger.error(f"Migration error: {e}")
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500


@app.route('/api/profile', methods=['GET', 'PUT'])
@login_required
def profile():
    """Get or update user profile"""
    user_id = session.get('user_id')

    if request.method == 'GET':
        profile = db.session.execute(
            select(Profile).filter_by(user_id=user_id)
        ).scalar_one_or_none()

        if not profile:
            profile = Profile(user_id=user_id)
            db.session.add(profile)
            db.session.commit()

        # FIXED: Get username from User table to return in response
        user = db.session.execute(
            select(User).filter_by(id=user_id)
        ).scalar_one_or_none()

        return jsonify({
            'username': user.username if user else '',  # FIXED: Added username field
            'birth_year': user.birth_year if user else 1985,  # PJ6001: Birth year field
            'bio': profile.bio or '',
            'interests': profile.interests or '',
            'occupation': profile.occupation or '',
            'goals': profile.goals or '',
            'favorite_hobbies': profile.favorite_hobbies or '',
            'mood_status': profile.mood_status or '',
            'avatar_url': profile.avatar_url or ''
        })

    elif request.method == 'PUT':
        data = request.json
        profile = db.session.execute(
            select(Profile).filter_by(user_id=user_id)
        ).scalar_one_or_none()

        if not profile:
            profile = Profile(user_id=user_id)
            db.session.add(profile)

        # Update fields
        if 'bio' in data:
            profile.bio = sanitize_input(data.get('bio', ''))[:1000]
        if 'interests' in data:
            profile.interests = sanitize_input(data.get('interests', ''))
        if 'occupation' in data:
            profile.occupation = sanitize_input(data.get('occupation', ''))
        if 'goals' in data:
            profile.goals = sanitize_input(data.get('goals', ''))
        if 'favorite_hobbies' in data:
            profile.favorite_hobbies = sanitize_input(data.get('favorite_hobbies', ''))
        if 'mood_status' in data:
            profile.mood_status = sanitize_input(data.get('mood_status', ''))[:50]
        if 'avatar_url' in data:
            profile.avatar_url = data.get('avatar_url', '')[:500]

        profile.updated_at = datetime.utcnow()

        # Update user's preferred language if provided
        if 'preferred_language' in data:
            user = db.session.get(User, user_id)
            if user and data['preferred_language'] in ['en', 'he', 'ar', 'ru']:
                user.preferred_language = data['preferred_language']

        # PJ6001: Update birth_year if provided
        if 'birth_year' in data:
            user = db.session.get(User, user_id)
            if user:
                try:
                    birth_year = int(data['birth_year'])
                    if 1900 <= birth_year <= 2025:
                        user.birth_year = birth_year
                except (ValueError, TypeError):
                    pass

        db.session.commit()
        return jsonify({'success': True, 'message': 'Profile updated'})


@app.route('/api/users/<int:user_id>/profile', methods=['GET'])
@login_required
def get_user_profile(user_id):
    """Get another user's profile - accessible by followers, circle members, or with allow_preview"""
    try:
        current_user_id = session.get('user_id')
        
        # PJ501: Check if user is blocked by target user first
        is_blocked_by_target = BlockedUser.query.filter_by(
            blocker_id=user_id,
            blocked_id=current_user_id
        ).first() is not None
        
        if is_blocked_by_target:
            return jsonify({'error': 'account_not_available', 'blocked': True}), 403

        # Check if following this user
        is_following = Follow.query.filter_by(
            follower_id=current_user_id,
            followed_id=user_id
        ).first() is not None

        # PJ401: Also check if current user is in target user's circles
        is_in_circle = Circle.query.filter_by(
            user_id=user_id,
            circle_user_id=current_user_id
        ).first() is not None
        
        # PJ501: Allow preview mode for recommended users (view basic profile without following)
        allow_preview = request.args.get('allow_preview', 'false').lower() == 'true'

        if not is_following and not is_in_circle and user_id != current_user_id and not allow_preview:
            return jsonify({'error': 'Must be following user or in their circles to view profile'}), 403

        user = User.query.get(user_id)
        if not user:
            return jsonify({'error': 'User not found'}), 404

        # Get profile data if it exists
        profile = user.profile if user.profile else None

        return jsonify({
            'id': user.id,
            'username': user.username,
            'email': user.email,
            'city': user.selected_city or '',
            'bio': profile.bio if profile else '',
            'avatar_url': profile.avatar_url if profile else '',
            'occupation': profile.occupation if profile else '',
            'interests': profile.interests if profile else '',
            'goals': profile.goals if profile else '',
            'favorite_hobbies': profile.favorite_hobbies if profile else '',
            'created_at': user.created_at.isoformat() if user.created_at else None,
            'is_following': is_following,  # PJ501: Include following status in response
            'is_preview': allow_preview and not is_following  # PJ501: Indicate if this is preview mode
        })
    except Exception as e:
        logger.error(f"Error loading user profile {user_id}: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/users/<int:user_id>/posts', methods=['GET'])
@login_required
def get_user_posts(user_id):
    """Get another user's feed posts with visibility-based access"""
    try:
        current_user_id = session.get('user_id')

        # Check if following this user
        is_following = Follow.query.filter_by(
            follower_id=current_user_id,
            followed_id=user_id
        ).first() is not None

        # PJ401: Also check if current user is in target user's circles
        membership = Circle.query.filter_by(
            user_id=user_id,
            circle_user_id=current_user_id
        ).first()
        
        is_in_circle = membership is not None

        if not is_following and not is_in_circle and user_id != current_user_id:
            return jsonify({'error': 'Must be following user or in their circles to view posts'}), 403

        # If viewing own posts, return all
        if user_id == current_user_id:
            posts = Post.query.filter_by(user_id=user_id).order_by(Post.created_at.desc()).all()

            # Calculate likes and comments for each post
            posts_data = []
            for post in posts:
                # Count likes from Reaction table
                likes_count = db.session.execute(
                    select(func.count(Reaction.id)).filter_by(post_id=post.id, type='like')
                ).scalar() or 0

                # Count comments
                comments_count = db.session.execute(
                    select(func.count(Comment.id)).filter_by(post_id=post.id)
                ).scalar() or 0

                # Check if current user liked this post
                user_liked = db.session.execute(
                    select(Reaction).filter_by(
                        post_id=post.id,
                        user_id=current_user_id,
                        type='like'
                    )
                ).scalar_one_or_none() is not None

                posts_data.append({
                    'id': post.id,
                    'content': post.content,
                    'created_at': post.created_at.isoformat() if post.created_at else None,
                    'likes_count': likes_count,
                    'comments_count': comments_count,
                    'user_liked': user_liked,
                    'visibility': post.visibility
                })

            return jsonify({'posts': posts_data})

        # PJ401: membership already checked above for access control
        # Determine which visibility levels current user can see
        visible_levels = ['general']  # Everyone can see public
        if membership:
            if membership.circle_type in ['class_a', 'family']:
                visible_levels = ['general', 'close_friends', 'family']
            elif membership.circle_type in ['class_b', 'close_friends']:
                visible_levels = ['general', 'close_friends']

        # Get posts filtered by visible visibility levels
        posts = Post.query.filter(
            Post.user_id == user_id,
            Post.visibility.in_(visible_levels)
        ).order_by(Post.created_at.desc()).all()

        # Calculate likes and comments for each post
        posts_data = []
        for post in posts:
            # Count likes from Reaction table
            likes_count = db.session.execute(
                select(func.count(Reaction.id)).filter_by(post_id=post.id, type='like')
            ).scalar() or 0

            # Count comments
            comments_count = db.session.execute(
                select(func.count(Comment.id)).filter_by(post_id=post.id)
            ).scalar() or 0

            # Check if current user liked this post
            user_liked = db.session.execute(
                select(Reaction).filter_by(
                    post_id=post.id,
                    user_id=current_user_id,
                    type='like'
                )
            ).scalar_one_or_none() is not None

            posts_data.append({
                'id': post.id,
                'content': post.content,
                'created_at': post.created_at.isoformat() if post.created_at else None,
                'likes_count': likes_count,
                'comments_count': comments_count,
                'user_liked': user_liked,
                'visibility': post.visibility
            })

        return jsonify({'posts': posts_data})

    except Exception as e:
        logger.error(f"Error getting user posts: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/users/<int:user_id>/circles', methods=['GET'])
@login_required
def get_user_circles(user_id):
    """Get another user's circles (read-only) - returns members in each circle"""
    try:
        current_user_id = session.get('user_id')

        # Check if following this user
        is_following = Follow.query.filter_by(
            follower_id=current_user_id,
            followed_id=user_id
        ).first() is not None

        # PJ401: Also check if current user is in target user's circles
        is_in_circle = Circle.query.filter_by(
            user_id=user_id,
            circle_user_id=current_user_id
        ).first() is not None

        if not is_following and not is_in_circle and user_id != current_user_id:
            return jsonify({'error': 'Must be following user or in their circles to view circles'}), 403

        # Initialize viewer_circle_type
        viewer_circle_type = None
        user_privacy_level = None

        # Check circles privacy settings if viewing another user's circles
        if user_id != current_user_id:
            target_user = db.session.get(User, user_id)
            if not target_user:
                return jsonify({'error': 'User not found'}), 404

            privacy_level = target_user.circles_privacy or 'private'
            user_privacy_level = privacy_level  # Store for return

            # Check viewer's circle membership with target user
            viewer_circle = db.session.execute(
                select(Circle).filter_by(
                    user_id=user_id,
                    circle_user_id=current_user_id
                )
            ).scalars().first()

            if viewer_circle:
                type_mapping_check = {
                    'general': 'public',
                    'close_friends': 'class_b',
                    'family': 'class_a',
                    'public': 'public',
                    'class_b': 'class_b',
                    'class_a': 'class_a'
                }
                viewer_circle_type = type_mapping_check.get(viewer_circle.circle_type, 'public')

            # Apply privacy filtering - return with user_privacy field
            if privacy_level == 'private':
                return jsonify({
                    'private': True,
                    'message': 'Circles set to private',
                    'public': [],
                    'class_b': [],
                    'class_a': [],
                    'can_view': {'public': False, 'class_b': False, 'class_a': False},
                    'viewer_circle_type': viewer_circle_type,
                    'user_privacy': privacy_level  # ADDED
                })

            if privacy_level == 'class_a' and viewer_circle_type != 'class_a':
                return jsonify({
                    'private': True,
                    'message': 'Circles set to private',
                    'public': [],
                    'class_b': [],
                    'class_a': [],
                    'can_view': {'public': False, 'class_b': False, 'class_a': False},
                    'viewer_circle_type': viewer_circle_type,
                    'user_privacy': privacy_level  # ADDED
                })

            if privacy_level == 'class_b' and viewer_circle_type not in ['class_a', 'class_b']:
                return jsonify({
                    'private': True,
                    'message': 'Circles set to private',
                    'public': [],
                    'class_b': [],
                    'class_a': [],
                    'can_view': {'public': False, 'class_b': False, 'class_a': False},
                    'viewer_circle_type': viewer_circle_type,
                    'user_privacy': privacy_level  # ADDED
                })

        # Get all circles for this user
        circles_stmt = select(Circle).filter_by(user_id=user_id)
        circles = db.session.execute(circles_stmt).scalars().all()

        # Organize by circle type
        result = {
            'public': [],
            'class_b': [],
            'class_a': []
        }

        type_mapping = {
            'public': 'public',
            'general': 'public',
            'class_b': 'class_b',
            'close_friends': 'class_b',
            'class_a': 'class_a',
            'family': 'class_a'
        }

        for circle in circles:
            # Get the user info for the circle member
            member_user = db.session.get(User, circle.circle_user_id)
            if member_user:
                user_info = {
                    'id': member_user.id,
                    'username': member_user.username,
                    'email': member_user.email
                }

                # Map to normalized circle type
                circle_type = type_mapping.get(circle.circle_type, circle.circle_type)
                if circle_type in result:
                    result[circle_type].append(user_info)

        # Add permission flags for frontend
        can_view = {
            'public': True,
            'class_b': True,
            'class_a': True
        }

        result['can_view'] = can_view
        result['viewer_circle_type'] = viewer_circle_type
        result['user_privacy'] = user_privacy_level  # ADDED

        return jsonify(result)

    except Exception as e:
        logger.error(f"Get user circles error: {str(e)}")
        return jsonify({'error': 'Failed to get circles'}), 500


@app.route('/api/users/<int:user_id>/parameters', methods=['GET'])
@login_required
def get_user_parameters(user_id):
    """Get another user's wellness parameters with circle-based privacy"""
    try:
        current_user_id = session.get('user_id')

        # Check if following this user
        is_following = Follow.query.filter_by(
            follower_id=current_user_id,
            followed_id=user_id
        ).first() is not None

        # PJ401: Also check if current user is in target user's circles
        is_in_circle = Circle.query.filter_by(
            user_id=user_id,
            circle_user_id=current_user_id
        ).first() is not None

        if not is_following and not is_in_circle and user_id != current_user_id:
            return jsonify({'error': 'Must be following user or in their circles to view parameters'}), 403

        # Get date range from query params (REQUIRED)
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')

        if not start_date or not end_date:
            return jsonify({'error': 'start_date and end_date are required'}), 400

        # Parse dates
        try:
            start = parse_date_as_local(start_date)
            end = parse_date_as_local(end_date)
        except ValueError:
            return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD'}), 400

        # Check what circle the current user is in for the target user
        circle_level = 'public'  # Default to public
        if current_user_id != user_id:
            circle_stmt = select(Circle).filter_by(
                user_id=user_id,
                circle_user_id=current_user_id
            )
            circle = db.session.execute(circle_stmt).scalar_one_or_none()

            if circle:
                # Map circle types to privacy levels
                type_mapping = {
                    'public': 'public',
                    'general': 'public',
                    'class_b': 'class_b',
                    'close_friends': 'class_b',
                    'class_a': 'class_a',
                    'family': 'class_a'
                }
                circle_level = type_mapping.get(circle.circle_type, 'public')
        else:
            # User viewing their own parameters - full access
            circle_level = 'class_a'

        # Get parameters with privacy settings
        query = text("""
            SELECT p.date, p.mood, p.energy, p.sleep_quality, 
                   p.physical_activity, p.anxiety, p.notes,
                   p.mood_privacy, p.energy_privacy, p.sleep_quality_privacy,
                   p.physical_activity_privacy, p.anxiety_privacy
            FROM saved_parameters p
            WHERE p.user_id = :user_id
              AND p.date >= :start_date 
              AND p.date <= :end_date
            ORDER BY p.date ASC
        """)

        result_proxy = db.session.execute(
            query,
            {
                'user_id': user_id,
                'start_date': start.isoformat(),
                'end_date': end.isoformat()
            }
        )

        parameters = result_proxy.fetchall()

        # Build result array with privacy filtering
        result = []
        for row in parameters:
            param_dict = {
                'date': row[0].isoformat() if hasattr(row[0], 'isoformat') else str(row[0])
            }

            # Check privacy for each parameter
            # mood
            mood_privacy = row[7] or 'public'
            if check_param_visibility(mood_privacy, circle_level):
                param_dict['mood'] = row[1]
            else:
                param_dict['mood'] = None

            # energy
            energy_privacy = row[8] or 'public'
            if check_param_visibility(energy_privacy, circle_level):
                param_dict['energy'] = row[2]
            else:
                param_dict['energy'] = None

            # sleep_quality
            sleep_privacy = row[9] or 'public'
            if check_param_visibility(sleep_privacy, circle_level):
                param_dict['sleep_quality'] = row[3]
            else:
                param_dict['sleep_quality'] = None

            # physical_activity
            activity_privacy = row[10] or 'public'
            if check_param_visibility(activity_privacy, circle_level):
                param_dict['physical_activity'] = row[4]
            else:
                param_dict['physical_activity'] = None

            # anxiety
            anxiety_privacy = row[11] or 'public'
            if check_param_visibility(anxiety_privacy, circle_level):
                param_dict['anxiety'] = row[5]
            else:
                param_dict['anxiety'] = None

            # Notes are always private unless user is in class_a
            if circle_level == 'class_a':
                param_dict['notes'] = row[6]
            else:
                param_dict['notes'] = None

            result.append(param_dict)

        return jsonify(result), 200

    except Exception as e:
        app.logger.error(f"Error loading user parameters: {str(e)}")
        return jsonify({'error': 'Failed to load parameters'}), 500


def check_param_visibility(param_privacy, viewer_circle_level):
    """Helper function to check if a parameter should be visible based on privacy and viewer's circle"""
    if param_privacy == 'public':
        return True
    elif param_privacy == 'class_b':
        return viewer_circle_level in ['class_b', 'class_a']
    elif param_privacy == 'class_a':
        return viewer_circle_level == 'class_a'
    return False


@app.route('/api/debug/parameters/<int:user_id>')
@login_required
def debug_parameters(user_id):
    """Temporary debug endpoint - DELETE after fixing"""
    try:
        # Get ALL parameters for this user (no date filter)
        query = text("""
            SELECT date, mood, energy, sleep_quality, 
                   physical_activity, anxiety, notes, user_id
            FROM saved_parameters
            WHERE user_id = :user_id
            ORDER BY date DESC
            LIMIT 10
        """)

        result = db.session.execute(query, {'user_id': user_id})
        rows = result.fetchall()

        parameters = []
        for row in rows:
            parameters.append({
                'date': str(row[0]),  # Convert to string to see exact format
                'date_type': type(row[0]).__name__,
                'mood': row[1],
                'energy': row[2],
                'sleep_quality': row[3],
                'physical_activity': row[4],
                'anxiety': row[5],
                'notes': row[6],
                'user_id': row[7]
            })

        return jsonify({
            'count': len(parameters),
            'data': parameters
        }), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/users/search')
@login_required
def search_users():
    """Search for users by username or email"""
    try:
        # Get and validate query
        query = request.args.get('q', '').strip().lower()
        logger.info(f"🔍 Search request: query='{query}'")

        if not query or len(query) < 1:
            logger.info(f"❌ Query too short: length={len(query)}")
            return jsonify({'users': []})

        # Get current user
        current_user_id = session.get('user_id')
        if not current_user_id:
            logger.error("❌ No user_id in session")
            return jsonify({'error': 'Not authenticated'}), 401

        logger.info(f"👤 Current user: {current_user_id}")

        # Perform search
        logger.info(f"🔎 Searching: username ILIKE '%{query}%' WHERE id != {current_user_id}")

        users = User.query.filter(
            User.id != current_user_id,
            User.username.ilike(f'%{query}%')
        ).limit(10).all()

        logger.info(f"✓ Found {len(users)} user(s): {[u.username for u in users]}")

        # PJ706: Get list of users the current user is following
        following_ids = set()
        try:
            following_result = db.session.execute(
                select(Follow.followed_id).filter_by(follower_id=current_user_id)
            ).scalars().all()
            following_ids = set(following_result)
            logger.info(f"👥 User is following {len(following_ids)} users")
        except Exception as follow_err:
            logger.warning(f"⚠️ Could not fetch following list: {follow_err}")

        # Format results with profile data
        results = []
        for user in users:
            try:
                profile = Profile.query.filter_by(user_id=user.id).first()

                # PJ706: Include is_following field
                is_following = user.id in following_ids

                user_data = {
                    'id': user.id,
                    'username': user.username,
                    'email': user.email,
                    'display_name': user.username,
                    'bio': profile.bio if profile else None,
                    'occupation': profile.occupation if profile else None,
                    'interests': profile.interests if profile else None,
                    'avatar_url': profile.avatar_url if profile else None,
                    'is_following': is_following  # PJ706: Added
                }

                results.append(user_data)
                logger.debug(f"  - Formatted user {user.id}: {user.username}, is_following={is_following}")

            except Exception as profile_error:
                # Include user even if profile loading fails
                logger.warning(f"⚠️  Profile error for user {user.id}: {profile_error}")
                results.append({
                    'id': user.id,
                    'username': user.username,
                    'email': user.email,
                    'display_name': user.username,
                    'bio': None,
                    'occupation': None,
                    'interests': None,
                    'avatar_url': None,
                    'is_following': user.id in following_ids  # PJ706: Added
                })

        logger.info(f"📤 Returning {len(results)} result(s)")
        return jsonify({'users': results})

    except Exception as e:
        logger.error(f"❌ Search failed: {type(e).__name__}: {e}", exc_info=True)
        logger.error(f"   Query: '{request.args.get('q', '')}'")
        logger.error(f"   Session: {dict(session)}")
        return jsonify({'users': [], 'error': 'Search failed'}), 500


# =====================
# ALERTS ROUTES
# =====================

@app.route('/api/alerts', methods=['GET'])
@login_required
def get_alerts():
    """Get user alerts - filters trigger/feed alerts to only show for followed users"""
    try:
        user_id = session['user_id']

        # PJ401: Get list of users this user follows
        following_ids = db.session.execute(
            select(Follow.followed_id).filter_by(follower_id=user_id)
        ).scalars().all()
        following_set = set(following_ids)

        # SQLAlchemy 2.0 style - get all unread alerts
        alerts_stmt = select(Alert).filter_by(
            user_id=user_id,
            is_read=False
        ).order_by(desc(Alert.created_at)).limit(100)

        all_alerts = db.session.execute(alerts_stmt).scalars().all()

        # PJ401: Filter alerts based on following status
        # PJ704: CRITICAL FIX - Only filter 'trigger' category alerts by following status
        # Follow and invite alerts (alert_category='follow') should ALWAYS show 
        # since they notify you about someone following/inviting YOU
        filtered_alerts = []
        for alert in all_alerts:
            # Check if this alert has a source_user_id AND is a trigger/feed alert
            if alert.source_user_id and alert.alert_category in ('trigger', 'feed'):
                # Only include trigger/feed alerts if user follows the source_user
                if alert.source_user_id in following_set:
                    filtered_alerts.append(alert)
            else:
                # Alerts without source_user_id OR with category 'follow', 'message', 'general' always show
                # This includes: new follower alerts, invite alerts, general notifications
                filtered_alerts.append(alert)

        # PJ812: Increased limit to 100 for full month of alerts
        filtered_alerts = filtered_alerts[:100]

        return jsonify({
            'alerts': [alert.to_dict() for alert in filtered_alerts],
            'unread_count': len(filtered_alerts)
        })

    except Exception as e:
        logger.error(f"Get alerts error: {str(e)}")
        return jsonify({'error': 'Failed to get alerts'}), 500


@app.route('/api/alerts/<int:alert_id>/read', methods=['PUT'])
@login_required
def mark_alert_read(alert_id):
    """Mark alert as read"""
    try:
        alert = db.session.get(Alert, alert_id)
        if alert and alert.user_id == session['user_id']:
            alert.is_read = True
            db.session.commit()
            return jsonify({'success': True})
        return jsonify({'error': 'Alert not found'}), 404
    except Exception as e:
        logger.error(f"Mark alert error: {str(e)}")
        return jsonify({'error': 'Failed to mark alert'}), 500


@app.route('/api/notification-settings', methods=['GET', 'PUT'])
@login_required
def notification_settings():
    """Get or update notification settings"""
    user_id = session.get('user_id')
    logger.info(f"[NOTIFICATION DEBUG] ========================================")
    logger.info(f"[NOTIFICATION DEBUG] /api/notification-settings called")
    logger.info(f"[NOTIFICATION DEBUG] Method: {request.method}")
    logger.info(f"[NOTIFICATION DEBUG] User ID: {user_id}")
    
    if request.method == 'GET':
        try:
            logger.info(f"[NOTIFICATION DEBUG] GET - Querying NotificationSettings for user {user_id}")
            settings = NotificationSettings.query.filter_by(user_id=user_id).first()
            
            if not settings:
                logger.info(f"[NOTIFICATION DEBUG] GET - No settings found, returning DEFAULTS")
                default_response = {
                    'email_on_alert': False,
                    'email_on_notification': True,  # PJ6001: Default to True for notifications
                    'email_daily_diary_reminder': False,
                    'email_on_new_message': True,
                    'follow_requests': True,
                    'parameter_triggers': True,
                    'daily_reminder': False,
                    'weekly_summary': False,
                    'diary_reminder_time': '09:00',
                    'diary_reminder_timezone': 'UTC'
                }
                logger.info(f"[NOTIFICATION DEBUG] GET - Returning: {default_response}")
                return jsonify(default_response)
            
            logger.info(f"[NOTIFICATION DEBUG] GET - Settings FOUND for user {user_id}")
            logger.info(f"[NOTIFICATION DEBUG] GET - Raw DB values:")
            logger.info(f"[NOTIFICATION DEBUG]   email_on_alert: {settings.email_on_alert} (type: {type(settings.email_on_alert)})")
            logger.info(f"[NOTIFICATION DEBUG]   email_daily_diary_reminder: {settings.email_daily_diary_reminder} (type: {type(settings.email_daily_diary_reminder)})")
            
            response_data = {
                'email_on_alert': settings.email_on_alert or False,
                'email_on_notification': settings.email_on_notification if hasattr(settings, 'email_on_notification') and settings.email_on_notification is not None else True,  # PJ6001
                'email_daily_diary_reminder': settings.email_daily_diary_reminder or False,
                'email_on_new_message': settings.email_on_new_message if settings.email_on_new_message is not None else True,
                'follow_requests': settings.follow_requests,
                'parameter_triggers': settings.parameter_triggers,
                'daily_reminder': settings.daily_reminder,
                'weekly_summary': settings.weekly_summary,
                'diary_reminder_time': settings.diary_reminder_time or '09:00',
                'diary_reminder_timezone': settings.diary_reminder_timezone or 'UTC'
            }
            logger.info(f"[NOTIFICATION DEBUG] GET - Returning: {response_data}")
            logger.info(f"[NOTIFICATION DEBUG] ========================================")
            return jsonify(response_data)
        except Exception as e:
            logger.error(f"[NOTIFICATION DEBUG] GET ERROR: {str(e)}")
            logger.error(f"[NOTIFICATION DEBUG] Traceback: {traceback.format_exc()}")
            return jsonify({'error': 'Failed to get settings'}), 500
    
    elif request.method == 'PUT':
        try:
            data = request.get_json()
            logger.info(f"[NOTIFICATION DEBUG] PUT - Received data: {data}")
            
            user = db.session.get(User, user_id)
            logger.info(f"[NOTIFICATION DEBUG] PUT - User found: {user is not None}")
            
            settings = NotificationSettings.query.filter_by(user_id=user_id).first()
            logger.info(f"[NOTIFICATION DEBUG] PUT - Existing settings found: {settings is not None}")
            
            if not settings:
                logger.info(f"[NOTIFICATION DEBUG] PUT - Creating NEW NotificationSettings for user {user_id}")
                settings = NotificationSettings(user_id=user_id)
                db.session.add(settings)
            
            # Track if email_on_alert was just enabled
            was_email_on_alert_enabled = settings.email_on_alert
            logger.info(f"[NOTIFICATION DEBUG] PUT - BEFORE update:")
            logger.info(f"[NOTIFICATION DEBUG]   email_on_alert was: {was_email_on_alert_enabled}")
            logger.info(f"[NOTIFICATION DEBUG]   email_daily_diary_reminder was: {settings.email_daily_diary_reminder}")
            
            # Update settings based on provided data
            if 'email_on_alert' in data:
                logger.info(f"[NOTIFICATION DEBUG] PUT - Setting email_on_alert to: {data['email_on_alert']}")
                settings.email_on_alert = data['email_on_alert']
            if 'email_on_notification' in data:  # PJ6001: Handle email_on_notification
                logger.info(f"[NOTIFICATION DEBUG] PUT - Setting email_on_notification to: {data['email_on_notification']}")
                settings.email_on_notification = data['email_on_notification']
            if 'email_daily_diary_reminder' in data:
                logger.info(f"[NOTIFICATION DEBUG] PUT - Setting email_daily_diary_reminder to: {data['email_daily_diary_reminder']}")
                settings.email_daily_diary_reminder = data['email_daily_diary_reminder']
            if 'email_on_new_message' in data:
                settings.email_on_new_message = data['email_on_new_message']
            if 'follow_requests' in data:
                settings.follow_requests = data['follow_requests']
            if 'parameter_triggers' in data:
                settings.parameter_triggers = data['parameter_triggers']
            if 'daily_reminder' in data:
                settings.daily_reminder = data['daily_reminder']
            if 'weekly_summary' in data:
                settings.weekly_summary = data['weekly_summary']
            if 'diary_reminder_time' in data:
                # Validate time format (HH:MM in 24-hour format)
                time_str = data['diary_reminder_time']
                import re
                if re.match(r'^([01]\d|2[0-3]):([0-5]\d)$', time_str):
                    settings.diary_reminder_time = time_str
                    logger.info(f"[NOTIFICATION DEBUG] PUT - Setting diary_reminder_time to: {time_str}")
                else:
                    logger.warning(f"[NOTIFICATION DEBUG] PUT - Invalid time format: {time_str}")
            if 'diary_reminder_timezone' in data:
                settings.diary_reminder_timezone = data['diary_reminder_timezone']
                logger.info(f"[NOTIFICATION DEBUG] PUT - Setting diary_reminder_timezone to: {data['diary_reminder_timezone']}")
            
            db.session.commit()
            logger.info(f"[NOTIFICATION DEBUG] PUT - Committed to database")
            logger.info(f"[NOTIFICATION DEBUG] PUT - AFTER update:")
            logger.info(f"[NOTIFICATION DEBUG]   email_on_alert is now: {settings.email_on_alert}")
            logger.info(f"[NOTIFICATION DEBUG]   email_daily_diary_reminder is now: {settings.email_daily_diary_reminder}")
            logger.info(f"[NOTIFICATION DEBUG]   diary_reminder_time is now: {settings.diary_reminder_time}")
            logger.info(f"[NOTIFICATION DEBUG]   diary_reminder_timezone is now: {settings.diary_reminder_timezone}")
            logger.info(f"[NOTIFICATION DEBUG] ========================================")
            
            # If email_on_alert was just turned ON, send all existing unread alerts as emails
            # PJ6011: EXCLUDE trigger alerts - they use consolidated batch emails only
            emails_sent = 0
            if 'email_on_alert' in data and data['email_on_alert'] and not was_email_on_alert_enabled:
                if user and user.email:
                    try:
                        # Get all unread alerts for this user, EXCLUDING trigger alerts
                        # Trigger alerts use consolidated batch emails (send_consolidated_wellness_alert_email)
                        unread_alerts = Alert.query.filter(
                            Alert.user_id == user_id,
                            Alert.is_read == False,
                            Alert.alert_category != 'trigger'  # PJ6011: Skip trigger alerts
                        ).order_by(Alert.created_at.desc()).limit(50).all()
                        logger.info(f"[PJ6011] Sending individual emails for {len(unread_alerts)} non-trigger alerts")
                        
                        user_language = user.preferred_language or 'en'
                        
                        for alert in unread_alerts:
                            try:
                                # Parse alert title (may be JSON with translation key)
                                alert_title = alert.title
                                try:
                                    title_data = json.loads(alert.title)
                                    if isinstance(title_data, dict) and 'key' in title_data:
                                        # It's a translation key, use a generic title
                                        alert_title = title_data.get('params', {}).get('username', 'Alert')
                                        if 'new_message' in title_data.get('key', ''):
                                            alert_title = f"New message from {alert_title}"
                                        elif 'started_following' in title_data.get('key', ''):
                                            alert_title = f"{alert_title} started following you"
                                        elif 'invitation' in title_data.get('key', '').lower():
                                            alert_title = "New invitation"
                                except:
                                    pass  # Not JSON, use as-is
                                
                                send_alert_notification_email(
                                    user.email,
                                    alert_title,
                                    alert.content or '',
                                    user_language
                                )
                                emails_sent += 1
                            except Exception as email_err:
                                logger.error(f"Error sending alert email: {str(email_err)}")
                                continue
                        
                        logger.info(f"Sent {emails_sent} alert emails to user {user_id}")
                    except Exception as batch_err:
                        logger.error(f"Error sending batch alert emails: {str(batch_err)}")
            
            return jsonify({
                'success': True,
                'email_on_alert': settings.email_on_alert or False,
                'email_daily_diary_reminder': settings.email_daily_diary_reminder or False,
                'email_on_new_message': settings.email_on_new_message if settings.email_on_new_message is not None else True,
                'emails_sent': emails_sent
            })
        except Exception as e:
            logger.error(f"Update notification settings error: {str(e)}")
            db.session.rollback()
            return jsonify({'error': 'Failed to update settings'}), 500


# =====================
# MESSAGES ROUTES
# =====================

@app.route('/api/messages', methods=['GET', 'POST'])
@login_required
def messages():
    """Get or send messages"""
    user_id = session.get('user_id')

    if request.method == 'GET':
        try:
            recipient_id = request.args.get('recipient_id', type=int)

            if recipient_id:
                # Get messages for specific conversation
                messages_stmt = select(Message).filter(
                    or_(
                        and_(Message.sender_id == user_id, Message.recipient_id == recipient_id),
                        and_(Message.sender_id == recipient_id, Message.recipient_id == user_id)
                    )
                ).order_by(Message.created_at)  # Chronological order for conversation view

                messages = db.session.execute(messages_stmt).scalars().all()

                def format_message(msg):
                    try:
                        sender = db.session.get(User, msg.sender_id)
                        recipient = db.session.get(User, msg.recipient_id)

                        # Ensure both users exist and have usernames
                        if not sender or not recipient:
                            logger.warning(f"Missing user data for message {msg.id}")
                            return None

                        return {
                            'id': msg.id,
                            'sender': {
                                'id': sender.id if sender else msg.sender_id,
                                'username': sender.username if sender else 'Unknown User'
                            },
                            'recipient': {
                                'id': recipient.id if recipient else msg.recipient_id,
                                'username': recipient.username if recipient else 'Unknown User'
                            },
                            'content': msg.content or '',
                            'is_read': msg.is_read,
                            'created_at': msg.created_at.isoformat() if msg.created_at else datetime.utcnow().isoformat()
                        }
                    except Exception as e:
                        logger.error(f"Error formatting message {msg.id}: {str(e)}")
                        return None

                formatted_messages = [m for msg in messages if (m := format_message(msg)) is not None]
                return jsonify({'messages': formatted_messages})

            else:
                # Get all messages for overview
                sent_stmt = select(Message).filter_by(sender_id=user_id).order_by(desc(Message.created_at)).limit(50)
                sent = db.session.execute(sent_stmt).scalars().all()

                received_stmt = select(Message).filter_by(recipient_id=user_id).order_by(
                    desc(Message.created_at)).limit(50)
                received = db.session.execute(received_stmt).scalars().all()

                def format_message(msg):
                    try:
                        sender = db.session.get(User, msg.sender_id)
                        recipient = db.session.get(User, msg.recipient_id)

                        if not sender or not recipient:
                            logger.warning(f"Missing user data for message {msg.id}")
                            return None

                        return {
                            'id': msg.id,
                            'sender': {
                                'id': sender.id if sender else msg.sender_id,
                                'username': sender.username if sender else 'Unknown User'
                            },
                            'recipient': {
                                'id': recipient.id if recipient else msg.recipient_id,
                                'username': recipient.username if recipient else 'Unknown User'
                            },
                            'content': msg.content or '',
                            'is_read': msg.is_read,
                            'created_at': msg.created_at.isoformat() if msg.created_at else datetime.utcnow().isoformat()
                        }
                    except Exception as e:
                        logger.error(f"Error formatting message {msg.id}: {str(e)}")
                        return None

                return jsonify({
                    'sent': [m for msg in sent if (m := format_message(msg)) is not None],
                    'received': [m for msg in received if (m := format_message(msg)) is not None]
                })

        except Exception as e:
            logger.error(f"Get messages error: {str(e)}")
            return jsonify({'error': 'Failed to get messages'}), 500

    elif request.method == 'POST':
        try:
            data = request.json
            recipient_id = data.get('recipient_id')
            content = data.get('content', '').strip()

            if not recipient_id or not content:
                return jsonify({'error': 'Missing recipient or content'}), 400

            # Check if recipient exists
            recipient = db.session.get(User, recipient_id)
            if not recipient:
                return jsonify({'error': 'Recipient not found'}), 404

            # Check content
            moderation = content_moderator.check_content(content)
            if not moderation['safe']:
                return jsonify({'error': moderation.get('message', 'Content not allowed')}), 400

            # Create message
            message = Message(
                sender_id=user_id,
                recipient_id=recipient_id,
                content=sanitize_input(content)
            )
            db.session.add(message)

            # Create alert for recipient - FIX 7: Safe sender name retrieval
            try:
                sender = db.session.get(User, user_id)
                sender_name = sender.username if sender and sender.username else 'Someone'
            except Exception as e:
                logger.error(f"Error getting sender username: {str(e)}")
                sender_name = 'Someone'

            # Create alert with optional email notification (for email_on_alert setting)
            # PJ6003: Messages should use create_notification_with_email, not create_alert_with_email
            alert = create_notification_with_email(
                user_id=recipient_id,
                title=json.dumps({
                    'key': 'alerts.new_message_from',
                    'params': {'username': sender_name}
                }),
                content=content[:100] + '...' if len(content) > 100 else content,
                alert_type='info',
                source_user_id=user_id,
                alert_category='message'
            )

            # PJ6005: Removed redundant dedicated message email (send_new_message_notification_email)
            # The notification email via create_notification_with_email above already includes
            # the message info ("New message from username" + message preview).
            # Previously there were TWO emails sent for each message:
            # 1. Notification email (via create_notification_with_email) - KEPT
            # 2. Dedicated message email (via send_new_message_notification_email) - REMOVED
            # Now only #1 is sent, controlled by "Email me on new notifications" setting.

            # Update activity for today
            today = datetime.utcnow().date()
            activity_stmt = select(Activity).filter_by(user_id=user_id, activity_date=today)
            activity = db.session.execute(activity_stmt).scalar_one_or_none()

            if not activity:
                activity = Activity(user_id=user_id, activity_date=today)
                db.session.add(activity)

            activity.message_count = (activity.message_count or 0) + 1

            db.session.commit()

            return jsonify({'success': True, 'message_id': message.id})

        except Exception as e:
            logger.error(f"Send message error: {str(e)}")
            db.session.rollback()
            return jsonify({'error': 'Failed to send message'}), 500


@app.route('/api/messages/read/<int:recipient_id>', methods=['POST'])
@login_required
def mark_messages_read(recipient_id):
    """Mark all messages from a recipient as read"""
    try:
        user_id = session.get('user_id')

        # Mark all unread messages from this sender as read
        unread_messages = Message.query.filter_by(
            sender_id=recipient_id,
            recipient_id=user_id,
            is_read=False
        ).all()

        for message in unread_messages:
            message.is_read = True

        db.session.commit()

        return jsonify({'success': True, 'marked_count': len(unread_messages)})
    except Exception as e:
        logger.error(f"Mark messages error: {str(e)}")
        db.session.rollback()
        return jsonify({'error': 'Failed to mark messages'}), 500


@app.route('/api/messages/conversations')
@login_required
def get_conversations():
    """Get conversation list with improved functionality"""
    try:
        user_id = session['user_id']

        # Get unique conversation partners - SQLAlchemy 2.0 style
        sent_stmt = select(Message.recipient_id).filter_by(sender_id=user_id).distinct()
        sent_partners = db.session.execute(sent_stmt).scalars().all()

        received_stmt = select(Message.sender_id).filter_by(recipient_id=user_id).distinct()
        received_partners = db.session.execute(received_stmt).scalars().all()

        # Combine and deduplicate
        partner_ids = set(sent_partners + received_partners)

        conversations = []
        for partner_id in partner_ids:
            try:
                partner = db.session.get(User, partner_id)

                # Skip if partner doesn't exist or has no username
                if not partner or not partner.username:
                    logger.warning(f"Skipping conversation with invalid partner {partner_id}")
                    continue

                # Get last message - SQLAlchemy 2.0 style
                last_msg_stmt = select(Message).filter(
                    or_(
                        and_(Message.sender_id == user_id, Message.recipient_id == partner_id),
                        and_(Message.sender_id == partner_id, Message.recipient_id == user_id)
                    )
                ).order_by(desc(Message.created_at))

                last_message = db.session.execute(last_msg_stmt).scalars().first()

                # Count unread messages from this partner
                unread_stmt = select(func.count(Message.id)).filter_by(
                    sender_id=partner_id,
                    recipient_id=user_id,
                    is_read=False
                )
                unread_count = db.session.execute(unread_stmt).scalar() or 0

                # Ensure all fields have safe fallback values
                conversations.append({
                    'user': {
                        'id': partner.id,
                        'username': partner.username or 'Unknown User'
                    },
                    'last_message': {
                        'content': last_message.content if last_message and last_message.content else '',
                        'created_at': last_message.created_at.isoformat() if last_message and last_message.created_at else None,
                        'is_own': last_message.sender_id == user_id if last_message else False
                    },
                    'unread_count': unread_count,
                    'timestamp': last_message.created_at.isoformat() if last_message and last_message.created_at else None
                })
            except Exception as e:
                logger.error(f"Error processing conversation with partner {partner_id}: {str(e)}")
                continue

        # Sort by last message time
        conversations.sort(
            key=lambda x: x['timestamp'] if x['timestamp'] else '',
            reverse=True
        )

        # ✅ FIXED: Wrap in object for frontend consistency
        return jsonify({'conversations': conversations})

    except Exception as e:
        logger.error(f"Get conversations error: {str(e)}")
        return jsonify({'conversations': []})  # Return empty array on error instead of error object


# =====================
# CIRCLES ROUTES
# =====================
@app.route('/api/circles', methods=['GET', 'POST'])
@login_required
def circles():
    """Manage user circles"""
    user_id = session.get('user_id')

    if request.method == 'GET':
        try:
            # NEW: Support viewing another user's circles
            viewing_user_id = request.args.get('user_id', type=int)

            # If no user_id specified, show logged-in user's circles
            target_user_id = viewing_user_id if viewing_user_id else user_id
            is_viewing_own = (target_user_id == user_id)

            # Get target user
            target_user = db.session.get(User, target_user_id)
            if not target_user:
                return jsonify({'error': 'User not found'}), 404

            circles_privacy = getattr(target_user, 'circles_privacy', 'public')

            logger.info(
                f"User {user_id} viewing circles of user {target_user_id} (privacy: {circles_privacy}, is_own: {is_viewing_own})")

            # If viewing someone else's circles, check follow status
            if not is_viewing_own:
                is_following = db.session.execute(
                    select(Follow).filter_by(
                        follower_id=user_id,
                        followed_id=target_user_id
                    )
                ).scalar_one_or_none()

                if not is_following:
                    return jsonify({'error': 'Must be following user to view circles'}), 403

                # If circles are private AND viewing someone else's, return empty
                if circles_privacy == 'private':
                    return jsonify({
                        'public': [],
                        'class_b': [],
                        'class_a': [],
                        'private': True
                    })

                # Determine viewer's circle level for this user
                viewer_circle_stmt = select(Circle).filter(
                    Circle.user_id == target_user_id,
                    Circle.circle_user_id == user_id
                )
                viewer_circle = db.session.execute(viewer_circle_stmt).scalar_one_or_none()

                if viewer_circle:
                    # Map to standardized circle type
                    type_mapping = {
                        'public': 'public',
                        'general': 'public',
                        'class_b': 'class_b',
                        'close_friends': 'class_b',
                        'class_a': 'class_a',
                        'family': 'class_a'
                    }
                    viewer_circle_level = type_mapping.get(viewer_circle.circle_type, 'public')
                else:
                    # Not in any circle, default to public
                    viewer_circle_level = 'public'

                logger.info(f"Viewer {user_id} is in '{viewer_circle_level}' circle for user {target_user_id}")
            else:
                viewer_circle_level = None  # Not used when viewing own

            # Get all circles - SQLAlchemy 2.0 style
            # Filter out circles with NULL circle_user_id to prevent SAWarning
            # Support both old and new naming conventions
            public_stmt = select(Circle).filter(
                Circle.user_id == target_user_id,
                Circle.circle_user_id.isnot(None),
                or_(Circle.circle_type == 'public', Circle.circle_type == 'general')
            )
            public = db.session.execute(public_stmt).scalars().all()

            class_b_stmt = select(Circle).filter(
                Circle.user_id == target_user_id,
                Circle.circle_user_id.isnot(None),
                or_(Circle.circle_type == 'class_b', Circle.circle_type == 'close_friends')
            )
            class_b = db.session.execute(class_b_stmt).scalars().all()

            class_a_stmt = select(Circle).filter(
                Circle.user_id == target_user_id,
                Circle.circle_user_id.isnot(None),
                or_(Circle.circle_type == 'class_a', Circle.circle_type == 'family')
            )
            class_a = db.session.execute(class_a_stmt).scalars().all()

            def get_user_info(circle):
                """Safely get user info, handling None/NULL circle_user_id"""
                if not circle.circle_user_id:
                    logger.warning(f"Circle {circle.id} has NULL circle_user_id for user {target_user_id}")
                    return None

                user = db.session.get(User, circle.circle_user_id)
                if user:
                    return {
                        'id': user.id,
                        'username': user.username,
                        'email': user.email,
                        'display_name': user.username
                    }
                logger.warning(f"User {circle.circle_user_id} not found for circle {circle.id}")
                return None

            # Apply privacy filtering based on circles_privacy AND viewer's circle level
            result = {
                'public': [],
                'class_b': [],
                'class_a': []
            }

            if is_viewing_own:
                # ✅ OWNER ALWAYS SEES ALL THEIR CIRCLES (regardless of privacy setting)
                result['public'] = [info for c in public if (info := get_user_info(c))]
                result['class_b'] = [info for c in class_b if (info := get_user_info(c))]
                result['class_a'] = [info for c in class_a if (info := get_user_info(c))]
            else:
                # ✅ VIEWING SOMEONE ELSE'S CIRCLES - Apply privacy rules

                if circles_privacy == 'private':
                    # Nobody sees anything (already returned earlier, but just in case)
                    result['private'] = True

                elif circles_privacy == 'public':
                    # Everyone sees all circles
                    result['public'] = [info for c in public if (info := get_user_info(c))]
                    result['class_b'] = [info for c in class_b if (info := get_user_info(c))]
                    result['class_a'] = [info for c in class_a if (info := get_user_info(c))]

                elif circles_privacy == 'class_b':
                    # Only Class B and Class A members can see
                    if viewer_circle_level in ['class_b', 'class_a']:
                        result['class_b'] = [info for c in class_b if (info := get_user_info(c))]
                        result['class_a'] = [info for c in class_a if (info := get_user_info(c))]
                        result['public'] = [info for c in public if (info := get_user_info(c))]
                    else:
                        # viewer is only in public circle, can't see anything
                        result['private'] = True

                elif circles_privacy == 'class_a':
                    # Only Class A members can see
                    if viewer_circle_level == 'class_a':
                        result['class_a'] = [info for c in class_a if (info := get_user_info(c))]
                        result['public'] = [info for c in public if (info := get_user_info(c))]
                        result['class_b'] = [info for c in class_b if (info := get_user_info(c))]
                    else:
                        # viewer is not in class_a, can't see anything
                        result['private'] = True

            return jsonify(result)

        except Exception as e:
            logger.error(f"Get circles error: {str(e)}")
            return jsonify({'error': 'Failed to get circles'}), 500

    elif request.method == 'POST':
        try:
            data = request.json
            circle_user_id = data.get('user_id')
            circle_type = data.get('circle_type')

            valid_types = ['public', 'class_b', 'class_a', 'general', 'close_friends', 'family']
            # Map old names to new names
            type_mapping = {
                'general': 'public',
                'close_friends': 'class_b',
                'family': 'class_a',
                'public': 'public',
                'class_b': 'class_b',
                'class_a': 'class_a'
            }
            circle_type = type_mapping.get(circle_type, circle_type)

            if circle_type not in ['public', 'class_b', 'class_a']:
                return jsonify({'error': 'Invalid circle type'}), 400

            # Check if user exists
            if not db.session.get(User, circle_user_id):
                return jsonify({'error': 'User not found'}), 404

            # Check if already in circle - SQLAlchemy 2.0 style
            existing_stmt = select(Circle).filter_by(
                user_id=user_id,
                circle_user_id=circle_user_id,
                circle_type=circle_type
            )
            existing = db.session.execute(existing_stmt).scalar_one_or_none()

            if existing:
                return jsonify({'error': 'User already in this circle'}), 400

            circle = Circle(
                user_id=user_id,
                circle_user_id=circle_user_id,
                circle_type=circle_type
            )
            db.session.add(circle)
            db.session.commit()

            logger.info(f"Added user {circle_user_id} to {circle_type} circle for user {user_id}")
            return jsonify({'success': True, 'message': 'User added to circle'})

        except Exception as e:
            logger.error(f"Add to circle error: {str(e)}")
            db.session.rollback()
            return jsonify({'error': 'Failed to add to circle'}), 500


@app.route('/api/circles/remove', methods=['DELETE'])
@login_required
def remove_from_circle():
    """Remove user from circle"""
    try:
        user_id = session.get('user_id')
        data = request.json

        # Validate data exists
        if not data:
            logger.error("No data provided to remove_from_circle")
            return jsonify({'error': 'No data provided'}), 400

        circle_user_id = data.get('user_id')
        circle_type = data.get('circle_type')

        # Validate required fields
        if not circle_user_id or not circle_type:
            logger.error(f"Missing required fields: user_id={circle_user_id}, circle_type={circle_type}")
            return jsonify({'error': 'Missing required fields'}), 400

        logger.info(f"Removing user {circle_user_id} from circle {circle_type} for user {user_id}")

        # Map old names to new names AND create list of possible matches
        type_mapping = {
            'general': ['public', 'general'],  # Check both new and old
            'public': ['public', 'general'],
            'close_friends': ['class_b', 'close_friends'],
            'class_b': ['class_b', 'close_friends'],
            'family': ['class_a', 'family'],
            'class_a': ['class_a', 'family'],
            'private': ['private']  # Private only has one name
        }

        # Get list of possible type names to check
        possible_types = type_mapping.get(circle_type, [circle_type])

        logger.info(f"Checking for circle types: {possible_types}")

        # Use SQLAlchemy to delete - check for ANY of the possible type names
        stmt = select(Circle).filter(
            Circle.user_id == user_id,
            Circle.circle_user_id == circle_user_id,
            Circle.circle_type.in_(possible_types)  # Check BOTH old and new names
        )
        circle = db.session.execute(stmt).scalar_one_or_none()

        if circle:
            db.session.delete(circle)
            db.session.commit()
            logger.info(f"Successfully removed user {circle_user_id} from circle {circle.circle_type}")
            return jsonify({'success': True})
        else:
            logger.warning(f"No circle membership found for types {possible_types}")
            return jsonify({'error': 'Not found'}), 404

    except Exception as e:
        logger.error(f"Remove from circle error: {str(e)}", exc_info=True)
        db.session.rollback()
        return jsonify({'error': 'Failed to remove from circle'}), 500


def migrate_circle_names():
    """Migrate old circle names to new ones in the database"""
    try:
        # Map old names to new names
        name_mapping = {
            'general': 'public',
            'close_friends': 'class_b',
            'family': 'class_a'
        }

        migrated_count = 0

        # Get all circles with old names and update them
        for old_name, new_name in name_mapping.items():
            stmt = select(Circle).filter_by(circle_type=old_name)
            old_circles = db.session.execute(stmt).scalars().all()

            for circle in old_circles:
                logger.info(f"Migrating circle {circle.id} from '{old_name}' to '{new_name}'")
                circle.circle_type = new_name
                migrated_count += 1

        if migrated_count > 0:
            db.session.commit()
            app.logger.info(f"Circle names migration completed: {migrated_count} circles updated")
        else:
            app.logger.info("Circle names check completed - no migration needed")

    except Exception as e:
        app.logger.error(f"Error checking circle names: {e}")
        db.session.rollback()


@app.route('/api/circles/membership/<int:check_user_id>', methods=['GET'])
@login_required
def check_circle_membership(check_user_id):
    """Check what circle the current user is in for another user"""
    try:
        current_user_id = session.get('user_id')

        # Check if current user is in any of check_user's circles
        circle_stmt = select(Circle).filter_by(
            user_id=check_user_id,
            circle_user_id=current_user_id
        )
        circle = db.session.execute(circle_stmt).scalar_one_or_none()

        if circle:
            # Map old types to new if needed
            type_mapping = {
                'general': 'public',
                'close_friends': 'class_b',
                'family': 'class_a'
            }
            circle_type = type_mapping.get(circle.circle_type, circle.circle_type)
            return jsonify({'circle': circle_type})

        return jsonify({'circle': None})

    except Exception as e:
        logger.error(f"Check circle membership error: {str(e)}")
        return jsonify({'error': 'Failed to check membership'}), 500


def get_my_circles():
    """Get all circles with proper member information"""
    try:
        user_id = session.get('user_id')

        # Check if requesting user's own circles or someone else's
        target_user_id = request.args.get('user_id', user_id, type=int)

        # Initialize viewer_circle_type for later use
        viewer_circle_type = None

        # If viewing someone else's circles, check their privacy settings
        if target_user_id != user_id:
            target_user = db.session.get(User, target_user_id)
            if not target_user:
                return jsonify({'error': 'User not found'}), 404

            privacy_level = target_user.circles_privacy or 'private'

            # Check viewer's circle membership with target user FIRST (before any privacy checks)
            viewer_circle = db.session.execute(
                select(Circle).filter_by(
                    user_id=target_user_id,
                    circle_user_id=user_id
                )
            ).scalars().first()

            if viewer_circle:
                type_mapping = {
                    'general': 'public',
                    'close_friends': 'class_b',
                    'family': 'class_a',
                    'public': 'public',
                    'class_b': 'class_b',
                    'class_a': 'class_a'
                }
                viewer_circle_type = type_mapping.get(viewer_circle.circle_type, 'public')

            # Apply privacy filtering with consistent response format
            if privacy_level == 'private':
                return jsonify({
                    'private': True,
                    'message': 'Circles set to private',
                    'public': [],
                    'class_b': [],
                    'class_a': [],
                    'viewer_circle_type': viewer_circle_type,
                    'viewing_user_id': target_user_id
                })

            if privacy_level == 'class_a' and viewer_circle_type != 'class_a':
                return jsonify({
                    'private': True,
                    'message': 'Circles set to private',
                    'public': [],
                    'class_b': [],
                    'class_a': [],
                    'viewer_circle_type': viewer_circle_type,
                    'viewing_user_id': target_user_id
                })

            if privacy_level == 'class_b' and viewer_circle_type not in ['class_a', 'class_b']:
                return jsonify({
                    'private': True,
                    'message': 'Circles set to private',
                    'public': [],
                    'class_b': [],
                    'class_a': [],
                    'viewer_circle_type': viewer_circle_type,
                    'viewing_user_id': target_user_id
                })

            # For public or allowed viewers, continue with circles
            circles_stmt = select(Circle).filter_by(user_id=target_user_id)
        else:
            # User viewing own circles - no restrictions
            circles_stmt = select(Circle).filter_by(user_id=user_id)

        circles = db.session.execute(circles_stmt).scalars().all()

        result = {
            'public': [],
            'class_b': [],
            'class_a': []
        }

        type_mapping = {
            'general': 'public',
            'close_friends': 'class_b',
            'family': 'class_a',
            'public': 'public',
            'class_b': 'class_b',
            'class_a': 'class_a'
        }

        for circle in circles:
            user = db.session.get(User, circle.circle_user_id)
            if user:
                user_info = {
                    'user_id': user.id,
                    'username': user.username
                }
                circle_type = type_mapping.get(circle.circle_type, circle.circle_type)
                if circle_type in result:
                    result[circle_type].append(user_info)

        # If viewing another user's circles, include the viewer's circle type
        if target_user_id != user_id:
            result['viewer_circle_type'] = viewer_circle_type
            result['viewing_user_id'] = target_user_id

        return jsonify(result)

    except Exception as e:
        logger.error(f"Get my circles error: {str(e)}")
        return jsonify({'error': 'Failed to get circles'}), 500


@app.route('/api/circles/privacy', methods=['GET'])
@login_required
def get_circles_privacy():
    """Get user's circles privacy setting - supports querying other users"""
    try:
        current_user_id = session.get('user_id')
        # Allow checking another user's privacy setting via query parameter
        target_user_id = request.args.get('user_id', current_user_id, type=int)

        user = db.session.get(User, target_user_id)

        if not user:
            return jsonify({'error': 'User not found'}), 404

        return jsonify({
            'privacy': user.circles_privacy or 'private',
            'circles_privacy': user.circles_privacy or 'private'  # Keep both for backwards compatibility
        })
    except Exception as e:
        logger.error(f"Get circles privacy error: {str(e)}")
        return jsonify({'error': 'Failed to get circles privacy'}), 500


@app.route('/api/circles/privacy', methods=['POST'])
@login_required
def update_circles_privacy():
    """Update user's circles privacy setting"""
    try:
        user_id = session.get('user_id')
        data = request.get_json()

        # Accept either 'privacy_level', 'privacy', or 'circles_privacy' from frontend
        privacy_level = data.get('privacy_level') or data.get('privacy') or data.get('circles_privacy')

        # Validate privacy level
        if not privacy_level or privacy_level not in ['public', 'class_b', 'class_a', 'private']:
            logger.error(f"Invalid privacy level received: {privacy_level}, data: {data}")
            return jsonify({'error': 'Invalid privacy level'}), 400
        user = db.session.get(User, user_id)
        if not user:
            return jsonify({'error': 'User not found'}), 404

        user.circles_privacy = privacy_level
        db.session.commit()

        return jsonify({
            'message': 'Privacy updated successfully',
            'circles_privacy': privacy_level
        })

    except Exception as e:
        logger.error(f"Update circles privacy error: {str(e)}")
        db.session.rollback()
        return jsonify({'error': 'Failed to update privacy'}), 500


@app.route('/api/circles/recommendations', methods=['GET'])
@login_required
def get_circle_recommendations():
    """Get recommended users to add to circles - prioritizes same city and mutual connections"""
    try:
        user_id = session.get('user_id')
        user = db.session.get(User, user_id)

        logger.info(f"[CIRCLE RECS] Starting recommendations for user {user_id}")

        if not user:
            logger.warning(f"[CIRCLE RECS] User {user_id} not found")
            return jsonify({'recommendations': [], 'debug': 'User not found'}), 200

        logger.info(f"[CIRCLE RECS] User city: {user.selected_city}")

        recommendations = []
        seen_ids = set([user_id])

        # Get users already in any circle
        existing_circle_users = db.session.execute(
            select(Circle.circle_user_id).filter_by(user_id=user_id)
        ).scalars().all()
        seen_ids.update(existing_circle_users)
        logger.info(f"[CIRCLE RECS] Users already in circles: {len(existing_circle_users)}")

        # PRIORITY 1: Mutual connections (users who follow me AND I follow them)
        # Get users I follow
        my_following_ids = db.session.execute(
            select(Follow.followed_id).filter_by(follower_id=user_id)
        ).scalars().all()
        logger.info(f"[CIRCLE RECS] Users I follow: {len(my_following_ids)}")

        # Get users who follow me
        my_follower_ids = db.session.execute(
            select(Follow.follower_id).filter_by(followed_id=user_id)
        ).scalars().all()
        logger.info(f"[CIRCLE RECS] Users following me: {len(my_follower_ids)}")

        # Find mutual connections
        mutual_ids = set(my_following_ids) & set(my_follower_ids)
        logger.info(f"[CIRCLE RECS] Mutual connections: {len(mutual_ids)}")

        for mutual_id in mutual_ids:
            if mutual_id in seen_ids or len(recommendations) >= 20:
                continue

            mutual_user = db.session.get(User, mutual_id)
            if mutual_user and mutual_user.is_active:
                seen_ids.add(mutual_id)
                reason = 'Mutual connection'
                if user.selected_city and mutual_user.selected_city == user.selected_city:
                    reason = 'Mutual connection & same city'

                recommendations.append({
                    'id': mutual_user.id,
                    'username': mutual_user.username,
                    'email': mutual_user.email,
                    'selected_city': mutual_user.selected_city,
                    'reason': reason,
                    'reason_key': 'circles.reason_mutual' if 'same city' not in reason else 'circles.reason_mutual_city'
                })

        logger.info(f"[CIRCLE RECS] After mutual connections: {len(recommendations)} recommendations")

        # PRIORITY 2: Users in same city that I follow (but not mutual)
        if user.selected_city and len(recommendations) < 20:
            for followed_id in my_following_ids:
                if followed_id in seen_ids or len(recommendations) >= 20:
                    continue

                followed_user = db.session.get(User, followed_id)
                if followed_user and followed_user.is_active and followed_user.selected_city == user.selected_city:
                    seen_ids.add(followed_id)
                    recommendations.append({
                        'id': followed_user.id,
                        'username': followed_user.username,
                        'email': followed_user.email,
                        'selected_city': followed_user.selected_city,
                        'reason': 'Same city (following)',
                        'reason_key': 'circles.reason_same_city'
                    })

        logger.info(f"[CIRCLE RECS] After same city following: {len(recommendations)} recommendations")

        # PRIORITY 3: Other users I follow
        if len(recommendations) < 20:
            for followed_id in my_following_ids:
                if followed_id in seen_ids or len(recommendations) >= 20:
                    continue

                followed_user = db.session.get(User, followed_id)
                if followed_user and followed_user.is_active:
                    seen_ids.add(followed_id)
                    recommendations.append({
                        'id': followed_user.id,
                        'username': followed_user.username,
                        'email': followed_user.email,
                        'selected_city': followed_user.selected_city,
                        'reason': 'You follow',
                        'reason_key': 'circles.reason_following'
                    })

        logger.info(f"[CIRCLE RECS] After other following: {len(recommendations)} recommendations")

        # PRIORITY 4: Same city users (not following yet) - LIKE "WHO TO FOLLOW"
        if user.selected_city and len(recommendations) < 20:
            logger.info(f"[CIRCLE RECS] Looking for same city users in: {user.selected_city}")
            same_city_users = User.query.filter(
                User.selected_city == user.selected_city,
                User.id != user_id,
                User.is_active == True
            ).limit(30).all()
            
            logger.info(f"[CIRCLE RECS] Found {len(same_city_users)} users in same city")
            
            for city_user in same_city_users:
                if city_user.id in seen_ids or len(recommendations) >= 20:
                    continue
                    
                seen_ids.add(city_user.id)
                recommendations.append({
                    'id': city_user.id,
                    'username': city_user.username,
                    'email': city_user.email,
                    'selected_city': city_user.selected_city,
                    'reason': 'Same city',
                    'reason_key': 'circles.reason_same_city'
                })

        logger.info(f"[CIRCLE RECS] After same city (all): {len(recommendations)} recommendations")

        # PRIORITY 5: Friends of friends (people in circles of people I follow)
        if len(recommendations) < 20:
            for circle_user_id in existing_circle_users[:10]:  # Limit to prevent slow queries
                if len(recommendations) >= 20:
                    break

                their_circles = db.session.execute(
                    select(Circle.circle_user_id).filter_by(
                        user_id=circle_user_id
                    ).filter(Circle.circle_user_id != user_id)
                ).scalars().all()

                for potential_id in their_circles[:5]:
                    if len(recommendations) >= 20:
                        break

                    if potential_id in seen_ids:
                        continue

                    potential_user = db.session.get(User, potential_id)
                    if potential_user and potential_user.is_active:
                        seen_ids.add(potential_user.id)

                        reason = 'Friend of friend'
                        if user.selected_city and potential_user.selected_city == user.selected_city:
                            reason = 'Same city & friend of friend'

                        recommendations.append({
                            'id': potential_user.id,
                            'username': potential_user.username,
                            'email': potential_user.email,
                            'selected_city': potential_user.selected_city,
                            'reason': reason,
                            'reason_key': 'circles.reason_friend_of_friend'
                        })

        logger.info(f"[CIRCLE RECS] Final count: {len(recommendations)} recommendations")
        
        return jsonify({
            'recommendations': recommendations[:20],
            'debug': {
                'user_city': user.selected_city,
                'following_count': len(my_following_ids),
                'followers_count': len(my_follower_ids),
                'mutual_count': len(mutual_ids),
                'in_circles_count': len(existing_circle_users)
            }
        })

    except Exception as e:
        logger.error(f"[CIRCLE RECS] ERROR: {str(e)}")
        import traceback
        logger.error(f"[CIRCLE RECS] Traceback: {traceback.format_exc()}")
        return jsonify({'error': 'Failed to get recommendations', 'debug': str(e)}), 500


# =====================
# FEED & POSTS ROUTES
# =====================
@app.route('/api/feed', methods=['GET'])
@login_required
@rate_limit_endpoint(max_requests=60, window=60)  # 60 requests per minute
def get_feed():
    """Get user feed with caching"""
    try:
        user_id = session['user_id']
        page = request.args.get('page', 1, type=int)

        # Try cache first
        cache_key = f'feed:{user_id}:{page}'
        if REDIS_URL:
            try:
                r = redis.from_url(REDIS_URL)
                cached_feed = r.get(cache_key)
                if cached_feed:
                    logger.debug(f'Cache hit for feed:{user_id}:{page}')
                    return jsonify(json.loads(cached_feed))
            except Exception as e:
                logger.warning(f'Cache read failed: {e}')

        # Get ONLY own posts for Feed page
        posts_stmt = select(Post).filter(
            Post.user_id == user_id,
            Post.is_published == True
        ).order_by(desc(Post.created_at)).limit(50)

        posts = db.session.execute(posts_stmt).scalars().all()

        feed = []
        for post in posts:
            # Count reactions
            reactions_count = db.session.execute(
                select(func.count(Reaction.id)).filter_by(post_id=post.id)
            ).scalar() or 0

            # Count comments
            comments_count = db.session.execute(
                select(func.count(Comment.id)).filter_by(post_id=post.id)
            ).scalar() or 0

            feed.append({
                'id': post.id,
                'content': post.content,
                'author': post.author.username,
                'author_id': post.author.id,
                'likes': post.likes,
                'reactions_count': reactions_count,
                'comments_count': comments_count,
                'created_at': post.created_at.isoformat()
            })

        result = {'posts': feed}

        # Cache result for 5 minutes
        if REDIS_URL:
            try:
                r = redis.from_url(REDIS_URL)
                r.setex(cache_key, 300, json.dumps(result))
                logger.debug(f'Cached feed:{user_id}:{page}')
            except Exception as e:
                logger.warning(f'Cache write failed: {e}')

        return jsonify(result)

    except Exception as e:
        logger.error(f"Feed error: {str(e)}")
        return jsonify({'posts': []})


# Add these new endpoints right here, after get_feed() function ends


@app.route('/api/feed/hierarchical', methods=['GET'])
@login_required
def get_hierarchical_feed():
    """Get feed posts with visibility hierarchy applied"""
    try:
        user_id = session.get('user_id')
        page = request.args.get('page', 1, type=int)
        per_page = 20

        # Get all posts the user can see
        visible_posts = []

        # 1. User's own posts (all visible)
        own_posts = Post.query.filter_by(user_id=user_id, is_published=True).all()
        visible_posts.extend(own_posts)

        # 2. Check circles user belongs to for other users' posts
        # Get circles where current user is a member
        circles_in = db.session.execute(
            select(Circle).filter_by(circle_user_id=user_id)
        ).scalars().all()

        for circle in circles_in:
            # Get posts from this circle owner
            owner_posts = Post.query.filter_by(
                user_id=circle.user_id,
                is_published=True
            ).all()

            for post in owner_posts:
                # Apply hierarchy rules
                if post.visibility == 'public':
                    # Public posts visible to all circle members
                    if post not in visible_posts:
                        visible_posts.append(post)
                elif post.visibility == 'class_b':
                    # Class B posts visible to Class B and Class A members
                    if circle.circle_type in ['class_b', 'class_a']:
                        if post not in visible_posts:
                            visible_posts.append(post)
                elif post.visibility == 'class_a':
                    # Class A posts only visible to Class A members
                    if circle.circle_type == 'class_a':
                        if post not in visible_posts:
                            visible_posts.append(post)

        # Sort by created_at descending
        visible_posts.sort(key=lambda x: x.created_at, reverse=True)

        # Paginate
        start = (page - 1) * per_page
        end = start + per_page
        paginated_posts = visible_posts[start:end]

        # Format response
        posts_data = []
        for post in paginated_posts:
            author = db.session.get(User, post.user_id)
            posts_data.append({
                'id': post.id,
                'author': {
                    'id': author.id,
                    'username': author.username
                } if author else None,
                'content': post.content,
                'visibility': post.visibility,
                'created_at': post.created_at.isoformat(),
                'likes': post.likes,
                'comments_count': len(post.comments)
            })

        return jsonify({
            'posts': posts_data,
            'has_more': len(visible_posts) > end
        })

    except Exception as e:
        logger.error(f"Hierarchical feed error: {str(e)}")
        return jsonify({'error': 'Failed to load feed'}), 500


@app.route('/api/parameters/hierarchical/<int:view_user_id>', methods=['GET'])
@login_required
def get_hierarchical_parameters(view_user_id):
    """Get parameters with visibility hierarchy applied"""
    try:
        current_user_id = session.get('user_id')

        # If viewing own parameters, return all
        if view_user_id == current_user_id:
            params = SavedParameters.query.filter_by(user_id=view_user_id).all()
            return jsonify({
                'parameters': [p.to_dict(viewer_id=current_user_id) for p in params]
            })

        # Check what circle current user is in for the viewed user
        circle = Circle.query.filter_by(
            user_id=view_user_id,
            circle_user_id=current_user_id
        ).first()

        if not circle:
            # Not in any circle - can only see public
            privacy_level = 'public'
        else:
            privacy_level = circle.circle_type

        # Get parameters and apply visibility rules
        params = SavedParameters.query.filter_by(user_id=view_user_id).all()
        visible_params = []

        for param in params:
            param_dict = param.to_dict(viewer_id=current_user_id, privacy_level=privacy_level)
            visible_params.append(param_dict)

        return jsonify({'parameters': visible_params})

    except Exception as e:
        logger.error(f"Hierarchical parameters error: {str(e)}")
        return jsonify({'error': 'Failed to load parameters'}), 500


@app.route('/api/feed/dates')
@login_required
def get_feed_saved_dates():
    """Get all dates with feed entries, organized by visibility level"""
    try:
        user_id = session['user_id']

        # Get all posts for this user
        posts = Post.query.filter_by(user_id=user_id).all()

        # Organize dates by visibility
        dates_by_visibility = {
            'general': [],
            'close_friends': [],
            'family': [],
            'private': []
        }

        for post in posts:
            date_str = post.created_at.strftime('%Y-%m-%d')
            visibility = post.visibility or 'private'  # Default to private if null

            if visibility in dates_by_visibility:
                if date_str not in dates_by_visibility[visibility]:
                    dates_by_visibility[visibility].append(date_str)

        # For backward compatibility, also return combined dates
        all_dates = set()
        for dates in dates_by_visibility.values():
            all_dates.update(dates)

        # Return BOTH combined dates AND dates separated by visibility
        return jsonify({
            'dates': {date: True for date in all_dates},
            'dates_by_visibility': dates_by_visibility
        })

    except Exception as e:
        logger.error(f"Error fetching feed dates: {e}")
        return jsonify({'dates': {}, 'dates_by_visibility': {}})

@app.route('/api/users/<int:user_id>/feed/dates')
@login_required
def get_user_feed_dates(user_id):
    """Get dates that have feed entries for a specific user with visibility-based access"""
    try:
        current_user_id = session.get('user_id')

        # If viewing own dates, return all
        if user_id == current_user_id:
            posts = db.session.query(
                db.func.date(Post.created_at).label('date'),
                Post.visibility
            ).filter_by(
                user_id=user_id
            ).group_by(
                db.func.date(Post.created_at),
                Post.visibility
            ).all()

            dates_with_visibility = {}

            for post in posts:
                date_str = post.date.strftime('%Y-%m-%d')
                if date_str not in dates_with_visibility:
                    dates_with_visibility[date_str] = []

                visibility = post.visibility or 'private'
                if visibility not in dates_with_visibility[date_str]:
                    dates_with_visibility[date_str].append(visibility)

            return jsonify({'dates': dates_with_visibility})

        # Check circle membership for other users
        membership = Circle.query.filter_by(
            user_id=user_id,
            circle_user_id=current_user_id
        ).first()

        # Determine which visibility levels current user can see
        visible_levels = ['general']  # Everyone can see public
        if membership:
            if membership.circle_type in ['class_a', 'family']:
                visible_levels = ['general', 'close_friends', 'family']
            elif membership.circle_type in ['class_b', 'close_friends']:
                visible_levels = ['general', 'close_friends']

        # Get posts filtered by visible visibility levels
        posts = db.session.query(
            db.func.date(Post.created_at).label('date'),
            Post.visibility
        ).filter(
            Post.user_id == user_id,
            Post.visibility.in_(visible_levels)
        ).group_by(
            db.func.date(Post.created_at),
            Post.visibility
        ).all()

        dates_with_visibility = {}

        for post in posts:
            date_str = post.date.strftime('%Y-%m-%d')
            if date_str not in dates_with_visibility:
                dates_with_visibility[date_str] = []

            visibility = post.visibility or 'general'
            if visibility not in dates_with_visibility[date_str]:
                dates_with_visibility[date_str].append(visibility)

        return jsonify({'dates': dates_with_visibility})

    except Exception as e:
        logger.error(f"Error getting user feed dates: {e}")
        return jsonify({'dates': {}})

@app.route('/api/users/<int:user_id>/feed/<date>')
@login_required
def get_user_feed_by_date(user_id, date):
    """Get a specific user's feed posts for a date with visibility-based access"""
    try:
        current_user_id = session.get('user_id')

        # Parse the date
        from datetime import datetime, timedelta
        feed_date = datetime.fromisoformat(date).date()

        # Get start and end of the day
        start_datetime = datetime.combine(feed_date, datetime.min.time())
        end_datetime = datetime.combine(feed_date, datetime.max.time())

        # If viewing own posts, return all for that date
        if user_id == current_user_id:
            posts = Post.query.filter(
                Post.user_id == user_id,
                Post.created_at >= start_datetime,
                Post.created_at <= end_datetime
            ).order_by(Post.created_at.desc()).all()

            # Calculate likes and comments for each post
            posts_data = []
            for post in posts:
                # Count likes from Reaction table
                likes_count = db.session.execute(
                    select(func.count(Reaction.id)).filter_by(post_id=post.id, type='like')
                ).scalar() or 0

                # Count comments
                comments_count = db.session.execute(
                    select(func.count(Comment.id)).filter_by(post_id=post.id)
                ).scalar() or 0

                # Check if current user liked this post
                user_liked = db.session.execute(
                    select(Reaction).filter_by(
                        post_id=post.id,
                        user_id=current_user_id,
                        type='like'
                    )
                ).scalar_one_or_none() is not None

                posts_data.append({
                    'id': post.id,
                    'content': post.content,
                    'created_at': post.created_at.isoformat() if post.created_at else None,
                    'likes_count': likes_count,
                    'comments_count': comments_count,
                    'user_liked': user_liked,
                    'visibility': post.visibility
                })

            return jsonify({'posts': posts_data})

        # Check if following this user
        is_following = Follow.query.filter_by(
            follower_id=current_user_id,
            followed_id=user_id
        ).first() is not None

        if not is_following:
            return jsonify({'error': 'Must be following user to view posts'}), 403

        # Check circle membership
        membership = Circle.query.filter_by(
            user_id=user_id,
            circle_user_id=current_user_id
        ).first()

        # Determine which visibility levels current user can see
        visible_levels = ['general']  # Everyone can see public
        if membership:
            if membership.circle_type in ['class_a', 'family']:
                visible_levels = ['general', 'close_friends', 'family']
            elif membership.circle_type in ['class_b', 'close_friends']:
                visible_levels = ['general', 'close_friends']

        # Get posts filtered by visible visibility levels for that date
        posts = Post.query.filter(
            Post.user_id == user_id,
            Post.visibility.in_(visible_levels),
            Post.created_at >= start_datetime,
            Post.created_at <= end_datetime
        ).order_by(Post.created_at.desc()).all()

        if not posts:
            return jsonify({'error': 'This update is not available to you based on your circle membership'}), 403

        # Calculate likes and comments for each post
        posts_data = []
        for post in posts:
            # Count likes from Reaction table
            likes_count = db.session.execute(
                select(func.count(Reaction.id)).filter_by(post_id=post.id, type='like')
            ).scalar() or 0

            # Count comments
            comments_count = db.session.execute(
                select(func.count(Comment.id)).filter_by(post_id=post.id)
            ).scalar() or 0

            # Check if current user liked this post
            user_liked = db.session.execute(
                select(Reaction).filter_by(
                    post_id=post.id,
                    user_id=current_user_id,
                    type='like'
                )
            ).scalar_one_or_none() is not None

            posts_data.append({
                'id': post.id,
                'content': post.content,
                'created_at': post.created_at.isoformat() if post.created_at else None,
                'likes_count': likes_count,
                'comments_count': comments_count,
                'user_liked': user_liked,
                'visibility': post.visibility
            })

        return jsonify({'posts': posts_data})

    except Exception as e:
        logger.error(f"Error getting user feed by date: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/posts', methods=['POST'])
@login_required
def save_feed_entry():
    """Save/update feed entry for a specific date and visibility with STRICT permissions"""
    try:
        data = request.get_json()
        user_id = session['user_id']

        # Get the date and visibility from request
        post_date = data.get('date', datetime.utcnow().strftime('%Y-%m-%d'))
        content = data.get('content', '').strip()
        visibility = data.get('visibility', 'general')  # general, close_friends, family, private

        if not content:
            return jsonify({'error': 'Content is required'}), 400

        # REMOVED: Map visibility to circle_id - no longer needed
        # We now use visibility field directly instead of circle_id

        # Delete any existing posts for this date first
        Post.query.filter_by(user_id=user_id).filter(
            db.func.date(Post.created_at) == post_date
        ).delete()

        # Create a SINGLE post with visibility field
        new_post = Post(
            user_id=user_id,
            content=content,
            circle_id=None,  # CHANGED: Always None, use visibility instead
            visibility=visibility,  # ADDED: Store visibility directly
            created_at=datetime.strptime(post_date, '%Y-%m-%d'),
            updated_at=datetime.utcnow(),
            is_published=True
        )
        db.session.add(new_post)

        db.session.commit()
        
        # PJ401: Create feed alerts for followers (only for public/general visibility posts)
        if visibility == 'general':
            try:
                user = User.query.get(user_id)
                username = user.username if user else 'Someone'
                
                # Get all followers of this user
                followers = Follow.query.filter_by(followed_id=user_id).all()
                
                for follower in followers:
                    # Create alert with source_user_id for filtering
                    alert = Alert(
                        user_id=follower.follower_id,
                        title=f"New post from {username}",
                        content=f"{username} shared a new feed post",
                        alert_type='feed',
                        source_user_id=user_id,
                        alert_category='feed'
                    )
                    db.session.add(alert)
                
                db.session.commit()
                logger.info(f"Created feed alerts for {len(followers)} followers")
            except Exception as alert_error:
                logger.warning(f"Failed to create feed alerts: {alert_error}")
                # Don't fail the main operation
        
        visibility_display = visibility.replace("_", " ").title()
        return jsonify({'success': True, 'message': f'Feed saved for {visibility_display} on {post_date}'})

    except Exception as e:
        logger.error(f"Feed save error: {e}")
        db.session.rollback()
        return jsonify({'error': 'Failed to save feed'}), 500


@app.route('/api/posts/<int:post_id>/like', methods=['POST'])
@login_required
def like_post(post_id):
    """Toggle like on a post"""
    try:
        user_id = session.get('user_id')

        # Check if post exists and user has access
        post = db.session.get(Post, post_id)
        if not post:
            return jsonify({'error': 'Post not found'}), 404

        # Check if user already liked this post
        existing_reaction = db.session.execute(
            select(Reaction).filter_by(
                post_id=post_id,
                user_id=user_id,
                type='like'
            )
        ).scalar_one_or_none()

        if existing_reaction:
            # Unlike - remove the reaction
            db.session.delete(existing_reaction)
            db.session.commit()

            # Invalidate feed cache for the post owner
            if REDIS_URL:
                try:
                    r = redis.from_url(REDIS_URL)
                    # Clear all pages of the post owner's feed
                    pattern = f'feed:{post.user_id}:*'
                    for key in r.scan_iter(match=pattern):
                        r.delete(key)
                    logger.debug(f'Invalidated feed cache for user {post.user_id}')
                except Exception as e:
                    logger.warning(f'Feed cache invalidation failed: {e}')

            # Get updated count
            likes_count = db.session.execute(
                select(func.count(Reaction.id)).filter_by(
                    post_id=post_id,
                    type='like'
                )
            ).scalar()

            return jsonify({
                'success': True,
                'liked': False,
                'likes_count': likes_count
            }), 200
        else:
            # Like - add the reaction
            new_reaction = Reaction(
                post_id=post_id,
                user_id=user_id,
                type='like'
            )
            db.session.add(new_reaction)
            db.session.commit()

            # Invalidate feed cache for the post owner
            if REDIS_URL:
                try:
                    r = redis.from_url(REDIS_URL)
                    # Clear all pages of the post owner's feed
                    pattern = f'feed:{post.user_id}:*'
                    for key in r.scan_iter(match=pattern):
                        r.delete(key)
                    logger.debug(f'Invalidated feed cache for user {post.user_id}')
                except Exception as e:
                    logger.warning(f'Feed cache invalidation failed: {e}')

            # Get updated count
            likes_count = db.session.execute(
                select(func.count(Reaction.id)).filter_by(
                    post_id=post_id,
                    type='like'
                )
            ).scalar()

            return jsonify({
                'success': True,
                'liked': True,
                'likes_count': likes_count
            }), 200

    except Exception as e:
        logger.error(f"Error toggling like: {e}")
        db.session.rollback()
        return jsonify({'error': 'Failed to toggle like'}), 500


@app.route('/api/posts/<int:post_id>/comments', methods=['GET'])
@login_required
def get_post_comments(post_id):
    """Get all comments for a post"""
    try:
        # Check if post exists
        post = db.session.get(Post, post_id)
        if not post:
            return jsonify({'error': 'Post not found'}), 404

        # Get comments with author information
        comments = db.session.execute(
            select(Comment).filter_by(post_id=post_id).order_by(Comment.created_at.asc())
        ).scalars().all()

        comments_data = []
        for comment in comments:
            author = db.session.get(User, comment.user_id)
            comments_data.append({
                'id': comment.id,
                'content': comment.content,
                'created_at': comment.created_at.isoformat(),
                'author': {
                    'id': author.id,
                    'username': author.username
                }
            })

        return jsonify({'comments': comments_data}), 200

    except Exception as e:
        logger.error(f"Error getting comments: {e}")
        return jsonify({'error': 'Failed to load comments'}), 500


@app.route('/api/posts/<int:post_id>/comments', methods=['POST'])
@login_required
def add_comment(post_id):
    """Add a comment to a post"""
    try:
        user_id = session.get('user_id')
        data = request.json
        content = data.get('content', '').strip()

        if not content:
            return jsonify({'error': 'Comment content required'}), 400

        # Check if post exists
        post = db.session.get(Post, post_id)
        if not post:
            return jsonify({'error': 'Post not found'}), 404

        # Create new comment
        new_comment = Comment(
            post_id=post_id,
            user_id=user_id,
            content=sanitize_input(content)
        )
        db.session.add(new_comment)
        db.session.commit()

        # Invalidate feed cache for the post owner
        if REDIS_URL:
            try:
                r = redis.from_url(REDIS_URL)
                # Clear all pages of the post owner's feed
                pattern = f'feed:{post.user_id}:*'
                for key in r.scan_iter(match=pattern):
                    r.delete(key)
                logger.debug(f'Invalidated feed cache for user {post.user_id}')
            except Exception as e:
                logger.warning(f'Feed cache invalidation failed: {e}')

        # Get author information
        author = db.session.get(User, user_id)

        # Get updated comment count
        comments_count = db.session.execute(
            select(func.count(Comment.id)).filter_by(post_id=post_id)
        ).scalar()

        return jsonify({
            'success': True,
            'comment': {
                'id': new_comment.id,
                'content': new_comment.content,
                'created_at': new_comment.created_at.isoformat(),
                'author': {
                    'id': author.id,
                    'username': author.username
                }
            },
            'comments_count': comments_count
        }), 201

    except Exception as e:
        logger.error(f"Error adding comment: {e}")
        db.session.rollback()
        return jsonify({'error': 'Failed to add comment'}), 500

@app.route('/api/feed/<date_str>')
@login_required
def load_feed_by_date(date_str):
    """Load feed entry for a specific date and visibility with STRICT matching"""
    try:
        user_id = session['user_id']
        visibility = request.args.get('visibility', 'general')

        # Get the post for this date that matches the requested visibility
        post = Post.query.filter_by(
            user_id=user_id,
            visibility=visibility  # Match by visibility field, not circle_id
        ).filter(
            db.func.date(Post.created_at) == date_str
        ).first()

        if post:
            return jsonify({
                'content': post.content,
                'date': date_str,
                'visibility': visibility,
                'updated_at': post.updated_at.isoformat() if post.updated_at else None
            })
        else:
            return jsonify({
                'content': '',
                'date': date_str,
                'visibility': visibility,
                'message': f'No {visibility.replace("_", " ")} feed entry for this date'
            })

    except Exception as e:
        logger.error(f"Load feed error: {e}")
        return jsonify({'error': 'Failed to load feed'}), 500


# =====================
# PARAMETERS ROUTES (Therapy Companion)
# =====================
# app.py - Fix for /api/parameters GET endpoint
# Location: Replace lines 3189-3235 in app.py

@app.route('/api/parameters', methods=['GET'])
@login_required
def get_parameters():
    """Get parameters for a specific date"""
    try:
        user_id = session.get('user_id')
        date_str = request.args.get('date')

        if not date_str:
            date_str = datetime.now().strftime('%Y-%m-%d')

        params = SavedParameters.query.filter_by(
            user_id=user_id,
            date=date_str
        ).first()

        if params:
            return jsonify({
                'success': True,
                'data': {
                    'parameters': {
                        'mood': int(params.mood) if params.mood else 0,
                        'energy': int(params.energy) if params.energy else 0,
                        'sleep_quality': int(params.sleep_quality) if params.sleep_quality else 0,
                        'physical_activity': int(params.physical_activity) if params.physical_activity else 0,
                        'anxiety': int(params.anxiety) if params.anxiety else 0
                    },
                    # ADD ALL PRIVACY SETTINGS HERE
                    'mood_privacy': params.mood_privacy or 'public',
                    'energy_privacy': params.energy_privacy or 'public',
                    'sleep_quality_privacy': params.sleep_quality_privacy or 'public',
                    'physical_activity_privacy': params.physical_activity_privacy or 'public',
                    'anxiety_privacy': params.anxiety_privacy or 'public',
                    'notes': params.notes or ''
                }
            })
        else:
            return jsonify({
                'success': True,
                'data': {
                    'parameters': {
                        'mood': 0,
                        'energy': 0,
                        'sleep_quality': 0,
                        'physical_activity': 0,
                        'anxiety': 0
                    },
                    # ADD DEFAULT PRIVACY SETTINGS
                    'mood_privacy': 'public',
                    'energy_privacy': 'public',
                    'sleep_quality_privacy': 'public',
                    'physical_activity_privacy': 'public',
                    'anxiety_privacy': 'public',
                    'notes': ''
                }
            })

    except Exception as e:
        logger.error(f"Get parameters error: {str(e)}")
        return jsonify({'error': 'Failed to get parameters'}), 500


@app.route('/api/parameters', methods=['POST'])
@login_required
def save_parameters():
    """Save user parameters with privacy settings"""
    try:
        data = request.get_json()
        user_id = session.get('user_id')
        date_str = data.get('date')
        
        # PJ809: Add logging to confirm save_parameters is called
        logger.info(f"[SAVE PARAMS] ========================================")
        logger.info(f"[SAVE PARAMS] save_parameters called for user_id={user_id}, date={date_str}")

        if not date_str:
            return jsonify({'error': 'Date is required'}), 400

            # Parse the date as local date without timezone conversion
        try:
            param_date = parse_date_as_local(date_str)
            date_str = param_date.isoformat()  # Ensure consistent YYYY-MM-DD format
        except ValueError:
            return jsonify({'error': 'Invalid date format'}), 400

        # Find or create parameter entry
        # Find or create parameter entry
        params = SavedParameters.query.filter_by(
            user_id=user_id,
            date=date_str
        ).first()

        if not params:
            params = SavedParameters(
                user_id=user_id,
                date=date_str
            )

        # Update values - ENSURE INTEGER CONVERSION
        for field in ['mood', 'energy', 'sleep_quality', 'physical_activity', 'anxiety']:
            if field in data:
                value = data[field]
                if value is not None:
                    try:
                        # Convert to integer
                        int_value = int(value)
                        if 1 <= int_value <= 4:
                            setattr(params, field, int_value)
                    except (ValueError, TypeError):
                        pass  # Skip invalid values

            # Handle privacy settings
            privacy_field = f"{field}_privacy"
            if privacy_field in data:
                privacy_value = data[privacy_field]
                if privacy_value in ['public', 'class_a', 'class_b', 'private']:
                    setattr(params, privacy_field, privacy_value)

        if 'notes' in data:
            params.notes = data['notes']

        params.updated_at = datetime.utcnow()
        db.session.add(params)
        db.session.commit()
        
        # PJ809: Log parameter values before trigger check
        logger.info(f"[SAVE PARAMS] Saved: mood={params.mood}, energy={params.energy}, sleep={params.sleep_quality}, activity={params.physical_activity}, anxiety={params.anxiety}")
        
        # PJ6006: Run trigger processing in background thread to avoid blocking response
        # This prevents the 5+ second delay when multiple alert emails need to be sent
        logger.info(f"[SAVE PARAMS] Creating background job for trigger processing user_id={user_id}")
        
        # Create a copy of param values for the background job
        param_snapshot = {
            'mood': params.mood,
            'energy': params.energy,
            'sleep_quality': params.sleep_quality,
            'physical_activity': params.physical_activity,
            'anxiety': params.anxiety,
            'mood_privacy': getattr(params, 'mood_privacy', 'private'),
            'energy_privacy': getattr(params, 'energy_privacy', 'private'),
            'sleep_quality_privacy': getattr(params, 'sleep_quality_privacy', 'private'),
            'physical_activity_privacy': getattr(params, 'physical_activity_privacy', 'private'),
            'anxiety_privacy': getattr(params, 'anxiety_privacy', 'private'),
            'date': params.date,
            'notes': params.notes
        }
        
        # Create a background job instead of spawning a thread
        # This provides: persistence, retries, and prevents thread exhaustion
        try:
            job = BackgroundJob(
                job_type='trigger_processing',
                payload={
                    'user_id': user_id,
                    'param_snapshot': param_snapshot
                },
                priority=1  # Normal priority
            )
            db.session.add(job)
            db.session.commit()
            logger.info(f"[SAVE PARAMS] Created background job {job.id} for trigger processing")
        except Exception as job_err:
            logger.error(f"[SAVE PARAMS] Failed to create background job: {str(job_err)}")
            # Don't fail the save if job creation fails - it's not critical
            db.session.rollback()

        import random
        encouragements = [
            "Great job tracking your wellness today! 🌟",
            "Your consistency is inspiring! Keep it up! 💪",
            "Every check-in is a step forward! 🚀",
            "Thank you for taking care of yourself! ❤️",
            "Your commitment to wellness is admirable! 🌈"
        ]

        # Return consistent format
        return jsonify({
            'success': True,
            'message': 'Parameters saved successfully',
            'encouragement': random.choice(encouragements),
            'data': {
                'parameters': {
                    'mood': int(params.mood) if params.mood else 0,
                    'energy': int(params.energy) if params.energy else 0,
                    'sleep_quality': int(params.sleep_quality) if params.sleep_quality else 0,
                    'physical_activity': int(params.physical_activity) if params.physical_activity else 0,
                    'anxiety': int(params.anxiety) if params.anxiety else 0
                },
                'notes': params.notes or ''
            }
        }), 200

    except Exception as e:
        logger.error(f"Error saving parameters: {str(e)}")
        db.session.rollback()
        return jsonify({'error': 'Failed to save parameters'}), 500


# =============================================================================
# BACKGROUND JOB QUEUE SYSTEM
# =============================================================================

def process_background_jobs(batch_size=10):
    """
    Process pending background jobs from the database queue.
    Called periodically by the scheduler.
    
    Features:
    - Processes jobs in priority order (higher first)
    - Automatic retries with exponential backoff
    - Job locking to prevent double-processing
    - Cleanup of completed jobs older than 24 hours
    
    Args:
        batch_size: Maximum number of jobs to process per run
    """
    import uuid
    worker_id = f"worker-{uuid.uuid4().hex[:8]}"
    
    try:
        logger.info(f"[JOB QUEUE] Processing background jobs (worker: {worker_id})")
        
        # Lock and fetch pending jobs
        jobs = BackgroundJob.query.filter(
            BackgroundJob.status == 'pending',
            # Don't pick up jobs that are locked (being processed by another worker)
            db.or_(
                BackgroundJob.locked_at.is_(None),
                BackgroundJob.locked_at < datetime.utcnow() - timedelta(minutes=5)  # Stale lock
            )
        ).order_by(
            BackgroundJob.priority.desc(),
            BackgroundJob.created_at.asc()
        ).limit(batch_size).all()
        
        if not jobs:
            logger.debug(f"[JOB QUEUE] No pending jobs")
            return
        
        logger.info(f"[JOB QUEUE] Found {len(jobs)} pending jobs")
        
        for job in jobs:
            try:
                # Lock the job
                job.locked_by = worker_id
                job.locked_at = datetime.utcnow()
                job.status = 'processing'
                job.started_at = datetime.utcnow()
                job.attempts += 1
                db.session.commit()
                
                logger.info(f"[JOB QUEUE] Processing job {job.id} type={job.job_type} attempt={job.attempts}")
                
                # Execute the job based on type
                if job.job_type == 'trigger_processing':
                    user_id = job.payload.get('user_id')
                    param_snapshot = job.payload.get('param_snapshot')
                    
                    if user_id and param_snapshot:
                        process_parameter_triggers_async(user_id, param_snapshot)
                        cleanup_stale_trigger_alerts_for_user(user_id)
                    else:
                        raise ValueError(f"Invalid payload for trigger_processing: {job.payload}")
                
                elif job.job_type == 'send_email':
                    # Future: Handle email sending jobs
                    email_data = job.payload
                    # send_email_from_job(email_data)
                    logger.info(f"[JOB QUEUE] Email job type - placeholder for future implementation")
                
                else:
                    logger.warning(f"[JOB QUEUE] Unknown job type: {job.job_type}")
                
                # Mark job as completed
                job.status = 'completed'
                job.completed_at = datetime.utcnow()
                job.locked_by = None
                job.locked_at = None
                db.session.commit()
                
                logger.info(f"[JOB QUEUE] Job {job.id} completed successfully")
                
            except Exception as e:
                logger.error(f"[JOB QUEUE] Job {job.id} failed: {str(e)}")
                logger.error(f"[JOB QUEUE] Traceback: {traceback.format_exc()}")
                
                # Handle retry or failure
                if job.attempts >= job.max_attempts:
                    job.status = 'failed'
                    job.error_message = str(e)[:1000]  # Truncate error message
                    logger.error(f"[JOB QUEUE] Job {job.id} permanently failed after {job.attempts} attempts")
                else:
                    job.status = 'pending'  # Will be retried
                    logger.info(f"[JOB QUEUE] Job {job.id} will be retried (attempt {job.attempts}/{job.max_attempts})")
                
                job.locked_by = None
                job.locked_at = None
                
                try:
                    db.session.commit()
                except:
                    db.session.rollback()
        
        # Cleanup old completed jobs (older than 24 hours)
        cleanup_cutoff = datetime.utcnow() - timedelta(hours=24)
        old_jobs = BackgroundJob.query.filter(
            BackgroundJob.status.in_(['completed', 'failed']),
            BackgroundJob.completed_at < cleanup_cutoff
        ).delete(synchronize_session=False)
        
        if old_jobs > 0:
            db.session.commit()
            logger.info(f"[JOB QUEUE] Cleaned up {old_jobs} old jobs")
        
    except Exception as e:
        logger.error(f"[JOB QUEUE] Error in job processor: {str(e)}")
        logger.error(f"[JOB QUEUE] Traceback: {traceback.format_exc()}")
        try:
            db.session.rollback()
        except:
            pass


def get_job_queue_stats():
    """
    Get statistics about the job queue for monitoring.
    
    Returns:
        Dict with queue statistics
    """
    try:
        stats = {
            'pending': BackgroundJob.query.filter_by(status='pending').count(),
            'processing': BackgroundJob.query.filter_by(status='processing').count(),
            'completed_24h': BackgroundJob.query.filter(
                BackgroundJob.status == 'completed',
                BackgroundJob.completed_at >= datetime.utcnow() - timedelta(hours=24)
            ).count(),
            'failed_24h': BackgroundJob.query.filter(
                BackgroundJob.status == 'failed',
                BackgroundJob.completed_at >= datetime.utcnow() - timedelta(hours=24)
            ).count()
        }
        return stats
    except Exception as e:
        logger.error(f"[JOB QUEUE] Error getting stats: {str(e)}")
        return {'error': str(e)}


def process_parameter_triggers_async(user_id, param_snapshot):
    """
    PJ6007: Async trigger processing with CONSOLIDATED emails.
    PJ6008: Fixed date parsing and minimum consecutive_days enforcement.
    
    Instead of sending one email per triggered parameter, collects all triggers
    per watcher and sends ONE consolidated email.
    
    Args:
        user_id: The user ID whose parameters were saved
        param_snapshot: Dict containing parameter values (since ORM objects can't cross threads)
    """
    try:
        logger.info(f"[TRIGGER PROCESS ASYNC] ========================================")
        logger.info(f"[TRIGGER PROCESS ASYNC] PJ6008: Starting with fixed date parsing for user_id={user_id}")
        
        # Find all triggers where someone is watching this user
        all_triggers = ParameterTrigger.query.filter_by(watched_id=user_id).all()
        logger.info(f"[TRIGGER PROCESS ASYNC] Found {len(all_triggers)} trigger rows watching user {user_id}")
        
        if len(all_triggers) == 0:
            logger.info(f"[TRIGGER PROCESS ASYNC] No triggers found - no one is watching user {user_id}")
            logger.info(f"[TRIGGER PROCESS ASYNC] ========================================")
            return
        
        watched_user = User.query.get(user_id)
        if not watched_user:
            logger.error(f"[TRIGGER PROCESS ASYNC] Watched user {user_id} not found")
            return
        
        # PJ6008: Helper to parse date string to date object
        def parse_date(date_val):
            """Convert date string or date object to date object"""
            if date_val is None:
                return None
            if isinstance(date_val, date):
                return date_val
            if isinstance(date_val, datetime):
                return date_val.date()
            if isinstance(date_val, str):
                try:
                    return datetime.strptime(date_val, '%Y-%m-%d').date()
                except ValueError:
                    try:
                        return datetime.fromisoformat(date_val.split('T')[0]).date()
                    except:
                        return None
            return None
        
        # Helper function to check privacy permissions
        def can_see_parameter(param_privacy, watcher_circle):
            if param_privacy == 'private':
                return False
            elif param_privacy == 'class_a':
                return watcher_circle == 'class_a'
            elif param_privacy == 'class_b':
                return watcher_circle in ['class_b', 'class_a']
            elif param_privacy == 'public':
                return True
            return False
        
        # Helper to safely convert values to numbers
        def to_number(val):
            if val is None:
                return None
            if isinstance(val, (int, float)):
                return float(val)
            if isinstance(val, str):
                if val.lower() in ['private', 'hidden', 'none', '']:
                    return None
                try:
                    return float(val)
                except ValueError:
                    return None
            return None
        
        # Get last 30 days of parameters from DB
        thirty_days_ago = datetime.now().date() - timedelta(days=30)
        all_params = SavedParameters.query.filter(
            SavedParameters.user_id == user_id,
            SavedParameters.date >= str(thirty_days_ago)  # PJ6008: Convert to string for comparison
        ).order_by(SavedParameters.date.asc()).all()
        
        logger.info(f"[TRIGGER PROCESS ASYNC] Found {len(all_params)} parameter entries in last 30 days")
        
        # PJ6007: Collect triggered params per watcher for consolidated email
        # Key: watcher_id, Value: list of {'param_name', 'days', 'date_range'}
        watcher_triggered_params = {}
        patterns_seen = set()
        alerts_created = 0
        alerts_skipped_duplicate = 0
        
        # Process each trigger row
        for trigger in all_triggers:
            watcher_id = trigger.watcher_id
            # PJ6008: Enforce minimum of 3 consecutive days - this is the key fix!
            raw_consecutive_days = trigger.consecutive_days
            consecutive_days = max(raw_consecutive_days or 3, 3)  # Minimum 3 days
            
            logger.info(f"[TRIGGER PROCESS ASYNC] Processing trigger for watcher {watcher_id}: raw_consecutive_days={raw_consecutive_days}, enforced={consecutive_days}")
            
            if len(all_params) < consecutive_days:
                logger.info(f"[TRIGGER PROCESS ASYNC] Skipping - not enough params ({len(all_params)} < {consecutive_days})")
                continue
            
            watcher_circle = get_watcher_circle_level(user_id, watcher_id)
            if not watcher_circle:
                logger.info(f"[TRIGGER PROCESS ASYNC] Skipping - watcher {watcher_id} not in any circle")
                continue
            
            # Initialize watcher's list if not exists
            if watcher_id not in watcher_triggered_params:
                watcher_triggered_params[watcher_id] = []
            
            # Determine which schema this trigger uses
            has_new_schema = any([
                trigger.mood_alert, trigger.energy_alert, trigger.sleep_alert,
                trigger.physical_alert, trigger.anxiety_alert
            ])
            has_old_schema = trigger.parameter_name is not None
            
            # Build list of parameters to check
            param_checks = []
            
            if has_new_schema:
                if trigger.mood_alert:
                    param_checks.append(('mood', 'mood', 'mood_privacy', lambda x: to_number(x) is not None and to_number(x) <= 2))
                if trigger.energy_alert:
                    param_checks.append(('energy', 'energy', 'energy_privacy', lambda x: to_number(x) is not None and to_number(x) <= 2))
                if trigger.sleep_alert:
                    param_checks.append(('sleep_quality', 'sleep_quality', 'sleep_quality_privacy', lambda x: to_number(x) is not None and to_number(x) <= 2))
                if trigger.physical_alert:
                    param_checks.append(('physical_activity', 'physical_activity', 'physical_activity_privacy', lambda x: to_number(x) is not None and to_number(x) <= 2))
                if trigger.anxiety_alert:
                    param_checks.append(('anxiety', 'anxiety', 'anxiety_privacy', lambda x: to_number(x) is not None and to_number(x) >= 3))
            elif has_old_schema:
                param_mapping = {
                    'mood': ('mood', 'mood', 'mood_privacy', lambda x: to_number(x) is not None and to_number(x) <= 2),
                    'energy': ('energy', 'energy', 'energy_privacy', lambda x: to_number(x) is not None and to_number(x) <= 2),
                    'sleep_quality': ('sleep_quality', 'sleep_quality', 'sleep_quality_privacy', lambda x: to_number(x) is not None and to_number(x) <= 2),
                    'physical_activity': ('physical_activity', 'physical_activity', 'physical_activity_privacy', lambda x: to_number(x) is not None and to_number(x) <= 2),
                    'anxiety': ('anxiety', 'anxiety', 'anxiety_privacy', lambda x: to_number(x) is not None and to_number(x) >= 3)
                }
                if trigger.parameter_name in param_mapping:
                    param_checks.append(param_mapping[trigger.parameter_name])
            
            logger.info(f"[TRIGGER PROCESS ASYNC] Checking {len(param_checks)} parameters for watcher {watcher_id}")
            
            # Check each parameter for consecutive day streaks
            for param_attr, param_name, privacy_attr, condition_func in param_checks:
                streak_dates = []
                last_date = None
                
                for param_entry in all_params:
                    param_value = getattr(param_entry, param_attr, None)
                    param_privacy = getattr(param_entry, privacy_attr, 'private')
                    # PJ6008: Parse date properly
                    entry_date = parse_date(param_entry.date)
                    
                    if entry_date is None:
                        logger.warning(f"[TRIGGER PROCESS ASYNC] Could not parse date: {param_entry.date}")
                        continue
                    
                    # Check privacy - process streak if long enough before reset
                    if not can_see_parameter(param_privacy, watcher_circle):
                        if len(streak_dates) >= consecutive_days:
                            start_date = streak_dates[0]
                            end_date = streak_dates[-1]
                            pattern_key = (watcher_id, param_name, start_date.isoformat(), end_date.isoformat())
                            
                            if pattern_key not in patterns_seen:
                                patterns_seen.add(pattern_key)
                                start_str = start_date.strftime('%b %d')
                                end_str = end_date.strftime('%b %d')
                                date_pattern = f"({start_str} - {end_str})"
                                
                                # Check for DB duplicate
                                existing = Alert.query.filter(
                                    Alert.user_id == watcher_id,
                                    Alert.alert_type == 'trigger',
                                    Alert.created_at >= datetime.now() - timedelta(hours=24),
                                    Alert.content.ilike(f"%{watched_user.username}'s {param_name}%{date_pattern}%")
                                ).first()
                                
                                if not existing:
                                    content = f"{watched_user.username}'s {param_name} has been at concerning levels for {len(streak_dates)} consecutive days {date_pattern}"
                                    logger.info(f"[TRIGGER PROCESS ASYNC] Creating alert: {content}")
                                    alert = create_alert_no_email(
                                        user_id=watcher_id,
                                        title=f"Wellness Alert for {watched_user.username}",
                                        content=content,
                                        alert_type='trigger',
                                        source_user_id=watched_user.id,
                                        alert_category='trigger'
                                    )
                                    if alert:
                                        alerts_created += 1
                                        watcher_triggered_params[watcher_id].append({
                                            'param_name': param_name,
                                            'days': len(streak_dates),
                                            'date_range': f"{start_str} - {end_str}"
                                        })
                                else:
                                    alerts_skipped_duplicate += 1
                        
                        streak_dates = []
                        last_date = None
                        continue
                    
                    if param_value is not None and condition_func(param_value):
                        if last_date is None:
                            streak_dates = [entry_date]
                            last_date = entry_date
                        elif (entry_date - last_date).days == 1:
                            streak_dates.append(entry_date)
                            last_date = entry_date
                        elif (entry_date - last_date).days == 0:
                            continue  # Same day, skip
                        else:
                            # Gap in streak - process if long enough
                            if len(streak_dates) >= consecutive_days:
                                start_date = streak_dates[0]
                                end_date = streak_dates[-1]
                                pattern_key = (watcher_id, param_name, start_date.isoformat(), end_date.isoformat())
                                
                                if pattern_key not in patterns_seen:
                                    patterns_seen.add(pattern_key)
                                    start_str = start_date.strftime('%b %d')
                                    end_str = end_date.strftime('%b %d')
                                    date_pattern = f"({start_str} - {end_str})"
                                    
                                    existing = Alert.query.filter(
                                        Alert.user_id == watcher_id,
                                        Alert.alert_type == 'trigger',
                                        Alert.created_at >= datetime.now() - timedelta(hours=24),
                                        Alert.content.ilike(f"%{watched_user.username}'s {param_name}%{date_pattern}%")
                                    ).first()
                                    
                                    if not existing:
                                        content = f"{watched_user.username}'s {param_name} has been at concerning levels for {len(streak_dates)} consecutive days {date_pattern}"
                                        logger.info(f"[TRIGGER PROCESS ASYNC] Creating alert: {content}")
                                        alert = create_alert_no_email(
                                            user_id=watcher_id,
                                            title=f"Wellness Alert for {watched_user.username}",
                                            content=content,
                                            alert_type='trigger',
                                            source_user_id=watched_user.id,
                                            alert_category='trigger'
                                        )
                                        if alert:
                                            alerts_created += 1
                                            watcher_triggered_params[watcher_id].append({
                                                'param_name': param_name,
                                                'days': len(streak_dates),
                                                'date_range': f"{start_str} - {end_str}"
                                            })
                                    else:
                                        alerts_skipped_duplicate += 1
                            
                            # Start new streak
                            streak_dates = [entry_date]
                            last_date = entry_date
                    else:
                        # Condition not met - process streak if long enough
                        if len(streak_dates) >= consecutive_days:
                            start_date = streak_dates[0]
                            end_date = streak_dates[-1]
                            pattern_key = (watcher_id, param_name, start_date.isoformat(), end_date.isoformat())
                            
                            if pattern_key not in patterns_seen:
                                patterns_seen.add(pattern_key)
                                start_str = start_date.strftime('%b %d')
                                end_str = end_date.strftime('%b %d')
                                date_pattern = f"({start_str} - {end_str})"
                                
                                existing = Alert.query.filter(
                                    Alert.user_id == watcher_id,
                                    Alert.alert_type == 'trigger',
                                    Alert.created_at >= datetime.now() - timedelta(hours=24),
                                    Alert.content.ilike(f"%{watched_user.username}'s {param_name}%{date_pattern}%")
                                ).first()
                                
                                if not existing:
                                    content = f"{watched_user.username}'s {param_name} has been at concerning levels for {len(streak_dates)} consecutive days {date_pattern}"
                                    logger.info(f"[TRIGGER PROCESS ASYNC] Creating alert: {content}")
                                    alert = create_alert_no_email(
                                        user_id=watcher_id,
                                        title=f"Wellness Alert for {watched_user.username}",
                                        content=content,
                                        alert_type='trigger',
                                        source_user_id=watched_user.id,
                                        alert_category='trigger'
                                    )
                                    if alert:
                                        alerts_created += 1
                                        watcher_triggered_params[watcher_id].append({
                                            'param_name': param_name,
                                            'days': len(streak_dates),
                                            'date_range': f"{start_str} - {end_str}"
                                        })
                                else:
                                    alerts_skipped_duplicate += 1
                        
                        streak_dates = []
                        last_date = None
                
                # Check final streak at end of loop
                if len(streak_dates) >= consecutive_days:
                    start_date = streak_dates[0]
                    end_date = streak_dates[-1]
                    pattern_key = (watcher_id, param_name, start_date.isoformat(), end_date.isoformat())
                    
                    if pattern_key not in patterns_seen:
                        patterns_seen.add(pattern_key)
                        start_str = start_date.strftime('%b %d')
                        end_str = end_date.strftime('%b %d')
                        date_pattern = f"({start_str} - {end_str})"
                        
                        existing = Alert.query.filter(
                            Alert.user_id == watcher_id,
                            Alert.alert_type == 'trigger',
                            Alert.created_at >= datetime.now() - timedelta(hours=24),
                            Alert.content.ilike(f"%{watched_user.username}'s {param_name}%{date_pattern}%")
                        ).first()
                        
                        if not existing:
                            content = f"{watched_user.username}'s {param_name} has been at concerning levels for {len(streak_dates)} consecutive days {date_pattern}"
                            logger.info(f"[TRIGGER PROCESS ASYNC] Creating alert: {content}")
                            alert = create_alert_no_email(
                                user_id=watcher_id,
                                title=f"Wellness Alert for {watched_user.username}",
                                content=content,
                                alert_type='trigger',
                                source_user_id=watched_user.id,
                                alert_category='trigger'
                            )
                            if alert:
                                alerts_created += 1
                                watcher_triggered_params[watcher_id].append({
                                    'param_name': param_name,
                                    'days': len(streak_dates),
                                    'date_range': f"{start_str} - {end_str}"
                                })
                        else:
                            alerts_skipped_duplicate += 1
        
        # Commit alerts to DB
        try:
            db.session.commit()
        except:
            db.session.rollback()
        
        # PJ6007: Send ONE consolidated email per watcher (instead of many individual emails)
        emails_sent = 0
        for watcher_id, triggered_params in watcher_triggered_params.items():
            if triggered_params:  # Only send if there are triggered params
                watcher = User.query.get(watcher_id)
                user_language = watcher.preferred_language if watcher else 'en'
                
                logger.info(f"[TRIGGER PROCESS ASYNC] Sending consolidated email to watcher {watcher_id} with {len(triggered_params)} params")
                
                if send_consolidated_wellness_alert_email(watcher_id, watched_user.username, triggered_params, user_language):
                    emails_sent += 1
        
        logger.info(f"[TRIGGER PROCESS ASYNC] PJ6008 Completed:")
        logger.info(f"[TRIGGER PROCESS ASYNC]   - {alerts_created} alerts created in DB")
        logger.info(f"[TRIGGER PROCESS ASYNC]   - {alerts_skipped_duplicate} duplicates skipped")
        logger.info(f"[TRIGGER PROCESS ASYNC]   - {emails_sent} consolidated emails sent")
        logger.info(f"[TRIGGER PROCESS ASYNC] ========================================")
        
    except Exception as e:
        logger.error(f"[TRIGGER PROCESS ASYNC] Error: {str(e)}")
        logger.error(f"[TRIGGER PROCESS ASYNC] Traceback: {traceback.format_exc()}")
        try:
            db.session.rollback()
        except:
            pass


def process_parameter_triggers(user_id, params):
    """Check triggers when parameters are saved - checks for N consecutive days based on trigger settings
    
    PJ815 FIX (v1705):
    - Process each trigger row individually (don't deduplicate upfront)
    - Handle BOTH old schema (parameter_name) and new schema (mood_alert, etc.)
    - Use patterns_seen SET to deduplicate results, not inputs
    - This runs when watched user saves parameters -> sends emails to watchers
    """
    try:
        logger.info(f"[TRIGGER PROCESS] ========================================")
        logger.info(f"[TRIGGER PROCESS] process_parameter_triggers called for user_id={user_id}")
        logger.info(f"[TRIGGER PROCESS] Parameter values: mood={params.mood}, energy={params.energy}, sleep={params.sleep_quality}, activity={params.physical_activity}, anxiety={params.anxiety}")
        
        # Find all triggers where someone is watching this user
        all_triggers = ParameterTrigger.query.filter_by(watched_id=user_id).all()
        logger.info(f"[TRIGGER PROCESS] Found {len(all_triggers)} trigger rows watching user {user_id}")
        
        if len(all_triggers) == 0:
            logger.info(f"[TRIGGER PROCESS] No triggers found - no one is watching user {user_id}")
            logger.info(f"[TRIGGER PROCESS] ========================================")
            return
        
        # Log each trigger's details
        for i, t in enumerate(all_triggers):
            watcher = User.query.get(t.watcher_id)
            watcher_name = watcher.username if watcher else f"user_{t.watcher_id}"
            # Check both new and old schema
            flags = []
            if t.mood_alert: flags.append('mood')
            if t.energy_alert: flags.append('energy')
            if t.sleep_alert: flags.append('sleep')
            if t.physical_alert: flags.append('activity')
            if t.anxiety_alert: flags.append('anxiety')
            logger.info(f"[TRIGGER PROCESS] Trigger {i+1}: watcher={watcher_name}, days={t.consecutive_days}, new_flags={flags}, old_param={t.parameter_name}")
        
        # PJ815: Helper function to check privacy permissions
        def can_see_parameter(param_privacy, watcher_circle):
            """Check if watcher can see this parameter based on privacy and circle level"""
            if param_privacy == 'private':
                return False
            elif param_privacy == 'class_a':
                return watcher_circle == 'class_a'
            elif param_privacy == 'class_b':
                return watcher_circle in ['class_b', 'class_a']
            elif param_privacy == 'public':
                return True
            return False
        
        # PJ815: Track patterns already alerted to prevent duplicates
        # Key = (watcher_id, param_name, start_date_iso, end_date_iso)
        patterns_seen = set()

        alerts_created = 0
        alerts_skipped_duplicate = 0
        alerts_skipped_privacy = 0
        alerts_emailed = 0
        watchers_processed = set()

        watched_user = User.query.get(user_id)
        
        # Get last 30 days of parameters
        thirty_days_ago = datetime.now().date() - timedelta(days=30)
        all_params = SavedParameters.query.filter(
            SavedParameters.user_id == user_id,
            SavedParameters.date >= thirty_days_ago
        ).order_by(SavedParameters.date.asc()).all()  # ASC for proper streak detection

        logger.info(f"[TRIGGER PROCESS] Found {len(all_params)} parameter entries in last 30 days")

        # PJ815: Process each trigger row individually
        for trigger in all_triggers:
            watcher_id = trigger.watcher_id
            consecutive_days = trigger.consecutive_days or 3
            
            if consecutive_days < 1:
                continue
                
            if len(all_params) < consecutive_days:
                continue
            
            # Get watcher's circle level for privacy check
            watcher_circle = get_watcher_circle_level(user_id, watcher_id)
            if not watcher_circle:
                logger.info(f"[TRIGGER PROCESS] Skipping watcher {watcher_id} - not in any circle")
                continue
            
            watchers_processed.add(watcher_id)
            
            # Determine which schema this trigger uses
            has_new_schema = any([
                trigger.mood_alert,
                trigger.energy_alert,
                trigger.sleep_alert,
                trigger.physical_alert,
                trigger.anxiety_alert
            ])
            
            has_old_schema = trigger.parameter_name is not None
            
            logger.info(f"[TRIGGER PROCESS] Processing watcher {watcher_id}: new_schema={has_new_schema}, old_schema={has_old_schema}")
            
            # PJ817: Helper to safely convert values to numbers (fixes TypeError with string comparisons)
            def to_number(val):
                """Convert value to number, handling string/int/float/None cases"""
                if val is None:
                    return None
                if isinstance(val, (int, float)):
                    return float(val)
                if isinstance(val, str):
                    if val.lower() in ['private', 'hidden', 'none', '']:
                        return None
                    try:
                        return float(val)
                    except ValueError:
                        return None
                return None
            
            # Build list of parameters to check
            param_checks = []
            
            if has_new_schema:
                # PJ817: Fixed lambdas to use to_number() to handle string values from database
                if trigger.mood_alert:
                    param_checks.append(('mood', 'mood', 'mood_privacy', lambda x: to_number(x) is not None and to_number(x) <= 2))
                if trigger.energy_alert:
                    param_checks.append(('energy', 'energy', 'energy_privacy', lambda x: to_number(x) is not None and to_number(x) <= 2))
                if trigger.sleep_alert:
                    param_checks.append(('sleep_quality', 'sleep_quality', 'sleep_quality_privacy', lambda x: to_number(x) is not None and to_number(x) <= 2))
                if trigger.physical_alert:
                    param_checks.append(('physical_activity', 'physical_activity', 'physical_activity_privacy', lambda x: to_number(x) is not None and to_number(x) <= 2))
                if trigger.anxiety_alert:
                    param_checks.append(('anxiety', 'anxiety', 'anxiety_privacy', lambda x: to_number(x) is not None and to_number(x) >= 3))
            elif has_old_schema:
                # Old schema - single parameter from trigger.parameter_name
                # PJ817: Fixed lambdas to use to_number() to handle string values from database
                param_mapping = {
                    'mood': ('mood', 'mood', 'mood_privacy', lambda x: to_number(x) is not None and to_number(x) <= 2),
                    'energy': ('energy', 'energy', 'energy_privacy', lambda x: to_number(x) is not None and to_number(x) <= 2),
                    'sleep_quality': ('sleep_quality', 'sleep_quality', 'sleep_quality_privacy', lambda x: to_number(x) is not None and to_number(x) <= 2),
                    'physical_activity': ('physical_activity', 'physical_activity', 'physical_activity_privacy', lambda x: to_number(x) is not None and to_number(x) <= 2),
                    'anxiety': ('anxiety', 'anxiety', 'anxiety_privacy', lambda x: to_number(x) is not None and to_number(x) >= 3)
                }
                if trigger.parameter_name in param_mapping:
                    param_checks.append(param_mapping[trigger.parameter_name])
            
            logger.info(f"[TRIGGER PROCESS] Watcher {watcher_id} checking {len(param_checks)} parameters")
            
            # Check each parameter
            for param_attr, param_name, privacy_attr, condition_func in param_checks:
                # Find all consecutive streaks
                streak_dates = []
                last_date = None
                
                for param_entry in all_params:
                    param_value = getattr(param_entry, param_attr, None)
                    param_privacy = getattr(param_entry, privacy_attr, 'private')
                    
                    # Check privacy
                    if not can_see_parameter(param_privacy, watcher_circle):
                        # Save streak if long enough before resetting
                        if len(streak_dates) >= consecutive_days:
                            start_date = streak_dates[0]
                            end_date = streak_dates[-1]
                            pattern_key = (watcher_id, param_name, start_date.isoformat(), end_date.isoformat())
                            
                            if pattern_key not in patterns_seen:
                                patterns_seen.add(pattern_key)
                                # Create alert for this pattern
                                start_str = start_date.strftime('%b %d')
                                end_str = end_date.strftime('%b %d')
                                date_pattern = f"({start_str} - {end_str})"
                                
                                # Check for DB duplicate
                                existing = Alert.query.filter(
                                    Alert.user_id == watcher_id,
                                    Alert.alert_type == 'trigger',
                                    Alert.created_at >= datetime.now() - timedelta(hours=24),
                                    Alert.content.ilike(f"%{watched_user.username}'s {param_name}%{date_pattern}%")
                                ).first()
                                
                                if not existing:
                                    content = f"{watched_user.username}'s {param_name} has been at concerning levels for {len(streak_dates)} consecutive days {date_pattern}"
                                    alert = create_alert_with_email(
                                        user_id=watcher_id,
                                        title=f"Wellness Alert for {watched_user.username}",
                                        content=content,
                                        alert_type='trigger',
                                        source_user_id=watched_user.id,
                                        alert_category='trigger'
                                    )
                                    if alert:
                                        alerts_created += 1
                                        logger.info(f"[TRIGGER PROCESS] ✅ Created alert for {param_name} {date_pattern}")
                                else:
                                    alerts_skipped_duplicate += 1
                        
                        streak_dates = []
                        last_date = None
                        continue
                    
                    if param_value is not None and condition_func(param_value):
                        # Condition met
                        if last_date is None:
                            streak_dates = [param_entry.date]
                            last_date = param_entry.date
                        elif (param_entry.date - last_date).days == 1:
                            streak_dates.append(param_entry.date)
                            last_date = param_entry.date
                        elif (param_entry.date - last_date).days == 0:
                            continue  # Same day
                        else:
                            # Gap - save and start new
                            if len(streak_dates) >= consecutive_days:
                                start_date = streak_dates[0]
                                end_date = streak_dates[-1]
                                pattern_key = (watcher_id, param_name, start_date.isoformat(), end_date.isoformat())
                                
                                if pattern_key not in patterns_seen:
                                    patterns_seen.add(pattern_key)
                                    start_str = start_date.strftime('%b %d')
                                    end_str = end_date.strftime('%b %d')
                                    date_pattern = f"({start_str} - {end_str})"
                                    
                                    existing = Alert.query.filter(
                                        Alert.user_id == watcher_id,
                                        Alert.alert_type == 'trigger',
                                        Alert.created_at >= datetime.now() - timedelta(hours=24),
                                        Alert.content.ilike(f"%{watched_user.username}'s {param_name}%{date_pattern}%")
                                    ).first()
                                    
                                    if not existing:
                                        content = f"{watched_user.username}'s {param_name} has been at concerning levels for {len(streak_dates)} consecutive days {date_pattern}"
                                        alert = create_alert_with_email(
                                            user_id=watcher_id,
                                            title=f"Wellness Alert for {watched_user.username}",
                                            content=content,
                                            alert_type='trigger',
                                            source_user_id=watched_user.id,
                                            alert_category='trigger'
                                        )
                                        if alert:
                                            alerts_created += 1
                                            logger.info(f"[TRIGGER PROCESS] ✅ Created alert for {param_name} {date_pattern}")
                                    else:
                                        alerts_skipped_duplicate += 1
                            
                            streak_dates = [param_entry.date]
                            last_date = param_entry.date
                    else:
                        # Condition not met - save streak if long enough
                        if len(streak_dates) >= consecutive_days:
                            start_date = streak_dates[0]
                            end_date = streak_dates[-1]
                            pattern_key = (watcher_id, param_name, start_date.isoformat(), end_date.isoformat())
                            
                            if pattern_key not in patterns_seen:
                                patterns_seen.add(pattern_key)
                                start_str = start_date.strftime('%b %d')
                                end_str = end_date.strftime('%b %d')
                                date_pattern = f"({start_str} - {end_str})"
                                
                                existing = Alert.query.filter(
                                    Alert.user_id == watcher_id,
                                    Alert.alert_type == 'trigger',
                                    Alert.created_at >= datetime.now() - timedelta(hours=24),
                                    Alert.content.ilike(f"%{watched_user.username}'s {param_name}%{date_pattern}%")
                                ).first()
                                
                                if not existing:
                                    content = f"{watched_user.username}'s {param_name} has been at concerning levels for {len(streak_dates)} consecutive days {date_pattern}"
                                    alert = create_alert_with_email(
                                        user_id=watcher_id,
                                        title=f"Wellness Alert for {watched_user.username}",
                                        content=content,
                                        alert_type='trigger',
                                        source_user_id=watched_user.id,
                                        alert_category='trigger'
                                    )
                                    if alert:
                                        alerts_created += 1
                                        logger.info(f"[TRIGGER PROCESS] ✅ Created alert for {param_name} {date_pattern}")
                                else:
                                    alerts_skipped_duplicate += 1
                        
                        streak_dates = []
                        last_date = None
                
                # Don't forget the last streak
                if len(streak_dates) >= consecutive_days:
                    start_date = streak_dates[0]
                    end_date = streak_dates[-1]
                    pattern_key = (watcher_id, param_name, start_date.isoformat(), end_date.isoformat())
                    
                    if pattern_key not in patterns_seen:
                        patterns_seen.add(pattern_key)
                        start_str = start_date.strftime('%b %d')
                        end_str = end_date.strftime('%b %d')
                        date_pattern = f"({start_str} - {end_str})"
                        
                        existing = Alert.query.filter(
                            Alert.user_id == watcher_id,
                            Alert.alert_type == 'trigger',
                            Alert.created_at >= datetime.now() - timedelta(hours=24),
                            Alert.content.ilike(f"%{watched_user.username}'s {param_name}%{date_pattern}%")
                        ).first()
                        
                        if not existing:
                            content = f"{watched_user.username}'s {param_name} has been at concerning levels for {len(streak_dates)} consecutive days {date_pattern}"
                            alert = create_alert_with_email(
                                user_id=watcher_id,
                                title=f"Wellness Alert for {watched_user.username}",
                                content=content,
                                alert_type='trigger',
                                source_user_id=watched_user.id,
                                alert_category='trigger'
                            )
                            if alert:
                                alerts_created += 1
                                logger.info(f"[TRIGGER PROCESS] ✅ Created alert for {param_name} {date_pattern}")
                        else:
                            alerts_skipped_duplicate += 1

        logger.info(f"[TRIGGER PROCESS] ========================================")
        logger.info(f"[TRIGGER PROCESS] process_parameter_triggers completed for user {user_id}")
        logger.info(f"[TRIGGER PROCESS] Summary: watchers={len(watchers_processed)}, alerts_created={alerts_created}, duplicates={alerts_skipped_duplicate}")
        logger.info(f"[TRIGGER PROCESS] ========================================")

    except Exception as e:
        logger.error(f"[TRIGGER PROCESS] ❌ Error processing parameter triggers: {str(e)}")
        logger.error(f"[TRIGGER PROCESS] Traceback: {traceback.format_exc()}")
        db.session.rollback()


def cleanup_stale_trigger_alerts_for_user(affected_user_id):
    """
    Automatically clean up trigger alerts that are no longer valid due to privacy changes.
    This runs whenever parameters are saved to ensure alerts respect current privacy settings.

    Args:
        affected_user_id: The user whose parameters were just updated
    """
    try:
        # Helper function to check privacy permissions
        def can_see_parameter(param_privacy, watcher_circle):
            """Check if watcher can see this parameter based on privacy and circle level"""
            if param_privacy == 'private':
                return False
            elif param_privacy == 'class_a':
                return watcher_circle == 'class_a'
            elif param_privacy == 'class_b':
                return watcher_circle in ['class_b', 'class_a']
            elif param_privacy == 'public':
                return True
            return False

        # Get the user who was just updated
        affected_user = User.query.get(affected_user_id)
        if not affected_user:
            return

        # Find all watchers who have triggers for this user
        triggers = ParameterTrigger.query.filter_by(watched_id=affected_user_id).all()

        for trigger in triggers:
            watcher_id = trigger.watcher_id

            # Get watcher's circle level
            watcher_circle = get_watcher_circle_level(affected_user_id, watcher_id)

            if not watcher_circle:
                # Watcher not in any circle - remove all their alerts for this user
                alerts_to_remove = Alert.query.filter(
                    Alert.user_id == watcher_id,
                    Alert.alert_type == 'trigger',
                    Alert.content.like(f"%{affected_user.username}%")
                ).all()

                for alert in alerts_to_remove:
                    db.session.delete(alert)
                    logger.info(
                        f"Removed alert {alert.id}: watcher {watcher_id} not in any circle of user {affected_user_id}")
                continue

            # Get current privacy settings for this user
            recent_param = SavedParameters.query.filter_by(
                user_id=affected_user_id
            ).order_by(SavedParameters.updated_at.desc()).first()

            if not recent_param:
                continue

            # Check each parameter type's privacy
            param_keywords = {
                'mood': 'mood_privacy',
                'anxiety': 'anxiety_privacy',
                'sleep_quality': 'sleep_quality_privacy',
                'sleep quality': 'sleep_quality_privacy',
                'physical_activity': 'physical_activity_privacy',
                'physical activity': 'physical_activity_privacy',
                'energy': 'energy_privacy'
            }

            # Get all trigger alerts for this watcher about this user
            watcher_alerts = Alert.query.filter(
                Alert.user_id == watcher_id,
                Alert.alert_type == 'trigger',
                Alert.content.like(f"%{affected_user.username}%")
            ).all()

            for alert in watcher_alerts:
                content = alert.content or ""

                # Find which parameter this alert is about
                privacy_attr = None
                for keyword, privacy_field in param_keywords.items():
                    if keyword in content.lower():
                        privacy_attr = privacy_field
                        break

                if not privacy_attr:
                    continue

                # Get current privacy setting for this parameter
                param_privacy = getattr(recent_param, privacy_attr, 'private')

                # Check if watcher should still see this alert
                if not can_see_parameter(param_privacy, watcher_circle):
                    db.session.delete(alert)
                    logger.info(
                        f"Auto-cleanup: Removed alert {alert.id} for watcher {watcher_id} - {param_privacy} vs {watcher_circle}")

        db.session.commit()
        logger.info(f"Auto-cleanup completed for user {affected_user_id}")

    except Exception as e:
        logger.error(f"Error in automatic cleanup: {str(e)}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        db.session.rollback()


def cleanup_all_stale_trigger_alerts():
    """
    One-time cleanup function to remove all existing stale trigger alerts.
    This should run on app startup after deployment to clean up historical alerts.
    """
    try:
        logger.info("Starting global trigger alerts cleanup...")

        # Helper function to check privacy permissions
        def can_see_parameter(param_privacy, watcher_circle):
            """Check if watcher can see this parameter based on privacy and circle level"""
            if param_privacy == 'private':
                return False
            elif param_privacy == 'class_a':
                return watcher_circle == 'class_a'
            elif param_privacy == 'class_b':
                return watcher_circle in ['class_b', 'class_a']
            elif param_privacy == 'public':
                return True
            return False

        # Get all trigger alerts
        all_trigger_alerts = Alert.query.filter_by(alert_type='trigger').all()

        total_checked = len(all_trigger_alerts)
        removed_count = 0
        kept_count = 0

        param_keywords = {
            'mood': 'mood_privacy',
            'anxiety': 'anxiety_privacy',
            'sleep_quality': 'sleep_quality_privacy',
            'sleep quality': 'sleep_quality_privacy',
            'physical_activity': 'physical_activity_privacy',
            'physical activity': 'physical_activity_privacy',
            'energy': 'energy_privacy'
        }

        for alert in all_trigger_alerts:
            watcher_id = alert.user_id
            content = alert.content or ""

            # Parse username from content
            if "'s " not in content:
                kept_count += 1
                continue

            username = content.split("'s ")[0]
            watched_user = User.query.filter_by(username=username).first()

            if not watched_user:
                kept_count += 1
                continue

            watched_id = watched_user.id

            # Get watcher's circle level
            watcher_circle = get_watcher_circle_level(watched_id, watcher_id)

            if not watcher_circle:
                # Watcher not in any circle - remove alert
                db.session.delete(alert)
                removed_count += 1
                logger.info(f"Global cleanup: Removed alert {alert.id} - watcher {watcher_id} not in circles")
                continue

            # Find which parameter this alert is about
            privacy_attr = None
            for keyword, privacy_field in param_keywords.items():
                if keyword in content.lower():
                    privacy_attr = privacy_field
                    break

            if not privacy_attr:
                kept_count += 1
                continue

            # Get current privacy settings
            recent_param = SavedParameters.query.filter_by(
                user_id=watched_id
            ).order_by(SavedParameters.updated_at.desc()).first()

            if not recent_param:
                kept_count += 1
                continue

            param_privacy = getattr(recent_param, privacy_attr, 'private')

            # Check if watcher should see this alert
            if not can_see_parameter(param_privacy, watcher_circle):
                db.session.delete(alert)
                removed_count += 1
                logger.info(
                    f"Global cleanup: Removed alert {alert.id} - privacy violation ({param_privacy} vs {watcher_circle})")
            else:
                kept_count += 1

        db.session.commit()

        logger.info(f"Global cleanup completed: Checked {total_checked}, Removed {removed_count}, Kept {kept_count}")
        return removed_count

    except Exception as e:
        logger.error(f"Error in global cleanup: {str(e)}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        db.session.rollback()
        return 0


@app.route('/api/parameters/dates', methods=['GET'])
@login_required
def get_parameter_dates():
    """Get all dates that have saved parameters for the current user"""
    try:
        user_id = session.get('user_id')

        # Get all parameter dates for this user using SQLAlchemy ORM
        params = SavedParameters.query.filter_by(user_id=user_id).all()

        # Extract just the dates
        dates = [param.date.strftime('%Y-%m-%d') if hasattr(param.date, 'strftime') else str(param.date)
                 for param in params if param.date]

        return jsonify({
            'success': True,
            'dates': dates
        })

    except Exception as e:
        logger.error(f"Get parameter dates error: {str(e)}")
        return jsonify({
            'success': False,
            'dates': []
        })


@app.route('/api/parameters/insights')
@login_required
def get_insights():
    """Get parameter insights"""
    try:
        user_id = session['user_id']

        # Get last 30 days - SQLAlchemy 2.0 style
        thirty_days_ago = datetime.now().date() - timedelta(days=30)
        params_stmt = select(SavedParameters).filter(
            SavedParameters.user_id == user_id,
            SavedParameters.date >= thirty_days_ago
        )
        params = db.session.execute(params_stmt).scalars().all()

        if not params:
            return jsonify({'message': 'No data available for insights'})

        # Calculate insights
        avg_sleep = sum(p.sleep_hours for p in params if p.sleep_hours) / len(params) if params else 0
        moods = [p.mood for p in params if p.mood]

        return jsonify({
            'average_sleep': round(avg_sleep, 1),
            'total_entries': len(params),
            'most_common_mood': max(moods, key=moods.count) if moods else 'N/A',
            'streak': calculate_streak(params)
        })

    except Exception as e:
        logger.error(f"Get insights error: {str(e)}")
        return jsonify({'message': 'Failed to get insights'})


@app.route('/api/parameters/today-status')
@login_required
def get_today_status():
    """FIX #5: Check if user has a COMPLETE diary entry for today (all 5 fields filled)
    
    Accepts optional 'client_date' query param for timezone-aware checking.
    """
    try:
        user_id = session['user_id']
        
        # FIX: Accept client's local date to handle timezone differences
        client_date = request.args.get('client_date')
        
        if client_date:
            today_str = client_date
        else:
            today_str = datetime.now().date().isoformat()
        
        # FIX: Query using string comparison since date column is String type
        params_stmt = select(SavedParameters).filter(
            SavedParameters.user_id == user_id,
            SavedParameters.date == today_str  # String comparison
        )
        entry = db.session.execute(params_stmt).scalar_one_or_none()
        
        if not entry:
            return jsonify({
                'has_entry_today': False,
                'has_complete_entry_today': False,
                'missing_fields': ['mood', 'energy', 'sleep_quality', 'physical_activity', 'anxiety'],
                'date': today_str
            })
        
        # FIX #5: Check which required fields are filled
        missing_fields = []
        
        mood_val = getattr(entry, 'mood', None)
        if not mood_val or mood_val == 0:
            missing_fields.append('mood')
        
        energy_val = getattr(entry, 'energy', None)
        if not energy_val or energy_val == 0:
            missing_fields.append('energy')
        
        sleep_val = getattr(entry, 'sleep_quality', None)
        if not sleep_val or sleep_val == 0:
            missing_fields.append('sleep_quality')
        
        activity_val = getattr(entry, 'physical_activity', None)
        if not activity_val or activity_val == 0:
            missing_fields.append('physical_activity')
        
        anxiety_val = getattr(entry, 'anxiety', None)
        if not anxiety_val or anxiety_val == 0:
            missing_fields.append('anxiety')
        
        has_complete = len(missing_fields) == 0
        
        return jsonify({
            'has_entry_today': True,
            'has_complete_entry_today': has_complete,
            'missing_fields': missing_fields,
            'date': today_str
        })
        
    except Exception as e:
        logger.error(f"Today status error: {str(e)}")
        return jsonify({'has_entry_today': False, 'has_complete_entry_today': False, 'error': str(e)})


@app.route('/api/parameters/should-redirect-to-diary')
@login_required
def should_redirect_to_diary():
    """
    FIX: Check if user should be redirected to diary page.
    This only returns True ONCE per session after login AND only if diary is incomplete.
    After the first check, the flag is cleared so users can navigate freely.
    
    Accepts optional 'client_date' query param for timezone-aware checking.
    """
    try:
        user_id = session['user_id']
        
        # FIX: Accept client's local date to handle timezone differences
        # Client sends their local date as YYYY-MM-DD string
        client_date = request.args.get('client_date')
        
        if client_date:
            # Use client's date (already a string in YYYY-MM-DD format)
            today_str = client_date
            logger.info(f"Using client date for diary check: {today_str}")
        else:
            # Fallback to server date as string
            today_str = datetime.now().date().isoformat()
            logger.info(f"Using server date for diary check: {today_str}")
        
        # Check if redirect is pending for this session
        redirect_pending = session.get('diary_redirect_pending', False)
        
        if not redirect_pending:
            # Redirect already done or not needed - user can navigate freely
            return jsonify({
                'should_redirect': False,
                'reason': 'redirect_already_done'
            })
        
        # FIX: Query using string comparison since date column is String type
        params_stmt = select(SavedParameters).filter(
            SavedParameters.user_id == user_id,
            SavedParameters.date == today_str  # String comparison
        )
        entry = db.session.execute(params_stmt).scalar_one_or_none()
        
        logger.info(f"Diary check for user {user_id}, date {today_str}: entry found = {entry is not None}")
        
        if not entry:
            # No entry today - should redirect
            # Clear the flag so we don't redirect again after this
            session['diary_redirect_pending'] = False
            return jsonify({
                'should_redirect': True,
                'reason': 'no_entry_today',
                'date': today_str
            })
        
        # Check which required fields are filled
        missing_fields = []
        
        mood_val = getattr(entry, 'mood', None)
        if not mood_val or mood_val == 0:
            missing_fields.append('mood')
        
        energy_val = getattr(entry, 'energy', None)
        if not energy_val or energy_val == 0:
            missing_fields.append('energy')
        
        sleep_val = getattr(entry, 'sleep_quality', None)
        if not sleep_val or sleep_val == 0:
            missing_fields.append('sleep_quality')
        
        activity_val = getattr(entry, 'physical_activity', None)
        if not activity_val or activity_val == 0:
            missing_fields.append('physical_activity')
        
        anxiety_val = getattr(entry, 'anxiety', None)
        if not anxiety_val or anxiety_val == 0:
            missing_fields.append('anxiety')
        
        has_complete = len(missing_fields) == 0
        
        logger.info(f"Diary check for user {user_id}: complete={has_complete}, missing={missing_fields}")
        logger.info(f"Entry values - mood:{mood_val}, energy:{energy_val}, sleep:{sleep_val}, activity:{activity_val}, anxiety:{anxiety_val}")
        
        # Clear the redirect flag regardless of completion status
        # (we only check once per session - user can navigate after this)
        session['diary_redirect_pending'] = False
        
        return jsonify({
            'should_redirect': not has_complete,
            'reason': 'incomplete_entry' if not has_complete else 'entry_complete',
            'missing_fields': missing_fields,
            'date': today_str,
            'entry_found': True,
            'values': {
                'mood': mood_val,
                'energy': energy_val,
                'sleep_quality': sleep_val,
                'physical_activity': activity_val,
                'anxiety': anxiety_val
            }
        })
        
    except Exception as e:
        logger.error(f"Should redirect check error: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        # On error, don't redirect (fail safe) and clear the flag
        session['diary_redirect_pending'] = False
        return jsonify({'should_redirect': False, 'reason': 'error', 'error': str(e)})


@app.route('/api/parameters/progress')
@login_required
def get_progress():
    """FIX #2 & #3: Get progress data for charts - includes Anxiety, uses 1-5 scale"""
    try:
        user_id = session['user_id']
        days = request.args.get('days', 30, type=int)  # Default to 30 days
        
        # Limit to reasonable ranges
        if days > 365:
            days = 365
        
        # FIX #3: Proper date range calculation
        end_date = datetime.now().date()
        start_date = end_date - timedelta(days=days)
        
        # Get parameters for the period
        params_stmt = select(SavedParameters).filter(
            SavedParameters.user_id == user_id,
            SavedParameters.date >= start_date,
            SavedParameters.date <= end_date
        ).order_by(SavedParameters.date)
        params = db.session.execute(params_stmt).scalars().all()
        
        dates = []
        mood_data = []
        energy_data = []
        sleep_data = []
        activity_data = []
        anxiety_data = []  # FIX #2: Added anxiety
        
        def safe_int(val):
            """Convert value to int, handling strings and None. Diary values are 1-4."""
            if val is None:
                return None
            try:
                v = int(val)
                return v if 1 <= v <= 4 else None  # Diary entries use 1-4 scale
            except (ValueError, TypeError):
                # Handle text values by mapping to 1-4 scale
                if isinstance(val, str):
                    val_lower = val.lower()
                    text_map = {
                        '1': 1, '2': 2, '3': 3, '4': 4,
                        'terrible': 1, 'bad': 1, 'poor': 2, 'low': 2,
                        'okay': 3, 'neutral': 3, 'moderate': 3, 'medium': 3,
                        'good': 4, 'high': 4, 'great': 4, 'excellent': 4
                    }
                    return text_map.get(val_lower, None)
                return None
        
        for p in params:
            dates.append(p.date.strftime('%m/%d'))
            
            # Get mood value (try different column names)
            mood_val = getattr(p, 'mood', None) or getattr(p, 'mood_rating', None)
            mood_data.append(safe_int(mood_val))
            
            # Get energy value
            energy_val = getattr(p, 'energy', None) or getattr(p, 'energy_rating', None)
            energy_data.append(safe_int(energy_val))
            
            # Get sleep value
            sleep_val = getattr(p, 'sleep_quality', None) or getattr(p, 'sleep_hours', None)
            sleep_data.append(safe_int(sleep_val))
            
            # Get activity value
            activity_val = getattr(p, 'physical_activity', None) or getattr(p, 'exercise', None)
            activity_data.append(safe_int(activity_val))
            
            # FIX #2: Get anxiety value
            anxiety_val = getattr(p, 'anxiety', None) or getattr(p, 'anxiety_level', None)
            anxiety_data.append(safe_int(anxiety_val))
        
        # Calculate averages
        def calc_avg(data):
            valid = [v for v in data if v is not None]
            return sum(valid) / len(valid) if valid else None
        
        avg_mood = calc_avg(mood_data)
        avg_energy = calc_avg(energy_data)
        avg_sleep = calc_avg(sleep_data)
        avg_activity = calc_avg(activity_data)
        avg_anxiety = calc_avg(anxiety_data)  # FIX #2: Added anxiety average
        
        # PJ6009: Get language from query parameter first, then fall back to user preference
        user_language = request.args.get('lang', None)
        if not user_language:
            try:
                user = db.session.get(User, user_id)
                if user and user.preferred_language:
                    user_language = user.preferred_language
                else:
                    user_language = 'en'
            except:
                user_language = 'en'
        
        # Validate language is supported
        if user_language not in ['en', 'he', 'ar', 'ru']:
            user_language = 'en'
        
        # Generate insights including anxiety with user's language
        insights = generate_progress_insights(avg_mood, avg_energy, avg_sleep, avg_activity, len(params), avg_anxiety, user_language)
        
        return jsonify({
            'dates': dates,
            'mood': mood_data,
            'energy': energy_data,
            'sleep': sleep_data,
            'activity': activity_data,
            'anxiety': anxiety_data,  # FIX #2: Added anxiety data
            'totalCheckins': len(params),
            'avgMood': avg_mood,
            'avgEnergy': avg_energy,
            'avgSleep': avg_sleep,
            'avgActivity': avg_activity,
            'avgAnxiety': avg_anxiety,  # FIX #2: Added anxiety average
            'insights': insights
        })
        
    except Exception as e:
        logger.error(f"Progress data error: {str(e)}")
        return jsonify({'error': str(e)}), 500


def generate_progress_insights(avg_mood, avg_energy, avg_sleep, avg_activity, total_entries, avg_anxiety=None, language='en'):
    """Generate personalized insights based on averages. Diary values are 1-4, chart Y-axis is 1-5."""
    
    # PJ6003: Insight translations
    insight_translations = {
        'en': {
            'start_tracking': 'Start tracking your daily wellness to receive personalized insights!',
            'logged_entries': "Great job tracking for {days} days!",
            'logged_few': "You've logged {days} entries. Keep tracking daily for better insights!",
            'mood_positive': 'Your mood has been generally positive. Keep up the good work!',
            'mood_low': 'Your mood has been lower than usual. Consider activities that boost your wellbeing.',
            'energy_strong': 'Your energy levels are strong!',
            'energy_low': 'Low energy detected. Consider prioritizing rest or exercise.',
            'sleep_great': 'Great sleep quality!',
            'sleep_improve': 'Your sleep quality could be improved. Consider better sleep hygiene.',
            'activity_excellent': 'Excellent activity levels!',
            'activity_low': 'Consider increasing your physical activity for better wellbeing.',
            'anxiety_high': 'Anxiety levels appear elevated. Consider relaxation techniques or speaking with a professional.',
            'anxiety_good': 'Good anxiety management!',
            'making_progress': "You're making progress! Continue tracking for more detailed insights."
        },
        'he': {
            'start_tracking': 'התחל/י לעקוב אחרי הבריאות שלך כדי לקבל תובנות מותאמות אישית!',
            'logged_entries': 'כל הכבוד על מעקב במשך {days} ימים!',
            'logged_few': 'רשמת {days} רשומות. המשך/י לעקוב יומיומית לתובנות טובות יותר!',
            'mood_positive': 'מצב הרוח שלך היה חיובי באופן כללי. המשך/י כך!',
            'mood_low': 'מצב הרוח שלך היה נמוך מהרגיל. שקול/י פעילויות שמשפרות את הרווחה שלך.',
            'energy_strong': 'רמות האנרגיה שלך חזקות!',
            'energy_low': 'זוהתה אנרגיה נמוכה. שקול/י לתעדף מנוחה או פעילות גופנית.',
            'sleep_great': 'איכות שינה מעולה!',
            'sleep_improve': 'איכות השינה שלך יכולה להשתפר. שקול/י היגיינת שינה טובה יותר.',
            'activity_excellent': 'רמות פעילות מצוינות!',
            'activity_low': 'שקול/י להגביר את הפעילות הגופנית שלך לרווחה טובה יותר.',
            'anxiety_high': 'רמות החרדה נראות מוגברות. שקול/י טכניקות הרגעה או שיחה עם מומחה.',
            'anxiety_good': 'ניהול חרדה טוב!',
            'making_progress': 'את/ה מתקדם/ת! המשך/י לעקוב לתובנות מפורטות יותר.'
        },
        'ar': {
            'start_tracking': 'ابدأ بتتبع صحتك اليومية لتلقي رؤى مخصصة!',
            'logged_entries': 'عمل رائع في التتبع لمدة {days} أيام!',
            'logged_few': 'لقد سجلت {days} إدخالات. استمر في التتبع يوميًا للحصول على رؤى أفضل!',
            'mood_positive': 'كان مزاجك إيجابيًا بشكل عام. استمر في ذلك!',
            'mood_low': 'كان مزاجك أقل من المعتاد. فكر في أنشطة تعزز رفاهيتك.',
            'energy_strong': 'مستويات طاقتك قوية!',
            'energy_low': 'تم اكتشاف طاقة منخفضة. فكر في إعطاء الأولوية للراحة أو التمارين.',
            'sleep_great': 'جودة نوم رائعة!',
            'sleep_improve': 'يمكن تحسين جودة نومك. فكر في نظافة نوم أفضل.',
            'activity_excellent': 'مستويات نشاط ممتازة!',
            'activity_low': 'فكر في زيادة نشاطك البدني لرفاهية أفضل.',
            'anxiety_high': 'يبدو أن مستويات القلق مرتفعة. فكر في تقنيات الاسترخاء أو التحدث مع متخصص.',
            'anxiety_good': 'إدارة قلق جيدة!',
            'making_progress': 'أنت تحرز تقدمًا! استمر في التتبع للحصول على رؤى أكثر تفصيلاً.'
        },
        'ru': {
            'start_tracking': 'Начните отслеживать своё здоровье, чтобы получать персонализированные советы!',
            'logged_entries': 'Отличная работа! Вы отслеживаете уже {days} дней!',
            'logged_few': 'Вы записали {days} записей. Продолжайте отслеживать ежедневно для лучших советов!',
            'mood_positive': 'Ваше настроение было в целом позитивным. Продолжайте в том же духе!',
            'mood_low': 'Ваше настроение было ниже обычного. Подумайте о занятиях, которые улучшат ваше самочувствие.',
            'energy_strong': 'Ваш уровень энергии высок!',
            'energy_low': 'Обнаружен низкий уровень энергии. Подумайте о приоритете отдыха или упражнений.',
            'sleep_great': 'Отличное качество сна!',
            'sleep_improve': 'Качество вашего сна можно улучшить. Подумайте о лучшей гигиене сна.',
            'activity_excellent': 'Отличный уровень активности!',
            'activity_low': 'Подумайте об увеличении физической активности для лучшего самочувствия.',
            'anxiety_high': 'Уровень тревожности повышен. Подумайте о техниках релаксации или консультации специалиста.',
            'anxiety_good': 'Хорошее управление тревожностью!',
            'making_progress': 'Вы делаете успехи! Продолжайте отслеживать для более детальных советов.'
        }
    }
    
    t = insight_translations.get(language, insight_translations['en'])
    insights = []
    
    if total_entries == 0:
        return t['start_tracking']
    
    if total_entries < 7:
        insights.append(t['logged_few'].replace('{days}', str(total_entries)))
    else:
        insights.append(t['logged_entries'].replace('{days}', str(total_entries)))
    
    # Thresholds based on 1-4 scale (midpoint is 2.5)
    if avg_mood is not None:
        if avg_mood >= 3:
            insights.append(t['mood_positive'])
        elif avg_mood <= 2:
            insights.append(t['mood_low'])
    
    if avg_energy is not None:
        if avg_energy >= 3:
            insights.append(t['energy_strong'])
        elif avg_energy <= 2:
            insights.append(t['energy_low'])
    
    if avg_sleep is not None:
        if avg_sleep >= 3:
            insights.append(t['sleep_great'])
        elif avg_sleep <= 2:
            insights.append(t['sleep_improve'])
    
    if avg_activity is not None:
        if avg_activity >= 3:
            insights.append(t['activity_excellent'])
        elif avg_activity <= 2:
            insights.append(t['activity_low'])
    
    # Anxiety insights (higher = worse for anxiety)
    if avg_anxiety is not None:
        if avg_anxiety >= 3:
            insights.append(t['anxiety_high'])
        elif avg_anxiety <= 2:
            insights.append(t['anxiety_good'])
    
    if not insights:
        insights.append(t['making_progress'])
    
    return " ".join(insights)


# =====================
# REPORT GENERATION - TRANSLATIONS
# =====================

def get_report_translations(lang='en'):
    """Get report translations for Excel/PDF generation"""
    translations = {
        'en': {
            'title': 'Weekly Progress Report - Patient',
            'week': 'Week',
            'daily_checkin': 'Daily Check-in',
            'date': 'Date',
            'day': 'Day',
            'checkin_time': 'Check-in Time',
            'mood': 'Mood',
            'mood_scale': '(1-4)',
            'mood_notes': 'Mood Notes',
            'energy': 'Energy',
            'energy_scale': '(1-4)',
            'energy_notes': 'Energy Notes',
            'social': 'Social Activity',
            'social_scale': '(1-4)',
            'social_notes': 'Social Notes',
            'sleep': 'Sleep',
            'sleep_scale': '(1-4)',
            'sleep_notes': 'Sleep Notes',
            'anxiety': 'Anxiety',
            'anxiety_scale': '(1-4)',
            'anxiety_notes': 'Anxiety Notes',
            'motivation': 'Motivation',
            'motivation_scale': '(1-4)',
            'motivation_notes': 'Motivation Notes',
            'medication': 'Medication',
            'medication_scale': '(1-4)',
            'medication_notes': 'Medication Notes',
            'physical': 'Activity',
            'physical_scale': '(1-4)',
            'physical_notes': 'Physical Notes',
            'completion': 'Completion',
            'no_checkin': 'No Check-in',
            'weekly_summary': 'Weekly Summary',
            'days_completed': 'days',
            'checkin_completion': 'Check-in Completion:',
            'mood_level': 'Mood Level:',
            'energy_level': 'Energy:',
            'social_activity': 'Social Activity:',
            'sleep_quality': 'Sleep Quality:',
            'anxiety_level': 'Anxiety Level:',
            'motivation_level': 'Motivation:',
            'medication_level': 'Medication:',
            'physical_level': 'Physical Activity:',
            'good': 'Good',
            'needs_support': 'Needs Support',
            'days': {
                0: 'Monday', 1: 'Tuesday', 2: 'Wednesday', 
                3: 'Thursday', 4: 'Friday', 5: 'Saturday', 6: 'Sunday'
            }
        },
        'he': {
            'title': 'דוח התקדמות שבועי - מטופל',
            'week': 'שבוע',
            'daily_checkin': "צ'ק-אין יומי",
            'date': 'תאריך',
            'day': 'יום',
            'checkin_time': "זמן צ'ק-אין",
            'mood': 'מצב רוח',
            'mood_scale': '(1-4)',
            'mood_notes': 'מצב רוח הערות',
            'energy': 'אנרגיה',
            'energy_scale': '(1-4)',
            'energy_notes': 'אנרגיה הערות',
            'social': 'פעילות חברתית',
            'social_scale': '(1-4)',
            'social_notes': 'פעילות חברתית הערות',
            'sleep': 'שינה',
            'sleep_scale': '(1-4)',
            'sleep_notes': 'שינה הערות',
            'anxiety': 'חרדה',
            'anxiety_scale': '(1-4)',
            'anxiety_notes': 'חרדה הערות',
            'motivation': 'מוטיבציה',
            'motivation_scale': '(1-4)',
            'motivation_notes': 'מוטיבציה הערות',
            'medication': 'תרופות',
            'medication_scale': '(1-4)',
            'medication_notes': 'תרופות הערות',
            'physical': 'פעילות',
            'physical_scale': '(1-4)',
            'physical_notes': 'פעילות גופנית הערות',
            'completion': 'השלמה',
            'no_checkin': "אין צ'ק-אין",
            'weekly_summary': 'סיכום שבועי',
            'days_completed': 'ימים',
            'checkin_completion': "השלמת צ'ק-אין:",
            'mood_level': 'מצב רוח:',
            'energy_level': 'אנרגיה:',
            'social_activity': 'פעילות חברתית:',
            'sleep_quality': 'איכות שינה:',
            'anxiety_level': 'רמת חרדה:',
            'motivation_level': 'מוטיבציה:',
            'medication_level': 'תרופות:',
            'physical_level': 'פעילות גופנית:',
            'good': 'טוב',
            'needs_support': 'זקוק לתמיכה',
            'days': {
                0: 'יום שני', 1: 'יום שלישי', 2: 'יום רביעי',
                3: 'יום חמישי', 4: 'יום שישי', 5: 'שבת', 6: 'יום ראשון'
            }
        },
        'ar': {
            'title': 'تقرير التقدم الأسبوعي - المريض',
            'week': 'الأسبوع',
            'daily_checkin': 'تسجيل الدخول اليومي',
            'date': 'التاريخ',
            'day': 'اليوم',
            'checkin_time': 'وقت التسجيل',
            'mood': 'المزاج',
            'mood_scale': '(1-4)',
            'mood_notes': 'ملاحظات المزاج',
            'energy': 'الطاقة',
            'energy_scale': '(1-4)',
            'energy_notes': 'ملاحظات الطاقة',
            'social': 'النشاط الاجتماعي',
            'social_scale': '(1-4)',
            'social_notes': 'ملاحظات اجتماعية',
            'sleep': 'النوم',
            'sleep_scale': '(1-4)',
            'sleep_notes': 'ملاحظات النوم',
            'anxiety': 'القلق',
            'anxiety_scale': '(1-4)',
            'anxiety_notes': 'ملاحظات القلق',
            'motivation': 'التحفيز',
            'motivation_scale': '(1-4)',
            'motivation_notes': 'ملاحظات التحفيز',
            'medication': 'الأدوية',
            'medication_scale': '(1-4)',
            'medication_notes': 'ملاحظات الأدوية',
            'physical': 'النشاط',
            'physical_scale': '(1-4)',
            'physical_notes': 'ملاحظات النشاط',
            'completion': 'الإكمال',
            'no_checkin': 'لا يوجد تسجيل',
            'weekly_summary': 'ملخص أسبوعي',
            'days_completed': 'أيام',
            'checkin_completion': 'إكمال التسجيل:',
            'mood_level': 'مستوى المزاج:',
            'energy_level': 'الطاقة:',
            'social_activity': 'النشاط الاجتماعي:',
            'sleep_quality': 'جودة النوم:',
            'anxiety_level': 'مستوى القلق:',
            'motivation_level': 'التحفيز:',
            'medication_level': 'الأدوية:',
            'physical_level': 'النشاط البدني:',
            'good': 'جيد',
            'needs_support': 'يحتاج دعم',
            'days': {
                0: 'الإثنين', 1: 'الثلاثاء', 2: 'الأربعاء',
                3: 'الخميس', 4: 'الجمعة', 5: 'السبت', 6: 'الأحد'
            }
        },
        'ru': {
            'title': 'Еженедельный отчет о прогрессе - Пациент',
            'week': 'Неделя',
            'daily_checkin': 'Ежедневная отметка',
            'date': 'Дата',
            'day': 'День',
            'checkin_time': 'Время отметки',
            'mood': 'Настроение',
            'mood_scale': '(1-4)',
            'mood_notes': 'Заметки о настроении',
            'energy': 'Энергия',
            'energy_scale': '(1-4)',
            'energy_notes': 'Заметки об энергии',
            'social': 'Социальная активность',
            'social_scale': '(1-4)',
            'social_notes': 'Социальные заметки',
            'sleep': 'Сон',
            'sleep_scale': '(1-4)',
            'sleep_notes': 'Заметки о сне',
            'anxiety': 'Тревога',
            'anxiety_scale': '(1-4)',
            'anxiety_notes': 'Заметки о тревоге',
            'motivation': 'Мотивация',
            'motivation_scale': '(1-4)',
            'motivation_notes': 'Заметки о мотивации',
            'medication': 'Лекарства',
            'medication_scale': '(1-4)',
            'medication_notes': 'Заметки о лекарствах',
            'physical': 'Активность',
            'physical_scale': '(1-4)',
            'physical_notes': 'Заметки об активности',
            'completion': 'Завершение',
            'no_checkin': 'Нет отметки',
            'weekly_summary': 'Еженедельный итог',
            'days_completed': 'дней',
            'checkin_completion': 'Выполнение отметок:',
            'mood_level': 'Настроение:',
            'energy_level': 'Энергия:',
            'social_activity': 'Социальная активность:',
            'sleep_quality': 'Качество сна:',
            'anxiety_level': 'Уровень тревоги:',
            'motivation_level': 'Мотивация:',
            'medication_level': 'Лекарства:',
            'physical_level': 'Физическая активность:',
            'good': 'Хорошо',
            'needs_support': 'Нужна поддержка',
            'days': {
                0: 'Понедельник', 1: 'Вторник', 2: 'Среда',
                3: 'Четверг', 4: 'Пятница', 5: 'Суббота', 6: 'Воскресенье'
            }
        }
    }
    return translations.get(lang, translations['en'])


def get_value_color(value, is_anxiety=False):
    """Get color for parameter value. Green=good, Red=bad.
    For anxiety, scale is inverted (4=bad, 1=good)"""
    if value is None or value == '':
        return None
    try:
        val = int(value)
    except (ValueError, TypeError):
        return None
    
    if is_anxiety:
        # Anxiety: 1=green (low anxiety=good), 4=red (high anxiety=bad)
        colors = {1: 'C8E6C9', 2: 'FFF9C4', 3: 'FFE0B2', 4: 'FFCDD2'}
    else:
        # Other metrics: 4=green (high=good), 1=red (low=bad)
        colors = {4: 'C8E6C9', 3: 'FFF9C4', 2: 'FFE0B2', 1: 'FFCDD2'}
    
    return colors.get(val, None)


def get_week_data(user_id, days=7):
    """Get parameter data for the past week"""
    end_date = datetime.now().date()
    start_date = end_date - timedelta(days=days-1)
    
    params_stmt = select(SavedParameters).filter(
        SavedParameters.user_id == user_id,
        SavedParameters.date >= start_date,
        SavedParameters.date <= end_date
    ).order_by(SavedParameters.date)
    params = db.session.execute(params_stmt).scalars().all()
    
    # Create a dict mapping date to params
    params_by_date = {p.date: p for p in params}
    
    # Build data for each day
    week_data = []
    current_date = start_date
    while current_date <= end_date:
        p = params_by_date.get(current_date)
        day_data = {
            'date': current_date,
            'day_of_week': current_date.weekday(),
            'has_checkin': p is not None,
            'checkin_time': p.created_at.strftime('%H:%M') if p and hasattr(p, 'created_at') and p.created_at else '',
            'mood': getattr(p, 'mood', None) if p else None,
            'energy': getattr(p, 'energy', None) if p else None,
            'social': getattr(p, 'social_activity', None) if p else None,
            'sleep': getattr(p, 'sleep_quality', None) if p else None,
            'anxiety': getattr(p, 'anxiety', None) if p else None,
            'motivation': getattr(p, 'motivation', None) if p else None,
            'medication': getattr(p, 'medication', None) if p else None,
            'physical': getattr(p, 'physical_activity', None) if p else None,
        }
        week_data.append(day_data)
        current_date += timedelta(days=1)
    
    return week_data, start_date, end_date


def calculate_summary(week_data):
    """Calculate weekly summary statistics"""
    def calc_avg(field):
        vals = [d[field] for d in week_data if d[field] is not None]
        try:
            vals = [int(v) for v in vals]
            return sum(vals) / len(vals) if vals else None
        except (ValueError, TypeError):
            return None
    
    checkin_count = sum(1 for d in week_data if d['has_checkin'])
    total_days = len(week_data)
    
    return {
        'checkin_count': checkin_count,
        'total_days': total_days,
        'completion_pct': (checkin_count / total_days * 100) if total_days > 0 else 0,
        'avg_mood': calc_avg('mood'),
        'avg_energy': calc_avg('energy'),
        'avg_social': calc_avg('social'),
        'avg_sleep': calc_avg('sleep'),
        'avg_anxiety': calc_avg('anxiety'),
        'avg_motivation': calc_avg('motivation'),
        'avg_medication': calc_avg('medication'),
        'avg_physical': calc_avg('physical'),
    }


# =====================
# EXCEL REPORT GENERATION
# =====================

@app.route('/api/reports/excel')
@login_required
def generate_excel_report():
    """Generate Excel report for past 7 days"""
    if not OPENPYXL_AVAILABLE:
        logger.error("Excel generation failed: openpyxl not available")
        return jsonify({'error': 'Excel generation not available - openpyxl not installed'}), 500
    
    try:
        user_id = session['user_id']
        user = User.query.get(user_id)
        lang = request.args.get('lang', 'en')
        t = get_report_translations(lang)
        
        week_data, start_date, end_date = get_week_data(user_id)
        summary = calculate_summary(week_data)
        
        # Get week number
        week_num = start_date.isocalendar()[1]
        year = start_date.year
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = t['daily_checkin']
        
        # RTL direction for Hebrew/Arabic
        if lang in ['he', 'ar']:
            ws.sheet_view.rightToLeft = True
        
        # Title
        ws['A1'] = f"{t['title']} {user.email}"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:J1')
        
        # Week info
        month_names = {
            'en': ['January', 'February', 'March', 'April', 'May', 'June',
                   'July', 'August', 'September', 'October', 'November', 'December'],
            'he': ['ינואר', 'פברואר', 'מרץ', 'אפריל', 'מאי', 'יוני',
                   'יולי', 'אוגוסט', 'ספטמבר', 'אוקטובר', 'נובמבר', 'דצמבר'],
            'ar': ['يناير', 'فبراير', 'مارس', 'أبريل', 'مايو', 'يونيو',
                   'يوليو', 'أغسطس', 'سبتمبر', 'أكتوبر', 'نوفمبر', 'ديسمبر'],
            'ru': ['Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
                   'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь']
        }
        months = month_names.get(lang, month_names['en'])
        ws['A2'] = f"{t['week']} {week_num}, {year} ({months[start_date.month-1]} {start_date.day} - {months[end_date.month-1]} {end_date.day}, {year})"
        ws.merge_cells('A2:J2')
        
        # Headers row - with label + scale format
        mood_header = f"{t['mood']}\n{t.get('mood_scale', '(1-4)')}"
        energy_header = f"{t['energy']}\n{t.get('energy_scale', '(1-4)')}"
        social_header = f"{t['social']}\n{t.get('social_scale', '(1-4)')}"
        sleep_header = f"{t['sleep']}\n{t.get('sleep_scale', '(1-4)')}"
        anxiety_header = f"{t['anxiety']}\n{t.get('anxiety_scale', '(1-4)')}"
        motivation_header = f"{t['motivation']}\n{t.get('motivation_scale', '(1-4)')}"
        medication_header = f"{t['medication']}\n{t.get('medication_scale', '(1-4)')}"
        physical_header = f"{t['physical']}\n{t.get('physical_scale', '(1-4)')}"
        
        headers = [t['date'], t['day'], t['checkin_time'],
                   mood_header, t['mood_notes'], energy_header, t['energy_notes'],
                   social_header, t['social_notes'], sleep_header, t['sleep_notes'],
                   anxiety_header, t['anxiety_notes'], motivation_header, t['motivation_notes'],
                   medication_header, t['medication_notes'], physical_header, t['physical_notes'],
                   t['completion']]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill('solid', fgColor='E3E3E3')
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        
        # Data rows
        for row_idx, day in enumerate(week_data, 5):
            ws.cell(row=row_idx, column=1, value=day['date'].strftime('%Y-%m-%d'))
            ws.cell(row=row_idx, column=2, value=t['days'].get(day['day_of_week'], ''))
            
            if day['has_checkin']:
                ws.cell(row=row_idx, column=3, value=day['checkin_time'] or '')
                
                # Mood
                val = day['mood']
                cell = ws.cell(row=row_idx, column=4, value=val)
                color = get_value_color(val, is_anxiety=False)
                if color:
                    cell.fill = PatternFill('solid', fgColor=color)
                
                # Energy
                val = day['energy']
                cell = ws.cell(row=row_idx, column=6, value=val)
                color = get_value_color(val, is_anxiety=False)
                if color:
                    cell.fill = PatternFill('solid', fgColor=color)
                
                # Social
                val = day['social']
                cell = ws.cell(row=row_idx, column=8, value=val)
                color = get_value_color(val, is_anxiety=False)
                if color:
                    cell.fill = PatternFill('solid', fgColor=color)
                
                # Sleep
                val = day['sleep']
                cell = ws.cell(row=row_idx, column=10, value=val)
                color = get_value_color(val, is_anxiety=False)
                if color:
                    cell.fill = PatternFill('solid', fgColor=color)
                
                # Anxiety (inverted scale)
                val = day['anxiety']
                cell = ws.cell(row=row_idx, column=12, value=val)
                color = get_value_color(val, is_anxiety=True)
                if color:
                    cell.fill = PatternFill('solid', fgColor=color)
                
                # Motivation
                val = day['motivation']
                cell = ws.cell(row=row_idx, column=14, value=val)
                color = get_value_color(val, is_anxiety=False)
                if color:
                    cell.fill = PatternFill('solid', fgColor=color)
                
                # Medication
                val = day['medication']
                cell = ws.cell(row=row_idx, column=16, value=val)
                color = get_value_color(val, is_anxiety=False)
                if color:
                    cell.fill = PatternFill('solid', fgColor=color)
                
                # Physical
                val = day['physical']
                cell = ws.cell(row=row_idx, column=18, value=val)
                color = get_value_color(val, is_anxiety=False)
                if color:
                    cell.fill = PatternFill('solid', fgColor=color)
                
                # Completion
                cell = ws.cell(row=row_idx, column=20, value='✓')
                cell.fill = PatternFill('solid', fgColor='C8E6C9')
            else:
                ws.cell(row=row_idx, column=3, value=t['no_checkin'])
                cell = ws.cell(row=row_idx, column=20, value='✗')
                cell.fill = PatternFill('solid', fgColor='FFCDD2')
        
        # Summary section
        summary_row = len(week_data) + 6
        ws.cell(row=summary_row, column=1, value=t['weekly_summary'])
        ws.cell(row=summary_row, column=1).font = Font(bold=True, size=12)
        ws.merge_cells(f'A{summary_row}:D{summary_row}')
        
        def format_avg(val, is_anxiety=False):
            if val is None:
                return '-'
            rating = f'{val:.1f}/4'
            if is_anxiety:
                status = t['good'] if val <= 2.5 else t['needs_support']
            else:
                status = t['good'] if val >= 2.5 else t['needs_support']
            return f'{rating} - {status}'
        
        summary_items = [
            (t['checkin_completion'], f"{summary['checkin_count']}/{summary['total_days']} {t['days_completed']} ({summary['completion_pct']:.0f}%)"),
            (t['mood_level'], format_avg(summary['avg_mood'])),
            (t['energy_level'], format_avg(summary['avg_energy'])),
            (t['social_activity'], format_avg(summary['avg_social'])),
            (t['sleep_quality'], format_avg(summary['avg_sleep'])),
            (t['anxiety_level'], format_avg(summary['avg_anxiety'], is_anxiety=True)),
            (t['motivation_level'], format_avg(summary['avg_motivation'])),
            (t['medication_level'], format_avg(summary['avg_medication'])),
            (t['physical_level'], format_avg(summary['avg_physical'])),
        ]
        
        for i, (label, value) in enumerate(summary_items):
            ws.cell(row=summary_row + 1 + i, column=1, value=label)
            ws.cell(row=summary_row + 1 + i, column=2, value=value)
        
        # Adjust column widths
        for col in range(1, 21):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Save to buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f'weekly_report_{start_date.strftime("%Y%m%d")}_{lang}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"Excel report error: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({'error': 'Failed to generate report'}), 500


# =====================
# PDF REPORT GENERATION
# =====================

def generate_pdf_with_weasyprint(user, lang, t, week_data, summary, start_date, end_date):
    """Generate PDF using WeasyPrint for full Unicode/RTL support"""
    week_num = start_date.isocalendar()[1]
    year = start_date.year
    
    month_names = {
        'en': ['January', 'February', 'March', 'April', 'May', 'June',
               'July', 'August', 'September', 'October', 'November', 'December'],
        'he': ['ינואר', 'פברואר', 'מרץ', 'אפריל', 'מאי', 'יוני',
               'יולי', 'אוגוסט', 'ספטמבר', 'אוקטובר', 'נובמבר', 'דצמבר'],
        'ar': ['يناير', 'فبراير', 'مارس', 'أبريل', 'مايو', 'يونيو',
               'يوليو', 'أغسطس', 'سبتمبر', 'أكتوبر', 'نوفمبر', 'ديسمبر'],
        'ru': ['Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
               'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь']
    }
    months = month_names.get(lang, month_names['en'])
    
    # RTL direction for Hebrew/Arabic
    direction = 'rtl' if lang in ['he', 'ar'] else 'ltr'
    text_align = 'right' if lang in ['he', 'ar'] else 'left'
    
    def get_color_class(val, is_anxiety=False):
        """Get CSS class for cell coloring"""
        if val is None:
            return ''
        if is_anxiety:
            # Anxiety: lower is better (1=green, 4=red)
            if val == 1:
                return 'bg-green'
            elif val == 2:
                return 'bg-yellow'
            elif val == 3:
                return 'bg-orange'
            else:
                return 'bg-red'
        else:
            # Others: higher is better (4=green, 1=red)
            if val == 4:
                return 'bg-green'
            elif val == 3:
                return 'bg-yellow'
            elif val == 2:
                return 'bg-orange'
            else:
                return 'bg-red'
    
    def format_avg(val, is_anxiety=False):
        if val is None:
            return '-'
        rating = f'{val:.1f}/4'
        if is_anxiety:
            status = t['good'] if val <= 2.5 else t['needs_support']
        else:
            status = t['good'] if val >= 2.5 else t['needs_support']
        return f'{rating} - {status}'
    
    # Build table rows
    table_rows = ''
    for day in week_data:
        if day['has_checkin']:
            table_rows += f'''
            <tr>
                <td>{day['date'].strftime('%Y-%m-%d')}</td>
                <td>{t['days'].get(day['day_of_week'], '')}</td>
                <td class="{get_color_class(day['mood'], False)}">{day['mood'] if day['mood'] else '-'}</td>
                <td class="{get_color_class(day['energy'], False)}">{day['energy'] if day['energy'] else '-'}</td>
                <td class="{get_color_class(day['sleep'], False)}">{day['sleep'] if day['sleep'] else '-'}</td>
                <td class="{get_color_class(day['anxiety'], True)}">{day['anxiety'] if day['anxiety'] else '-'}</td>
                <td class="{get_color_class(day['physical'], False)}">{day['physical'] if day['physical'] else '-'}</td>
                <td class="bg-green">✓</td>
            </tr>
            '''
        else:
            table_rows += f'''
            <tr>
                <td>{day['date'].strftime('%Y-%m-%d')}</td>
                <td>{t['days'].get(day['day_of_week'], '')}</td>
                <td>-</td>
                <td>-</td>
                <td>-</td>
                <td>-</td>
                <td>-</td>
                <td class="bg-red">{t['no_checkin']}</td>
            </tr>
            '''
    
    # HTML template with embedded fonts
    html_content = f'''
    <!DOCTYPE html>
    <html dir="{direction}" lang="{lang}">
    <head>
        <meta charset="UTF-8">
        <style>
            @import url('https://fonts.googleapis.com/css2?family=Noto+Sans:wght@400;700&family=Noto+Sans+Hebrew:wght@400;700&family=Noto+Sans+Arabic:wght@400;700&display=swap');
            
            * {{
                font-family: 'Noto Sans', 'Noto Sans Hebrew', 'Noto Sans Arabic', Arial, sans-serif;
            }}
            
            body {{
                direction: {direction};
                text-align: {text_align};
                padding: 20px;
                font-size: 12px;
                line-height: 1.4;
            }}
            
            h1 {{
                text-align: center;
                font-size: 18px;
                margin-bottom: 5px;
            }}
            
            .week-info {{
                text-align: center;
                margin-bottom: 20px;
                color: #666;
            }}
            
            h2 {{
                font-size: 14px;
                margin-top: 20px;
                margin-bottom: 10px;
                border-bottom: 1px solid #ccc;
                padding-bottom: 5px;
            }}
            
            table {{
                width: 100%;
                border-collapse: collapse;
                margin-bottom: 20px;
            }}
            
            th, td {{
                border: 1px solid #333;
                padding: 6px 4px;
                text-align: center;
                font-size: 10px;
            }}
            
            th {{
                background-color: #666;
                color: white;
                font-weight: bold;
            }}
            
            th .scale {{
                display: block;
                font-size: 9px;
                font-weight: normal;
            }}
            
            .bg-green {{ background-color: #C8E6C9; }}
            .bg-yellow {{ background-color: #FFF9C4; }}
            .bg-orange {{ background-color: #FFE0B2; }}
            .bg-red {{ background-color: #FFCDD2; }}
            
            .summary-item {{
                margin: 5px 0;
            }}
        </style>
    </head>
    <body>
        <h1>{t['title']} {user.email}</h1>
        <div class="week-info">{t['week']} {week_num}, {year} ({months[start_date.month-1]} {start_date.day} - {months[end_date.month-1]} {end_date.day}, {year})</div>
        
        <h2>{t['daily_checkin']}</h2>
        <table>
            <thead>
                <tr>
                    <th>{t['date']}</th>
                    <th>{t['day']}</th>
                    <th>{t['mood']}<span class="scale">{t.get('mood_scale', '(1-4)')}</span></th>
                    <th>{t['energy']}<span class="scale">{t.get('energy_scale', '(1-4)')}</span></th>
                    <th>{t['sleep']}<span class="scale">{t.get('sleep_scale', '(1-4)')}</span></th>
                    <th>{t['anxiety']}<span class="scale">{t.get('anxiety_scale', '(1-4)')}</span></th>
                    <th>{t['physical']}<span class="scale">{t.get('physical_scale', '(1-4)')}</span></th>
                    <th>{t['completion']}</th>
                </tr>
            </thead>
            <tbody>
                {table_rows}
            </tbody>
        </table>
        
        <h2>{t['weekly_summary']}</h2>
        <div class="summary-item">{t['checkin_completion']} {summary['checkin_count']}/{summary['total_days']} {t['days_completed']} ({summary['completion_pct']:.0f}%)</div>
        <div class="summary-item">{t['mood_level']} {format_avg(summary['avg_mood'])}</div>
        <div class="summary-item">{t['energy_level']} {format_avg(summary['avg_energy'])}</div>
        <div class="summary-item">{t['sleep_quality']} {format_avg(summary['avg_sleep'])}</div>
        <div class="summary-item">{t['anxiety_level']} {format_avg(summary['avg_anxiety'], is_anxiety=True)}</div>
        <div class="summary-item">{t['physical_level']} {format_avg(summary['avg_physical'])}</div>
    </body>
    </html>
    '''
    
    # Generate PDF with WeasyPrint
    output = io.BytesIO()
    HTML(string=html_content).write_pdf(output)
    output.seek(0)
    
    return output


@app.route('/api/reports/pdf')
@login_required
def generate_pdf_report():
    """Generate PDF report for past 7 days - uses WeasyPrint if available, falls back to reportlab"""
    
    # Check if any PDF library is available
    if not WEASYPRINT_AVAILABLE and not REPORTLAB_AVAILABLE:
        logger.error("PDF generation failed: no PDF library available")
        return jsonify({'error': 'PDF generation not available - install weasyprint or reportlab'}), 500
    
    try:
        user_id = session['user_id']
        user = User.query.get(user_id)
        lang = request.args.get('lang', 'en')
        t = get_report_translations(lang)
        
        week_data, start_date, end_date = get_week_data(user_id)
        summary = calculate_summary(week_data)
        
        filename = f'weekly_report_{start_date.strftime("%Y%m%d")}_{lang}.pdf'
        
        # Try WeasyPrint first (better Unicode/RTL support)
        if WEASYPRINT_AVAILABLE:
            try:
                logger.info(f"Generating PDF with WeasyPrint for lang={lang}")
                output = generate_pdf_with_weasyprint(user, lang, t, week_data, summary, start_date, end_date)
                return send_file(
                    output,
                    mimetype='application/pdf',
                    as_attachment=True,
                    download_name=filename
                )
            except Exception as e:
                logger.warning(f"WeasyPrint failed, trying reportlab: {str(e)}")
                if not REPORTLAB_AVAILABLE:
                    raise
        
        # Fallback to reportlab
        if REPORTLAB_AVAILABLE:
            logger.info(f"Generating PDF with reportlab for lang={lang}")
            
            week_num = start_date.isocalendar()[1]
            year = start_date.year
            
            month_names = {
                'en': ['January', 'February', 'March', 'April', 'May', 'June',
                       'July', 'August', 'September', 'October', 'November', 'December'],
                'he': ['ינואר', 'פברואר', 'מרץ', 'אפריל', 'מאי', 'יוני',
                       'יולי', 'אוגוסט', 'ספטמבר', 'אוקטובר', 'נובמבר', 'דצמבר'],
                'ar': ['يناير', 'فبراير', 'مارس', 'أبريل', 'مايو', 'يونيو',
                       'يوليو', 'أغسطس', 'سبتمبر', 'أكتوبر', 'نوفمبر', 'ديسمبر'],
                'ru': ['Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
                       'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь']
            }
            months = month_names.get(lang, month_names['en'])
            
            output = io.BytesIO()
            doc = SimpleDocTemplate(output, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
            
            styles = getSampleStyleSheet()
            elements = []
            
            # Title
            title_style = ParagraphStyle('Title', parent=styles['Title'], alignment=1)
            elements.append(Paragraph(f"{t['title']} {user.email}", title_style))
            
            # Week info
            week_info = f"{t['week']} {week_num}, {year} ({months[start_date.month-1]} {start_date.day} - {months[end_date.month-1]} {end_date.day}, {year})"
            elements.append(Paragraph(week_info, styles['Normal']))
            elements.append(Spacer(1, 20))
            
            # Daily Check-in header
            elements.append(Paragraph(t['daily_checkin'], styles['Heading2']))
            elements.append(Spacer(1, 10))
            
            # Table header with label + scale on two lines
            mood_header = f"{t['mood']}\n{t.get('mood_scale', '(1-4)')}"
            energy_header = f"{t['energy']}\n{t.get('energy_scale', '(1-4)')}"
            sleep_header = f"{t['sleep']}\n{t.get('sleep_scale', '(1-4)')}"
            anxiety_header = f"{t['anxiety']}\n{t.get('anxiety_scale', '(1-4)')}"
            physical_header = f"{t['physical']}\n{t.get('physical_scale', '(1-4)')}"
            
            table_header = [t['date'], t['day'], mood_header, energy_header, sleep_header, anxiety_header, physical_header, t['completion']]
            table_data = [table_header]
            
            def get_pdf_color(val, is_anxiety=False):
                hex_color = get_value_color(val, is_anxiety)
                if hex_color:
                    return HexColor('#' + hex_color)
                return None
            
            cell_colors = []
            
            for day in week_data:
                if day['has_checkin']:
                    row = [
                        day['date'].strftime('%Y-%m-%d'),
                        t['days'].get(day['day_of_week'], ''),
                        str(day['mood']) if day['mood'] else '-',
                        str(day['energy']) if day['energy'] else '-',
                        str(day['sleep']) if day['sleep'] else '-',
                        str(day['anxiety']) if day['anxiety'] else '-',
                        str(day['physical']) if day['physical'] else '-',
                        '✓'
                    ]
                    row_colors = [
                        None, None,
                        get_pdf_color(day['mood'], False),
                        get_pdf_color(day['energy'], False),
                        get_pdf_color(day['sleep'], False),
                        get_pdf_color(day['anxiety'], True),
                        get_pdf_color(day['physical'], False),
                        HexColor('#C8E6C9')
                    ]
                else:
                    row = [
                        day['date'].strftime('%Y-%m-%d'),
                        t['days'].get(day['day_of_week'], ''),
                        '-', '-', '-', '-', '-',
                        t['no_checkin']
                    ]
                    row_colors = [None, None, None, None, None, None, None, HexColor('#FFCDD2')]
                
                table_data.append(row)
                cell_colors.append(row_colors)
            
            col_widths = [70, 80, 50, 50, 50, 50, 60, 70]
            table = Table(table_data, colWidths=col_widths)
            
            style_commands = [
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]
            
            for row_idx, row_colors in enumerate(cell_colors):
                for col_idx, color in enumerate(row_colors):
                    if color:
                        style_commands.append(('BACKGROUND', (col_idx, row_idx + 1), (col_idx, row_idx + 1), color))
            
            table.setStyle(TableStyle(style_commands))
            elements.append(table)
            elements.append(Spacer(1, 20))
            
            # Summary section
            elements.append(Paragraph(t['weekly_summary'], styles['Heading2']))
            elements.append(Spacer(1, 10))
            
            def format_avg(val, is_anxiety=False):
                if val is None:
                    return '-'
                rating = f'{val:.1f}/4'
                if is_anxiety:
                    status = t['good'] if val <= 2.5 else t['needs_support']
                else:
                    status = t['good'] if val >= 2.5 else t['needs_support']
                return f'{rating} - {status}'
            
            summary_text = f"""
            {t['checkin_completion']} {summary['checkin_count']}/{summary['total_days']} {t['days_completed']} ({summary['completion_pct']:.0f}%)
            {t['mood_level']} {format_avg(summary['avg_mood'])}
            {t['energy_level']} {format_avg(summary['avg_energy'])}
            {t['sleep_quality']} {format_avg(summary['avg_sleep'])}
            {t['anxiety_level']} {format_avg(summary['avg_anxiety'], is_anxiety=True)}
            {t['physical_level']} {format_avg(summary['avg_physical'])}
            """
            
            for line in summary_text.strip().split('\n'):
                if line.strip():
                    elements.append(Paragraph(line.strip(), styles['Normal']))
            
            doc.build(elements)
            output.seek(0)
            
            return send_file(
                output,
                mimetype='application/pdf',
                as_attachment=True,
                download_name=filename
            )
        
        return jsonify({'error': 'No PDF library available'}), 500
        
    except Exception as e:
        logger.error(f"PDF report error: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({'error': 'Failed to generate report'}), 500


# =====================
# EMAIL REPORT
# =====================

@app.route('/api/reports/email', methods=['POST'])
@login_required
def email_report():
    """Send weekly report via email"""
    try:
        user_id = session['user_id']
        user = User.query.get(user_id)
        data = request.get_json() or {}
        lang = data.get('lang', 'en')
        recipient = data.get('email', user.email)
        t = get_report_translations(lang)
        
        week_data, start_date, end_date = get_week_data(user_id)
        summary = calculate_summary(week_data)
        
        week_num = start_date.isocalendar()[1]
        year = start_date.year
        
        month_names = {
            'en': ['January', 'February', 'March', 'April', 'May', 'June',
                   'July', 'August', 'September', 'October', 'November', 'December'],
            'he': ['ינואר', 'פברואר', 'מרץ', 'אפריל', 'מאי', 'יוני',
                   'יולי', 'אוגוסט', 'ספטמבר', 'אוקטובר', 'נובמבר', 'דצמבר'],
            'ar': ['يناير', 'فبراير', 'مارس', 'أبريل', 'مايو', 'يونيو',
                   'يوليو', 'أغسطس', 'سبتمبر', 'أكتوبر', 'نوفمبر', 'ديسمبر'],
            'ru': ['Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
                   'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь']
        }
        months = month_names.get(lang, month_names['en'])
        
        def format_avg(val, is_anxiety=False):
            if val is None:
                return '-'
            rating = f'{val:.1f}/4'
            if is_anxiety:
                status = t['good'] if val <= 2.5 else t['needs_support']
            else:
                status = t['good'] if val >= 2.5 else t['needs_support']
            return f'{rating} - {status}'
        
        # Build HTML email
        direction = 'rtl' if lang in ['he', 'ar'] else 'ltr'
        
        html_content = f"""
        <html dir="{direction}">
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; direction: {direction}; }}
                table {{ border-collapse: collapse; width: 100%; margin: 20px 0; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: center; }}
                th {{ background-color: #667eea; color: white; }}
                .green {{ background-color: #C8E6C9; }}
                .yellow {{ background-color: #FFF9C4; }}
                .orange {{ background-color: #FFE0B2; }}
                .red {{ background-color: #FFCDD2; }}
                .summary {{ margin-top: 20px; }}
                .summary-item {{ padding: 5px 0; }}
            </style>
        </head>
        <body>
            <h2>{t['title']} {user.email}</h2>
            <p>{t['week']} {week_num}, {year} ({months[start_date.month-1]} {start_date.day} - {months[end_date.month-1]} {end_date.day}, {year})</p>
            
            <h3>{t['daily_checkin']}</h3>
            <table>
                <tr>
                    <th>{t['date']}</th>
                    <th>{t['day']}</th>
                    <th>{t['mood']}</th>
                    <th>{t['energy']}</th>
                    <th>{t['sleep']}</th>
                    <th>{t['anxiety']}</th>
                    <th>{t['physical']}</th>
                    <th>{t['completion']}</th>
                </tr>
        """
        
        def get_css_class(val, is_anxiety=False):
            try:
                v = int(val)
            except:
                return ''
            if is_anxiety:
                return {1: 'green', 2: 'yellow', 3: 'orange', 4: 'red'}.get(v, '')
            else:
                return {4: 'green', 3: 'yellow', 2: 'orange', 1: 'red'}.get(v, '')
        
        for day in week_data:
            if day['has_checkin']:
                html_content += f"""
                <tr>
                    <td>{day['date'].strftime('%Y-%m-%d')}</td>
                    <td>{t['days'].get(day['day_of_week'], '')}</td>
                    <td class="{get_css_class(day['mood'])}">{day['mood'] or '-'}</td>
                    <td class="{get_css_class(day['energy'])}">{day['energy'] or '-'}</td>
                    <td class="{get_css_class(day['sleep'])}">{day['sleep'] or '-'}</td>
                    <td class="{get_css_class(day['anxiety'], True)}">{day['anxiety'] or '-'}</td>
                    <td class="{get_css_class(day['physical'])}">{day['physical'] or '-'}</td>
                    <td class="green">✓</td>
                </tr>
                """
            else:
                html_content += f"""
                <tr>
                    <td>{day['date'].strftime('%Y-%m-%d')}</td>
                    <td>{t['days'].get(day['day_of_week'], '')}</td>
                    <td>-</td><td>-</td><td>-</td><td>-</td><td>-</td>
                    <td class="red">{t['no_checkin']}</td>
                </tr>
                """
        
        html_content += f"""
            </table>
            
            <div class="summary">
                <h3>{t['weekly_summary']}</h3>
                <div class="summary-item">{t['checkin_completion']} {summary['checkin_count']}/{summary['total_days']} {t['days_completed']} ({summary['completion_pct']:.0f}%)</div>
                <div class="summary-item">{t['mood_level']} {format_avg(summary['avg_mood'])}</div>
                <div class="summary-item">{t['energy_level']} {format_avg(summary['avg_energy'])}</div>
                <div class="summary-item">{t['sleep_quality']} {format_avg(summary['avg_sleep'])}</div>
                <div class="summary-item">{t['anxiety_level']} {format_avg(summary['avg_anxiety'], is_anxiety=True)}</div>
                <div class="summary-item">{t['physical_level']} {format_avg(summary['avg_physical'])}</div>
            </div>
        </body>
        </html>
        """
        
        # Send email using SMTP (Resend.com)
        smtp_pass = os.environ.get('SMTP_PASSWORD')
        if not smtp_pass:
            return jsonify({'error': 'Email service not configured'}), 500
        
        subject_translations = {
            'en': 'Your Weekly Wellness Report',
            'he': 'דוח הבריאות השבועי שלך',
            'ar': 'تقرير صحتك الأسبوعي',
            'ru': 'Ваш еженедельный отчет о здоровье'
        }
        
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject_translations.get(lang, subject_translations['en'])
        msg['From'] = os.environ.get('FROM_EMAIL', 'TheraSocial <onboarding@resend.dev>')
        msg['To'] = recipient
        msg.attach(MIMEText(html_content, 'html'))
        
        smtp_server = os.environ.get('SMTP_SERVER', 'smtp.resend.com')
        smtp_port = int(os.environ.get('SMTP_PORT', '465'))
        smtp_user = os.environ.get('SMTP_USERNAME', 'resend')
        
        with smtplib.SMTP_SSL(smtp_server, smtp_port) as server:
            server.login(smtp_user, smtp_pass)
            server.sendmail(msg['From'], recipient, msg.as_string())
        
        return jsonify({'success': True, 'message': 'Report sent successfully'})
        
    except Exception as e:
        logger.error(f"Email report error: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({'error': 'Failed to send report'}), 500
    try:
        watcher_id = session.get('user_id')
        trigger = db.session.execute(
            select(ParameterTrigger).filter_by(
                watcher_id=watcher_id,
                watched_id=user_id
            )
        ).scalar_one_or_none()

        if trigger:
            return jsonify({
                'mood_alert': trigger.mood_alert,
                'energy_alert': trigger.energy_alert,
                'sleep_alert': trigger.sleep_alert,
                'physical_alert': trigger.physical_alert,
                'anxiety_alert': trigger.anxiety_alert
            })
        else:
            return jsonify({
                'mood_alert': False,
                'energy_alert': False,
                'sleep_alert': False,
                'physical_alert': False,
                'anxiety_alert': False
            })

    except Exception as e:
        logger.error(f"Get triggers error: {e}")
        return jsonify({'error': 'Failed to get triggers'}), 500


@app.route('/api/parameters/triggers/<int:user_id>', methods=['POST'])
@login_required
def set_parameter_triggers(user_id):
    """Set trigger alerts for a user being watched"""
    try:
        watcher_id = session.get('user_id')
        data = request.json

        follow = db.session.execute(
            select(Follow).filter_by(
                follower_id=watcher_id,
                followed_id=user_id
            )
        ).scalar_one_or_none()

        if not follow:
            return jsonify({'error': 'You must be following this user'}), 400

        trigger = db.session.execute(
            select(ParameterTrigger).filter_by(
                watcher_id=watcher_id,
                watched_id=user_id
            )
        ).scalar_one_or_none()

        if not trigger:
            trigger = ParameterTrigger(
                watcher_id=watcher_id,
                watched_id=user_id
            )
            db.session.add(trigger)

        trigger.mood_alert = data.get('mood_alert', False)
        trigger.energy_alert = data.get('energy_alert', False)
        trigger.sleep_alert = data.get('sleep_alert', False)
        trigger.physical_alert = data.get('physical_alert', False)
        trigger.anxiety_alert = data.get('anxiety_alert', False)

        db.session.commit()
        return jsonify({'success': True})

    except Exception as e:
        logger.error(f"Set triggers error: {e}")
        db.session.rollback()
        return jsonify({'error': 'Failed to set triggers'}), 500


def get_watcher_circle_level(watched_id, watcher_id):
    """
    Get the circle level that the watcher belongs to for the watched user.

    Args:
        watched_id: User ID of the person being watched
        watcher_id: User ID of the person watching

    Returns:
        str: 'class_a', 'class_b', 'public', or None if not in any circle
    """
    # Check if watcher is in any of watched user's circles
    circle = db.session.execute(
        select(Circle).filter_by(
            user_id=watched_id,
            circle_user_id=watcher_id
        )
    ).scalar_one_or_none()

    if circle:
        return circle.circle_type

    # Not in any circle
    return None


@app.route('/api/parameters/check-triggers', methods=['GET'])
@login_required
def check_parameter_triggers():
    """Check for parameter alerts - respects privacy settings
    
    PJ815 FIX (v1705):
    - Process each trigger row individually (don't deduplicate upfront)
    - Use a SET to track unique (user, param, start_date, end_date) patterns already seen
    - Only add NEW patterns to results - prevents duplicate alerts from duplicate trigger rows
    - Extensive logging showing each trigger's flags and each pattern found
    """
    try:
        watcher_id = session.get('user_id')
        triggers = db.session.execute(
            select(ParameterTrigger).filter_by(watcher_id=watcher_id)
        ).scalars().all()
        
        logger.info(f"[PJ815 DEBUG] ========================================")
        logger.info(f"[PJ815 DEBUG] check_parameter_triggers called for watcher_id={watcher_id}")
        logger.info(f"[PJ815 DEBUG] Found {len(triggers)} trigger rows")
        
        # Log each trigger's actual flags
        for i, t in enumerate(triggers):
            watched_user = db.session.get(User, t.watched_id)
            username = watched_user.username if watched_user else f"user_{t.watched_id}"
            flags = []
            if t.mood_alert: flags.append('mood')
            if t.energy_alert: flags.append('energy')
            if t.sleep_alert: flags.append('sleep')
            if t.physical_alert: flags.append('activity')
            if t.anxiety_alert: flags.append('anxiety')
            logger.info(f"[PJ815 DEBUG] Trigger {i+1}: watched={username}, consecutive_days={t.consecutive_days}, flags={flags}, mood_alert_raw={t.mood_alert}")

        alerts = []
        thirty_days_ago = datetime.utcnow() - timedelta(days=30)
        
        # PJ815: Track unique patterns to prevent duplicates from multiple trigger rows
        # Key = (username, param_name, start_date_iso, end_date_iso)
        patterns_seen = set()

        # Helper function to convert values to numbers (for OLD schema)
        def to_number(val):
            if val is None:
                return None
            if isinstance(val, (int, float)):
                return float(val)
            if isinstance(val, str):
                if val.lower() in ['private', 'hidden', 'none']:
                    return None
                try:
                    return float(val)
                except ValueError:
                    return None
            return None

        def can_see_parameter(param_privacy, watcher_circle):
            """Check if watcher can see this parameter based on privacy and circle level"""
            if param_privacy == 'private':
                return False
            elif param_privacy == 'class_a':
                return watcher_circle == 'class_a'
            elif param_privacy == 'class_b':
                return watcher_circle in ['class_b', 'class_a']
            elif param_privacy == 'public':
                return True
            return False

        for trigger in triggers:
            # Skip if no consecutive_days setting
            if not trigger.consecutive_days or trigger.consecutive_days < 1:
                continue

            watcher_circle = get_watcher_circle_level(trigger.watched_id, watcher_id)
            if not watcher_circle:
                continue

            parameters = db.session.execute(
                select(SavedParameters).filter(
                    SavedParameters.user_id == trigger.watched_id,
                    SavedParameters.date >= thirty_days_ago
                ).order_by(SavedParameters.date.asc())
            ).scalars().all()

            if len(parameters) < trigger.consecutive_days:
                continue

            watched_user = db.session.get(User, trigger.watched_id)
            if not watched_user:
                continue

            consecutive_days = trigger.consecutive_days
            
            # Determine which schema this trigger uses
            has_new_schema = any([
                trigger.mood_alert,
                trigger.energy_alert,
                trigger.sleep_alert,
                trigger.physical_alert,
                trigger.anxiety_alert
            ])

            has_old_schema = trigger.parameter_name is not None
            
            logger.info(f"[PJ815 DEBUG] Processing trigger for {watched_user.username}: has_new_schema={has_new_schema}, has_old_schema={has_old_schema}")

            # ===== NEW SCHEMA CODE =====
            if has_new_schema:
                def check_consecutive_pattern(param_attr, privacy_attr, condition_func, alert_level_func):
                    found_patterns = []
                    
                    # Collect all valid entries that meet the condition
                    valid_entries = []
                    for param in parameters:
                        param_value = getattr(param, param_attr, None)
                        param_privacy = getattr(param, privacy_attr, 'private')

                        if not can_see_parameter(param_privacy, watcher_circle):
                            continue
                        
                        if param_value is not None and condition_func(param_value):
                            valid_entries.append({
                                'date': param.date,
                                'value': param_value
                            })
                    
                    logger.info(f"[PJ815 PATTERN] {watched_user.username}/{param_attr}: {len(valid_entries)} entries meet condition")
                    
                    # Find ALL consecutive streaks of required length
                    if len(valid_entries) >= consecutive_days:
                        valid_entries.sort(key=lambda x: x['date'])
                        current_streak = []
                        
                        for entry in valid_entries:
                            if not current_streak:
                                current_streak = [entry]
                            else:
                                last_date = current_streak[-1]['date']
                                curr_date = entry['date']
                                days_diff = (curr_date - last_date).days
                                
                                if days_diff == 1:
                                    current_streak.append(entry)
                                elif days_diff == 0:
                                    continue  # Same day, skip
                                else:
                                    # Gap found - save streak if long enough
                                    if len(current_streak) >= consecutive_days:
                                        start_date = current_streak[0]['date']
                                        end_date = current_streak[-1]['date']
                                        
                                        # PJ815: Check if this pattern is already seen
                                        pattern_key = (watched_user.username, param_attr, 
                                                      start_date.isoformat(), end_date.isoformat())
                                        
                                        if pattern_key not in patterns_seen:
                                            patterns_seen.add(pattern_key)
                                            pattern_values = [e['value'] for e in current_streak]
                                            # Convert dates to ISO strings for JSON serialization
                                            date_strings = []
                                            for e in current_streak:
                                                if hasattr(e['date'], 'isoformat'):
                                                    date_strings.append(e['date'].isoformat())
                                                else:
                                                    date_strings.append(str(e['date']))
                                            pattern = {
                                                'level': alert_level_func(pattern_values),
                                                'user': watched_user.username,
                                                'parameter': param_attr,
                                                'dates': date_strings,
                                                'values': pattern_values,
                                                'consecutive_days': len(current_streak)
                                            }
                                            found_patterns.append(pattern)
                                            logger.info(f"[PJ815 PATTERN] NEW pattern: {watched_user.username}/{param_attr} {start_date} to {end_date}")
                                        else:
                                            logger.info(f"[PJ815 PATTERN] DUPLICATE (from another trigger row): {watched_user.username}/{param_attr} {start_date} to {end_date}")
                                    
                                    current_streak = [entry]
                        
                        # Don't forget the last streak
                        if len(current_streak) >= consecutive_days:
                            start_date = current_streak[0]['date']
                            end_date = current_streak[-1]['date']
                            pattern_key = (watched_user.username, param_attr,
                                          start_date.isoformat(), end_date.isoformat())
                            
                            if pattern_key not in patterns_seen:
                                patterns_seen.add(pattern_key)
                                pattern_values = [e['value'] for e in current_streak]
                                date_strings = []
                                for e in current_streak:
                                    if hasattr(e['date'], 'isoformat'):
                                        date_strings.append(e['date'].isoformat())
                                    else:
                                        date_strings.append(str(e['date']))
                                pattern = {
                                    'level': alert_level_func(pattern_values),
                                    'user': watched_user.username,
                                    'parameter': param_attr,
                                    'dates': date_strings,
                                    'values': pattern_values,
                                    'consecutive_days': len(current_streak)
                                }
                                found_patterns.append(pattern)
                                logger.info(f"[PJ815 PATTERN] NEW pattern (final): {watched_user.username}/{param_attr} {start_date} to {end_date}")
                            else:
                                logger.info(f"[PJ815 PATTERN] DUPLICATE (final): {watched_user.username}/{param_attr} {start_date} to {end_date}")
                    
                    return found_patterns

                # Check each parameter type if its alert flag is enabled
                if trigger.mood_alert:
                    logger.info(f"[PJ815 DEBUG] Checking mood for {watched_user.username}")
                    results = check_consecutive_pattern('mood', 'mood_privacy', 
                                                        lambda val: val <= 2,
                                                        lambda vals: 'critical' if sum(vals)/len(vals) == 1 else ('high' if sum(vals)/len(vals) <= 1.5 else 'warning'))
                    alerts.extend(results)

                if trigger.energy_alert:
                    logger.info(f"[PJ815 DEBUG] Checking energy for {watched_user.username}")
                    results = check_consecutive_pattern('energy', 'energy_privacy',
                                                        lambda val: val <= 2,
                                                        lambda vals: 'critical' if sum(vals)/len(vals) == 1 else ('high' if sum(vals)/len(vals) <= 1.5 else 'warning'))
                    alerts.extend(results)

                if trigger.sleep_alert:
                    logger.info(f"[PJ815 DEBUG] Checking sleep_quality for {watched_user.username}")
                    results = check_consecutive_pattern('sleep_quality', 'sleep_quality_privacy',
                                                        lambda val: val <= 2,
                                                        lambda vals: 'critical' if sum(vals)/len(vals) == 1 else ('high' if sum(vals)/len(vals) <= 1.5 else 'warning'))
                    alerts.extend(results)

                if trigger.physical_alert:
                    logger.info(f"[PJ815 DEBUG] Checking physical_activity for {watched_user.username}")
                    results = check_consecutive_pattern('physical_activity', 'physical_activity_privacy',
                                                        lambda val: val <= 2,
                                                        lambda vals: 'critical' if sum(vals)/len(vals) == 1 else ('high' if sum(vals)/len(vals) <= 1.5 else 'warning'))
                    alerts.extend(results)

                if trigger.anxiety_alert:
                    logger.info(f"[PJ815 DEBUG] Checking anxiety for {watched_user.username}")
                    results = check_consecutive_pattern('anxiety', 'anxiety_privacy',
                                                        lambda val: val >= 3,
                                                        lambda vals: 'critical' if sum(vals)/len(vals) == 4 else ('high' if sum(vals)/len(vals) >= 3.5 else 'warning'))
                    alerts.extend(results)

            # ===== OLD SCHEMA CODE =====
            elif has_old_schema:
                param_name = trigger.parameter_name
                condition = trigger.trigger_condition
                threshold = trigger.trigger_value
                
                logger.info(f"[PJ815 DEBUG] OLD SCHEMA: param_name={param_name}, condition={condition}, threshold={threshold}")

                # Map parameter name to model attribute
                param_mapping = {
                    'mood': ('mood', 'mood_privacy'),
                    'anxiety': ('anxiety', 'anxiety_privacy'),
                    'sleep_quality': ('sleep_quality', 'sleep_quality_privacy'),
                    'physical_activity': ('physical_activity', 'physical_activity_privacy'),
                    'energy': ('energy', 'energy_privacy')
                }

                if param_name not in param_mapping:
                    logger.info(f"[PJ815 DEBUG] Skipping unknown param_name: {param_name}")
                    continue

                param_attr, privacy_attr = param_mapping[param_name]

                # Create condition function
                if condition == 'less_than':
                    def condition_func(val, t=threshold):
                        num = to_number(val)
                        return num is not None and num < t

                    condition_text = f"less than {threshold}"
                elif condition == 'greater_than':
                    def condition_func(val, t=threshold):
                        num = to_number(val)
                        return num is not None and num > t

                    condition_text = f"greater than {threshold}"
                elif condition == 'equals':
                    def condition_func(val, t=threshold):
                        num = to_number(val)
                        return num is not None and num == t

                    condition_text = f"equal to {threshold}"
                else:
                    logger.info(f"[PJ815 DEBUG] Skipping unknown condition: {condition}")
                    continue

                # PJ815: Find ALL consecutive patterns with proper dates array
                streak_dates = []  # Track dates in current streak
                streak_values = []  # Track values in current streak
                last_date = None

                for param in parameters:
                    param_value = getattr(param, param_attr, None)
                    param_privacy = getattr(param, privacy_attr, 'private')

                    # Check if watcher can see this parameter
                    if not can_see_parameter(param_privacy, watcher_circle):
                        # Reset streak if we hit a private parameter
                        if len(streak_dates) >= consecutive_days:
                            # Save the streak before resetting
                            start_date = streak_dates[0]
                            end_date = streak_dates[-1]
                            pattern_key = (watched_user.username, param_name, 
                                          start_date.isoformat(), end_date.isoformat())
                            
                            if pattern_key not in patterns_seen:
                                patterns_seen.add(pattern_key)
                                date_strings = [d.isoformat() for d in streak_dates]
                                alert_data = {
                                    'user': watched_user.username,
                                    'parameter': param_name,
                                    'consecutive_days': len(streak_dates),
                                    'dates': date_strings,
                                    'values': streak_values[:],
                                    'end_date': end_date,
                                    'condition_text': condition_text,
                                    'is_old_schema': True
                                }
                                alerts.append(alert_data)
                                logger.info(f"[PJ815 PATTERN] OLD SCHEMA NEW: {watched_user.username}/{param_name} {start_date} to {end_date}")
                        streak_dates = []
                        streak_values = []
                        last_date = None
                        continue

                    if condition_func(param_value):
                        # Condition met
                        if last_date is None:
                            # Start new streak
                            streak_dates = [param.date]
                            streak_values = [param_value]
                            last_date = param.date
                        elif (param.date - last_date).days == 1:
                            # Consecutive - extend streak
                            streak_dates.append(param.date)
                            streak_values.append(param_value)
                            last_date = param.date
                        elif (param.date - last_date).days == 0:
                            # Same day - skip
                            continue
                        else:
                            # Gap found - save streak if long enough, then start new one
                            if len(streak_dates) >= consecutive_days:
                                start_date = streak_dates[0]
                                end_date = streak_dates[-1]
                                pattern_key = (watched_user.username, param_name,
                                              start_date.isoformat(), end_date.isoformat())
                                
                                if pattern_key not in patterns_seen:
                                    patterns_seen.add(pattern_key)
                                    date_strings = [d.isoformat() for d in streak_dates]
                                    alert_data = {
                                        'user': watched_user.username,
                                        'parameter': param_name,
                                        'consecutive_days': len(streak_dates),
                                        'dates': date_strings,
                                        'values': streak_values[:],
                                        'end_date': end_date,
                                        'condition_text': condition_text,
                                        'is_old_schema': True
                                    }
                                    alerts.append(alert_data)
                                    logger.info(f"[PJ815 PATTERN] OLD SCHEMA NEW: {watched_user.username}/{param_name} {start_date} to {end_date}")
                            
                            # Start new streak
                            streak_dates = [param.date]
                            streak_values = [param_value]
                            last_date = param.date
                    else:
                        # Condition not met - save streak if long enough
                        if len(streak_dates) >= consecutive_days:
                            start_date = streak_dates[0]
                            end_date = streak_dates[-1]
                            pattern_key = (watched_user.username, param_name,
                                          start_date.isoformat(), end_date.isoformat())
                            
                            if pattern_key not in patterns_seen:
                                patterns_seen.add(pattern_key)
                                date_strings = [d.isoformat() for d in streak_dates]
                                alert_data = {
                                    'user': watched_user.username,
                                    'parameter': param_name,
                                    'consecutive_days': len(streak_dates),
                                    'dates': date_strings,
                                    'values': streak_values[:],
                                    'end_date': end_date,
                                    'condition_text': condition_text,
                                    'is_old_schema': True
                                }
                                alerts.append(alert_data)
                                logger.info(f"[PJ815 PATTERN] OLD SCHEMA NEW: {watched_user.username}/{param_name} {start_date} to {end_date}")
                        
                        streak_dates = []
                        streak_values = []
                        last_date = None
                
                # Don't forget the last streak after loop ends
                if len(streak_dates) >= consecutive_days:
                    start_date = streak_dates[0]
                    end_date = streak_dates[-1]
                    pattern_key = (watched_user.username, param_name,
                                  start_date.isoformat(), end_date.isoformat())
                    
                    if pattern_key not in patterns_seen:
                        patterns_seen.add(pattern_key)
                        date_strings = [d.isoformat() for d in streak_dates]
                        alert_data = {
                            'user': watched_user.username,
                            'parameter': param_name,
                            'consecutive_days': len(streak_dates),
                            'dates': date_strings,
                            'values': streak_values[:],
                            'end_date': end_date,
                            'condition_text': condition_text,
                            'is_old_schema': True
                        }
                        alerts.append(alert_data)
                        logger.info(f"[PJ815 PATTERN] OLD SCHEMA FINAL: {watched_user.username}/{param_name} {start_date} to {end_date}")

        # ===== PJ811 FIX: CREATE DATABASE ALERTS WITH PROPER DUPLICATE DETECTION =====
        # PJ815: Extensive debugging and proper pattern deduplication
        # This endpoint now creates ONE-TIME database alerts when trigger patterns are found.
        # Duplicate detection uses 24-hour window per (watcher, watched_user, parameter, date_range) combo.
        # Alerts persist in the database and don't vanish on page refresh.
        # Emails are sent if user has email_on_alert enabled.
        
        logger.info(f"[PJ815 DEBUG] ========================================")
        logger.info(f"[PJ815 DEBUG] Alert processing starting")
        logger.info(f"[PJ815 DEBUG] Total unique patterns found: {len(alerts)}")
        
        # Log each pattern's date range
        for i, a in enumerate(alerts[:20]):  # First 20 for logging
            dates = a.get('dates', [])
            if dates:
                logger.info(f"[PJ815 DEBUG] Pattern {i+1}: {a.get('user')}/{a.get('parameter')} dates={dates[0]} to {dates[-1]}")
            else:
                logger.info(f"[PJ815 DEBUG] Pattern {i+1}: {a.get('user')}/{a.get('parameter')} NO DATES (end_date={a.get('end_date')})")
        
        logger.info(f"[TRIGGER CHECK] ========================================")
        logger.info(f"[TRIGGER CHECK] Found {len(alerts)} trigger patterns for watcher {watcher_id}")
        
        alerts_created = 0
        alerts_skipped_duplicate = 0
        alerts_emailed = 0
        
        for alert_data in alerts:
            try:
                watched_username = alert_data.get('user', 'Unknown')
                parameter = alert_data.get('parameter', 'unknown')
                consecutive_days = alert_data.get('consecutive_days', 0)
                
                # Normalize parameter name for consistent duplicate detection
                normalized_param = parameter.replace(' ', '_').lower()
                
                # Build unique key for this alert
                alert_key = f"{watched_username.lower()}_{normalized_param}"
                
                # PJ813 FIX: Build date range string for unique key - each date range should be a separate alert
                date_range_str = ""
                date_pattern = ""
                if alert_data.get('dates') and len(alert_data['dates']) >= 2:
                    try:
                        from datetime import datetime as dt
                        start_date = dt.fromisoformat(alert_data['dates'][0])
                        end_date = dt.fromisoformat(alert_data['dates'][-1])
                        date_range_str = f"_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}"
                        start_str = start_date.strftime('%b %d')
                        end_str = end_date.strftime('%b %d')
                        date_pattern = f"({start_str} - {end_str})"
                    except Exception as e:
                        logger.warning(f"[PJ815 DEBUG] Could not parse dates: {e}")
                
                # PJ813 FIX: Include date range in alert key so different patterns create separate alerts
                alert_key_with_dates = f"{alert_key}{date_range_str}"
                
                logger.info(f"[TRIGGER CHECK] Processing pattern: user={watched_username}, param={parameter}, days={consecutive_days}, dates={date_pattern}, key={alert_key_with_dates}")
                
                # PJ813 FIX: Check for existing alert with EXACT same date range in content
                # Changed from broad ilike match to exact date range match
                existing_alert = None
                if date_pattern:
                    # If we have dates, check for alert with exact date range in content
                    try:
                        existing_alert = Alert.query.filter(
                            Alert.user_id == watcher_id,
                            Alert.alert_type == 'trigger',
                            Alert.created_at >= datetime.now() - timedelta(hours=24),
                            Alert.content.ilike(f"%{watched_username}'s {parameter}%{date_pattern}%")
                        ).first()
                        
                        # Also check with underscore version
                        if not existing_alert:
                            param_with_underscore = parameter.replace(' ', '_')
                            existing_alert = Alert.query.filter(
                                Alert.user_id == watcher_id,
                                Alert.alert_type == 'trigger',
                                Alert.created_at >= datetime.now() - timedelta(hours=24),
                                Alert.content.ilike(f"%{watched_username}'s {param_with_underscore}%{date_pattern}%")
                            ).first()
                        
                        if existing_alert:
                            logger.info(f"[PJ815 DEBUG] Found existing alert with date pattern '{date_pattern}': ID={existing_alert.id}")
                        else:
                            logger.info(f"[PJ815 DEBUG] No existing alert with date pattern '{date_pattern}' - will create new")
                    except Exception as date_err:
                        logger.warning(f"[TRIGGER CHECK] Could not check date-specific duplicate: {date_err}")
                
                # Fallback: if no date range available, use broader check
                if not date_range_str:
                    existing_alert = Alert.query.filter(
                        Alert.user_id == watcher_id,
                        Alert.alert_type == 'trigger',
                        Alert.created_at >= datetime.now() - timedelta(hours=24),
                        Alert.content.ilike(f"%{watched_username}'s {parameter}%")
                    ).first()
                    
                    if not existing_alert:
                        param_with_underscore = parameter.replace(' ', '_')
                        existing_alert = Alert.query.filter(
                            Alert.user_id == watcher_id,
                            Alert.alert_type == 'trigger',
                            Alert.created_at >= datetime.now() - timedelta(hours=24),
                            Alert.content.ilike(f"%{watched_username}'s {param_with_underscore}%")
                        ).first()
                
                if existing_alert:
                    logger.info(f"[TRIGGER DUPLICATE] Skipping {watched_username}/{parameter} - alert already exists for this date range (ID: {existing_alert.id}, created: {existing_alert.created_at})")
                    alerts_skipped_duplicate += 1
                    continue
                
                # Get the watched user's ID for source_user_id
                watched_user = User.query.filter_by(username=watched_username).first()
                source_user_id = watched_user.id if watched_user else None
                
                # Create alert content with nicely formatted dates
                if alert_data.get('dates') and len(alert_data['dates']) >= 2:
                    try:
                        # Parse ISO dates and format nicely (e.g., "Dec 5 - Dec 7")
                        from datetime import datetime as dt
                        start_date = dt.fromisoformat(alert_data['dates'][0])
                        end_date = dt.fromisoformat(alert_data['dates'][-1])
                        start_str = start_date.strftime('%b %d')  # e.g., "Dec 05"
                        end_str = end_date.strftime('%b %d')      # e.g., "Dec 07"
                        content = f"{watched_username}'s {parameter} has been at concerning levels for {consecutive_days} consecutive days ({start_str} - {end_str})"
                    except Exception as date_err:
                        logger.warning(f"[TRIGGER CHECK] Could not parse dates: {date_err}")
                        content = f"{watched_username}'s {parameter} has been at concerning levels for {consecutive_days} consecutive days (from {alert_data['dates'][0]} to {alert_data['dates'][-1]})"
                elif alert_data.get('end_date'):
                    # Old schema format - has end_date object
                    try:
                        end_date = alert_data['end_date']
                        if hasattr(end_date, 'strftime'):
                            end_str = end_date.strftime('%b %d')
                            start_date = end_date - timedelta(days=consecutive_days - 1)
                            start_str = start_date.strftime('%b %d')
                            content = f"{watched_username}'s {parameter} has been at concerning levels for {consecutive_days} consecutive days ({start_str} - {end_str})"
                        else:
                            content = f"{watched_username}'s {parameter} has been at concerning levels for {consecutive_days} consecutive days"
                    except Exception as date_err:
                        logger.warning(f"[TRIGGER CHECK] Could not format end_date: {date_err}")
                        content = f"{watched_username}'s {parameter} has been at concerning levels for {consecutive_days} consecutive days"
                else:
                    content = f"{watched_username}'s {parameter} has been at concerning levels for {consecutive_days} consecutive days"
                
                # Create the alert with email notification
                logger.info(f"[TRIGGER CREATE] Creating alert for watcher {watcher_id}: {content[:100]}...")
                
                alert = create_alert_with_email(
                    user_id=watcher_id,
                    title=f"Wellness Alert for {watched_username}",
                    content=content,
                    alert_type='trigger',
                    source_user_id=source_user_id,
                    alert_category='trigger'
                )
                
                if alert:
                    alerts_created += 1
                    logger.info(f"[TRIGGER CREATE] ✅ Created alert ID {alert.id} for {watched_username}/{parameter}")
                    
                    # Check if email was sent
                    try:
                        settings = NotificationSettings.query.filter_by(user_id=watcher_id).first()
                        if settings and settings.email_on_alert:
                            alerts_emailed += 1
                            logger.info(f"[TRIGGER EMAIL] Email was sent for alert ID {alert.id}")
                        else:
                            logger.info(f"[TRIGGER EMAIL] No email sent - email_on_alert is {'disabled' if settings else 'no settings found'}")
                    except Exception as email_check_err:
                        logger.error(f"[TRIGGER EMAIL] Error checking email status: {email_check_err}")
                        
            except Exception as pattern_err:
                logger.error(f"[TRIGGER CHECK] Error processing pattern: {pattern_err}")
                logger.error(f"[TRIGGER CHECK] Pattern data: {alert_data}")
                continue
        
        # Commit all created alerts
        try:
            db.session.commit()
            logger.info(f"[TRIGGER CHECK] Committed {alerts_created} new alerts to database")
        except Exception as commit_err:
            logger.error(f"[TRIGGER CHECK] Error committing alerts: {commit_err}")
            db.session.rollback()
        
        logger.info(f"[TRIGGER CHECK] Summary: patterns={len(alerts)}, created={alerts_created}, duplicates_skipped={alerts_skipped_duplicate}, emailed={alerts_emailed}")
        logger.info(f"[TRIGGER CHECK] ========================================")

        return jsonify({
            'success': True,
            'alerts': alerts,
            'count': len(alerts),
            'alerts_created': alerts_created,
            'alerts_skipped_duplicate': alerts_skipped_duplicate,
            'alerts_emailed': alerts_emailed
        })

    except Exception as e:
        logger.error(f"Check triggers error: {str(e)}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e),
            'alerts': [],
            'count': 0
        })


# ==========================================
# PART 4: Automatic Cleanup Endpoint
# ==========================================
# Add this endpoint to app.py to handle retroactive cleanup automatically
# Insert anywhere after the get_watcher_circle_level() function

@app.route('/api/admin/cleanup-trigger-privacy', methods=['POST'])
@login_required
def cleanup_trigger_privacy():
    """
    Retroactively clean up trigger alerts that violate privacy settings.
    This should be run once after deploying the privacy fix.

    Only the logged-in user's alerts are cleaned up (not admin-only).
    """
    try:
        user_id = session.get('user_id')

        # Helper function (copy from main fix)
        def can_see_parameter(param_privacy, watcher_circle):
            """Check if watcher can see this parameter based on privacy and circle level"""
            if param_privacy == 'private':
                return False
            elif param_privacy == 'class_a':
                return watcher_circle == 'class_a'
            elif param_privacy == 'class_b':
                return watcher_circle in ['class_b', 'class_a']
            elif param_privacy == 'public':
                return True
            return False

        # Get all trigger alerts for this user
        trigger_alerts = Alert.query.filter_by(
            user_id=user_id,
            alert_type='trigger'
        ).all()

        total_checked = len(trigger_alerts)
        removed_count = 0
        kept_count = 0

        for alert in trigger_alerts:
            watcher_id = alert.user_id

            # Parse alert content to find watched user and parameter
            # Format: "username's parameter_name has been..."
            content = alert.content or ""

            if "'s " not in content:
                kept_count += 1
                continue

            username = content.split("'s ")[0]

            # Find watched user
            watched_user = User.query.filter_by(username=username).first()
            if not watched_user:
                kept_count += 1
                continue

            watched_id = watched_user.id

            # Get watcher's circle level using the helper function
            watcher_circle = get_watcher_circle_level(watched_id, watcher_id)

            if not watcher_circle:
                # Watcher not in any circle - should not have this alert
                db.session.delete(alert)
                removed_count += 1
                logger.info(f"Removed alert {alert.id}: watcher not in any circle")
                continue

            # Extract parameter name from content
            param_keywords = {
                'mood': 'mood_privacy',
                'anxiety': 'anxiety_privacy',
                'sleep_quality': 'sleep_quality_privacy',
                'sleep quality': 'sleep_quality_privacy',
                'physical_activity': 'physical_activity_privacy',
                'physical activity': 'physical_activity_privacy',
                'energy': 'energy_privacy'
            }

            privacy_attr = None
            for keyword, privacy_field in param_keywords.items():
                if keyword in content.lower():
                    privacy_attr = privacy_field
                    break

            if not privacy_attr:
                # Can't determine parameter - keep alert to be safe
                kept_count += 1
                continue

            # Get the most recent parameter entry for this user to check privacy
            recent_param = SavedParameters.query.filter_by(
                user_id=watched_id
            ).order_by(SavedParameters.updated_at.desc()).first()

            if not recent_param:
                kept_count += 1
                continue

            param_privacy = getattr(recent_param, privacy_attr, 'private')

            # Check if watcher should see this parameter
            if not can_see_parameter(param_privacy, watcher_circle):
                db.session.delete(alert)
                removed_count += 1
                logger.info(f"Removed alert {alert.id}: privacy violation ({param_privacy} vs {watcher_circle})")
            else:
                kept_count += 1

        db.session.commit()

        return jsonify({
            'success': True,
            'message': 'Privacy cleanup completed',
            'total_checked': total_checked,
            'removed': removed_count,
            'kept': kept_count
        })

    except Exception as e:
        logger.error(f"Cleanup error: {str(e)}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ==========================================
# ALTERNATIVE: Admin-Only Global Cleanup
# ==========================================
# If you want an admin endpoint that cleans up ALL users' alerts at once:

@app.route('/api/admin/cleanup-all-trigger-privacy', methods=['POST'])
@login_required
def cleanup_all_trigger_privacy():
    """
    Clean up ALL trigger alerts across all users that violate privacy.
    Requires admin privileges.
    """
    try:
        # Check if user is admin (you'll need to add admin flag to User model)
        user_id = session.get('user_id')
        current_user = db.session.get(User, user_id)

        # For now, skip admin check - remove this in production!
        # if not current_user.is_admin:
        #     return jsonify({'error': 'Admin access required'}), 403

        def can_see_parameter(param_privacy, watcher_circle):
            if param_privacy == 'private':
                return False
            elif param_privacy == 'class_a':
                return watcher_circle == 'class_a'
            elif param_privacy == 'class_b':
                return watcher_circle in ['class_b', 'class_a']
            elif param_privacy == 'public':
                return True
            return False

        # Get ALL trigger alerts
        trigger_alerts = Alert.query.filter_by(alert_type='trigger').all()

        total_checked = len(trigger_alerts)
        removed_count = 0
        kept_count = 0
        removed_by_user = {}  # Track removals per user

        for alert in trigger_alerts:
            watcher_id = alert.user_id
            content = alert.content or ""

            if "'s " not in content:
                kept_count += 1
                continue

            username = content.split("'s ")[0]
            watched_user = User.query.filter_by(username=username).first()

            if not watched_user:
                kept_count += 1
                continue

            watched_id = watched_user.id
            watcher_circle = get_watcher_circle_level(watched_id, watcher_id)

            if not watcher_circle:
                db.session.delete(alert)
                removed_count += 1
                removed_by_user[watcher_id] = removed_by_user.get(watcher_id, 0) + 1
                continue

            param_keywords = {
                'mood': 'mood_privacy',
                'anxiety': 'anxiety_privacy',
                'sleep_quality': 'sleep_quality_privacy',
                'sleep quality': 'sleep_quality_privacy',
                'physical_activity': 'physical_activity_privacy',
                'physical activity': 'physical_activity_privacy',
                'energy': 'energy_privacy'
            }

            privacy_attr = None
            for keyword, privacy_field in param_keywords.items():
                if keyword in content.lower():
                    privacy_attr = privacy_field
                    break

            if not privacy_attr:
                kept_count += 1
                continue

            recent_param = SavedParameters.query.filter_by(
                user_id=watched_id
            ).order_by(SavedParameters.updated_at.desc()).first()

            if not recent_param:
                kept_count += 1
                continue

            param_privacy = getattr(recent_param, privacy_attr, 'private')

            if not can_see_parameter(param_privacy, watcher_circle):
                db.session.delete(alert)
                removed_count += 1
                removed_by_user[watcher_id] = removed_by_user.get(watcher_id, 0) + 1
            else:
                kept_count += 1

        db.session.commit()

        # Build user breakdown
        user_breakdown = []
        for uid, count in removed_by_user.items():
            user = db.session.get(User, uid)
            if user:
                user_breakdown.append({
                    'username': user.username,
                    'removed': count
                })

        return jsonify({
            'success': True,
            'message': 'Global privacy cleanup completed',
            'total_checked': total_checked,
            'removed': removed_count,
            'kept': kept_count,
            'affected_users': len(removed_by_user),
            'breakdown': user_breakdown
        })

    except Exception as e:
        logger.error(f"Global cleanup error: {str(e)}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


def calculate_streak(params):
    """Calculate consecutive days streak"""
    if not params:
        return 0

    dates = sorted([p.date for p in params], reverse=True)
    streak = 1
    today = datetime.now().date()

    # Check if most recent date is today or yesterday
    if dates[0] < today - timedelta(days=1):
        return 0

    for i in range(1, len(dates)):
        if (dates[i - 1] - dates[i]).days == 1:
            streak += 1
        else:
            break

    return streak


# =====================
# ACTIVITY ROUTES (Calendar Feature)
# =====================

@app.route('/api/activity/<date_str>')
@login_required
def get_activity(date_str):
    """Get activity data for specific date"""
    try:
        user_id = session['user_id']

        # Parse date
        try:
            activity_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'error': 'Invalid date format'}), 400

        # Get activity record
        activity_stmt = select(Activity).filter_by(
            user_id=user_id,
            activity_date=activity_date
        )
        activity = db.session.execute(activity_stmt).scalar_one_or_none()

        if not activity:
            return jsonify({
                'date': date_str,
                'post_count': 0,
                'comment_count': 0,
                'message_count': 0,
                'mood_entries': []
            })

        return jsonify({
            'date': date_str,
            'post_count': activity.post_count or 0,
            'comment_count': activity.comment_count or 0,
            'message_count': activity.message_count or 0,
            'mood_entries': activity.mood_entries or []
        })

    except Exception as e:
        logger.error(f"Get activity error: {str(e)}")
        return jsonify({'error': 'Failed to get activity'}), 500


@app.route('/api/activity/<date_str>', methods=['POST'])
@login_required
def update_activity(date_str):
    """Update activity data for specific date"""
    try:
        user_id = session['user_id']
        data = request.json

        # Parse date
        try:
            activity_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'error': 'Invalid date format'}), 400

        # Get or create activity
        activity_stmt = select(Activity).filter_by(
            user_id=user_id,
            activity_date=activity_date
        )
        activity = db.session.execute(activity_stmt).scalar_one_or_none()

        if not activity:
            activity = Activity(
                user_id=user_id,
                activity_date=activity_date
            )
            db.session.add(activity)

        # Update fields
        if 'mood_entry' in data:
            mood_entries = activity.mood_entries or []
            mood_entries.append({
                'mood': data['mood_entry'].get('mood'),
                'note': data['mood_entry'].get('note'),
                'timestamp': datetime.utcnow().isoformat()
            })
            activity.mood_entries = mood_entries

        if 'post_count' in data:
            activity.post_count = data['post_count']
        if 'comment_count' in data:
            activity.comment_count = data['comment_count']
        if 'message_count' in data:
            activity.message_count = data['message_count']

        activity.updated_at = datetime.utcnow()
        db.session.commit()

        return jsonify({'success': True, 'message': 'Activity updated'})

    except Exception as e:
        logger.error(f"Update activity error: {str(e)}")
        db.session.rollback()
        return jsonify({'error': 'Failed to update activity'}), 500


@app.route('/api/activity/dates')
@login_required
def get_activity_dates():
    """Get dates with activity data"""
    try:
        user_id = session['user_id']
        # SQLAlchemy 2.0 style
        activities_stmt = select(Activity).filter_by(user_id=user_id)
        activities = db.session.execute(activities_stmt).scalars().all()
        dates = [a.activity_date.strftime('%Y-%m-%d') for a in activities]
        return jsonify({'dates': dates})
    except Exception as e:
        logger.error(f"Get activity dates error: {str(e)}")
        return jsonify({'dates': []})


# =====================
# ADMIN ROUTES
# =====================

@app.route('/api/admin/users')
@admin_required
def admin_get_users():
    """Get all users (admin only)"""
    try:
        # SQLAlchemy 2.0 style
        users_stmt = select(User).order_by(desc(User.created_at))
        users = db.session.execute(users_stmt).scalars().all()

        return jsonify([{
            'id': u.id,
            'username': u.username,
            'email': u.email,
            'role': u.role,
            'is_active': u.is_active,
            'created_at': u.created_at.isoformat()
        } for u in users])

    except Exception as e:
        logger.error(f"Admin get users error: {str(e)}")
        return jsonify({'error': 'Failed to get users'}), 500


@app.route('/api/admin/stats')
@admin_required
def admin_stats():
    """Get platform statistics"""
    try:
        # SQLAlchemy 2.0 style
        total_users = db.session.execute(select(func.count(User.id))).scalar() or 0
        total_posts = db.session.execute(select(func.count(Post.id))).scalar() or 0
        total_messages = db.session.execute(select(func.count(Message.id))).scalar() or 0

        # Active users today
        today_start = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        active_stmt = select(func.count(User.id)).filter(
            User.last_login >= today_start
        )
        active_users_today = db.session.execute(active_stmt).scalar() or 0

        return jsonify({
            'total_users': total_users,
            'total_posts': total_posts,
            'total_messages': total_messages,
            'active_users_today': active_users_today
        })

    except Exception as e:
        logger.error(f"Admin stats error: {str(e)}")
        return jsonify({'error': 'Failed to get stats'}), 500


@app.route('/api/admin/job-queue-stats')
@admin_required
def admin_job_queue_stats():
    """
    Get job queue statistics for monitoring.
    
    Returns:
        - pending: Jobs waiting to be processed
        - processing: Jobs currently being processed
        - completed_24h: Jobs completed in last 24 hours
        - failed_24h: Jobs that failed in last 24 hours
        - recent_jobs: Last 10 jobs with details
    """
    try:
        stats = get_job_queue_stats()
        
        # Get recent jobs for debugging
        recent_jobs = BackgroundJob.query.order_by(
            BackgroundJob.created_at.desc()
        ).limit(10).all()
        
        recent_list = [{
            'id': job.id,
            'job_type': job.job_type,
            'status': job.status,
            'attempts': job.attempts,
            'created_at': job.created_at.isoformat() if job.created_at else None,
            'completed_at': job.completed_at.isoformat() if job.completed_at else None,
            'error_message': job.error_message[:100] if job.error_message else None
        } for job in recent_jobs]
        
        return jsonify({
            **stats,
            'recent_jobs': recent_list
        })

    except Exception as e:
        logger.error(f"Job queue stats error: {str(e)}")
        return jsonify({'error': 'Failed to get job queue stats'}), 500


# =====================
# SAMPLE DATA ROUTE
# =====================

@app.route('/api/setup/sample-users', methods=['POST'])
def create_sample_users():
    """Create sample users for testing"""
    if is_production:
        return jsonify({'error': 'Not available in production'}), 403

    try:
        sample_users = [
            {'username': 'alice_wonder', 'email': 'alice@example.com', 'name': 'Alice Wonder'},
            {'username': 'bob_builder', 'email': 'bob@example.com', 'name': 'Bob Builder'},
            {'username': 'charlie_day', 'email': 'charlie@example.com', 'name': 'Charlie Day'},
            {'username': 'diana_prince', 'email': 'diana@example.com', 'name': 'Diana Prince'},
            {'username': 'edward_snow', 'email': 'edward@example.com', 'name': 'Edward Snow'},
            {'username': 'fiona_green', 'email': 'fiona@example.com', 'name': 'Fiona Green'}
        ]

        created = []
        for user_data in sample_users:
            # Check if exists - SQLAlchemy 2.0 style
            existing_stmt = select(User).filter_by(email=user_data['email'])
            if db.session.execute(existing_stmt).scalar_one_or_none():
                continue

            user = User(
                username=user_data['username'],
                email=user_data['email']
            )
            user.set_password('password123')
            db.session.add(user)
            db.session.flush()

            # Create profile
            profile = Profile(
                user_id=user.id,
                bio=f"Hi, I'm {user_data['name']}!",
                interests='Reading, Travel, Technology',
                occupation='Professional',
                goals='Connect with interesting people',
                favorite_hobbies='Hiking, Photography, Cooking'
            )
            db.session.add(profile)
            created.append(user_data['username'])

        db.session.commit()

        return jsonify({
            'success': True,
            'created': created,
            'message': f'Created {len(created)} sample users with default password: password123'
        })

    except Exception as e:
        logger.error(f"Create sample users error: {str(e)}")
        db.session.rollback()
        return jsonify({'error': 'Failed to create sample users'}), 500


# =====================
# CITY/TIMEZONE ROUTES
# =====================

@app.route('/api/user/city', methods=['GET', 'POST'])
@login_required
def user_city():
    """Get or update user's selected city"""
    user_id = session.get('user_id')
    user = db.session.get(User, user_id)

    if not user:
        return jsonify({'error': 'User not found'}), 404

    if request.method == 'GET':
        return jsonify({
            'selected_city': user.selected_city or 'Jerusalem, Israel'
        })

    elif request.method == 'POST':
        try:
            data = request.json
            city = data.get('selected_city')

            # List of valid cities - Israel, USA, and UK only (MVP focus)
            valid_cities = [
                # Israel (10 cities)
                'Jerusalem, Israel', 'Tel Aviv, Israel', 'Haifa, Israel',
                'Beer Sheva, Israel', 'Netanya, Israel', 'Rishon LeZion, Israel',
                'Petah Tikva, Israel', 'Ashdod, Israel', 'Eilat, Israel', 'Herzliya, Israel',
                # USA (25 cities)
                'New York City, USA', 'Los Angeles, USA', 'Chicago, USA',
                'Washington, USA', 'Houston, USA', 'San Francisco, USA',
                'Boston, USA', 'Philadelphia, USA', 'Phoenix, USA', 'San Diego, USA',
                'Dallas, USA', 'Seattle, USA', 'Miami, USA', 'Atlanta, USA',
                'Denver, USA', 'Austin, USA', 'San Jose, USA', 'Portland, USA',
                'Las Vegas, USA', 'Minneapolis, USA', 'Detroit, USA', 'Baltimore, USA',
                'Nashville, USA', 'Charlotte, USA', 'Orlando, USA',
                # UK (15 cities)
                'London, UK', 'Manchester, UK', 'Birmingham, UK', 'Edinburgh, UK',
                'Glasgow, UK', 'Bristol, UK', 'Liverpool, UK', 'Leeds, UK',
                'Sheffield, UK', 'Newcastle, UK', 'Nottingham, UK', 'Southampton, UK',
                'Cardiff, UK', 'Belfast, UK', 'Cambridge, UK'
            ]

            if city not in valid_cities:
                return jsonify({'error': 'Invalid city'}), 400

            user.selected_city = city
            user.updated_at = datetime.utcnow()
            db.session.commit()

            return jsonify({
                'success': True,
                'selected_city': user.selected_city
            })
        except Exception as e:
            logger.error(f"City update error: {str(e)}")
            db.session.rollback()
            return jsonify({'error': 'Failed to update city'}), 500


# =====================
# FOLLOWING/FOLLOWERS ROUTES
# =====================
@app.route('/api/follow/<int:user_id>', methods=['POST'])
@login_required
def follow_user(user_id):
    """Follow another user with optional note and trigger"""
    try:
        current_user_id = session.get('user_id')
        current_user = db.session.get(User, current_user_id)
        user_to_follow = db.session.get(User, user_id)

        # Use get_json(silent=True) to handle missing JSON body
        data = request.get_json(silent=True) or {}
        # Support both 'note' and 'follow_note' keys for compatibility
        follow_note = (data.get('note') or data.get('follow_note') or '').strip()[:300] if data else ''
        follow_trigger = data.get('trigger', False) if data else False

        if not user_to_follow:
            return jsonify({'error': 'User not found'}), 404

        if current_user_id == user_id:
            return jsonify({'error': 'Cannot follow yourself'}), 400

        # Check if already following
        existing = Follow.query.filter_by(
            follower_id=current_user_id,
            followed_id=user_id
        ).first()

        if existing:
            # Update existing follow with note
            if hasattr(existing, 'follow_note'):
                existing.follow_note = follow_note
            # Add trigger support if your Follow model has this field
            if hasattr(existing, 'follow_trigger'):
                existing.follow_trigger = follow_trigger
            db.session.commit()
            return jsonify({'success': True, 'message': 'Follow updated'}), 200

        # Create new follow
        current_user.follow(user_to_follow, note=follow_note)

        # Update the newly created follow record with note if model supports it
        follow = Follow.query.filter_by(
            follower_id=current_user_id,
            followed_id=user_id
        ).first()

        if follow:
            if hasattr(follow, 'follow_note'):
                follow.follow_note = follow_note
            if hasattr(follow, 'follow_trigger'):
                follow.follow_trigger = follow_trigger

        db.session.commit()

        # Create alert for followed user
        alert_content = 'You have a new follower!'
        if follow_note:
            alert_content += f' They said: "{follow_note}"'
        if follow_trigger:
            alert_content += ' (Following your parameters)'

        # PJ6001: Use create_notification_with_email for follow notifications (not wellness alerts)
        alert = create_notification_with_email(
            user_id=user_id,
            title=f'{current_user.username} started following you',
            content=alert_content,
            alert_type='info',
            source_user_id=current_user_id,
            alert_category='follow'
        )

        if follow_note:
            message = Message(
                sender_id=current_user_id,
                recipient_id=user_id,
                content=f"Follow note: {follow_note}"
            )
            db.session.add(message)

        db.session.commit()
        return jsonify({'success': True, 'message': 'User followed'})

    except Exception as e:
        logger.error(f"Follow error: {str(e)}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        db.session.rollback()
        return jsonify({'error': 'Failed to follow user'}), 500


@app.route('/api/unfollow/<int:user_id>', methods=['POST'])
@login_required
def unfollow_user(user_id):
    """Unfollow a user"""
    try:
        current_user_id = session.get('user_id')
        current_user = db.session.get(User, current_user_id)
        user_to_unfollow = db.session.get(User, user_id)

        if not user_to_unfollow:
            return jsonify({'error': 'User not found'}), 404

        current_user.unfollow(user_to_unfollow)
        db.session.commit()

        return jsonify({'success': True, 'message': 'User unfollowed'})

    except Exception as e:
        logger.error(f"Unfollow error: {str(e)}")
        db.session.rollback()
        return jsonify({'error': 'Failed to unfollow user'}), 500


@app.route('/api/following')
@login_required
@rate_limit_endpoint(max_requests=60, window=60)  # 60 requests per minute
def get_following():
    """Get list of users the current user is following with pagination"""
    try:
        user_id = session.get('user_id')
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 50, type=int)

        # Limit per_page to prevent abuse
        per_page = min(per_page, 100)

        # Get total count for pagination
        total = db.session.execute(
            select(func.count()).select_from(Follow).filter_by(follower_id=user_id)
        ).scalar()

        # Get follows directly from Follow model with pagination
        follows_query = select(Follow).filter_by(follower_id=user_id).limit(per_page).offset((page - 1) * per_page)
        follows = db.session.execute(follows_query).scalars().all()

        following = []
        for follow in follows:
            followed_user = db.session.get(User, follow.followed_id)
            if followed_user:
                following.append({
                    'id': followed_user.id,
                    'username': followed_user.username,
                    'email': followed_user.email,
                    'note': follow.follow_note,  # ADD THIS FIELD
                    'selected_city': followed_user.selected_city,
                    'created_at': follow.created_at.isoformat()
                })

        return jsonify({
            'following': following,
            'page': page,
            'per_page': per_page,
            'total': total,
            'pages': (total + per_page - 1) // per_page
        })

    except Exception as e:
        logger.error(f"Get following error: {str(e)}")
        return jsonify({'error': 'Failed to get following'}), 500


@app.route('/api/parameters/user/<int:user_id>', methods=['GET'])
@login_required
def get_user_parameters_for_triggers(user_id):
    """Get parameters for viewing in trigger modal"""
    try:
        viewer_id = session.get('user_id')

        # Check if following
        if viewer_id != user_id:
            follow = Follow.query.filter_by(
                follower_id=viewer_id,
                followed_id=user_id
            ).first()
            if not follow:
                return jsonify({'error': 'Not authorized'}), 403

        # Get last 30 days
        thirty_days_ago = datetime.utcnow() - timedelta(days=30)

        parameters = SavedParameters.query.filter(
            SavedParameters.user_id == user_id,
            SavedParameters.date >= thirty_days_ago
        ).order_by(SavedParameters.date.desc()).all()

        result = {'parameters': []}
        for param in parameters:
            # Add each parameter as separate entry
            for param_name in ['mood', 'energy', 'sleep_quality', 'physical_activity', 'anxiety']:
                value = getattr(param, param_name, None)
                if value:
                    result['parameters'].append({
                        'date': param.date.isoformat() if hasattr(param.date, 'isoformat') else str(param.date),
                        'parameter_name': param_name,
                        'value': value
                    })

        return jsonify(result), 200

    except Exception as e:
        logger.error(f"Error getting user parameters: {e}")
        return jsonify({'error': 'Failed to load parameters'}), 500


@app.route('/api/followers')
@login_required
@rate_limit_endpoint(max_requests=60, window=60)  # 60 requests per minute
def get_followers():
    """Get list of users following the current user with pagination"""
    try:
        user_id = session.get('user_id')
        user = db.session.get(User, user_id)
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 50, type=int)

        # Limit per_page to prevent abuse
        per_page = min(per_page, 100)

        # Get total count
        total = db.session.execute(
            select(func.count()).select_from(Follow).filter_by(followed_id=user_id)
        ).scalar()

        # Get followers with pagination
        follows_query = select(Follow).filter_by(followed_id=user_id).limit(per_page).offset((page - 1) * per_page)
        follows = db.session.execute(follows_query).scalars().all()

        followers = []
        for follow in follows:
            follower_user = db.session.get(User, follow.follower_id)
            if follower_user:
                followers.append({
                    'id': follower_user.id,
                    'username': follower_user.username,
                    'email': follower_user.email,
                    'selected_city': follower_user.selected_city,
                    'created_at': follow.created_at.isoformat(),
                    'is_following_back': user.is_following(follower_user)
                })

        return jsonify({
            'followers': followers,
            'page': page,
            'per_page': per_page,
            'total': total,
            'pages': (total + per_page - 1) // per_page
        })

    except Exception as e:
        logger.error(f"Get followers error: {str(e)}")
        return jsonify({'error': 'Failed to get followers'}), 500


# =====================
# BLOCKING ENDPOINTS
# =====================

@app.route('/api/users/<int:user_id>/block', methods=['POST'])
@login_required
def block_user(user_id):
    """Block a user - they won't be able to view your profile"""
    try:
        current_user_id = session.get('user_id')
        current_user = db.session.get(User, current_user_id)
        user_to_block = db.session.get(User, user_id)
        
        if not user_to_block:
            return jsonify({'error': 'User not found'}), 404
        
        if user_id == current_user_id:
            return jsonify({'error': 'Cannot block yourself'}), 400
        
        # Check if already blocked
        existing_block = BlockedUser.query.filter_by(
            blocker_id=current_user_id,
            blocked_id=user_id
        ).first()
        
        if existing_block:
            return jsonify({'error': 'User already blocked'}), 400
        
        # Create the block
        block = BlockedUser(blocker_id=current_user_id, blocked_id=user_id)
        db.session.add(block)
        db.session.commit()
        
        logger.info(f"User {current_user_id} blocked user {user_id}")
        return jsonify({'success': True, 'message': 'User blocked successfully'})
        
    except Exception as e:
        logger.error(f"Block user error: {str(e)}")
        db.session.rollback()
        return jsonify({'error': 'Failed to block user'}), 500


@app.route('/api/users/<int:user_id>/unblock', methods=['POST'])
@login_required
def unblock_user(user_id):
    """Unblock a previously blocked user"""
    try:
        current_user_id = session.get('user_id')
        
        block = BlockedUser.query.filter_by(
            blocker_id=current_user_id,
            blocked_id=user_id
        ).first()
        
        if not block:
            return jsonify({'error': 'User is not blocked'}), 400
        
        db.session.delete(block)
        db.session.commit()
        
        logger.info(f"User {current_user_id} unblocked user {user_id}")
        return jsonify({'success': True, 'message': 'User unblocked successfully'})
        
    except Exception as e:
        logger.error(f"Unblock user error: {str(e)}")
        db.session.rollback()
        return jsonify({'error': 'Failed to unblock user'}), 500


@app.route('/api/users/<int:user_id>/check-blocked', methods=['GET'])
@login_required
def check_if_blocked(user_id):
    """Check if the current user is blocked by the specified user OR has blocked them"""
    try:
        current_user_id = session.get('user_id')
        
        # Check if current user is blocked by user_id (they blocked me)
        is_blocked_by = BlockedUser.query.filter_by(
            blocker_id=user_id,
            blocked_id=current_user_id
        ).first() is not None
        
        # Check if current user has blocked user_id (I blocked them)
        has_blocked = BlockedUser.query.filter_by(
            blocker_id=current_user_id,
            blocked_id=user_id
        ).first() is not None
        
        return jsonify({
            'is_blocked': is_blocked_by,  # Kept for backward compatibility
            'blockedBy': is_blocked_by,   # PJ501: New field - target user has blocked current user
            'blocked': has_blocked,        # PJ501: Current user has blocked target user
            'blocker_id': user_id,
            'blocked_id': current_user_id
        })
        
    except Exception as e:
        logger.error(f"Check blocked error: {str(e)}")
        return jsonify({'error': 'Failed to check blocked status'}), 500


@app.route('/api/blocked-users', methods=['GET'])
@login_required
def get_blocked_users():
    """Get list of users the current user has blocked"""
    try:
        current_user_id = session.get('user_id')
        
        blocks = BlockedUser.query.filter_by(blocker_id=current_user_id).all()
        
        blocked_list = []
        for block in blocks:
            blocked_user = db.session.get(User, block.blocked_id)
            if blocked_user:
                blocked_list.append({
                    'id': blocked_user.id,
                    'username': blocked_user.username,
                    'email': blocked_user.email,
                    'blocked_at': block.created_at.isoformat() if block.created_at else None
                })
        
        return jsonify({'blocked_users': blocked_list})
        
    except Exception as e:
        logger.error(f"Get blocked users error: {str(e)}")
        return jsonify({'error': 'Failed to get blocked users'}), 500


# =====================
# CITY TIMEZONE MAPPING
# =====================

CITY_TIMEZONE_MAP = {
    # Israel
    'Jerusalem': 'Asia/Jerusalem',
    'Tel Aviv': 'Asia/Jerusalem',
    'Haifa': 'Asia/Jerusalem',
    'Rishon LeZion': 'Asia/Jerusalem',
    'Petah Tikva': 'Asia/Jerusalem',
    'Ashdod': 'Asia/Jerusalem',
    'Netanya': 'Asia/Jerusalem',
    'Beer Sheva': 'Asia/Jerusalem',
    'Holon': 'Asia/Jerusalem',
    'Ramat Gan': 'Asia/Jerusalem',
    
    # UK
    'London': 'Europe/London',
    'Manchester': 'Europe/London',
    'Birmingham': 'Europe/London',
    'Liverpool': 'Europe/London',
    'Leeds': 'Europe/London',
    'Sheffield': 'Europe/London',
    'Bristol': 'Europe/London',
    'Edinburgh': 'Europe/London',
    'Glasgow': 'Europe/London',
    'Cardiff': 'Europe/London',
    
    # US - Eastern
    'New York': 'America/New_York',
    'Boston': 'America/New_York',
    'Philadelphia': 'America/New_York',
    'Washington DC': 'America/New_York',
    'Miami': 'America/New_York',
    'Atlanta': 'America/New_York',
    
    # US - Central
    'Chicago': 'America/Chicago',
    'Houston': 'America/Chicago',
    'Dallas': 'America/Chicago',
    'San Antonio': 'America/Chicago',
    'Austin': 'America/Chicago',
    
    # US - Mountain
    'Denver': 'America/Denver',
    'Phoenix': 'America/Phoenix',
    'Salt Lake City': 'America/Denver',
    
    # US - Pacific
    'Los Angeles': 'America/Los_Angeles',
    'San Francisco': 'America/Los_Angeles',
    'San Diego': 'America/Los_Angeles',
    'Seattle': 'America/Los_Angeles',
    'Portland': 'America/Los_Angeles',
    'Las Vegas': 'America/Los_Angeles',
    
    # Default
    'default': 'UTC'
}


def get_timezone_for_city(city_name):
    """Get the timezone for a given city.
    Handles both 'City' and 'City, Country' formats.
    """
    if not city_name:
        logger.info(f"[CITY TIMEZONE] No city name provided, returning UTC")
        return 'UTC'
    
    # First try exact match
    if city_name in CITY_TIMEZONE_MAP:
        timezone = CITY_TIMEZONE_MAP[city_name]
        logger.info(f"[CITY TIMEZONE] Exact match for '{city_name}' -> {timezone}")
        return timezone
    
    # Try extracting just the city part (before comma)
    if ',' in city_name:
        city_only = city_name.split(',')[0].strip()
        if city_only in CITY_TIMEZONE_MAP:
            timezone = CITY_TIMEZONE_MAP[city_only]
            logger.info(f"[CITY TIMEZONE] City-only match for '{city_only}' (from '{city_name}') -> {timezone}")
            return timezone
    
    logger.warning(f"[CITY TIMEZONE] No match found for '{city_name}', returning UTC")
    return 'UTC'


@app.route('/api/diary-reminder/send-test', methods=['POST'])
@login_required
def send_test_diary_reminder():
    """Send a test diary reminder email to the current user"""
    try:
        user_id = session.get('user_id')
        user = db.session.get(User, user_id)
        
        if not user or not user.email:
            return jsonify({'error': 'User or email not found'}), 404
        
        logger.info(f"[DAILY REMINDER TEST] Sending test reminder to user {user_id} ({user.email})")
        
        user_language = user.preferred_language or 'en'
        success = send_daily_diary_reminder_email(user.email, user_language)
        
        if success:
            return jsonify({'success': True, 'message': 'Test reminder email sent successfully'})
        else:
            return jsonify({'success': False, 'message': 'Failed to send test reminder email'}), 500
            
    except Exception as e:
        logger.error(f"Test diary reminder error: {str(e)}")
        return jsonify({'error': 'Failed to send test reminder'}), 500


@app.route('/api/diary-reminder/process', methods=['POST', 'GET'])
def process_diary_reminders():
    """
    Process and send diary reminders for users whose reminder time has arrived.
    This should be called by a cron job or scheduler every minute.
    
    Security: Requires CRON_SECRET header or query parameter to prevent unauthorized calls.
    
    OPTIMIZED: Uses database-level filtering to only query users who need reminders NOW,
    rather than loading all users and checking in Python.
    """
    # Security check - verify CRON_SECRET
    cron_secret = os.environ.get('CRON_SECRET')
    if cron_secret:
        # Check header first, then query parameter
        provided_secret = request.headers.get('X-Cron-Secret') or request.args.get('secret')
        if provided_secret != cron_secret:
            logger.warning(f"[DAILY REMINDER CRON] Unauthorized access attempt - invalid secret")
            return jsonify({'error': 'Unauthorized'}), 401
    else:
        logger.warning(f"[DAILY REMINDER CRON] CRON_SECRET not configured - endpoint is unprotected!")
    
    try:
        logger.info(f"[DAILY REMINDER CRON] ========================================")
        logger.info(f"[DAILY REMINDER CRON] Starting diary reminder processing (OPTIMIZED)")
        
        now_utc = datetime.utcnow().replace(tzinfo=pytz.UTC)
        logger.info(f"[DAILY REMINDER CRON] Current UTC time: {now_utc}")
        
        # Get all unique timezones from users with reminders enabled
        unique_timezones = db.session.query(
            NotificationSettings.diary_reminder_timezone
        ).filter(
            NotificationSettings.email_daily_diary_reminder == True
        ).distinct().all()
        
        unique_timezones = [tz[0] or 'UTC' for tz in unique_timezones]
        logger.info(f"[DAILY REMINDER CRON] Unique timezones to check: {unique_timezones}")
        
        emails_sent = 0
        emails_failed = 0
        total_users_matched = 0
        
        # Process each timezone - only query users whose time matches NOW
        for tz_str in unique_timezones:
            try:
                user_tz = pytz.timezone(tz_str)
            except:
                logger.warning(f"[DAILY REMINDER CRON] Invalid timezone {tz_str}, using UTC")
                user_tz = pytz.UTC
                tz_str = 'UTC'
            
            # Get current time in this timezone
            now_in_tz = now_utc.astimezone(user_tz)
            current_time_str = now_in_tz.strftime('%H:%M')
            
            logger.info(f"[DAILY REMINDER CRON] Timezone {tz_str}: current time = {current_time_str}")
            
            # DATABASE-LEVEL FILTER: Only get users in this timezone whose reminder time is NOW
            matching_settings = NotificationSettings.query.filter(
                NotificationSettings.email_daily_diary_reminder == True,
                NotificationSettings.diary_reminder_timezone == tz_str,
                NotificationSettings.diary_reminder_time == current_time_str
            ).all()
            
            if not matching_settings:
                logger.debug(f"[DAILY REMINDER CRON] No users to remind in {tz_str} at {current_time_str}")
                continue
            
            logger.info(f"[DAILY REMINDER CRON] Found {len(matching_settings)} users to remind in {tz_str}")
            total_users_matched += len(matching_settings)
            
            # Process matching users in batches to avoid overwhelming email service
            BATCH_SIZE = 50
            for i in range(0, len(matching_settings), BATCH_SIZE):
                batch = matching_settings[i:i + BATCH_SIZE]
                
                for settings in batch:
                    try:
                        user = db.session.get(User, settings.user_id)
                        if not user or not user.email:
                            logger.warning(f"[DAILY REMINDER CRON] User {settings.user_id} not found or no email")
                            continue
                        
                        logger.info(f"[DAILY REMINDER CRON] Sending reminder to user {user.id} ({user.email})")
                        
                        user_language = user.preferred_language or 'en'
                        success = send_daily_diary_reminder_email(user.email, user_language)
                        
                        if success:
                            emails_sent += 1
                            logger.info(f"[DAILY REMINDER CRON] Successfully sent reminder to user {user.id}")
                        else:
                            emails_failed += 1
                            logger.error(f"[DAILY REMINDER CRON] Failed to send reminder to user {user.id}")
                            
                    except Exception as user_error:
                        logger.error(f"[DAILY REMINDER CRON] Error processing user {settings.user_id}: {str(user_error)}")
                        emails_failed += 1
                
                # Small delay between batches to avoid rate limiting
                if i + BATCH_SIZE < len(matching_settings):
                    import time
                    time.sleep(0.5)
        
        logger.info(f"[DAILY REMINDER CRON] Processing complete: matched={total_users_matched}, sent={emails_sent}, failed={emails_failed}")
        logger.info(f"[DAILY REMINDER CRON] ========================================")
        
        return jsonify({
            'success': True,
            'emails_sent': emails_sent,
            'emails_failed': emails_failed,
            'total_users_matched': total_users_matched,
            'timezones_checked': len(unique_timezones)
        })
        
    except Exception as e:
        logger.error(f"[DAILY REMINDER CRON] CRITICAL ERROR: {str(e)}")
        logger.error(f"[DAILY REMINDER CRON] Traceback: {traceback.format_exc()}")
        return jsonify({'error': 'Failed to process reminders'}), 500


# =====================
# BACKGROUND DIARY REMINDER SCHEDULER
# =====================

# Global flag to track if scheduler is running
_diary_reminder_scheduler_started = False
_diary_reminder_scheduler_lock = threading.Lock()

def run_diary_reminder_scheduler():
    """
    Background thread that checks for diary reminders every minute.
    This runs automatically on deployment - no external cron job needed.
    """
    global _diary_reminder_scheduler_started
    
    logger.info("[DIARY SCHEDULER] Background diary reminder scheduler started")
    
    while True:
        try:
            # Sleep for 60 seconds between checks
            time.sleep(60)
            
            # Process diary reminders within app context
            with app.app_context():
                try:
                    now_utc = datetime.utcnow().replace(tzinfo=pytz.UTC)
                    logger.info(f"[DIARY SCHEDULER] ========================================")
                    logger.info(f"[DIARY SCHEDULER] Checking reminders at UTC: {now_utc.strftime('%Y-%m-%d %H:%M:%S')}")
                    
                    # Get all users with diary reminders enabled and log their settings
                    all_diary_settings = NotificationSettings.query.filter(
                        NotificationSettings.email_daily_diary_reminder == True
                    ).all()
                    
                    logger.info(f"[DIARY SCHEDULER] Found {len(all_diary_settings)} users with diary reminders enabled")
                    
                    for s in all_diary_settings:
                        user = db.session.get(User, s.user_id)
                        user_email = user.email if user else 'NO_USER'
                        user_city = user.selected_city if user else 'NO_CITY'
                        logger.info(f"[DIARY SCHEDULER] User {s.user_id} ({user_email}): time={s.diary_reminder_time}, timezone={s.diary_reminder_timezone}, city={user_city}")
                    
                    # Get all unique timezones from users with reminders enabled
                    unique_timezones = db.session.query(
                        NotificationSettings.diary_reminder_timezone
                    ).filter(
                        NotificationSettings.email_daily_diary_reminder == True
                    ).distinct().all()
                    
                    unique_timezones = [tz[0] or 'UTC' for tz in unique_timezones]
                    logger.info(f"[DIARY SCHEDULER] Unique timezones to check: {unique_timezones}")
                    
                    if not unique_timezones:
                        logger.info("[DIARY SCHEDULER] No users with diary reminders enabled - skipping")
                        logger.info(f"[DIARY SCHEDULER] ========================================")
                        continue
                    
                    emails_sent = 0
                    emails_failed = 0
                    
                    # Process each timezone
                    for tz_str in unique_timezones:
                        try:
                            user_tz = pytz.timezone(tz_str)
                            logger.info(f"[DIARY SCHEDULER] Processing timezone: {tz_str}")
                        except Exception as tz_error:
                            logger.warning(f"[DIARY SCHEDULER] Invalid timezone '{tz_str}': {tz_error}, using UTC")
                            user_tz = pytz.UTC
                            tz_str = 'UTC'
                        
                        # Get current time in this timezone
                        now_in_tz = now_utc.astimezone(user_tz)
                        current_time_str = now_in_tz.strftime('%H:%M')
                        logger.info(f"[DIARY SCHEDULER] Timezone {tz_str}: current local time = {current_time_str}")
                        
                        # DATABASE-LEVEL FILTER: Only get users whose reminder time is NOW
                        matching_settings = NotificationSettings.query.filter(
                            NotificationSettings.email_daily_diary_reminder == True,
                            NotificationSettings.diary_reminder_timezone == tz_str,
                            NotificationSettings.diary_reminder_time == current_time_str
                        ).all()
                        
                        logger.info(f"[DIARY SCHEDULER] Query: timezone='{tz_str}', time='{current_time_str}' -> {len(matching_settings)} matches")
                        
                        if not matching_settings:
                            logger.info(f"[DIARY SCHEDULER] No users to remind in {tz_str} at {current_time_str}")
                            continue
                        
                        logger.info(f"[DIARY SCHEDULER] Found {len(matching_settings)} users to remind in {tz_str} at {current_time_str}")
                        
                        for settings in matching_settings:
                            try:
                                user = db.session.get(User, settings.user_id)
                                if not user or not user.email:
                                    logger.warning(f"[DIARY SCHEDULER] User {settings.user_id} has no email - skipping")
                                    continue
                                
                                logger.info(f"[DIARY SCHEDULER] Sending reminder to user {user.id} ({user.email})...")
                                user_language = user.preferred_language or 'en'
                                success = send_daily_diary_reminder_email(user.email, user_language)
                                
                                if success:
                                    emails_sent += 1
                                    logger.info(f"[DIARY SCHEDULER] SUCCESS: Sent reminder to {user.email}")
                                else:
                                    emails_failed += 1
                                    logger.error(f"[DIARY SCHEDULER] FAILED: Could not send reminder to {user.email}")
                                    
                            except Exception as user_error:
                                logger.error(f"[DIARY SCHEDULER] Error for user {settings.user_id}: {str(user_error)}")
                                logger.error(f"[DIARY SCHEDULER] Traceback: {traceback.format_exc()}")
                                emails_failed += 1
                    
                    logger.info(f"[DIARY SCHEDULER] Completed: sent={emails_sent}, failed={emails_failed}")
                    logger.info(f"[DIARY SCHEDULER] ========================================")
                        
                except Exception as inner_error:
                    logger.error(f"[DIARY SCHEDULER] Processing error: {str(inner_error)}")
                    logger.error(f"[DIARY SCHEDULER] Traceback: {traceback.format_exc()}")
                    db.session.rollback()
                    
        except Exception as e:
            logger.error(f"[DIARY SCHEDULER] Scheduler error: {str(e)}")
            logger.error(f"[DIARY SCHEDULER] Traceback: {traceback.format_exc()}")
            # Continue running even if there's an error
            time.sleep(60)


def start_diary_reminder_scheduler():
    """
    Start the background diary reminder scheduler.
    This is called automatically on app startup/deployment.
    Uses a lock to ensure only one scheduler runs per process.
    """
    global _diary_reminder_scheduler_started
    
    with _diary_reminder_scheduler_lock:
        if _diary_reminder_scheduler_started:
            logger.info("[DIARY SCHEDULER] Scheduler already running, skipping")
            return
        
        _diary_reminder_scheduler_started = True
        
        # Start the background thread
        scheduler_thread = threading.Thread(
            target=run_diary_reminder_scheduler,
            daemon=True,  # Thread will stop when main process stops
            name="DiaryReminderScheduler"
        )
        scheduler_thread.start()
        logger.info("[DIARY SCHEDULER] Background scheduler thread started successfully")


# =====================
# PJ816: BACKGROUND TRIGGER SCHEDULER
# =====================
# This scheduler checks triggers for ALL watchers every 5 minutes
# WITHOUT requiring them to be logged in. Emails are sent automatically
# when trigger conditions are met, just like message notifications.

_trigger_scheduler_started = False
_trigger_scheduler_lock = threading.Lock()

def run_background_trigger_check_for_watcher(watcher_id):
    """
    Check triggers for a specific watcher - same logic as check_parameter_triggers()
    but runs without requiring login. Used by background scheduler.
    
    PJ816: This is extracted from check_parameter_triggers() to run for any watcher.
    """
    try:
        triggers = db.session.execute(
            select(ParameterTrigger).filter_by(watcher_id=watcher_id)
        ).scalars().all()
        
        if not triggers:
            return {'alerts_created': 0, 'duplicates_skipped': 0}
        
        alerts = []
        thirty_days_ago = datetime.utcnow() - timedelta(days=30)
        patterns_seen = set()

        def to_number(val):
            if val is None:
                return None
            if isinstance(val, (int, float)):
                return float(val)
            if isinstance(val, str):
                if val.lower() in ['private', 'hidden', 'none']:
                    return None
                try:
                    return float(val)
                except ValueError:
                    return None
            return None

        def can_see_parameter(param_privacy, watcher_circle):
            if param_privacy == 'private':
                return False
            elif param_privacy == 'class_a':
                return watcher_circle == 'class_a'
            elif param_privacy == 'class_b':
                return watcher_circle in ['class_b', 'class_a']
            elif param_privacy == 'public':
                return True
            return False

        for trigger in triggers:
            if not trigger.consecutive_days or trigger.consecutive_days < 1:
                continue

            watcher_circle = get_watcher_circle_level(trigger.watched_id, watcher_id)
            if not watcher_circle:
                continue

            parameters = db.session.execute(
                select(SavedParameters).filter(
                    SavedParameters.user_id == trigger.watched_id,
                    SavedParameters.date >= thirty_days_ago
                ).order_by(SavedParameters.date.asc())
            ).scalars().all()

            if len(parameters) < trigger.consecutive_days:
                continue

            watched_user = db.session.get(User, trigger.watched_id)
            if not watched_user:
                continue

            consecutive_days = trigger.consecutive_days
            
            has_new_schema = any([
                trigger.mood_alert,
                trigger.energy_alert,
                trigger.sleep_alert,
                trigger.physical_alert,
                trigger.anxiety_alert
            ])

            has_old_schema = trigger.parameter_name is not None

            # NEW SCHEMA processing
            if has_new_schema:
                def check_consecutive_pattern(param_attr, privacy_attr, condition_func):
                    found_patterns = []
                    valid_entries = []
                    for param in parameters:
                        param_value = getattr(param, param_attr, None)
                        param_privacy = getattr(param, privacy_attr, 'private')
                        if not can_see_parameter(param_privacy, watcher_circle):
                            continue
                        if param_value is not None and condition_func(param_value):
                            valid_entries.append({'date': param.date, 'value': param_value})
                    
                    if len(valid_entries) >= consecutive_days:
                        valid_entries.sort(key=lambda x: x['date'])
                        current_streak = []
                        
                        for entry in valid_entries:
                            if not current_streak:
                                current_streak = [entry]
                            else:
                                last_date = current_streak[-1]['date']
                                curr_date = entry['date']
                                days_diff = (curr_date - last_date).days
                                
                                if days_diff == 1:
                                    current_streak.append(entry)
                                elif days_diff == 0:
                                    continue
                                else:
                                    if len(current_streak) >= consecutive_days:
                                        start_date = current_streak[0]['date']
                                        end_date = current_streak[-1]['date']
                                        pattern_key = (watched_user.username, param_attr, 
                                                      start_date.isoformat(), end_date.isoformat())
                                        if pattern_key not in patterns_seen:
                                            patterns_seen.add(pattern_key)
                                            found_patterns.append({
                                                'user': watched_user.username,
                                                'parameter': param_attr,
                                                'consecutive_days': len(current_streak),
                                                'dates': [e['date'].isoformat() for e in current_streak],
                                                'values': [e['value'] for e in current_streak]
                                            })
                                    current_streak = [entry]
                        
                        # Don't forget last streak
                        if len(current_streak) >= consecutive_days:
                            start_date = current_streak[0]['date']
                            end_date = current_streak[-1]['date']
                            pattern_key = (watched_user.username, param_attr, 
                                          start_date.isoformat(), end_date.isoformat())
                            if pattern_key not in patterns_seen:
                                patterns_seen.add(pattern_key)
                                found_patterns.append({
                                    'user': watched_user.username,
                                    'parameter': param_attr,
                                    'consecutive_days': len(current_streak),
                                    'dates': [e['date'].isoformat() for e in current_streak],
                                    'values': [e['value'] for e in current_streak]
                                })
                    
                    return found_patterns

                # PJ817: Fixed lambdas to use to_number() to handle string values from database
                if trigger.mood_alert:
                    alerts.extend(check_consecutive_pattern('mood', 'mood_privacy', lambda val: to_number(val) is not None and to_number(val) <= 2))
                if trigger.energy_alert:
                    alerts.extend(check_consecutive_pattern('energy', 'energy_privacy', lambda val: to_number(val) is not None and to_number(val) <= 2))
                if trigger.sleep_alert:
                    alerts.extend(check_consecutive_pattern('sleep_quality', 'sleep_quality_privacy', lambda val: to_number(val) is not None and to_number(val) <= 2))
                if trigger.physical_alert:
                    alerts.extend(check_consecutive_pattern('physical_activity', 'physical_activity_privacy', lambda val: to_number(val) is not None and to_number(val) <= 2))
                if trigger.anxiety_alert:
                    alerts.extend(check_consecutive_pattern('anxiety', 'anxiety_privacy', lambda val: to_number(val) is not None and to_number(val) >= 3))

            # OLD SCHEMA processing
            elif has_old_schema:
                param_name = trigger.parameter_name
                condition = trigger.trigger_condition
                threshold = trigger.trigger_value

                param_mapping = {
                    'mood': ('mood', 'mood_privacy'),
                    'anxiety': ('anxiety', 'anxiety_privacy'),
                    'sleep_quality': ('sleep_quality', 'sleep_quality_privacy'),
                    'physical_activity': ('physical_activity', 'physical_activity_privacy'),
                    'energy': ('energy', 'energy_privacy')
                }

                if param_name not in param_mapping:
                    continue

                param_attr, privacy_attr = param_mapping[param_name]

                if condition == 'less_than':
                    def condition_func(val, t=threshold):
                        num = to_number(val)
                        return num is not None and num < t
                    condition_text = f"less than {threshold}"
                elif condition == 'greater_than':
                    def condition_func(val, t=threshold):
                        num = to_number(val)
                        return num is not None and num > t
                    condition_text = f"greater than {threshold}"
                elif condition == 'equals':
                    def condition_func(val, t=threshold):
                        num = to_number(val)
                        return num is not None and num == t
                    condition_text = f"equal to {threshold}"
                else:
                    continue

                streak_dates = []
                streak_values = []
                last_date = None

                for param in parameters:
                    param_value = getattr(param, param_attr, None)
                    param_privacy = getattr(param, privacy_attr, 'private')

                    if not can_see_parameter(param_privacy, watcher_circle):
                        if len(streak_dates) >= consecutive_days:
                            start_date = streak_dates[0]
                            end_date = streak_dates[-1]
                            pattern_key = (watched_user.username, param_name, 
                                          start_date.isoformat(), end_date.isoformat())
                            if pattern_key not in patterns_seen:
                                patterns_seen.add(pattern_key)
                                alerts.append({
                                    'user': watched_user.username,
                                    'parameter': param_name,
                                    'consecutive_days': len(streak_dates),
                                    'dates': [d.isoformat() for d in streak_dates],
                                    'values': streak_values[:],
                                    'condition_text': condition_text
                                })
                        streak_dates = []
                        streak_values = []
                        last_date = None
                        continue

                    if condition_func(param_value):
                        if last_date is None:
                            streak_dates = [param.date]
                            streak_values = [param_value]
                            last_date = param.date
                        elif (param.date - last_date).days == 1:
                            streak_dates.append(param.date)
                            streak_values.append(param_value)
                            last_date = param.date
                        elif (param.date - last_date).days == 0:
                            continue
                        else:
                            if len(streak_dates) >= consecutive_days:
                                start_date = streak_dates[0]
                                end_date = streak_dates[-1]
                                pattern_key = (watched_user.username, param_name,
                                              start_date.isoformat(), end_date.isoformat())
                                if pattern_key not in patterns_seen:
                                    patterns_seen.add(pattern_key)
                                    alerts.append({
                                        'user': watched_user.username,
                                        'parameter': param_name,
                                        'consecutive_days': len(streak_dates),
                                        'dates': [d.isoformat() for d in streak_dates],
                                        'values': streak_values[:],
                                        'condition_text': condition_text
                                    })
                            streak_dates = [param.date]
                            streak_values = [param_value]
                            last_date = param.date
                    else:
                        if len(streak_dates) >= consecutive_days:
                            start_date = streak_dates[0]
                            end_date = streak_dates[-1]
                            pattern_key = (watched_user.username, param_name,
                                          start_date.isoformat(), end_date.isoformat())
                            if pattern_key not in patterns_seen:
                                patterns_seen.add(pattern_key)
                                alerts.append({
                                    'user': watched_user.username,
                                    'parameter': param_name,
                                    'consecutive_days': len(streak_dates),
                                    'dates': [d.isoformat() for d in streak_dates],
                                    'values': streak_values[:],
                                    'condition_text': condition_text
                                })
                        streak_dates = []
                        streak_values = []
                        last_date = None
                
                # Last streak
                if len(streak_dates) >= consecutive_days:
                    start_date = streak_dates[0]
                    end_date = streak_dates[-1]
                    pattern_key = (watched_user.username, param_name,
                                  start_date.isoformat(), end_date.isoformat())
                    if pattern_key not in patterns_seen:
                        patterns_seen.add(pattern_key)
                        alerts.append({
                            'user': watched_user.username,
                            'parameter': param_name,
                            'consecutive_days': len(streak_dates),
                            'dates': [d.isoformat() for d in streak_dates],
                            'values': streak_values[:],
                            'condition_text': condition_text
                        })

        # Now create database alerts for found patterns
        alerts_created = 0
        alerts_skipped_duplicate = 0
        
        # PJ6012: Collect triggered params per watched user for consolidated email
        # Key: watched_username, Value: list of {'param_name', 'days', 'date_range'}
        triggered_params_by_user = {}
        
        for alert_data in alerts:
            try:
                watched_username = alert_data.get('user', 'Unknown')
                parameter = alert_data.get('parameter', 'unknown')
                consecutive_days = alert_data.get('consecutive_days', 0)
                
                date_pattern = ""
                # PJ817: Generate date pattern for ALL streaks, including single-day
                # This ensures duplicate detection works properly
                if alert_data.get('dates') and len(alert_data['dates']) >= 1:
                    try:
                        from datetime import datetime as dt
                        start_date = dt.fromisoformat(alert_data['dates'][0])
                        end_date = dt.fromisoformat(alert_data['dates'][-1])
                        start_str = start_date.strftime('%b %d')
                        end_str = end_date.strftime('%b %d')
                        if len(alert_data['dates']) == 1:
                            # Single day - use format like "(Dec 10)"
                            date_pattern = f"({start_str})"
                        else:
                            # Multi-day - use format like "(Dec 07 - Dec 09)"
                            date_pattern = f"({start_str} - {end_str})"
                    except:
                        pass
                
                # Check for existing alert with exact date pattern
                existing_alert = None
                if date_pattern:
                    existing_alert = Alert.query.filter(
                        Alert.user_id == watcher_id,
                        Alert.alert_type == 'trigger',
                        Alert.content.ilike(f"%{watched_username}'s {parameter}%{date_pattern}%")
                    ).first()
                    
                    if not existing_alert:
                        param_with_underscore = parameter.replace(' ', '_')
                        existing_alert = Alert.query.filter(
                            Alert.user_id == watcher_id,
                            Alert.alert_type == 'trigger',
                            Alert.content.ilike(f"%{watched_username}'s {param_with_underscore}%{date_pattern}%")
                        ).first()
                
                if existing_alert:
                    alerts_skipped_duplicate += 1
                    continue
                
                # Create alert content
                watched_user = User.query.filter_by(username=watched_username).first()
                source_user_id = watched_user.id if watched_user else None
                
                # PJ817: Generate content with date pattern for ALL streak lengths
                if alert_data.get('dates') and len(alert_data['dates']) >= 1:
                    try:
                        from datetime import datetime as dt
                        start_date = dt.fromisoformat(alert_data['dates'][0])
                        end_date = dt.fromisoformat(alert_data['dates'][-1])
                        start_str = start_date.strftime('%b %d')
                        end_str = end_date.strftime('%b %d')
                        if len(alert_data['dates']) == 1:
                            content = f"{watched_username}'s {parameter} has been at concerning levels for {consecutive_days} consecutive days ({start_str})"
                        else:
                            content = f"{watched_username}'s {parameter} has been at concerning levels for {consecutive_days} consecutive days ({start_str} - {end_str})"
                    except:
                        content = f"{watched_username}'s {parameter} has been at concerning levels for {consecutive_days} consecutive days"
                else:
                    content = f"{watched_username}'s {parameter} has been at concerning levels for {consecutive_days} consecutive days"
                
                # Create alert with email notification
                alert = create_alert_with_email(
                    user_id=watcher_id,
                    title=f"Wellness Alert for {watched_username}",
                    content=content,
                    alert_type='trigger',
                    source_user_id=source_user_id,
                    alert_category='trigger'
                )
                
                if alert:
                    alerts_created += 1
                    
                    # PJ6012: Collect triggered params for consolidated email
                    if watched_username not in triggered_params_by_user:
                        triggered_params_by_user[watched_username] = []
                    
                    # Build date_range string
                    date_range_str = ""
                    if alert_data.get('dates') and len(alert_data['dates']) >= 1:
                        try:
                            from datetime import datetime as dt
                            start_date_obj = dt.fromisoformat(alert_data['dates'][0])
                            end_date_obj = dt.fromisoformat(alert_data['dates'][-1])
                            start_str = start_date_obj.strftime('%b %d')
                            end_str = end_date_obj.strftime('%b %d')
                            date_range_str = f"{start_str} - {end_str}"
                        except:
                            date_range_str = "recent"
                    
                    triggered_params_by_user[watched_username].append({
                        'param_name': parameter,
                        'days': consecutive_days,
                        'date_range': date_range_str
                    })
                    
            except Exception as pattern_err:
                logger.error(f"[TRIGGER SCHEDULER] Error processing pattern: {pattern_err}")
                continue
        
        # Commit alerts
        try:
            db.session.commit()
        except Exception as commit_err:
            logger.error(f"[TRIGGER SCHEDULER] Error committing alerts: {commit_err}")
            db.session.rollback()
        
        # PJ6012: Send consolidated emails for each watched user that had alerts created
        emails_sent = 0
        if alerts_created > 0:
            watcher = db.session.get(User, watcher_id)
            user_language = watcher.preferred_language if watcher else 'en'
            
            for watched_username, triggered_params in triggered_params_by_user.items():
                if triggered_params:
                    logger.info(f"[TRIGGER SCHEDULER] Sending consolidated email to watcher {watcher_id} for {watched_username} with {len(triggered_params)} params")
                    if send_consolidated_wellness_alert_email(watcher_id, watched_username, triggered_params, user_language):
                        emails_sent += 1
                        logger.info(f"[TRIGGER SCHEDULER] ✅ Consolidated email sent successfully")
                    else:
                        logger.info(f"[TRIGGER SCHEDULER] ⚠️ Consolidated email not sent (user may have email_on_alert disabled)")
        
        return {'alerts_created': alerts_created, 'duplicates_skipped': alerts_skipped_duplicate, 'emails_sent': emails_sent}
        
    except Exception as e:
        logger.error(f"[TRIGGER SCHEDULER] Error checking triggers for watcher {watcher_id}: {e}")
        db.session.rollback()
        return {'alerts_created': 0, 'duplicates_skipped': 0, 'error': str(e)}


def run_trigger_scheduler():
    """
    Background thread that checks triggers for ALL watchers every 5 minutes.
    This runs automatically on deployment - no external cron job needed.
    Emails are sent automatically when trigger conditions are met.
    
    PJ816: This is the key fix - trigger emails no longer require watcher login.
    """
    global _trigger_scheduler_started
    
    logger.info("[TRIGGER SCHEDULER] Background trigger scheduler started")
    
    # Initial delay to let the app fully start
    time.sleep(30)
    
    while True:
        try:
            # Process triggers every 5 minutes
            with app.app_context():
                try:
                    now_utc = datetime.utcnow()
                    logger.info(f"[TRIGGER SCHEDULER] ========================================")
                    logger.info(f"[TRIGGER SCHEDULER] Starting trigger check at UTC: {now_utc.strftime('%Y-%m-%d %H:%M:%S')}")
                    
                    # Get all unique watchers who have triggers configured
                    all_watcher_ids = db.session.query(
                        ParameterTrigger.watcher_id
                    ).distinct().all()
                    
                    all_watcher_ids = [w[0] for w in all_watcher_ids]
                    
                    logger.info(f"[TRIGGER SCHEDULER] Found {len(all_watcher_ids)} watchers with triggers configured")
                    
                    if not all_watcher_ids:
                        logger.info("[TRIGGER SCHEDULER] No watchers with triggers - skipping")
                        logger.info(f"[TRIGGER SCHEDULER] ========================================")
                    else:
                        total_created = 0
                        total_skipped = 0
                        
                        for watcher_id in all_watcher_ids:
                            try:
                                watcher = db.session.get(User, watcher_id)
                                watcher_name = watcher.username if watcher else f"user_{watcher_id}"
                                
                                result = run_background_trigger_check_for_watcher(watcher_id)
                                
                                created = result.get('alerts_created', 0)
                                skipped = result.get('duplicates_skipped', 0)
                                total_created += created
                                total_skipped += skipped
                                
                                if created > 0:
                                    logger.info(f"[TRIGGER SCHEDULER] Watcher {watcher_name}: created={created}, skipped={skipped}")
                                    
                            except Exception as watcher_err:
                                logger.error(f"[TRIGGER SCHEDULER] Error processing watcher {watcher_id}: {watcher_err}")
                                continue
                        
                        logger.info(f"[TRIGGER SCHEDULER] Completed: total_created={total_created}, total_skipped={total_skipped}")
                        logger.info(f"[TRIGGER SCHEDULER] ========================================")
                        
                except Exception as inner_error:
                    logger.error(f"[TRIGGER SCHEDULER] Processing error: {str(inner_error)}")
                    logger.error(f"[TRIGGER SCHEDULER] Traceback: {traceback.format_exc()}")
                    db.session.rollback()
            
            # Sleep for 5 minutes between checks
            time.sleep(300)
                    
        except Exception as e:
            logger.error(f"[TRIGGER SCHEDULER] Scheduler error: {str(e)}")
            logger.error(f"[TRIGGER SCHEDULER] Traceback: {traceback.format_exc()}")
            time.sleep(300)


def start_trigger_scheduler():
    """
    Start the background trigger scheduler.
    This is called automatically on app startup/deployment.
    Uses a lock to ensure only one scheduler runs per process.
    """
    global _trigger_scheduler_started
    
    with _trigger_scheduler_lock:
        if _trigger_scheduler_started:
            logger.info("[TRIGGER SCHEDULER] Scheduler already running, skipping")
            return
        
        _trigger_scheduler_started = True
        
        # Start the background thread
        scheduler_thread = threading.Thread(
            target=run_trigger_scheduler,
            daemon=True,
            name="TriggerScheduler"
        )
        scheduler_thread.start()
        logger.info("[TRIGGER SCHEDULER] Background scheduler thread started successfully")


# =====================
# JOB QUEUE SCHEDULER
# =====================
# This scheduler processes background jobs from the database queue
# every 10 seconds for near-immediate processing without blocking requests.

_job_queue_scheduler_started = False
_job_queue_scheduler_lock = threading.Lock()

def run_job_queue_scheduler():
    """
    Background thread that processes pending jobs from the database queue.
    Runs every 10 seconds for near-immediate job processing.
    
    This replaces per-request threading with a centralized job processor
    that provides: retries, persistence, and prevents thread exhaustion.
    """
    global _job_queue_scheduler_started
    
    logger.info("[JOB QUEUE SCHEDULER] Background job queue scheduler started")
    
    # Initial delay to let the app fully start
    time.sleep(15)
    
    while True:
        try:
            with app.app_context():
                try:
                    # Process up to 10 jobs per cycle
                    process_background_jobs(batch_size=10)
                except Exception as inner_error:
                    logger.error(f"[JOB QUEUE SCHEDULER] Processing error: {str(inner_error)}")
                    db.session.rollback()
            
            # Sleep for 10 seconds between job processing cycles
            time.sleep(10)
                    
        except Exception as e:
            logger.error(f"[JOB QUEUE SCHEDULER] Scheduler error: {str(e)}")
            time.sleep(10)


def start_job_queue_scheduler():
    """
    Start the background job queue scheduler.
    This is called automatically on app startup/deployment.
    Uses a lock to ensure only one scheduler runs per process.
    """
    global _job_queue_scheduler_started
    
    with _job_queue_scheduler_lock:
        if _job_queue_scheduler_started:
            logger.info("[JOB QUEUE SCHEDULER] Scheduler already running, skipping")
            return
        
        _job_queue_scheduler_started = True
        
        # Start the background thread
        scheduler_thread = threading.Thread(
            target=run_job_queue_scheduler,
            daemon=True,
            name="JobQueueScheduler"
        )
        scheduler_thread.start()
        logger.info("[JOB QUEUE SCHEDULER] Background job queue scheduler started successfully")


@app.route('/api/city-timezone', methods=['GET'])
@login_required
def get_city_timezone():
    """Get the timezone for the current user's selected city"""
    try:
        user_id = session.get('user_id')
        logger.info(f"[CITY TIMEZONE] Getting timezone for user {user_id}")
        
        user = db.session.get(User, user_id)
        
        if not user:
            logger.warning(f"[CITY TIMEZONE] User {user_id} not found, returning UTC")
            return jsonify({'timezone': 'UTC', 'city': None})
        
        logger.info(f"[CITY TIMEZONE] User {user_id} selected_city: '{user.selected_city}'")
        timezone = get_timezone_for_city(user.selected_city)
        logger.info(f"[CITY TIMEZONE] Resolved timezone: '{timezone}'")
        
        return jsonify({
            'city': user.selected_city,
            'timezone': timezone
        })
        
    except Exception as e:
        logger.error(f"Get city timezone error: {str(e)}")
        return jsonify({'timezone': 'UTC'})


@app.route('/api/recommendations')
@login_required
def get_recommendations():
    """Get follow recommendations prioritizing same city, then common connections"""
    try:
        user_id = session.get('user_id')
        user = db.session.get(User, user_id)

        if not user:
            return jsonify({'recommendations': []}), 200

        recommendations = []
        seen_ids = set()

        # PRIORITY 1: Users in same city (only if user has selected a city)
        if user.selected_city:
            same_city_users = User.query.filter(
                User.selected_city == user.selected_city,
                User.id != user_id,
                User.is_active == True
            ).limit(15).all()

            for city_user in same_city_users:
                if not user.is_following(city_user) and city_user.id not in seen_ids:
                    seen_ids.add(city_user.id)
                    recommendations.append({
                        'id': city_user.id,
                        'username': city_user.username,
                        'email': city_user.email,
                        'selected_city': city_user.selected_city,
                        'reason': 'Same city'
                    })

        # PRIORITY 2: Friends of friends (only if not enough same-city users)
        if len(recommendations) < 20:
            # Get users from circles
            circle_users = db.session.execute(
                select(Circle.circle_user_id).filter_by(user_id=user_id)
            ).scalars().all()

            # Get friends of friends
            for circle_user_id in circle_users:
                if len(recommendations) >= 20:
                    break

                their_circles = db.session.execute(
                    select(Circle.circle_user_id).filter_by(
                        user_id=circle_user_id
                    ).filter(Circle.circle_user_id != user_id)
                ).scalars().all()

                for potential_id in their_circles[:5]:
                    if len(recommendations) >= 20:
                        break

                    if potential_id in seen_ids:
                        continue

                    potential_user = db.session.get(User, potential_id)
                    if potential_user and not user.is_following(potential_user):
                        seen_ids.add(potential_user.id)

                        # Check if also same city for enhanced reason
                        reason = 'Friend of friend'
                        if potential_user.selected_city == user.selected_city:
                            reason = 'Same city & friend of friend'

                        recommendations.append({
                            'id': potential_user.id,
                            'username': potential_user.username,
                            'email': potential_user.email,
                            'selected_city': potential_user.selected_city,
                            'reason': reason
                        })

        return jsonify({'recommendations': recommendations[:20]})

    except Exception as e:
        logger.error(f"Get recommendations error: {str(e)}")
        return jsonify({'error': 'Failed to get recommendations'}), 500


# ADD THESE TWO ENDPOINTS TO app.py AFTER LINE 4235
# (Right after the existing get_recommendations() function)

@app.route('/api/users/recommendations', methods=['GET'])
@login_required
def get_user_recommendations():
    """Get recommended users to INVITE to follow YOU (not people for you to follow)"""
    try:
        # Set query timeout to prevent hanging (PostgreSQL only)
        try:
            db.session.execute(text("SET LOCAL statement_timeout = '5000'"))
        except Exception:
            pass  # Ignore if not PostgreSQL

        user_id = session.get('user_id')
        current_user = db.session.get(User, user_id)

        if not current_user:
            return jsonify({
                'error': 'User not found',
                'recommendations': [],
                'count': 0
            }), 200

        # Get IDs of users who are ALREADY FOLLOWING ME (my current followers)
        # These should be EXCLUDED because they're already following me
        my_followers = []
        try:
            my_followers = db.session.execute(
                select(Follow.follower_id).filter_by(followed_id=user_id)
            ).scalars().all()
        except Exception as e:
            logger.warning(f"Followers query failed: {e}")

        # Get pending follow requests I RECEIVED
        # These should be EXCLUDED because they already want to follow me
        received_requests = []
        try:
            received_requests = db.session.execute(
                select(FollowRequest.requester_id).filter_by(
                    target_id=user_id,
                    status='pending'
                )
            ).scalars().all()
        except Exception as e:
            logger.warning(f"Received requests query failed: {e}")

        # Get pending follow requests I SENT
        # These should be EXCLUDED because I already invited them
        sent_requests = []
        try:
            sent_requests = db.session.execute(
                select(FollowRequest.target_id).filter_by(
                    requester_id=user_id,
                    status='pending'
                )
            ).scalars().all()
        except Exception as e:
            logger.warning(f"Sent requests query failed: {e}")

        # Exclude: myself, people already following me, people who sent me requests, and people I sent requests to
        exclude_ids = set(my_followers + received_requests + sent_requests + [user_id])

        logger.info(f"User {user_id} ({current_user.username}) invite recommendations - "
                    f"excluding {len(exclude_ids)} users: "
                    f"followers={len(my_followers)}, received_requests={len(received_requests)}, sent_requests={len(sent_requests)}")

        # Get users with similar location (potential people to invite)
        location_matches = []
        try:
            if hasattr(current_user, 'selected_city') and current_user.selected_city:
                location_matches = db.session.execute(
                    select(User).filter(
                        User.selected_city == current_user.selected_city,
                        ~User.id.in_(exclude_ids)
                    ).limit(10)
                ).scalars().all()
                logger.info(f"Found {len(location_matches)} location matches in "
                            f"{current_user.selected_city} for user {current_user.username}")
        except Exception as e:
            logger.warning(f"Location query failed: {e}")
            location_matches = []

        # Get recently active users (potential people to invite)
        recent_users = []
        try:
            recent_users = db.session.execute(
                select(User).filter(
                    ~User.id.in_(exclude_ids)
                ).order_by(User.created_at.desc()).limit(10)
            ).scalars().all()
            logger.info(f"Found {len(recent_users)} recent users for user {current_user.username}")
        except Exception as e:
            logger.warning(f"Recent users query failed: {e}")
            recent_users = []

        # Combine and deduplicate
        all_recommendations = []
        seen_ids = set()

        for user_list in [location_matches, recent_users]:
            for user in user_list:
                if user.id not in seen_ids:
                    seen_ids.add(user.id)
                    all_recommendations.append({
                        'id': user.id,
                        'username': user.username,
                        'location': getattr(user, 'selected_city', None)
                    })

        # Limit to 20 recommendations
        all_recommendations = all_recommendations[:20]

        logger.info(f"Returning {len(all_recommendations)} recommendations for user "
                    f"{current_user.username}: {[r['username'] for r in all_recommendations]}")

        return jsonify({
            'recommendations': all_recommendations,
            'count': len(all_recommendations)
        })

    except Exception as e:
        logger.error(f"Recommendations error: {str(e)}")
        return jsonify({
            'error': 'Failed to load recommendations',
            'recommendations': [],
            'count': 0
        }), 200


@app.route('/invite/<username>')
def public_invite_page(username):
    """Public invite page for a user - accessible without login"""
    try:
        # Find user by username
        user = db.session.execute(
            select(User).filter_by(username=username)
        ).scalar_one_or_none()

        if not user:
            return """
            <!DOCTYPE html>
            <html><head><title>User Not Found</title></head>
            <body style="font-family: Arial; text-align: center; padding: 50px;">
            <h1>User Not Found</h1><p>The user you're looking for doesn't exist.</p>
            <a href="/">Go to TheraSocial</a>
            </body></html>
            """, 404

        # Get user's public stats
        follower_count = db.session.execute(
            select(func.count(Follow.id)).filter_by(followed_id=user.id)
        ).scalar() or 0

        following_count = db.session.execute(
            select(func.count(Follow.id)).filter_by(follower_id=user.id)
        ).scalar() or 0

        # Check if current user is logged in
        current_user_id = session.get('user_id')
        is_logged_in = current_user_id is not None

        # Check follow status if logged in
        already_following = False
        pending_request = False

        if is_logged_in:
            follow_exists = db.session.execute(
                select(Follow).filter_by(
                    follower_id=current_user_id,
                    followed_id=user.id
                )
            ).scalar_one_or_none()
            already_following = follow_exists is not None

            request_exists = db.session.execute(
                select(FollowRequest).filter_by(
                    requester_id=current_user_id,
                    target_id=user.id,
                    status='pending'
                )
            ).scalar_one_or_none()
            pending_request = request_exists is not None

        # PJ6003: Check for language query parameter first, then browser language
        lang_param = request.args.get('lang')
        if lang_param and lang_param in ['en', 'he', 'ar', 'ru']:
            default_language = lang_param
        else:
            # Detect browser language from Accept-Language header
            browser_lang = request.accept_languages.best_match(['en', 'he', 'ar', 'ru'], default='en')
            default_language = browser_lang

        # Translations for invite page
        translations = {
            'en': {
                'title': f"Join {username}'s Wellness Journey",
                'subtitle': 'Follow their progress on TheraSocial',
                'followers': 'Followers',
                'following': 'Following',
                'description': f'{username} is tracking their wellness journey and wants to share it with you.',
                'join_text': 'Join TheraSocial to follow their progress and support their mental health goals.',
                'follow_btn': f'Follow {username}',
                'dashboard_btn': 'Go to Dashboard',
                'already_following': 'Already Following',
                'request_pending': 'Request Pending'
            },
            'he': {
                'title': f'הצטרף/י למסע הבריאות של {username}',
                'subtitle': 'עקוב/י אחרי ההתקדמות שלו/ה ב-TheraSocial',
                'followers': 'עוקבים',
                'following': 'עוקב/ת אחרי',
                'description': f'{username} עוקב/ת אחרי מסע הבריאות שלו/ה ורוצה לשתף אותך.',
                'join_text': 'הצטרף/י ל-TheraSocial כדי לעקוב אחרי ההתקדמות שלו/ה ולתמוך ביעדי הבריאות הנפשית שלו/ה.',
                'follow_btn': f'עקוב/י אחרי {username}',
                'dashboard_btn': 'עבור ללוח הבקרה',
                'already_following': 'כבר עוקב/ת',
                'request_pending': 'בקשה ממתינה'
            },
            'ar': {
                'title': f'انضم إلى رحلة {username} الصحية',
                'subtitle': 'تابع تقدمهم على TheraSocial',
                'followers': 'المتابعون',
                'following': 'يتابع',
                'description': f'{username} يتتبع رحلته الصحية ويريد مشاركتها معك.',
                'join_text': 'انضم إلى TheraSocial لمتابعة تقدمهم ودعم أهدافهم الصحية النفسية.',
                'follow_btn': f'تابع {username}',
                'dashboard_btn': 'اذهب إلى لوحة التحكم',
                'already_following': 'متابع بالفعل',
                'request_pending': 'طلب قيد الانتظار'
            },
            'ru': {
                'title': f'Присоединяйтесь к пути здоровья {username}',
                'subtitle': 'Следите за их прогрессом на TheraSocial',
                'followers': 'Подписчики',
                'following': 'Подписки',
                'description': f'{username} отслеживает свой путь к здоровью и хочет поделиться им с вами.',
                'join_text': 'Присоединяйтесь к TheraSocial, чтобы следить за их прогрессом и поддерживать их цели психического здоровья.',
                'follow_btn': f'Подписаться на {username}',
                'dashboard_btn': 'Перейти к панели',
                'already_following': 'Уже подписаны',
                'request_pending': 'Запрос ожидает'
            }
        }

        t = translations.get(default_language, translations['en'])
        is_rtl = default_language in ['he', 'ar']
        text_dir = 'rtl' if is_rtl else 'ltr'

        # Determine button state
        if already_following:
            button_html = f'<button class="btn btn-secondary" disabled>{t["already_following"]}</button>'
        elif pending_request:
            button_html = f'<button class="btn btn-secondary" disabled>{t["request_pending"]}</button>'
        elif is_logged_in:
            button_html = f'<a href="/api/follow/{user.id}" class="btn btn-primary">{t["follow_btn"]}</a>'
        else:
            button_html = f'<a href="/" class="btn btn-primary">{t["follow_btn"]}</a>'

        dashboard_btn = f'<a href="/" class="btn btn-outline">{t["dashboard_btn"]}</a>' if is_logged_in else ''

        html = f"""
        <!DOCTYPE html>
        <html lang="{default_language}" dir="{text_dir}">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>{t['title']} - TheraSocial</title>
            <style>
                * {{ margin: 0; padding: 0; box-sizing: border-box; }}
                body {{
                    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                    min-height: 100vh;
                    display: flex;
                    flex-direction: column;
                    align-items: center;
                    padding: 20px;
                    direction: {text_dir};
                }}
                .language-selector {{
                    position: fixed;
                    top: 20px;
                    {'left' if is_rtl else 'right'}: 20px;
                    z-index: 1000;
                }}
                .language-selector select {{
                    padding: 8px 16px;
                    border-radius: 20px;
                    border: none;
                    background: white;
                    font-size: 14px;
                    cursor: pointer;
                    box-shadow: 0 2px 10px rgba(0,0,0,0.1);
                }}
                .card {{
                    background: white;
                    border-radius: 20px;
                    padding: 40px;
                    max-width: 500px;
                    width: 100%;
                    margin-top: 60px;
                    box-shadow: 0 10px 40px rgba(0,0,0,0.2);
                    text-align: center;
                }}
                .avatar {{
                    width: 100px;
                    height: 100px;
                    border-radius: 50%;
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                    color: white;
                    display: flex;
                    align-items: center;
                    justify-content: center;
                    font-size: 40px;
                    font-weight: bold;
                    margin: 0 auto 20px;
                }}
                h1 {{
                    color: #667eea;
                    font-size: 28px;
                    margin-bottom: 10px;
                }}
                .subtitle {{
                    color: #666;
                    margin-bottom: 25px;
                }}
                .stats {{
                    display: flex;
                    justify-content: center;
                    gap: 40px;
                    padding: 20px;
                    background: #f8f9fa;
                    border-radius: 12px;
                    margin-bottom: 25px;
                }}
                .stat {{
                    text-align: center;
                }}
                .stat-value {{
                    font-size: 28px;
                    font-weight: bold;
                    color: #667eea;
                }}
                .stat-label {{
                    color: #888;
                    font-size: 14px;
                }}
                .description {{
                    color: #666;
                    line-height: 1.6;
                    margin-bottom: 30px;
                }}
                .btn {{
                    display: inline-block;
                    padding: 14px 40px;
                    border-radius: 25px;
                    text-decoration: none;
                    font-weight: bold;
                    margin: 5px;
                    transition: all 0.3s;
                    border: none;
                    cursor: pointer;
                    font-size: 16px;
                }}
                .btn-primary {{
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                    color: white;
                    box-shadow: 0 4px 15px rgba(102, 126, 234, 0.4);
                }}
                .btn-primary:hover {{
                    transform: translateY(-2px);
                    box-shadow: 0 6px 20px rgba(102, 126, 234, 0.5);
                }}
                .btn-outline {{
                    background: transparent;
                    color: #667eea;
                    border: 2px solid #667eea;
                }}
                .btn-outline:hover {{
                    background: #667eea;
                    color: white;
                }}
                .btn-secondary {{
                    background: #e0e0e0;
                    color: #666;
                }}
                .buttons {{
                    display: flex;
                    flex-direction: column;
                    gap: 10px;
                    align-items: center;
                }}
            </style>
        </head>
        <body>
            <div class="language-selector">
                <select id="langSelect" onchange="changeLanguage(this.value)">
                    <option value="en" {'selected' if default_language == 'en' else ''}>English</option>
                    <option value="he" {'selected' if default_language == 'he' else ''}>עברית</option>
                    <option value="ar" {'selected' if default_language == 'ar' else ''}>العربية</option>
                    <option value="ru" {'selected' if default_language == 'ru' else ''}>Русский</option>
                </select>
            </div>
            <div class="card">
                <div class="avatar">{username[0].upper()}</div>
                <h1>{t['title']}</h1>
                <p class="subtitle">{t['subtitle']}</p>
                <div class="stats">
                    <div class="stat">
                        <div class="stat-value">{follower_count}</div>
                        <div class="stat-label">{t['followers']}</div>
                    </div>
                    <div class="stat">
                        <div class="stat-value">{following_count}</div>
                        <div class="stat-label">{t['following']}</div>
                    </div>
                </div>
                <p class="description">
                    {t['description']}<br><br>
                    {t['join_text']}
                </p>
                <div class="buttons">
                    {button_html}
                    {dashboard_btn}
                </div>
            </div>
            <script>
                function changeLanguage(lang) {{
                    // Reload page with language parameter
                    const url = new URL(window.location.href);
                    url.searchParams.set('lang', lang);
                    window.location.href = url.toString();
                }}
            </script>
        </body>
        </html>
        """

        return html

    except Exception as e:
        logger.error(f"Invite page error: {str(e)}")
        return """
        <!DOCTYPE html>
        <html><head><title>Error</title></head>
        <body style="font-family: Arial; text-align: center; padding: 50px;">
        <h1>Something went wrong</h1><p>Please try again later.</p>
        <a href="/">Go to TheraSocial</a>
        </body></html>
        """, 500


@app.route('/api/user/<int:user_id>/feed/<date_str>')
@login_required
def get_user_feed(user_id, date_str):
    """Get another user's feed for a specific date (read-only with circle permissions)"""
    try:
        current_user_id = session.get('user_id')
        current_user = db.session.get(User, current_user_id)

        # Check if following or is self
        target_user = db.session.get(User, user_id)
        if not target_user:
            return jsonify({'error': 'User not found'}), 404

        if user_id != current_user_id and not current_user.is_following(target_user):
            return jsonify({'error': 'You must follow this user to view their feed'}), 403

        post = Post.query.filter_by(user_id=user_id).filter(
            db.func.date(Post.created_at) == date_str
        ).first()

        if not post:
            return jsonify({'content': '', 'date': date_str})

        # Check circle permission if viewing another user's post
        if user_id != current_user_id:
            post_visibility = post.visibility if post.visibility else 'general'

            if post_visibility == 'family':
                # Must be in family circle
                in_circle = Circle.query.filter_by(
                    user_id=user_id,
                    circle_user_id=current_user_id,
                    circle_type='family'
                ).first() is not None

                if not in_circle:
                    return jsonify({'error': 'This post is only visible to family members'}), 403

            elif post_visibility == 'close_friends':
                # Must be in family OR close_friends circle
                in_circle = Circle.query.filter(
                    Circle.user_id == user_id,
                    Circle.circle_user_id == current_user_id,
                    Circle.circle_type.in_(['family', 'close_friends'])
                ).first() is not None

                if not in_circle:
                    return jsonify({'error': 'This post is only visible to close friends'}), 403

        return jsonify({
            'content': post.content,
            'date': date_str,
            'visibility': post.visibility if post.visibility else 'general',
            'updated_at': post.updated_at.isoformat() if post.updated_at else None
        })

    except Exception as e:
        logger.error(f"Get user feed error: {str(e)}")
        return jsonify({'error': 'Failed to get feed'}), 500


@app.route('/api/user/<int:user_id>/parameters/<date_str>')
@login_required
def get_user_parameters_by_date(user_id, date_str):
    """Get another user's parameters for a specific date (read-only)"""
    try:
        current_user_id = session.get('user_id')
        current_user = db.session.get(User, current_user_id)

        # Check if following or is self
        target_user = db.session.get(User, user_id)
        if not target_user:
            return jsonify({'error': 'User not found'}), 404

        if user_id != current_user_id and not current_user.is_following(target_user):
            return jsonify({'error': 'You must follow this user to view their parameters'}), 403

        params = SavedParameters.query.filter_by(
            user_id=user_id,
            date=date_str
        ).first()

        if params:
            return jsonify({
                'success': True,
                'data': {
                    'parameters': {
                        'mood': params.mood,
                        'energy': params.energy,
                        'sleep_quality': params.sleep_quality,
                        'physical_activity': params.physical_activity,
                        'anxiety': params.anxiety
                    },
                    'notes': params.notes or ''
                }
            })
        else:
            return jsonify({'success': False, 'message': 'No parameters for this date'})

    except Exception as e:
        logger.error(f"Get user parameters error: {str(e)}")
        return jsonify({'error': 'Failed to get parameters'}), 500


@app.route('/api/follow-requests', methods=['POST'])
@login_required
def create_follow_request():
    try:
        data = request.get_json()
        requester_id = session.get('user_id')
        target_id = data.get('target_id')

        if not target_id or requester_id == target_id:
            return jsonify({'error': 'Invalid target'}), 400

        existing = FollowRequest.query.filter_by(
            requester_id=requester_id,
            target_id=target_id
        ).first()

        if existing:
            if existing.status == 'pending':
                return jsonify({'error': 'Request already pending'}), 400
            existing.status = 'pending'
            existing.created_at = datetime.utcnow()
        else:
            existing = FollowRequest(
                requester_id=requester_id,
                target_id=target_id
            )
            db.session.add(existing)

        db.session.commit()

        # Create alert
        requester = db.session.get(User, requester_id)
        requester_username = requester.username if requester else "Someone"

        # PJ6001: Use create_notification_with_email for invite notifications (not wellness alerts)
        alert = create_notification_with_email(
            user_id=target_id,
            title="invite.alert_title",
            content=f"{requester_username}|invite.alert_content",
            alert_type='follow_request',
            source_user_id=requester_id,
            alert_category='follow'
        )
        db.session.commit()

        return jsonify({'message': 'Follow request sent'}), 200

    except Exception as e:
        logger.error(f"Error: {str(e)}")
        db.session.rollback()
        return jsonify({'error': 'Failed'}), 500


@app.route('/api/follow-requests/received', methods=['GET'])
@login_required
def get_received_follow_requests():
    user_id = session.get('user_id')
    requests = FollowRequest.query.filter_by(
        target_id=user_id,
        status='pending'
    ).all()

    return jsonify({
        'requests': [{
            'id': req.id,
            'requester_id': req.requester_id,
            'requester_name': req.requester.username,
            'created_at': req.created_at.isoformat()
        } for req in requests]
    })


@app.route('/api/follow-requests/<int:request_id>/respond', methods=['POST'])
@login_required
def respond_to_follow_request(request_id):
    try:
        data = request.get_json()
        user_id = session.get('user_id')
        action = data.get('action')
        privacy_level = data.get('privacy_level', 'public')

        follow_request = FollowRequest.query.get(request_id)

        if not follow_request or follow_request.target_id != user_id:
            return jsonify({'error': 'Not found'}), 404

        if follow_request.status != 'pending':
            return jsonify({'error': 'Already processed'}), 400

        if action == 'accept':
            follow_request.status = 'accepted'
            follow_request.privacy_level = privacy_level
            follow_request.responded_at = datetime.utcnow()

            follow = Follow(
                follower_id=follow_request.requester_id,
                followed_id=follow_request.target_id
            )
            db.session.add(follow)

        elif action == 'reject':
            follow_request.status = 'rejected'
            follow_request.responded_at = datetime.utcnow()

        db.session.commit()
        return jsonify({'message': f'Request {action}ed'}), 200

    except Exception as e:
        logger.error(f"Error: {str(e)}")
        db.session.rollback()
        return jsonify({'error': 'Failed'}), 500


@app.route('/api/triggers', methods=['GET'])
@login_required
def get_triggers():
    user_id = session.get('user_id')
    triggers = ParameterTrigger.query.filter_by(
        watcher_id=user_id,
        is_active=True
    ).all()

    return jsonify({
        'triggers': [{
            'id': t.id,
            'watched_user': t.watched.username,
            'parameter': t.parameter_name,
            'condition': t.trigger_condition,
            'value': t.trigger_value,
            'consecutive_days': t.consecutive_days
        } for t in triggers]
    })


@app.route('/api/triggers', methods=['POST'])
@login_required
def create_trigger():
    try:
        data = request.get_json()
        trigger = ParameterTrigger(
            watcher_id=session.get('user_id'),
            watched_id=data.get('watched_id'),
            parameter_name=data.get('parameter_name'),
            trigger_condition=data.get('condition'),
            trigger_value=data.get('value'),
            consecutive_days=data.get('consecutive_days')
        )
        db.session.add(trigger)
        db.session.commit()
        return jsonify({'message': 'Trigger created'}), 201
    except Exception as e:
        logger.error(f"Error: {str(e)}")
        db.session.rollback()
        return jsonify({'error': 'Failed'}), 500


@app.route('/api/triggers/<int:trigger_id>', methods=['DELETE'])
@login_required
def delete_trigger(trigger_id):
    trigger = ParameterTrigger.query.get(trigger_id)
    if trigger and trigger.watcher_id == session.get('user_id'):
        trigger.is_active = False
        db.session.commit()
        return jsonify({'message': 'Deleted'}), 200
    return jsonify({'error': 'Not found'}), 404


@app.route('/api/users/shareable-link', methods=['GET'])
@login_required
def get_shareable_link():
    user = User.query.get(session.get('user_id'))

    if not user.shareable_link_token:
        import uuid
        user.shareable_link_token = str(uuid.uuid4())
        db.session.commit()

    base_url = request.host_url.rstrip('/')
    shareable_link = f"{base_url}/profile/{user.shareable_link_token}"

    return jsonify({'link': shareable_link})


@app.route('/api/profile/<token>', methods=['GET'])
def get_profile_by_token(token):
    user = User.query.filter_by(shareable_link_token=token).first()
    if not user:
        return jsonify({'error': 'Not found'}), 404
    return jsonify({'display_name': user.username, 'id': user.id})


# =====================
# ERROR HANDLERS
# =====================

@app.errorhandler(404)
def not_found(error):
    """Handle 404 errors"""
    if request.path.startswith('/api/'):
        return jsonify({'error': 'Not found'}), 404
    return jsonify({'error': 'Page not found'}), 404


@app.errorhandler(500)
def internal_error(error):
    """Handle 500 errors"""
    db.session.rollback()
    logger.error(f"Internal error: {str(error)}")
    return jsonify({'error': 'Internal server error'}), 500


# =====================
# CLI COMMANDS
# =====================

@app.cli.command()
def init_db():
    """Initialize the database"""
    db.create_all()  # Let SQLAlchemy create from models.py

    # Only do migration for existing data
    with app.app_context():
        db_conn = get_db()
        try:
            migrate_parameters_data(db_conn)  # Use your migration function
        except:
            pass  # New installation, no migration needed

    print("Database initialized.")


def migrate_parameters_data(db):
    """Migrate existing text-based parameters to numeric values"""
    try:
        # Migrate mood from text to numeric
        db.execute('''
            UPDATE parameters 
            SET mood = CASE 
                WHEN mood = 'very_bad' OR mood = '1' THEN 1
                WHEN mood = 'bad' OR mood = '2' THEN 2
                WHEN mood IN ('ok', 'neutral', '3', 'moderate') THEN 3
                WHEN mood IN ('good', '4', 'excellent') THEN 4
                WHEN CAST(mood AS INTEGER) BETWEEN 1 AND 4 THEN CAST(mood AS INTEGER)
                ELSE NULL
            END
            WHERE mood IS NOT NULL AND typeof(mood) = 'text'
        ''')

        # Migrate exercise to physical_activity
        db.execute('''
            UPDATE parameters 
            SET physical_activity = CASE 
                WHEN exercise IN ('none', '1', 'no') THEN 1
                WHEN exercise IN ('light', '2', 'mild') THEN 2
                WHEN exercise IN ('moderate', '3', 'medium') THEN 3
                WHEN exercise IN ('intense', 'high', '4', 'heavy') THEN 4
                WHEN CAST(exercise AS INTEGER) BETWEEN 1 AND 4 THEN CAST(exercise AS INTEGER)
                ELSE exercise
            END
            WHERE exercise IS NOT NULL
        ''')

        # Migrate anxiety from text to numeric
        db.execute('''
            UPDATE parameters 
            SET anxiety = CASE 
                WHEN anxiety IN ('none', '1', 'no') THEN 1
                WHEN anxiety IN ('low', 'mild', '2') THEN 2
                WHEN anxiety IN ('moderate', '3', 'medium') THEN 3
                WHEN anxiety IN ('high', 'severe', '4') THEN 4
                WHEN CAST(anxiety AS INTEGER) BETWEEN 1 AND 4 THEN CAST(anxiety AS INTEGER)
                ELSE NULL
            END
            WHERE anxiety IS NOT NULL AND typeof(anxiety) = 'text'
        ''')

        # Handle old energy column - ensure it's numeric
        db.execute('''
            UPDATE parameters 
            SET energy = CASE 
                WHEN energy IN ('very_low', '1') THEN 1
                WHEN energy IN ('low', '2') THEN 2
                WHEN energy IN ('moderate', '3', 'medium', 'ok') THEN 3
                WHEN energy IN ('high', '4', 'good') THEN 4
                WHEN CAST(energy AS INTEGER) BETWEEN 1 AND 4 THEN CAST(energy AS INTEGER)
                ELSE NULL
            END
            WHERE energy IS NOT NULL AND typeof(energy) = 'text'
        ''')

        # Convert sleep_hours to sleep_quality if needed
        cursor = db.execute("PRAGMA table_info(parameters)")
        columns = [row[1] for row in cursor.fetchall()]

        if 'sleep_hours' in columns:
            db.execute('''
                UPDATE parameters 
                SET sleep_quality = CASE 
                    WHEN sleep_hours <= 4 THEN 1
                    WHEN sleep_hours > 4 AND sleep_hours <= 6 THEN 2
                    WHEN sleep_hours > 6 AND sleep_hours <= 8 THEN 3
                    WHEN sleep_hours > 8 THEN 4
                    ELSE NULL
                END
                WHERE sleep_hours IS NOT NULL AND sleep_quality IS NULL
            ''')

        db.commit()
        print("Parameters data migration completed successfully")
    except Exception as e:
        print(f"Error during migration: {e}")


@app.cli.command()
def fix_alerts():
    """Fix alerts table schema"""
    fix_alerts_table()
    print("Alerts table fixed.")


# =====================
# MAIN INITIALIZATION
# =====================

# Flag to track if background initialization has completed
_init_complete = False
_init_lock = threading.Lock()

def _background_init():
    """Run heavy initialization in background so health checks pass immediately"""
    global _init_complete
    try:
        init_database()
        with _init_lock:
            _init_complete = True
        logger.info("Background initialization completed successfully")
    except Exception as e:
        logger.error(f"Background initialization error: {e}")
        # Still mark as complete so app doesn't hang
        with _init_lock:
            _init_complete = True

if __name__ == '__main__':
    # Initialize database
    init_database()
    migrate_circle_names()
    # data time

    # Get port from environment
    port = int(os.environ.get('PORT', 5000))

    # Run application
    logger.info(f"Starting Social Social Platform on port {port}")
    logger.info(f"Production mode: {is_production}")

    app.run(
        host='0.0.0.0',
        port=port,
        debug=not is_production
    )
else:
    # For production servers (gunicorn, etc.)
    # Run initialization in background thread so health checks pass immediately
    # This allows Render to mark the service as "live" much faster
    init_thread = threading.Thread(target=_background_init, daemon=True)
    init_thread.start()
    logger.info("Gunicorn started - background initialization running")